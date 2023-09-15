import asyncio
import openpyxl
import os
from docx import Document
from docx.shared import RGBColor
import win32com.client as win32
from playwright.async_api import async_playwright


result = 0
page = None
browser = None
col_n = None
col_s = None
col_g = None
col_l = None
col_a = None



async def 엑셀():
    global col_n,col_s,col_g,col_l,col_a

    desktop_path = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(desktop_path, '역량향상도학생리스트.xlsx')
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = wb.active

    col_n = [sheet[f"A{i}"].value for i in range(2, sheet.max_row + 1)]
    col_s = [sheet[f"B{i}"].value for i in range(2, sheet.max_row + 1)]
    col_g = [sheet[f"C{i}"].value for i in range(2, sheet.max_row + 1)]
    col_l = [sheet[f"D{i}"].value for i in range(2, sheet.max_row + 1)]
    col_a = [sheet[f"E{i}"].value for i in range(2, sheet.max_row + 1)]

    wb.close()
     

async def 시험():
    global page, col_n
    for i in range(len(col_n)):
        await page.goto(col_l[i])
        await page.wait_for_selector('span.NPEfkd.RveJvd.snByac', timeout=10000)
        await page.fill('input[aria-labelledby="i1"]',str(col_n[i]))
        await page.fill('input[aria-labelledby="i5"]',str(col_s[i]))
        if col_g[i] > 6 :
            col_g[i] = col_g[i]-6
        await page.fill('input[aria-labelledby="i9"]',str(col_g[i]))
        col_a_list = [int(char) for char in str(col_a[i])]
        
        for j in range(1, len(col_a_list)):
            col_a_list[j] += 4*j
        elements = await page.query_selector_all('.AB7Lab.Id5V1')

        for k in range(0,15):
            await elements[col_a_list[k]-1].click()
            print(col_a_list[k]-1)
        
        await page.click('span.NPEfkd.RveJvd.snByac')
        



async def 동작():
    global browser, page  # 전역 변수로 사용
    
    await 엑셀()
    
    async with async_playwright() as playwright: # 크로미움 브라우저 열기
        browser = await playwright.chromium.launch(
            headless=False,
        )
        page = await browser.new_page(accept_downloads=True) # 새로운 창이 열리면 page에 저장
        await 시험()
    
asyncio.run(동작())
