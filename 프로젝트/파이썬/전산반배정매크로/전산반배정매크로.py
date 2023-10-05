import asyncio
import re
import os
from playwright.async_api import async_playwright
import openpyxl


강의날짜 = 202310


아이디 = None
반이름 = None
학생이름 = None
page = None
browser = None
new_handle = None

폴더경로 = "os.path.dirname(os.path.abspath(__file__))"                      # 지금 코드 파일이 있는 위치를 저장

async def 로그인():
    global page, new_handle                                                  # 전역 변수로 사용
    await page.goto('https://enozsw-bukgu.enoz.kr/Admin')                    # 로그인 사이트로 이동
    await page.wait_for_load_state()
    await asyncio.sleep(2)
    await page.fill('input[name="tbAdminId"]', 'admin')                      # 아이디 및 비밀번호 입력하고 로그인
    await page.fill('input[name="tbAdminPass"]', 'admin55&&')
    await asyncio.sleep(2)
    await page.press('input[name="tbAdminPass"]', 'Enter')
    await asyncio.sleep(2)
    await page.goto('https://enozsw-bukgu.enoz.kr/Admin/Class/StudyList.asp') # 수강 관리로 이동
    await asyncio.sleep(2)
    await page.select_option('select[name="ddlKeyField"]', value='b.m_id')    # 아이디로 검색으로 바꾸기
    await asyncio.sleep(2)




async def 반배정(대상):
    global page, new_handle                                                                  # 전역 변수로 사용
  
    excel_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '반배정.xlsx') # 이 코드파일이 있는 폴더에 반배정.xlsx 파일을 엶
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    if 대상 == '학생배정':                                           # 대상별 엑셀 최대 길이 값 구함 (헤더 제외)
        column_index = 1  # A열
    elif 대상 == '망령출동':
        column_index = 6  # F열
    elif 대상 == '망령퇴장':
        column_index = 11  # K열
    for row_index in range(2, ws.max_row + 1):
        if not ws.cell(row=row_index, column=column_index).value:
            break
        
    마지막_행 = row_index - 1
    
    for row in ws.iter_rows(min_row=2, max_row=마지막_행, values_only=True): # 대상별 엑셀 최대 길이 값만큼 반복
        if 대상 == '학생배정': # A ~ C
            반이름 = row[0]
            아이디 = row[1]
            학생이름 = row[2]  
        if 대상 == '망령출동': # F ~ H
            반이름 = row[5]
            아이디 = row[6]
            학생이름 = row[7]
        if 대상 == '망령퇴장': # K ~ M
            반이름 = row[10]
            아이디 = row[11]
            학생이름 = row[12]             
               
        await page.fill('input[name="tbKeyWord"].font_blue', 아이디)   # 입력란에 아이디 입력
        await page.press('input[name="tbKeyWord"].font_blue', 'Enter')  # 엔터
        await asyncio.sleep(1)
        await page.click(f'a[href*="{강의날짜}"].button_red_small')     # 강의 날짜에 맞는 것으로 선택 몇 년 몇 월인지 적으면 됨
        await asyncio.sleep(2)
        new_handle = None # 새로 열린 창 핸들 얻기
        while not new_handle:
            for handle in browser.contexts[0].pages:
                if handle != page:
                    new_handle = handle
                    break
            await asyncio.sleep(2)
            
            await new_handle.evaluate('''(name) => { 
            const selectElement = document.querySelector('select[name="ddlTargetGroupNo"]');
            for (const option of selectElement.options) {
                if (option.textContent.includes(name)) {
                    option.selected = true;
                    const event = new Event('change', { bubbles: true });
                    selectElement.dispatchEvent(event);
                    break;
                }
            }
        }''', 반이름)                                                                                   # 반이름에 해당되는 항목 선택
        
        new_handle.on('dialog', lambda dialog: asyncio.ensure_future(handle_alert(new_handle, dialog))) # Alert 팝업창 핸들링
        await new_handle.wait_for_selector('a.button_yellow.bold:has-text("수강 변경"), a.button_red.bold:has-text("수강 인원이 모두 찼습니다. (변경불가 => 가능)")', timeout=10000) # 쿠키 허용 버튼 클릭
        combined_selector = 'a.button_yellow.bold:has-text("수강 변경"), a.button_red.bold:has-text("수강 인원이 모두 찼습니다. (변경불가 => 가능)")'
        await new_handle.locator(combined_selector).click()
        await asyncio.sleep(1)
        await new_handle.close()                                                                        # 변경 실패하면 창이 안 닫히므로 수동으로 닫음
        await page.bring_to_front()                                                                     # 만약 창이 안 닫히면 무시하고 다음 동작하도록 함
        await asyncio.sleep(1)
        print(학생이름,"(",아이디,")",반이름,"로 배정 완료")                                               # 정상적으로 배정됐는지 터미널로 확인
        
    wb.close() # 엑셀 닫음


async def handle_alert(page, dialog):
    await dialog.accept()                                                                                # Alert 팝업 창이 열리면 확인을 누름


async def 동작():
    global browser, page                                        # 전역 변수로 사용
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(
            headless=False,
            args=['--disable-popup-blocking']
        )
        page = await browser.new_page(accept_downloads=True)
        await 로그인()
        await 반배정('망령출동')
        await 반배정('학생배정')
        await 반배정('망령퇴장')
        print("반배정 완료")
        await browser.close()                                   # 브라우저 종료

asyncio.run(동작())

