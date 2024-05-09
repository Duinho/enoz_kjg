import asyncio
import os
from playwright.async_api import async_playwright
import openpyxl

강의날짜 = 202310
아이디 = None
반이름 = None
학생이름 = None
page = None
browser = None
new_handle = None

폴더경로 = "os.path.dirname(os.path.abspath(__file__))"                      # 지금 코드 파일이 있는 위치를 저장
handled_dialogs = set()

async def 수강신청():
    global page, new_handle, 아이디, handled_dialogs
    
    def handle_dialog(dialog):
        if dialog not in handled_dialogs:
            asyncio.ensure_future(handle_alert(new_handle, dialog))
            handled_dialogs.add(dialog)
    page.on('dialog', handle_dialog)                                     
    await page.goto('https://www.bgswcoding.co.kr/Account/Login')    
    await page.wait_for_selector('input[name="tbID"]', timeout=10000)    
    await page.fill('input[name="tbID"]', 아이디)                     
    await page.fill('input[name="tbPass"]', '1234')
    await page.press('input[name="tbPass"]', 'Enter')
    await page.goto('https://www.bgswcoding.co.kr/Course/CourseList')     
    await page.wait_for_selector('span:text("수강신청")', timeout=10000)    
    await page.click('label.cc-cc.check[for="1_7"]')
    await page.click('span:text("수강신청")')
    await page.wait_for_selector('input#btn_poll', timeout=10000) 
    await page.click('label:has-text("매우 그렇다")[for="q1_5"]')
    await page.click('label:has-text("매우 그렇다")[for="q2_5"]')
    await page.click('label:has-text("매우 그렇다")[for="q3_5"]')
    await page.click('label:has-text("매우 그렇다")[for="q4_5"]')
    await page.click('label:has-text("매우 그렇다")[for="q5_5"]')
    await page.click('input#btn_poll')
    await page.click('input#cbAddress')
    checkboxes = await page.query_selector_all('input[name="cbDevice"][type="checkbox"]')
    for checkbox in checkboxes:
        await checkbox.click()
    await asyncio.sleep(30)
    await page.click('a:has(span:text("신청"))')
    await page.wait_for_selector('a:text("LOGOUT")', timeout=10000)
    await page.click('a:text("LOGOUT")')      
    await page.wait_for_selector('a:text("온라인 SW 코딩 교실")', timeout=10000)    


async def 로그인():
    global page, new_handle  
    await page.goto('https://enozsw-bukgu.enoz.kr/Admin')                    # 로그인 사이트로 이동
    await page.wait_for_load_state()
    await asyncio.sleep(2)
    await page.fill('input[name="tbAdminId"]', 'admin')                      # 아이디 및 비밀번호 입력하고 로그인
    await page.fill('input[name="tbAdminPass"]', 'admin55&&')
    await asyncio.sleep(2)
    await page.press('input[name="tbAdminPass"]', 'Enter')
    await asyncio.sleep(2)
    await page.goto('https://enozsw-bukgu.enoz.kr/Admin/Class/StudyList.asp') # 수강 관리로 이동
    await asyncio.sleep(2)


async def 수강취소():
    global page, new_handle, 아이디, handled_dialogs,학생이름
    try:
        def handle_dialog(dialog):
            if dialog not in handled_dialogs:
                asyncio.ensure_future(handle_alert(new_handle, dialog))
                handled_dialogs.add(dialog)# 전역 변수로 사용
        page.on('dialog', handle_dialog)
        await page.select_option('select[name="ddlTargetDate"]', value=str(강의날짜))    # 강의날짜로 검색으로 바꾸기
        await page.select_option('select[name="ddlKeyField"]', value='b.m_id')          # 아이디로 검색으로 바꾸기
        await asyncio.sleep(1)
        await page.fill('input[name="tbKeyWord"].font_blue', 아이디)                    # 입력란에 아이디 입력
        await page.press('input[name="tbKeyWord"].font_blue', 'Enter')                  # 엔터
        await asyncio.sleep(1)
        await page.click('a.button_red_small:has-text("삭제")')
        await asyncio.sleep(1)
    except Exception as e:
        print(f"{학생이름} 처리 과정에서 오류 발생: {e}")                                 # 디버깅을 위한 오류 위치 출력    
    


async def 엑셀(모드):
    global page, new_handle,아이디,학생이름                                             # 전역 변수로 사용
    
    # 엑셀 파일 열기
    excel_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '반배정.xlsx') # 이 코드파일이 있는 폴더에 반배정.xlsx 파일을 엶
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active                                                                            # 대상별 엑셀 최대 길이 값 구함 (헤더 제외)

    column_index = 6  # F열
    for row_index in range(2, ws.max_row + 1):
        if not ws.cell(row=row_index, column=column_index).value:
            break
    if 모드 == '취소':
        await 로그인()   
    마지막_행 = row_index - 1
    
    for row in ws.iter_rows(min_row=2, max_row=마지막_행, values_only=True): # 대상별 엑셀 최대 길이 값만큼 반복
        아이디 = row[6]
        학생이름 = row[7] 
        if 모드 == '신청':
            await 수강신청()    
            print(학생이름, "수강신청 완료")                                               # 정상적으로 배정됐는지 터미널로 확인
        if 모드 == '취소':
            await 수강취소()
            print(학생이름, "수강취소 완료")
        
    wb.close() # 엑셀 닫음


async def handle_alert(page, dialog):
    try:
        await dialog.accept()
    except Exception as e:
        print(f"Error handling dialog: {e}")

async def 동작():
    global browser, page                                        # 전역 변수로 사용
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(
            headless=False,
            args=['--disable-popup-blocking']
        )
        page = await browser.new_page(accept_downloads=True)
        await 엑셀('취소')
        print("처리 완료")
        await browser.close()                                   # 브라우저 종료

asyncio.run(동작())

