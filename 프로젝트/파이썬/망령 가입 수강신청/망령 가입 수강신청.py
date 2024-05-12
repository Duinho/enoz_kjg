import os
import openpyxl
from playwright.sync_api import sync_playwright
import random
import time
# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '망령 가입 수강신청.xlsx')  # 엑셀 파일 경로
사용자_데이터_디렉토리 = os.path.join(폴더경로, 'UserData')


클릭 = 0
학교_리스트 = []
아파트_리스트 = []
동_리스트 = []
아이디1_리스트 = []
아이디2_리스트 = []
아이디3_리스트 = []
남아_리스트 = []
여아_리스트 = []
부_리스트 = []
모_리스트 = []
성 = [
    ("김", 0.21782), ("이", 0.14882), ("박", 0.08536), ("최", 0.04754), ("정", 0.04402),
    ("강", 0.02355), ("조", 0.02161), ("윤", 0.02077), ("장", 0.02017), ("임", 0.01674),
    ("한", 0.01560), ("오", 0.01548), ("서", 0.01524), ("신", 0.01526), ("권", 0.01427),
    ("황", 0.01412), ("안", 0.01393), ("송", 0.01389), ("류", 0.01300), ("전", 0.01098),
    ("홍", 0.01137), ("고", 0.00957), ("문", 0.00938), ("양", 0.00884), ("손", 0.00919),
    ("배", 0.00812), ("백", 0.00773), ("허", 0.00660), ("유", 0.00566), ("남", 0.00559),
    ("심", 0.00551), ("노", 0.00498), ("정", 0.00442), ("하", 0.00463), ("곽", 0.00410),
    ("성", 0.00405), ("차", 0.00396), ("주", 0.00389), ("우", 0.00390), ("구", 0.00392),
    ("신", 0.00369), ("임", 0.00379), ("라", 0.00377), ("전", 0.00386), ("민", 0.00350),
    ("유", 0.00381), ("진", 0.00315), ("지", 0.00311), ("엄", 0.00291), ("채", 0.00254),
    ("원", 0.00262), ("천", 0.00239), ("방", 0.00191), ("공", 0.00183), ("강", 0.00201),
    ("현", 0.00180), ("함", 0.00164), ("변", 0.00164), ("염", 0.00139), ("양", 0.00181),
    ("변", 0.00121), ("여", 0.00123), ("추", 0.00121), ("노", 0.00134), ("도", 0.00115),
    ("소", 0.00098), ("신", 0.00103), ("석", 0.00100), ("선", 0.00085), ("설", 0.00086),
    ("마", 0.00078), ("길", 0.00075), ("주", 0.00081), ("연", 0.00068), ("방", 0.00068),
    ("위", 0.00062), ("표", 0.00062), ("명", 0.00059), ("기", 0.00056), ("반", 0.00054),
    ("왕", 0.00051), ("금", 0.00051), ("옥", 0.00051), ("육", 0.00047), ("인", 0.00045),
    ("맹", 0.00044), ("제", 0.00044), ("모", 0.00042), ("장", 0.00041), ("남", 0.00042),
    ("탁", 0.00043), ("국", 0.00039), ("여", 0.00039), ("진", 0.00042), ("어", 0.00038),
    ("은", 0.00034), ("편", 0.00033), ("구", 0.00029), ("용", 0.00016), ("남궁", 0.00015)
]
성씨, weights = zip(*성)

비밀번호 = 'enoz7223'
def 초기화():
    global wb, ws,회원가입사이트,수강신청사이트,학교_리스트,아파트_리스트,동_리스트,아이디1_리스트,아이디2_리스트,아이디3_리스트,남아_리스트,여아_리스트,부_리스트,모_리스트,반복횟수,랜덤최소,랜덤최대
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    학교마지막행 = 31
    아이디마지막행 = 92
    자녀마지막행 = 501
    부모마지막행 = 78
    회원가입사이트 = ws.cell(row=1, column=20).value
    수강신청사이트 = ws.cell(row=2, column=20).value
    반복횟수 = ws.cell(row=3, column=20).value
    랜덤최소 = ws.cell(row=4, column=20).value
    랜덤최대 = ws.cell(row=5, column=20).value
    for row in ws.iter_rows(min_row=2, max_row=학교마지막행, values_only=True):
        학교_리스트.append(row[0])
        아파트_리스트.append(row[1])
        동_리스트.append(row[2])        
    for row in ws.iter_rows(min_row=2, max_row=아이디마지막행, values_only=True):
        아이디1_리스트.append(row[5])
        아이디2_리스트.append(row[6])
        아이디3_리스트.append(row[7])
    for row in ws.iter_rows(min_row=2, max_row=자녀마지막행, values_only=True):
        남아_리스트.append(row[10])
        여아_리스트.append(row[11])
    for row in ws.iter_rows(min_row=2, max_row=부모마지막행, values_only=True):
        부_리스트.append(row[14])
        모_리스트.append(row[15])
    
def 회원가입(page):
    # 로그인 사이트 정보를 읽어옴
    global 강의날짜,클릭


    # 로그인 프로세스
    page.goto(회원가입사이트)
    logout_link = page.locator('a[href="/Account/LoginProc"]:has-text("LOGOUT")')
    if logout_link.count() > 0:
        logout_link.click()
        time.sleep(1)
        page.goto(회원가입사이트)
    else:
        print("Logout link not found.")
    page.wait_for_load_state('networkidle')
    page.locator("#cbAgree1").click()
    page.locator("#cbAgree2").click()
    page.click(".btn_type4.c1:has-text('다음')")
    page.wait_for_selector("input[name='tbMemName'].type2.w1.IsKor") 
    이름 = random.choices(성씨, weights=weights, k=1)[0]
    아이성 = 이름 
    성별 = random.randint(0, 6)
    if 성별 < 5 :
        page.select_option('select[name="ddlSex"]', value='남자')
        이름 = f"{이름}{남아_리스트[random.randint(0, 499)]}"
    else :
        page.select_option('select[name="ddlSex"]', value='여자') 
        이름 = f"{이름}{여아_리스트[random.randint(0, 499)]}"  
    page.fill('input[name="tbMemName"]', 이름)
    아이디 = 0
    랜덤 = random.randint(0, 2)
    if 랜덤 == 0 :
        아이디 = 아이디1_리스트[random.randint(0, 90)]
    if 랜덤 == 1 :
        아이디 = 아이디2_리스트[random.randint(0, 90)]
    if 랜덤 == 2 :
        아이디 = 아이디3_리스트[random.randint(0, 90)]
    랜덤 = random.randint(0, 11)
    if 랜덤 < 3 :
        아이디 = f"{아이디}{아이디1_리스트[random.randint(0, 90)]}"
    elif 랜덤 < 6 :
        아이디 = f"{아이디}{아이디2_리스트[random.randint(0, 90)]}"
    elif 랜덤 < 9 :
        아이디 = f"{아이디}{아이디3_리스트[random.randint(0, 90)]}"
    else :
        아이디 += str(random.randint(0, 9))
    랜덤 = random.randint(0, 1)
    if 랜덤 == 0 :
        아이디 += str(random.randint(0, 9))
    if len(아이디) > 15:
        아이디 = 아이디[:15]
    page.fill('input[name="tbMemID"]', 아이디)
    page.locator("text=중복확인").click()
    page.fill('input[name="tbMemPass"]', 비밀번호)
    page.fill('input[name="tbMemPass2"]', 비밀번호)
    page.fill('input[name="tbMobile1"]', '010')
    중간번호 = f"{str(random.randint(2, 9))}{str(random.randint(0, 9))}{str(random.randint(0, 9))}{str(random.randint(0, 9))}"
    page.fill('input[name="tbMobile2"]', 중간번호)
    끝번호 = f"{str(random.randint(0, 9))}{str(random.randint(0, 9))}{str(random.randint(0, 9))}{str(random.randint(0, 9))}"
    page.fill('input[name="tbMobile3"]', 끝번호)
    전화번호 = f"{'010'}{중간번호}{끝번호}"
    나이 = random.randint(11, 13)
    출생년도 = 2025 - 나이
    page.select_option('select[name="ddlBirthDay1"]', value=str(출생년도))
    month = f"{random.randint(1, 12):02}"
    page.select_option('select[name="ddlBirthDay2"]', value=month)
    day = f"{random.randint(1, 28):02}"
    page.select_option('select[name="ddlBirthDay3"]', value=day)
    if 성별 < 5 :
        page.select_option('select[name="ddlSex"]', value='남자')
    else :
        page.select_option('select[name="ddlSex"]', value='여자')
    학교선택 = random.randint(0,29)
    학교 = 학교_리스트[학교선택]
    page.fill('input[name="tbAcaName5"]', 학교)
    page.fill('input[name="tbGrade5"]', str(나이 - 7))
    page.fill('input[name="tbClass5"]', str(random.randint(1, 5)))
    page.click("span:text('우편번호 찾기')")
    time.sleep(2)
    page.keyboard.type(아파트_리스트[학교선택])
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    for _ in range(3):
        page.keyboard.press('Tab')
    page.keyboard.press('Enter')
    time.sleep(2)
    호 = f"{str(random.randint(1, 5))}{'0'}{str(random.randint(1, 9))}"
    주소 = f"{동_리스트[학교선택]}{' '}{호}{'호'}"
    page.fill('input[name="tbAddr2"]', 주소)
    부모성별 = random.randint(0, 2)
    if 부모성별 < 1 :
        부모이름 = f"{아이성}{부_리스트[random.randint(0, 76)]}"
    else :
        부모이름 = f"{random.choices(성씨, weights=weights, k=1)[0]}{모_리스트[random.randint(0, 76)]}"
    page.fill('input[name="tbPName"]', 부모이름)
    page.fill('input[name="tbPMobile1"]', '010')
    page.fill('input[name="tbPMobile2"]', 중간번호)
    page.fill('input[name="tbPMobile3"]', 끝번호)
    포털 = ["@gmail.com","@naver.com","@daum.net","@nate.com"]
    page.fill('input[name="tbMemEmail"]', f"{아이디}{random.choice(포털)}")
    page.click("#rdRoute1")
    time.sleep(2)
    page.wait_for_selector("span:text('회원가입')", state="visible")
    page.once('dialog', handle_alert)
    page.click("span:text('회원가입')")
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    엑셀_초기화_및_데이터_저장(이름, 아이디, 나이, 전화번호, 학교)
    page.goto(수강신청사이트)
    page.locator('button.btn_week[data-week="2,4"]').click() 
    time.sleep(2)
    page.locator('label.cc-cc.check[for="1_39"]').click()
    page.once('dialog', handle_alert)
    page.locator('a.btn_type5.mb30:has-text("수강신청")').click()
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    page.wait_for_selector('input#btn_poll', timeout=10000) 
    page.locator('input#q1_1').click()
    page.locator('input#q2_1').click()
    page.locator('input#q3_1').click()
    page.locator('input#q4_1').click()
    page.locator('input#q5_1').click()
    page.locator('input#q6_1').click()
    page.locator('input#q7_1').click()
    page.click('input#btn_poll')
    time.sleep(2)
    page.click('input#cbAddress')
    time.sleep(2)
    page.click('input#cbPc')
    time.sleep(2)
    page.once('dialog', handle_alert)
    page.click('a:has(span:text("신청"))')
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    page.keyboard.press('Enter')
    print(f"{이름}{' 신청 완료'}")
    time.sleep(random.randint(랜덤최소, 랜덤최대))


def handle_alert(dialog):
    global 클릭
    time.sleep(3)
    클릭 = 1
    dialog.accept()

def 엑셀_초기화_및_데이터_저장(이름, 아이디, 나이, 전화번호, 학교):
    파일_경로 = os.path.join(폴더경로, '망령 가입 현황.xlsx')
    if not os.path.exists(파일_경로):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "가입 정보"
        ws['A1'] = "이름"
        ws['B1'] = "아이디"
        ws['C1'] = "전화번호"
        ws['D1'] = "주소"
    else:
        wb = openpyxl.load_workbook(파일_경로)
        ws = wb.active
    
    # 새 데이터 행 추가
    new_row = ws.max_row + 1
    ws[f'A{new_row}'] = 이름
    ws[f'B{new_row}'] = 아이디
    ws[f'C{new_row}'] = 나이
    ws[f'D{new_row}'] = 전화번호
    ws[f'D{new_row}'] = 학교
    wb.save(파일_경로)
    wb.close()


def 동작():
    if not os.path.exists(사용자_데이터_디렉토리):
        os.makedirs(사용자_데이터_디렉토리)
    with sync_playwright() as playwright:
        # 파이어폭스를 사용하여 launch_persistent_context 호출
        context = playwright.firefox.launch_persistent_context(
            user_data_dir=사용자_데이터_디렉토리,
            headless=False,
            args=['--disable-popup-blocking']
        )
        page = context.new_page()
        초기화()
        for _ in range(반복횟수):
            회원가입(page)
        time.sleep(5)
        context.close()


동작()
