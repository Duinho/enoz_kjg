import os
import openpyxl
import time
from playwright.sync_api import sync_playwright
from googleapiclient.discovery import build
import re
from datetime import datetime, timedelta
import pytz

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '영상보내기.xlsx')  # 엑셀 파일 경로

def find_folder_id(드라이브, folder_name, parent_folder_id=None):
    query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder'"
    if parent_folder_id:
        query += f" and '{parent_folder_id}' in parents"
    results = 드라이브.files().list(q=query, fields="files(id, name, webViewLink)").execute()
    items = results.get('files', [])
    if not items:
        raise ValueError(f"No folder named '{folder_name}' found.")
    return items[0]['id'], items[0].get('webViewLink')

def get_subfolders(드라이브, parent_folder_id):
    query = f"'{parent_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
    results = 드라이브.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', [])

def format_date(date_str):
    match = re.findall(r'\d+', date_str)
    if not match or len(match) < 2:
        raise ValueError("Invalid date format")
    return ''.join(match[-2:])

def extract_spreadsheet_id(url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match:
        raise ValueError("Invalid Google Sheets URL")
    return match.group(1)

def 드라이브링크복사():
    global 영상링크, 문자박스내용
    드라이브 = build('drive', 'v3', developerKey=api키)
    FOLDER_ID = re.search(r"/folders/([a-zA-Z0-9-_]+)", 드라이브링크).group(1)
    날짜가공 = format_date(날짜)
    반이름 = f' {a_value}반'
    try:
        folder_id, _ = find_folder_id(드라이브, 요일, parent_folder_id=FOLDER_ID)
        subfolders = get_subfolders(드라이브, folder_id)
        target_folder_id = None
        for subfolder in subfolders:
            try:
                if format_date(subfolder['name']) == 날짜가공:
                    target_folder_id = subfolder['id']
                    break
            except ValueError:
                continue
        if target_folder_id:
            반_subfolders = get_subfolders(드라이브, target_folder_id)
            for subfolder in 반_subfolders:
                if 반이름 in subfolder['name']:
                    _, 영상링크 = find_folder_id(드라이브, subfolder['name'], parent_folder_id=target_folder_id)
                    break
            if 영상링크:
                문자박스내용 = f"{멘트1}\n영상 링크 : {영상링크}\n{멘트2}"
            else:
                print("No matching 반 folder found.")
        else:
            print("No matching 날짜 folder found.")
            
    except ValueError as e:
        print(e)

def 엑셀():
    global 날짜, api키, 시트링크, 시트이름, 드라이브링크, wb, ws, 요일, 멘트1, 멘트2, 문자박스링크, 문자박스아이디, 문자박스비번
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    날짜 = ws.cell(row=1, column=17).value    
    시트이름 = ws.cell(row=2, column=17).value
    if 시트이름 == '출석부(월/수)':
        요일 = '월수'
    elif 시트이름 == '출석부(화/목)':
        요일 = '화목'
    api키 = ws.cell(row=3, column=17).value
    시트링크 = ws.cell(row=4, column=17).value
    드라이브링크 = ws.cell(row=5, column=17).value
    멘트1 = ws.cell(row=6, column=17).value
    멘트2 = ws.cell(row=7, column=17).value
    문자박스링크 = ws.cell(row=8, column=17).value
    문자박스아이디 = ws.cell(row=9, column=17).value
    문자박스비번 = ws.cell(row=10, column=17).value
    kst = pytz.timezone('Asia/Seoul')
    now = datetime.now(kst)
    yesterday = now - timedelta(1)
    month = yesterday.month  # 월 (앞에 0을 제거한 형태)
    day = yesterday.day  # 일
    weekday_dict = {
        'Monday': '월',
        'Tuesday': '화',
        'Wednesday': '수',
        'Thursday': '목',
        'Friday': '금',
        'Saturday': '토',
        'Sunday': '일'
    }
    weekday_eng = yesterday.strftime('%A')
    weekday_kor = weekday_dict[weekday_eng]
    formatted_date = f"{month}/{day}"
    어제날짜 = f"{formatted_date}({weekday_kor})"
    if 날짜 == '자동' :
        날짜 = 어제날짜
        if weekday_kor == '월' or  weekday_kor == '수' :
            요일 = '월수'
        elif weekday_kor == '화' or  weekday_kor == '목' :
            요일 = '화목'

def 날짜검색(시트데이터, search_value):
    for col_index, value in enumerate(시트데이터[1], start=1):
        if value == search_value:
            return col_index
    return None  

def 영상검색(시트데이터, column_index, search_value):
    rows = []
    for row_index, row in enumerate(시트데이터[1:], start=2):
        if len(row) >= column_index and row[column_index - 1] == search_value:
            rows.append(row_index)
    return rows

def 특정열값검색(시트데이터, rows, page):
    global a_value, j_value, l_value, e_value
    for row in rows:
        # A 열에서 최초로 나오는 값을 찾기 위한 로직 추가
        a_value = None
        current_row = row
        while current_row > 0:
            if len(시트데이터[current_row - 1]) > 0:
                a_value = 시트데이터[current_row - 1][0]
                if a_value:
                    break
            current_row -= 1

        # E 열, J 열, L 열의 값을 찾기
        e_value = 시트데이터[row - 1][4] if len(시트데이터[row - 1]) > 4 else None
        j_value = 시트데이터[row - 1][9] if len(시트데이터[row - 1]) > 9 else None
        l_value = 시트데이터[row - 1][11] if len(시트데이터[row - 1]) > 11 else None

        드라이브링크복사()
        page.fill('textarea#recvList', j_value)
        if j_value == l_value:
            page.fill('textarea#recvList', j_value)
        else:
            page.fill('textarea#recvList', f'{j_value}\n{l_value}')
        page.fill('textarea#msg', 문자박스내용)
        page.click('div.num_select')
        time.sleep(0.3)
        frame = page.frame(name='callbackFrame')
        frame.click("a:has-text('01053006552')")
        page.click('a:has(img[src*="send_btn.gif"])')
        time.sleep(0.3)
        page.keyboard.press('Enter')
        time.sleep(0.3)
        page.keyboard.press('Enter')
        print(f'{a_value}반 {e_value} 학생 영상 발송 완료')
        max_row = 1
        for cell in ws['A']:
            if cell.value is not None:
                max_row = cell.row + 1

        ws[f'A{max_row}'] = f'{a_value}반 {e_value}'
        wb.save(excel_file_path)

def 시트확인(page):
    시트아이디 = extract_spreadsheet_id(시트링크)   
    시트 = build('sheets', 'v4', developerKey=api키)
    
    # 시트 데이터 한 번에 가져오기
    range_name = f'{시트이름}'
    result = 시트.spreadsheets().values().get(spreadsheetId=시트아이디, range=range_name).execute()
    시트데이터 = result.get('values', [])
    if not 시트데이터:
        print("No data found.")
        return

    날짜위치 = 날짜검색(시트데이터, 날짜)  # 셀 주소 찾기
    print(날짜위치)
    if 날짜위치:
        영상_행목록 = 영상검색(시트데이터, 날짜위치, '영상')
        특정열값검색(시트데이터, 영상_행목록, page)
    else:
        print(f'날짜 "{날짜}" not found in the second row')

def 로그인(page):
    page.goto(문자박스링크)
    page.fill('input[name="id"]', 문자박스아이디)
    page.fill('input[name="pwd"]', 문자박스비번)
    page.press('input[name="pwd"]', 'Enter')
    time.sleep(1) 
    try:
        page.click("a[onclick*='contentsLayerClose']")  # 닫기 버튼 클릭
    except:
        pass 

def 동작():
    global browser
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()
        엑셀()
        로그인(page)
        시트확인(page)
        browser.close()

동작()
