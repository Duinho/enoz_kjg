import asyncio
import re
import os
import sys
import openpyxl
from playwright.async_api import async_playwright

fm = None
err = None
강의이름 = None
강의날짜 = None
폴더이름 = None
폴더정보 = None
download = None
item = None
page = None
browser = None
new_handle = None

폴더경로 = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(폴더경로, '영상보내기.xlsx')  # 엑셀 파일 경로

def 엑셀():
    global 줌로그인, 줌링크, 줌아이디, 줌비번, 알람받을번호, 문자박스링크, 문자박스아이디, 문자박스비번 
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active    
    줌로그인 = ws.cell(row=1, column=15).value
    줌링크 = ws.cell(row=2, column=15).value
    줌아이디 = ws.cell(row=3, column=15).value
    줌비번 = ws.cell(row=4, column=15).value
    알람받을번호 = ws.cell(row=5, column=15).value
    문자박스링크 = ws.cell(row=7, column=17).value
    문자박스아이디 = ws.cell(row=8, column=17).value
    문자박스비번 = ws.cell(row=9, column=17).value

async def 문자박스():
    await page.goto(문자박스링크)
    await page.fill('input[name="id"]', 문자박스아이디)
    await page.fill('input[name="pwd"]', 문자박스비번)
    await page.press('input[name="pwd"]', 'Enter')
    await asyncio.sleep(3)    
    try:
        await page.click("a[onclick*='contentsLayerClose']")  # 닫기 버튼 클릭
    except:
        pass
    await asyncio.sleep(3)
    await page.fill('textarea#recvList', 알람받을번호)
    await page.fill('textarea#msg', '영상 다운로드 완료')
    await page.click('div.num_select')
    await asyncio.sleep(3)
    frame = page.frame(name='callbackFrame')
    await frame.wait_for_selector("a:has-text('01053006552')", timeout=30000)
    await frame.click("a:has-text('01053006552')")
    await page.click('a:has(img[src*="send_btn.gif"])')
    await asyncio.sleep(3)
    await page.keyboard.press('Enter')
    await asyncio.sleep(3)
    await page.keyboard.press('Enter')
    
async def 로그인():
    global page                                                                       # 전역 변수로 사용
    await page.goto(줌로그인)                                 # Zoom 로그인 페이지로 이동
    await page.wait_for_load_state()
    await page.wait_for_selector('button#onetrust-accept-btn-handler', timeout=10000) # 쿠키 허용 버튼 클릭
    await page.click('button#onetrust-accept-btn-handler')
    await asyncio.sleep(3)
    await page.fill('input#email', 줌아이디)                             # 이메일 및 비밀번호 입력하고 로그인
    await page.fill('input#password', 줌비번)
    await asyncio.sleep(1)
    await page.press('#password', 'Enter')     
    await page.wait_for_selector('span.font-bold:has-text("이앤오즈 주식회사")', timeout=30000)
    await page.goto(줌링크)                   # 기록관리 페이지로 이동
    await asyncio.sleep(1)
    await page.wait_for_selector('//h1[contains(text(), "기록 관리") and contains(@style, "font-size: 24px;")]', timeout=30000)
    await asyncio.sleep(5)

async def 정보가져오기():
    global page, 강의이름, 강의날짜, 폴더정보, fm, err
    try:
        fm = await page.query_selector('span[role="alert"]:has-text("검색과 일치하는 결과를 찾을 수 없습니다")')
        if fm:
            await 문자박스()
            print("모든 영상 다운로드 완료")
            await browser.close()
            sys.exit()

        link_element = await page.query_selector('a.mgb-0.cursor-pointer.topic-actived')
        await link_element.click()
        await asyncio.sleep(1)
        error_count = await page.locator('div.error-wrap').count()
        if error_count > 0 and await page.locator('div.error-wrap').text_content() == "녹화가 너무 짧아 저장되지 않았습니다.":
            await page.goto('https://us02web.zoom.us/recording/management')
            await asyncio.sleep(1)
            first_button = await page.locator('button.grey.zm-button--icon.zm-button--plain.zm-button--icon-ghost.zm-button--icon.zm-button--mini.is-ghost.zm-button.zm-dropdown-selfdefine').first
            await first_button.click()
            await page.locator('div.textRed', has_text="삭제").click()
            await asyncio.sleep(3)
            await page.goto('https://us02web.zoom.us/recording/management')
            link_element = await page.query_selector('a.mgb-0.cursor-pointer.topic-actived')
            await link_element.click()
            await asyncio.sleep(1)
        await page.wait_for_selector('span.meeting-topic', timeout=30000)
        await asyncio.sleep(1)
        topic_element = await page.query_selector('span.meeting-topic')
        if topic_element:
            강의이름 = await topic_element.inner_text()
        date_element = await page.query_selector('span.mgr-md')
        if date_element:
            date_text = await date_element.inner_text()
            pattern = r'(\d+년 \d+월 \d+일)'
            matches = re.search(pattern, date_text)
            강의날짜 = matches.group(1)
            폴더이름 = f"{강의이름}_{강의날짜}"
            if "경산" in 강의이름:
                폴더정보 = os.path.join(폴더경로, '녹화영상', '경산', 강의날짜, 폴더이름)
            elif "포항" in 강의이름:
                폴더정보 = os.path.join(폴더경로, '녹화영상', '포항', 강의날짜, 폴더이름)
            else:
                폴더정보 = os.path.join(폴더경로, '녹화영상', 강의날짜, 폴더이름)
            os.makedirs(폴더정보, exist_ok=True)
    except Exception as e:
        print(f"{강의이름}정보가져오기에서 오류 발생: {e}")
        err = 1

async def 영상다운(item, 이름):
    global browser, 폴더정보, 강의이름, 강의날짜, err
    try:
        await item.click()
        new_handle = None
        while not new_handle:
            for handle in browser.contexts[0].pages:
                if handle != page:
                    new_handle = handle
                    break
                await asyncio.sleep(1)
        await new_handle.wait_for_selector('a.download-btn', timeout=30000)
        dl_element = await new_handle.query_selector('a.download-btn')
        if dl_element:
            async with new_handle.expect_download() as download_info:
                await dl_element.click()
                download = await download_info.value
                await download.path()
            await download.save_as(os.path.join(폴더정보, f"{강의이름}_{이름}_{강의날짜}.mp4"))
            await new_handle.close()
            new_handle = None
            await page.wait_for_selector('span.meeting-topic', timeout=30000)
            await asyncio.sleep(1)
    except Exception as e:
        print(f"{강의이름}의 {이름} 영상다운에서 오류 발생: {e}")
        if new_handle:
            await new_handle.close()
            err = 1

async def 삭제():
    global page, err
    try:
        if err:
            await page.goto('https://us02web.zoom.us/recording/management')
            await asyncio.sleep(1)
            await page.wait_for_selector('//h1[contains(text(), "기록 관리") and contains(@style, "font-size: 24px;")]', timeout=30000)
            print(f"{강의이름}설치 재시작")
            await asyncio.sleep(1)
            err = None
        else:
            delete_element = await page.query_selector('button:nth-child(4)')
            if delete_element:
                await delete_element.click()
                await asyncio.sleep(1)
                delete1_element = await page.query_selector('button.zm-button--primary.zm-button--small.zm-button > span')
                await asyncio.sleep(1)
                if delete1_element:
                    await page.get_by_role("button", name="휴지통으로 이동").click()
                    await asyncio.sleep(1)
                    await page.wait_for_load_state()
            await asyncio.sleep(1)
            await page.reload()
            await page.wait_for_selector('text=휴지통', timeout=30000)
            await asyncio.sleep(2)
    except Exception as e:
        print(f"{강의이름}삭제에서 오류 발생: {e}")
        await page.goto('https://us02web.zoom.us/recording/management')
        await asyncio.sleep(1)
        await page.wait_for_selector('//h1[contains(text(), "기록 관리") and contains(@style, "font-size: 24px;")]', timeout=30000)
        print(f"{강의이름}설치 재시작")
        await asyncio.sleep(3)

async def 동작():
    global browser, page, item, err
    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(
            headless=False,
        )
        page = await browser.new_page(accept_downloads=True)
        엑셀()
        await 로그인()
        while(1):
            await 정보가져오기()
            downloadlist = await page.query_selector_all('div.item_list_header.relative > a')
            tasks = []
            for item in downloadlist:
                inner_text = await item.inner_text()
                if inner_text == "발표자 보기가 포함된 공유 화면":
                    tasks.append(영상다운(item, '발표자 공유화면'))
                if inner_text == "갤러리 보기가 포함된 공유 화면":
                    tasks.append(영상다운(item, '갤러리 공유화면'))
                if inner_text == "갤러리 보기":
                    tasks.append(영상다운(item, '갤러리'))
            await asyncio.gather(*tasks)
            await 삭제()

asyncio.run(동작())
