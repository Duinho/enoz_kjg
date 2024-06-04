import os
import time
import openpyxl
from playwright.sync_api import sync_playwright

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '전산강사배정.xlsx')  # 엑셀 파일 경로
로그인사이트 = None
adminID = None
adminPW = None
회원관리링크 = None
아이디 = None
학년 = None
이름 = None
page = None
browser = None
new_handle = None
wb = None
ws = None
클릭 = 0
# 엑셀 초기화 함수
def 초기화():
    global wb, ws
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

def 로그인():
    global page, new_handle, 로그인사이트, adminID, adminPW, 강의날짜, 월수, 화목, 월수반수, 화목반수, 월수시작번호, 화목시작번호
    로그인사이트 = ws.cell(row=1, column=17).value
    adminID = ws.cell(row=2, column=17).value
    adminPW = ws.cell(row=3, column=17).value
    강의날짜 = ws.cell(row=4, column=17).value
    월수 = ws.cell(row=5, column=17).value
    화목 = ws.cell(row=6, column=17).value
    월수반수 = ws.cell(row=7, column=17).value
    화목반수 = ws.cell(row=8, column=17).value
    월수시작번호 = ws.cell(row=9, column=17).value
    화목시작번호 = ws.cell(row=10, column=17).value
    page.goto(로그인사이트)
    page.wait_for_load_state()
    page.fill('input[name="tbAdminId"]', adminID)
    page.fill('input[name="tbAdminPass"]', adminPW)
    page.press('input[name="tbAdminPass"]', 'Enter')
    time.sleep(1)

from playwright.sync_api import TimeoutError  # TimeoutError를 명시적으로 임포트

def 강사배정():
    global page, new_handle, 아이디, 학년, 이름, 클릭

    for i in range(1, 월수반수+1):
        번호 = str(i).zfill(2)
        사이트 = "https://enozsw.enoz.kr/Admin/Class/popScheduleReg.asp?sOrderNo="
        사이트이름 = f"{사이트}{강의날짜}24_{번호}"
        강사이름 = f"{월수}{번호}"
        page.goto(사이트이름)
        time.sleep(1)
        element = page.query_selector(f"a.button_yellow_small.bold[href=\"javascript:popScheduleRegProc('{강사이름}')\"]")
        if element:
            page.once('dialog', handle_alert)
            page.click(f"a.button_yellow_small.bold[href=\"javascript:popScheduleRegProc('{강사이름}')\"]")
            time.sleep(0.5)
            page.keyboard.press('Enter')
            time.sleep(0.5)
        else:
            print(f"{강사이름}이 이미 배정되었거나 찾을 수 없습니다.")
    for i in range(1, 화목반수+1):
        번호 = str(i).zfill(2)
        사이트 = "https://enozsw.enoz.kr/Admin/Class/popScheduleReg.asp?sOrderNo="
        사이트이름 = f"{사이트}{강의날짜}35_{번호}"
        강사이름 = f"{화목}{번호}"
        page.goto(사이트이름)
        time.sleep(1)
        element = page.query_selector(f"a.button_yellow_small.bold[href=\"javascript:popScheduleRegProc('{강사이름}')\"]")
        if element:
            page.once('dialog', handle_alert)
            page.click(f"a.button_yellow_small.bold[href=\"javascript:popScheduleRegProc('{강사이름}')\"]")
            time.sleep(0.5)
            page.keyboard.press('Enter')
        else:
            print(f"{강사이름}이 이미 배정되었거나 찾을 수 없습니다.")

def 강사링크():
    for i in range(1, 월수반수+1):
        id번호 = str(i+월수시작번호-1)
        번호 = str(i).zfill(2)
        사이트 = f"https://enozsw.enoz.kr/Admin/Teacher/popTeacherReg.asp?teacher_chk=mod&idx={id번호}&id="
        강사이름 = f"{월수}{번호}"
        사이트이름 = f"{사이트}{강사이름}"
        page.goto(사이트이름)
        줌링크 = str(ws.cell(row=i+1, column=1).value)
        page.fill('input[name="video_link"]', 줌링크)
        page.click('input[name="Submit"]')
        time.sleep(0.5)
        page.keyboard.press('Enter')
    for i in range(1, 화목반수+1):
        id번호 = str(i+화목시작번호-1)
        번호 = str(i).zfill(2)
        사이트 = f"https://enozsw.enoz.kr/Admin/Teacher/popTeacherReg.asp?teacher_chk=mod&idx={id번호}&id="
        강사이름 = f"{화목}{번호}"
        사이트이름 = f"{사이트}{강사이름}"
        page.goto(사이트이름)
        줌링크 = str(ws.cell(row=i+1, column=1).value)
        page.fill('input[name="video_link"]', 줌링크)
        page.click('input[name="Submit"]')
        time.sleep(0.5)
        page.keyboard.press('Enter')
   
def handle_alert(dialog):
    global 클릭
    time.sleep(0.3)
    클릭 = 1
    dialog.accept()

def 동작():
    global browser, page
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=False, channel="chrome",args=['--disable-popup-blocking'])
    page = browser.new_page(accept_downloads=True)
    초기화()
    로그인()
    강사배정()
    print("강사배정 완료")
    강사링크()
    browser.close()

if __name__ == "__main__":
    동작()
