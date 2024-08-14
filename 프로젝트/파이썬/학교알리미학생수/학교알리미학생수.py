import os
import openpyxl
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
import time

폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 폴더경로를 코드가 있는 디렉토리로 저장
excel_file_path = os.path.join(폴더경로, '학교알리미학생수.xlsx')  # 엑셀 파일 경로

def 엑셀(): # 엑셀을 열어 정보를 가져오는 함수
    global 학교알리미주소, 학교, 시도, 시군구, 학년
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb['정보']
    학교알리미주소 = ws.cell(row=1, column=17).value
    학교 = ws.cell(row=2, column=17).value
    시도 = ws.cell(row=3, column=17).value
    시군구 = ws.cell(row=4, column=17).value
    if "초등학교" in 학교: 
        학년 = ['1', '2', '3', '4', '5', '6']
    else:
        학년 = ['1', '2', '3']
    wb.close()
    return ws

def 헤더추가(ws): # 엑셀에 제일 위에 헤더를 미리 추가하는 함수
    headers = ['학교명', '1학년', '2학년', '3학년', '4학년', '5학년', '6학년']
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

def 학교받아오기(page): # 내가 지정한 학교 종류와 시,구에 있는 학교의 명단을 리스트에 저장
    page.goto(학교알리미주소)
    radio_buttons = page.query_selector_all('input[name="level1"]') # level1에 학교가 있는지 확인 후 클릭
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 학교 in label.inner_text():
            label.click()
            page.wait_for_load_state('load')   
            break
    radio_buttons = page.query_selector_all('input[name="level2"]') # level2에 시도가 있는지 확인 후 클릭
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 시도 in label.inner_text():
            label.click()
            page.wait_for_load_state('load')   
            break
    radio_buttons = page.query_selector_all('input[name="level3"]') # level3에 시군구가 있는지 확인 후 클릭
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 시군구 in label.inner_text():
            label.click()
            page.wait_for_load_state('load') 
            break 
    학교명 = []
    labels = page.query_selector_all('label[for^="shlCd"]')
    for label in labels:
        학교명.append(label.inner_text())  # shlCd가 포함된 값들을 전부 가져와서 학교명 리스트에 순차적으로 작성
    return 학교명

def 학교정보검색(page, 학교명, 학년, ws): # 가져온 학교 명단을 순서대로 하나씩 값을 가져옴 
    row_index = 2  # 첫 번째 줄은 헤더이므로 두 번째 줄부터 시작
    for 학교 in 학교명:
        label = page.query_selector(f'label:has-text("{학교}")')
        if label:
            label.click()
            page.wait_for_load_state('load')   
            page.click('a[data-tab-id="tabSel"]')
            page.wait_for_load_state('load')   
            page.click('a.accordian_title:has-text("학생현황")')
            page.wait_for_load_state('load')   
            page.click('label[for="hangmok03"]')
            page.wait_for_load_state('load')   
            page.click('#webSearchButton')
            time.sleep(0.1)
            page.wait_for_load_state('load')  # 선택한 학교의 '학년별·학급별 학생수' 필터 선택 후 검색
            values = []
            row_header = page.query_selector('th:has-text("합 계")')
            if row_header:
                parent_row = row_header.evaluate_handle('el => el.parentElement')
                for grade in 학년:
                    index = (int(grade) * 3) + 1  # 마지막에 있는 합 계라는 표에서 학년별로 합계의 셀을 선택
                    selector = f'td:nth-child({index})' # 셀에 적혀있는 값을 가져옴
                    value = parent_row.query_selector(selector).inner_text()
                    values.append(int(value.replace(',', '')))  # 값을 숫자로 변환하여 리스트에 추가
            ws.cell(row=row_index, column=1, value=학교)
            for col_index, value in enumerate(values, start=2):
                ws.cell(row=row_index, column=col_index, value=value) #엑셀에 해당하는 값을 입력
            row_index += 1  # 다음 줄로 이동
            print(f'{학교} 학생수 저장 완료')
            page.click('a.slidedown[href="javascript:research();"]')
            time.sleep(0.1)
            page.wait_for_load_state('load')   

def 동작():
    global browser, page  # 전역 변수로 사용
    with sync_playwright() as p:  # 크로미움 브라우저 열기
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()  # 새로운 창이 열리면 page에 저장
        ws = 엑셀()
        헤더추가(ws)
        학교명 = 학교받아오기(page)
        학교정보검색(page, 학교명, 학년, ws)
        ws.parent.save(excel_file_path)  # 동작이 끝나면 엑셀 파일 저장
        print('모든 학교 저장 완료')
        browser.close()  # 브라우저 닫기

동작()