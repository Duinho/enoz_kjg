import os
import time
import openpyxl
from playwright.sync_api import sync_playwright

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '학년변경.xlsx')  # 엑셀 파일 경로
로그인사이트 = None
adminID = None
adminPW = None
회원관리링크 = None
아이디 = None
학년 = None
이름 = None
page = None
browser = None
new_handle = None
wb = None
ws = None
클릭 = 0
# 엑셀 초기화 함수
def 초기화():
    global wb, ws
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

def 로그인():
    global page, new_handle, 로그인사이트, adminID, adminPW, 회원관리링크
    로그인사이트 = ws.cell(row=1, column=17).value
    adminID = ws.cell(row=2, column=17).value
    adminPW = ws.cell(row=3, column=17).value
    회원관리링크 = ws.cell(row=4, column=17).value
    
    page.goto(로그인사이트)
    page.wait_for_load_state()
    page.fill('input[name="tbAdminId"]', adminID)
    page.fill('input[name="tbAdminPass"]', adminPW)
    page.press('input[name="tbAdminPass"]', 'Enter')
    page.goto(회원관리링크)
    page.select_option('select[name="ddlKeyField"]', value='a.m_id')

from playwright.sync_api import TimeoutError  # TimeoutError를 명시적으로 임포트

def 반배정():
    global page, new_handle, 아이디, 학년, 이름, 클릭, 반복

    마지막_행 = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=마지막_행, values_only=True):
        이름 = row[0]
        아이디 = row[1]
        학년 = row[2] + 1
        if 학년 == 7:
            학년 = 1  # 학년이 7이면 1로 설정
            
        반복 = 1
        while 반복:
            클릭 = 0
            try:
                page.goto(f'https://enozsw.enoz.kr/Admin/Member/popMemReg.asp?sMemID={아이디}')
                time.sleep(0.1)
                page.fill('input[name="tbMemEName"]', 'a')
                if 학년 == 1:
                    page.fill('input[name="tbAcaName"]', '수정필요')
                page.fill('input[name="tbGrade"]', str(학년))

                # 첫 번째 팝업 처리
                page.once('dialog', handle_alert)
                page.click('a.button_yellow:has-text("수정")')
                for i in range(10) :
                    time.sleep(0.05)
                    if(클릭 == 1):
                        반복 = 0 
                        i = 10
                
            except TimeoutError:
                print(f"팝업이 시간 내에 나타나지 않아 페이지를 다시 로드합니다. 아이디: {아이디}")
                continue  # 타임아웃 발생 시 다시 시도
            except Exception as e:
                print(f"기타 예외 발생: {e}, 페이지를 다시 로드합니다.")
                continue  # 기타 예외 발생 시 다시 시도

        print(f"{이름}({아이디})이(가) {학년}학년으로 수정 완료")

def handle_alert(dialog):
    global 클릭
    time.sleep(0.3)
    클릭 = 1
    dialog.accept()

def 동작():
    global browser, page
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=False, channel="chrome",args=['--disable-popup-blocking'])
    page = browser.new_page(accept_downloads=True)
    초기화()
    로그인()
    반배정()
    print("학년변경 완료")
    browser.close()

if __name__ == "__main__":
    동작()
