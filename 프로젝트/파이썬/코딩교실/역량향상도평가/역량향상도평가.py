import asyncio
import openpyxl
import os
from docx import Document
from docx.shared import RGBColor
import win32com.client as win32
from playwright.async_api import async_playwright


result = 0
page = None
browser = None
col_n = None
col_s = None
col_g = None
col_l = None
col_a = None


async def 엑셀():
    global col_n,col_s,col_g,col_l,col_a
    desktop_path = os.path.dirname(os.path.abspath(__file__))             #현재 코드가 있는 디렉토리 위치 저장
    excel_file = os.path.join(desktop_path, '역량향상도학생리스트.xlsx')   # 디렉토리에 있는 역량향상도학생리스트.xlsx 위치 저장
    wb = openpyxl.load_workbook(excel_file, data_only=True)               # 엑셀 열기
    sheet = wb.active                                                     # sheet에 엑셀 값 저장
    col_n = [sheet[f"A{i}"].value for i in range(2, sheet.max_row + 1)]  # 이름 학교 학년 링크 정답 값 받아와 각 리스트에 저장
    col_s = [sheet[f"B{i}"].value for i in range(2, sheet.max_row + 1)]
    col_g = [sheet[f"C{i}"].value for i in range(2, sheet.max_row + 1)]
    col_l = [sheet[f"D{i}"].value for i in range(2, sheet.max_row + 1)]
    col_a = [sheet[f"E{i}"].value for i in range(2, sheet.max_row + 1)]
    wb.close()                                                           # 엑셀 닫음
     

async def 자동화():
    global page, col_n
    for i in range(len(col_n)):                                                   # col_n의 저장된 개수만큼 반복
        await page.goto(col_l[i])                                                 # 해당 학생의 링크로 이동
        await page.wait_for_selector('span.NPEfkd.RveJvd.snByac', timeout=10000)  # 제출 버튼이 나올 때까지 기다림
        await page.fill('input[aria-labelledby="i1"]',str(col_n[i]))              # 이름란에 이름 기입
        await page.fill('input[aria-labelledby="i5"]',str(col_s[i]))              # 학교란에 학교 기입
        if col_g[i] > 6 :                                                         # 만약 6보다 크다면 (중학생이라면)                
            col_g[i] = col_g[i]-6                                                 # 6을 빼고 저장(중1은 7이 저장되어 있으므료 6을 빼서 1로 바꿈)
        await page.fill('input[aria-labelledby="i9"]',str(col_g[i]))              # 학년란에 학년 기입
        col_a_list = [int(char) for char in str(col_a[i])]                        # 정답란에 적힌 숫자를 각각 쪼개어 리스트에 저장
        
        for j in range(1, len(col_a_list)):                                       # col_a_list의 길이만큼 반복
            col_a_list[j] += 4*j                                                  # col_a_list에 0 4 8 12 16로 순차적으로 더함(4개 학목)
        elements = await page.query_selector_all('.AB7Lab.Id5V1')                 # 현재 페이지에 있는 모든 클릭 항목을 찾음

        for k in range(0,15):                                                     # 15번 반복함(총 15문제)
            await elements[col_a_list[k]-1].click()                               # 정답에 해당하는 항목을 클릭함
 
        print(f'{col_n[i]} 제출 완료')                                             # 제출한 학생 확인
        await page.click('span.NPEfkd.RveJvd.snByac')                             # 제출 클릭
        

async def 동작():
    global browser, page                             # 전역 변수로 사용
    await 엑셀()    
    async with async_playwright() as playwright:     # 크로미움 브라우저 열기
        browser = await playwright.chromium.launch(
            headless=False,
        )
        page = await browser.new_page()              # 새로운 창이 열리면 page에 저장
        await 자동화()
    
asyncio.run(동작())
