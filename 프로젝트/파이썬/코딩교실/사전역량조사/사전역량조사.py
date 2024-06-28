import openpyxl
import os
from playwright.sync_api import sync_playwright
import time

result = 0
page = None
browser = None
col_a = None


def 엑셀():
    global wb, col_a, 링크
    desktop_path = os.path.dirname(os.path.abspath(__file__))  # 현재 코드가 있는 디렉토리 위치 저장
    excel_file = os.path.join(desktop_path, '사전역량조사.xlsx')  # 디렉토리에 있는 사전역량조사.xlsx 위치 저장
    wb = openpyxl.load_workbook(excel_file, data_only=True)  # 엑셀 열기
    sheet = wb.active  # sheet에 엑셀 값 저장
    col_a = [sheet[f"A{i}"].value for i in range(2, sheet.max_row + 1)]  # 이름 학교 학년 링크 정답 값 받아와 각 리스트에 저장
    링크 = sheet.cell(row=1, column=17).value
    wb.close()  # 엑셀 닫음
     

def 자동화():
    global page, col_a
    for col_a_value in col_a:  # col_a의 저장된 개수만큼 반복
        page.goto(링크)  # 해당 학생의 링크로 이동
        page.wait_for_selector('span.NPEfkd.RveJvd.snByac', timeout=10000)  # 제출 버튼이 나올 때까지 기다림
        col_a_list = [int(char) for char in str(col_a_value)]  # 정답란에 적힌 숫자를 각각 쪼개어 리스트에 저장
        
        for j in range(1, len(col_a_list)):  # col_a_list의 길이만큼 반복
            col_a_list[j] += 5 * j  # col_a_list에 0 5 10 15 20로 순차적으로 더함(5개 항목)
        elements = page.query_selector_all('.AB7Lab.Id5V1')  # 현재 페이지에 있는 모든 클릭 항목을 찾음

        for k in range(0, 7):  # 7문제
            elements[col_a_list[k] - 1].click()  # 정답에 해당하는 항목을 클릭함
 
        page.click('span.NPEfkd.RveJvd.snByac')  # 제출 클릭
        

def 동작():
    global browser, page  # 전역 변수로 사용
    엑셀()
    with sync_playwright() as playwright:  # 크로미움 브라우저 열기
        browser = playwright.chromium.launch(
            headless=False,
        )
        page = browser.new_page()  # 새로운 창이 열리면 page에 저장
        자동화()
        browser.close()
    
동작()
