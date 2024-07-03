import asyncio
import openpyxl
import os
from docx import Document
from docx.shared import RGBColor
import win32com.client as win32
from playwright.async_api import async_playwright
import numpy as np


page = None
browser = None
context = None


async def 엑셀():
    global page, context                                                                  # 전역 변수로 사용
  
    excel_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '코스페이시스 가입.xlsx') # 이 코드파일이 있는 폴더에 반배정.xlsx 파일을 엶
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

    for row_index in range(2, ws.max_row + 1):
        if not ws.cell(row=row_index, column=4).value:
            break
        
    마지막_행 = row_index - 1
    
    for row in ws.iter_rows(min_row=2, max_row=마지막_행, values_only=True): # 대상별 엑셀 최대 길이 값만큼 반복

        학급코드 = row[3]
        학급코드_리스트 = list(학급코드)

        이름 = row[4]
        아이디 = row[5]
        비밀번호 = row[6]    
        context = await browser.new_context()
        page = await context.new_page()
        
        await page.goto('https://edu.cospaces.io/Auth/Signup/Student')                    # 로그인 사이트로 이동
        await asyncio.sleep(3)
        elements = await page.query_selector_all('.dxTextFieldText')
        for i, element in enumerate(elements):
            if i < len(학급코드_리스트):  # 학급코드_리스트의 길이를 넘지 않도록 확인
                await element.fill(학급코드_리스트[i])
        await element.press('Enter')
        await asyncio.sleep(1)
        await page.wait_for_selector('input[name="password"]', timeout=10000) # 비밀번호란 찾기
        await page.fill('input[name="name"]', 이름)
        await page.fill('input[name="username"]', 아이디)
        await page.fill('input[name="password"]', 비밀번호)
        await page.press('input[name="password"]', 'Enter')
        await asyncio.sleep(1)
        await page.wait_for_selector('input[name="password"]', state="detached")
        await asyncio.sleep(3)                                          
        await context.close()
        print(이름," 가입 완료")                                               # 정상적으로 배정됐는지 터미널로 확인
        
    wb.close() # 엑셀 닫음

                                                    
 
async def 동작():
    global browser, page  # 전역 변수로 사용
    
    async with async_playwright() as playwright: 
        browser = await playwright.chromium.launch(
            headless=False,
        )
        await 엑셀()
        await browser.close()
        
asyncio.run(동작())