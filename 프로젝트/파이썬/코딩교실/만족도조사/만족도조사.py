import openpyxl
import os
from docx import Document
from docx.shared import RGBColor
import win32com.client as win32
from playwright.sync_api import sync_playwright
import numpy as np
import time
result = 0
page = None
browser = None
col_9 = NotImplemented

def 엑셀():
    global col_9, 만족도조사링크
    desktop_path = os.path.dirname(os.path.abspath(__file__))  # 현재 코드가 있는 디렉토리 위치 저장
    excel_file = os.path.join(desktop_path, '만족도조사리스트.xlsx')  # 디렉토리에 있는 만족도조사리스트.xlsx 위치 저장
    wb = openpyxl.load_workbook(excel_file, data_only=True)  # 엑셀 열기
    sheet = wb.active  # sheet에 엑셀 값 저장
    만족도조사링크 = sheet.cell(row=1, column=17).value    
    col_9 = [sheet[f"A{i}"].value for i in range(2, sheet.max_row + 1)]  # col_9에 A2부터 값 순차적으로 저장
    wb.close()  # 엑셀 닫음

def 자동화(page):
    global col_9
    for i in range(len(col_9)):  # col_9에 저장된 개수만큼 반복
        page.goto(만족도조사링크)  # 만족도조사 링크로 이동
        time.sleep(1)
        page.click('span.NPEfkd.RveJvd.snByac')
        time.sleep(1)
        page.wait_for_selector('span.NPEfkd.RveJvd.snByac', timeout=10000)  # 제출 버튼이 나올 때까지 대기
        col_a = np.random.choice([5, 4], size=8, p=[0.95, 0.05])  # 95%로 5점과 5%로 4점을 col_a에 8개 저장
        for j in range(1, len(col_a)):
            col_a[j] += 5 * j  # col_a에 0 5 10 15로 순차적으로 더함(5개 항목)
        time.sleep(1)
        elements = page.query_selector_all('.AB7Lab.Id5V1')  # 현재 페이지에 있는 모든 클릭 항목들을 찾음
        
        for k in range(0, 8):  # 8번 반복함 (총 8문제)
            elements[col_a[k] - 1].click()  # 클릭 항목을 순차적으로 체크함

        page.fill('textarea[aria-labelledby="i33"]', str(col_9[i]))  # 마지막 주관식에 엑셀에 col_9에 있는 값을 적음
        elements = page.query_selector_all('span.NPEfkd.RveJvd.snByac')
        for element in elements:
            if element.inner_text() == '다음':
                element.click()
                break
        time.sleep(1)
        col_a = np.random.choice([5, 4], size=8, p=[0.95, 0.05])  # 두 번째 섹션을 위해 col_a를 다시 초기화
        for j in range(1, len(col_a)):
            col_a[j] += 5 * j  # col_a에 0 5 10 15로 순차적으로 더함(5개 항목)
            
        elements = page.query_selector_all('.AB7Lab.Id5V1')  # 현재 페이지에 있는 모든 클릭 항목들을 찾음
        for k in range(0, 7): 
            elements[col_a[k] - 1].click()  # 클릭 항목을 순차적으로 체크함
        
        elements = page.query_selector_all('span.NPEfkd.RveJvd.snByac')
        for element in elements:
            if element.inner_text() == '다음':
                element.click()
                break
        time.sleep(1)
        elements = page.query_selector_all('span.NPEfkd.RveJvd.snByac')
        for element in elements:
            if element.inner_text() == '제출':
                element.click()
                break
        time.sleep(1)


def 동작():
    global browser, page  # 전역 변수로 사용
    엑셀()
    with sync_playwright() as playwright:  # 크로미움 브라우저 열기
        browser = playwright.chromium.launch(
            headless=False,
        )
        page = browser.new_page()  # 새로운 창이 열리면 page에 저장
        자동화(page)
        browser.close()

동작()
