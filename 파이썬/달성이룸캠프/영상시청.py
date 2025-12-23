# -*- coding: utf-8 -*-
# pip install playwright openpyxl && python -m playwright install
# optional (image matching): pip install opencv-python numpy
"""
Video auto player: login -> start -> play 16x
Excel file path is EXCEL_FILE.
Each video: 87.5% watch. Student skip: 1%.
"""

from __future__ import annotations

import asyncio
import math
import random
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from playwright.async_api import TimeoutError as PlaywrightTimeoutError
from playwright.async_api import async_playwright

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / "\uc601\uc0c1\uc2dc\uccad.xlsx"

# Browser settings
HEADLESS = False
KEEP_OPEN = False
USE_PERSISTENT = True
BROWSER_CHANNEL = "chrome"  # or "msedge" or None
EXECUTABLE_PATH = None
VIEWPORT_WIDTH = 1280
VIEWPORT_HEIGHT = 720
DEVICE_SCALE_FACTOR = 1

# Playback / probabilities
PLAYBACK_RATE = 12.0
WATCH_PROB = 0.85
SKIP_STUDENT_PROB = 0.01
SKIP_IF_PROGRESS_COMPLETE = 0  # 스킵할지 말지

# Extension (Video Speed Controller unpacked dir)
EXTENSION_ENABLED = False
EXTENSION_DIR = BASE_DIR / "extension_video_speed_controller"

# Image matching (optional)
IMAGE_MATCH_ENABLED = True
IMAGE_DIR = BASE_DIR / "pic"
START_IMAGE = IMAGE_DIR / "ss.png"
IMAGE_MATCH_THRESHOLD = 0.92
IMAGE_SCROLL_OFFSETS = (0,)
IMAGE_SCAN_DELAY_MS = 50

# Timeouts (ms)
NAV_TIMEOUT = 5000
FAST_TIMEOUT = 1000
RETRY_WAIT_MS = 500
START_WAIT_TIMEOUT = 300
CANCEL_WAIT_TIMEOUT = 300
PLAY_CLICK_TIMEOUT = 400
VIDEO_WAIT_TIMEOUT = 400
SPEED_RETRY_COUNT = 1
SPEED_RETRY_DELAY_MS = 100
SPEED_APPLY_DELAY_MS = 1000
SPEED_CHECK_DELAY_MS = 600
SPEED_RECHECK_COUNT = 2
SPEED_LOW_RATE_THRESHOLD = 2.1
FINAL_WAIT_MS = 50


@dataclass(frozen=True)
class Account:
    user_id: str
    password: str
    name: str


# ---------------- Excel ---------------- #
def _to_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, int):
        return str(v).strip()
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def _derive_password(user_id: str, password: str) -> str:
    """Always derive password from the last 6 chars of user_id."""
    return user_id[-6:] if len(user_id) >= 6 else password


def _derive_name(user_id: str, name: str) -> str:
    """If name is empty, derive it from user_id by stripping the last 6 chars."""
    if name:
        return name
    return user_id[:-6] if len(user_id) > 6 else user_id


def load_config_from_excel(excel_path: str | Path) -> dict:
    excel_path = Path(excel_path)
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active

    site_url = _to_str(ws["N1"].value)
    if not site_url:
        raise ValueError("N1 (site_url) is empty.")

    video_urls: list[str] = []
    for r in range(2, 8):
        val = _to_str(ws[f"N{r}"].value)
        if val:
            video_urls.append(val)
    if not video_urls:
        raise ValueError("N2~N7 video_urls are empty.")

    accounts: list[Account] = []
    row = 2
    while True:
        uid = _to_str(ws.cell(row=row, column=1).value)
        pw = _to_str(ws.cell(row=row, column=2).value)
        name = _to_str(ws.cell(row=row, column=3).value)
        if not uid:
            break
        pw = _derive_password(uid, pw)
        name = _derive_name(uid, name)
        accounts.append(Account(user_id=uid, password=pw, name=name))
        row += 1

    if not accounts:
        raise ValueError("No accounts in column A.")

    return {"site_url": site_url, "video_urls": video_urls, "accounts": accounts}


# ---------------- utils ---------------- #
def _all_frames(page) -> list:
    return [page] + [f for f in page.frames if f is not page]


async def _click_first(frames: Iterable, selectors: list[str], timeout: int) -> bool:
    for fr in frames:
        for sel in selectors:
            try:
                await fr.wait_for_selector(sel, state="visible", timeout=timeout)
                await fr.click(sel, timeout=timeout)
                return True
            except Exception:
                continue
    return False


async def _click_if_exists(page, selector: str, timeout: int) -> bool:
    try:
        await page.wait_for_selector(selector, state="visible", timeout=timeout)
        await page.click(selector, timeout=timeout)
        return True
    except Exception:
        return False


def _load_cv2():
    try:
        import cv2  # type: ignore
        import numpy as np  # type: ignore
    except Exception:
        return None, None
    return cv2, np


async def _click_by_image(page, image_path: Path, threshold: float = IMAGE_MATCH_THRESHOLD) -> bool:
    if not IMAGE_MATCH_ENABLED:
        return False
    if not image_path.exists():
        return False
    cv2, np = _load_cv2()
    if cv2 is None or np is None:
        return False
    try:
        # try a few scroll positions to find the image in viewport
        for y in IMAGE_SCROLL_OFFSETS:
            await page.evaluate("window.scrollTo(0, arguments[0]);", y)
            await page.wait_for_timeout(IMAGE_SCAN_DELAY_MS)
            png_bytes = await page.screenshot()
            screen = cv2.imdecode(np.frombuffer(png_bytes, np.uint8), cv2.IMREAD_COLOR)
            tpl_bytes = image_path.read_bytes()
            template = cv2.imdecode(np.frombuffer(tpl_bytes, np.uint8), cv2.IMREAD_COLOR)
            if screen is None or template is None:
                return False
            sh, sw = screen.shape[:2]
            th, tw = template.shape[:2]
            if th > sh or tw > sw:
                continue
            res = cv2.matchTemplate(screen, template, cv2.TM_CCOEFF_NORMED)
            _, max_val, _, max_loc = cv2.minMaxLoc(res)
            if max_val < threshold:
                continue
            x = int(max_loc[0] + tw / 2)
            y_click = int(max_loc[1] + th / 2)
            await page.mouse.click(x, y_click)
            return True
        return False
    except Exception:
        return False


# ---------------- login ---------------- #
async def _do_login(page, site_url: str, account: Account) -> None:
    await page.goto(site_url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)

    # logout if visible
    await _click_if_exists(page, 'text=\ub85c\uadf8\uc544\uc6c3', timeout=FAST_TIMEOUT)
    await _click_if_exists(page, 'a:has-text("\ub85c\uadf8\uc544\uc6c3")', timeout=FAST_TIMEOUT)

    # open login popup
    await page.click("li.login_btn a", timeout=FAST_TIMEOUT)
    await page.wait_for_selector("#lUserID", timeout=FAST_TIMEOUT)
    await page.fill("#lUserID", account.user_id)
    await page.fill("#lPasswd", account.password)
    await page.click("button.keybg", timeout=FAST_TIMEOUT)
    try:
        await page.wait_for_selector('text=\ub85c\uadf8\uc544\uc6c3', timeout=FAST_TIMEOUT)
    except PlaywrightTimeoutError:
        pass


# ---------------- start ---------------- #
async def _click_start_button(page) -> bool:
    selectors = [
        'a:has-text("\uc218\uac15\uc2dc\uc791")',
        "a.course_request_btn",
        "a.request.keybg.keybd.course_request_btn.noCartBtn",
        "a.course_request_btn.noCartBtn",
    ]
    # fast check: no waiting
    for fr in _all_frames(page):
        for sel in selectors:
            try:
                loc = fr.locator(sel)
                if await loc.count() > 0:
                    await loc.first.click(timeout=START_WAIT_TIMEOUT)
                    return True
            except Exception:
                continue

    # short wait on the main selector
    for fr in _all_frames(page):
        try:
            await fr.wait_for_selector(selectors[0], state="visible", timeout=START_WAIT_TIMEOUT)
            await fr.click(selectors[0], timeout=START_WAIT_TIMEOUT)
            return True
        except Exception:
            continue

    return await _click_by_image(page, START_IMAGE)


async def _ensure_started(page, video_url: str) -> bool:
    """Try start; if not, reload and retry."""
    if await _click_start_button(page):
        return True

    try:
        await page.reload(wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        await page.wait_for_timeout(RETRY_WAIT_MS)
    except Exception:
        pass
    if await _click_start_button(page):
        return True

    try:
        await page.goto(video_url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
        await page.wait_for_timeout(RETRY_WAIT_MS)
    except Exception:
        pass
    return await _click_start_button(page)


async def _click_cancel_popup(page) -> bool:
    selectors = [
        "button.btn-cancel",
        "button.vjs-control.btn-cancel",
        'button[title="Cancel"]',
        'button:has-text("\\ucde8\\uc18c")',
        "text=\\ucde8\\uc18c",
    ]
    for fr in _all_frames(page):
        for sel in selectors:
            try:
                loc = fr.locator(sel)
                if await loc.count() > 0:
                    await loc.first.click(timeout=CANCEL_WAIT_TIMEOUT)
                    return True
            except Exception:
                continue
    return False


# ---------------- playback ---------------- #
async def _click_big_play(page) -> bool:
    selectors = [
        "button.vjs-big-play-button",
        ".vjs-big-play-button",
        "svg.svg-big-play-button-dims",
        "svg.svg-big-play-button-dims use",
    ]
    return await _click_first(_all_frames(page), selectors, timeout=PLAY_CLICK_TIMEOUT)


async def _boost_video_speed(page, rate: float) -> bool:
    for fr in _all_frames(page):
        try:
            video = await fr.wait_for_selector("video", state="visible", timeout=VIDEO_WAIT_TIMEOUT)
        except Exception:
            continue
        if not video:
            continue
        try:
            await video.evaluate("(v, r) => { v.muted = true; v.play(); v.playbackRate = r; }", rate)
            return True
        except Exception:
            continue
    return False


async def _force_videojs_speed(page, rate: float) -> bool:
    for fr in _all_frames(page):
        try:
            applied = await fr.evaluate(
                """(r) => {
                    let ok = false;
                    if (window.videojs && videojs.players) {
                        Object.values(videojs.players).forEach(p => {
                            try { p.muted(true); p.play(); p.playbackRate(r); ok = true; } catch(e){}
                        });
                    }
                    const vids = Array.from(document.querySelectorAll('video'));
                    vids.forEach(v => { try { v.muted = true; v.play(); v.playbackRate = r; ok = true; } catch(e){} });
                    return ok;
                }""",
                rate,
            )
            if applied:
                return True
        except Exception:
            continue
    return False


async def _apply_extension_speed(page, rate: float) -> None:
    if not EXTENSION_ENABLED or not EXTENSION_DIR.exists():
        return
    try:
        await page.keyboard.press("g")
        await page.keyboard.type(str(rate))
        await page.keyboard.press("Enter")
    except Exception:
        pass


async def _get_max_playback_rate(page) -> float:
    max_rate = 0.0
    for fr in _all_frames(page):
        try:
            rate = await fr.evaluate(
                """() => {
                    const vids = Array.from(document.querySelectorAll('video'));
                    if (!vids.length) return 0;
                    return Math.max(...vids.map(v => Number(v.playbackRate || 0) || 0));
                }"""
            )
            if rate and rate > max_rate:
                max_rate = float(rate)
        except Exception:
            continue
    return max_rate


async def _reapply_speed(page, rate: float = PLAYBACK_RATE, retries: int = 3, delay_ms: int = 300) -> None:
    for _ in range(retries):
        current = await _get_max_playback_rate(page)
        if current >= rate - 0.1:
            break
        boosted = await _boost_video_speed(page, rate=rate)
        if not boosted:
            await _force_videojs_speed(page, rate=rate)
        await _apply_extension_speed(page, rate=rate)
        await page.wait_for_timeout(delay_ms)


async def _ensure_speed(page, target_rate: float) -> None:
    for _ in range(SPEED_RECHECK_COUNT):
        current = await _get_max_playback_rate(page)
        if current >= target_rate - 0.1:
            return
        # If it is stuck at ~2x, force another pass
        if current <= SPEED_LOW_RATE_THRESHOLD:
            await _reapply_speed(
                page,
                rate=target_rate,
                retries=max(2, SPEED_RETRY_COUNT),
                delay_ms=SPEED_RETRY_DELAY_MS,
            )
        else:
            await _reapply_speed(
                page,
                rate=target_rate,
                retries=SPEED_RETRY_COUNT,
                delay_ms=SPEED_RETRY_DELAY_MS,
            )
        await page.wait_for_timeout(SPEED_CHECK_DELAY_MS)


async def _is_progress_complete(page, timeout_ms: int = 200) -> bool:
    try:
        prog = await page.locator("div.video_progress").first.inner_text(timeout=timeout_ms)
        return bool(prog and "100%" in prog)
    except Exception:
        return False


async def _wait_by_duration(page, rate: float = PLAYBACK_RATE, extra_sec: float = 5.0, min_sec: float = 5.0, max_cap_sec: float = 1800.0) -> None:
    # If already 100% complete, skip waiting
    if SKIP_IF_PROGRESS_COMPLETE and await _is_progress_complete(page, timeout_ms=200):
        return

    duration = 0.0
    for fr in _all_frames(page):
        try:
            d = await fr.evaluate(
                """() => {
                    const vids = Array.from(document.querySelectorAll('video'));
                    if (!vids.length) return 0;
                    return Math.max(...vids.map(v => Number(v.duration) || 0));
                }"""
            )
            if d and math.isfinite(d):
                duration = max(duration, float(d))
        except Exception:
            continue

    if duration <= 0:
        wait_sec = min_sec + extra_sec
    else:
        wait_sec = (duration / max(rate, 0.1)) + extra_sec
        wait_sec = max(wait_sec, min_sec)
        wait_sec = min(wait_sec, max_cap_sec)

    await page.wait_for_timeout(int(wait_sec * 1000))


# ---------------- account flow ---------------- #
async def _watch_one_video(page, video_url: str) -> bool:
    await page.goto(video_url, wait_until="domcontentloaded", timeout=NAV_TIMEOUT)
    await _click_cancel_popup(page)
    if SKIP_IF_PROGRESS_COMPLETE and await _is_progress_complete(page, timeout_ms=300):
        return True
    started = await _ensure_started(page, video_url=video_url)
    if not started:
        return False
    await _click_cancel_popup(page)
    await _click_big_play(page)
    await _click_cancel_popup(page)
    await page.wait_for_timeout(SPEED_APPLY_DELAY_MS)
    await _reapply_speed(page, rate=PLAYBACK_RATE, retries=SPEED_RETRY_COUNT, delay_ms=SPEED_RETRY_DELAY_MS)
    await _ensure_speed(page, target_rate=PLAYBACK_RATE)
    await _wait_by_duration(page, rate=PLAYBACK_RATE, extra_sec=2, min_sec=2)
    return True


async def login_once(
    site_url: str,
    account: Account,
    video_urls: list[str],
    headless: bool = HEADLESS,
    keep_open: bool = KEEP_OPEN,
    browser_channel: str | None = BROWSER_CHANNEL,
    executable_path: str | None = EXECUTABLE_PATH,
    use_persistent: bool = USE_PERSISTENT,
) -> None:
    async with async_playwright() as p:
        launch_kwargs = {}
        if browser_channel:
            launch_kwargs["channel"] = browser_channel
        if executable_path:
            launch_kwargs["executable_path"] = executable_path

        extension_args = []
        if EXTENSION_ENABLED and EXTENSION_DIR.exists():
            extension_args = [
                f"--disable-extensions-except={EXTENSION_DIR}",
                f"--load-extension={EXTENSION_DIR}",
            ]

        args_common = ["--autoplay-policy=no-user-gesture-required", *extension_args]

        if use_persistent:
            user_data_dir = BASE_DIR / "pw_profile"
            user_data_dir.mkdir(parents=True, exist_ok=True)
            context = await p.chromium.launch_persistent_context(
                user_data_dir,
                headless=headless,
                args=args_common,
                viewport={"width": VIEWPORT_WIDTH, "height": VIEWPORT_HEIGHT},
                device_scale_factor=DEVICE_SCALE_FACTOR,
                **launch_kwargs,
            )
            page = context.pages[0] if context.pages else await context.new_page()
        else:
            browser = await p.chromium.launch(headless=headless, args=args_common, **launch_kwargs)
            context = await browser.new_context(
                viewport={"width": VIEWPORT_WIDTH, "height": VIEWPORT_HEIGHT},
                device_scale_factor=DEVICE_SCALE_FACTOR,
            )
            page = await context.new_page()

        try:
            try:
                await page.set_viewport_size({"width": VIEWPORT_WIDTH, "height": VIEWPORT_HEIGHT})
            except Exception:
                pass
            await _do_login(page, site_url, account)


            for video_url in video_urls:
                if random.random() > WATCH_PROB:
                    continue
                await _watch_one_video(page, video_url)

            await page.wait_for_timeout(FINAL_WAIT_MS)
        finally:
            if keep_open:
                try:
                    await asyncio.to_thread(input, "Press Enter to close: ")
                except Exception:
                    pass
            await context.close()


# ---------------- main ---------------- #
async def _run_all():
    config = load_config_from_excel(EXCEL_FILE)
    accounts: list[Account] = config["accounts"]
    video_urls: list[str] = config["video_urls"]
    site_url: str = config["site_url"]

    total = len(accounts)
    completed = 0

    for idx, acc in enumerate(accounts, start=1):
        if random.random() < SKIP_STUDENT_PROB:
            print(f"[SKIP] account {idx}/{total} {acc.name} skip(1%)")
            continue
        try:
            await login_once(
                site_url,
                acc,
                video_urls=video_urls,
                headless=HEADLESS,
                keep_open=KEEP_OPEN,
                browser_channel=BROWSER_CHANNEL,
                executable_path=EXECUTABLE_PATH,
                use_persistent=USE_PERSISTENT,
            )
            completed += 1
            print(f"[DONE] account {idx}/{total} {acc.name} done")
        except Exception as e:
            print(f"[ERROR] account {idx}/{total} {acc.name} fail: {e}")

    print(f"[DONE] total done: {completed}/{total}")


if __name__ == "__main__":
    asyncio.run(_run_all())
