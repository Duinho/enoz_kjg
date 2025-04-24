import pyautogui
import os
import time as pytime
from PIL import Image
import numpy as np
import keyboard  # 키보드 이벤트를 감지하기 위한 라이브러리
import pyperclip  # 클립보드에서 텍스트를 가져오기 위한 라이브러리
import re
import sys
import openpyxl
from datetime import datetime, timedelta, time 
import random
import calendar


# 엑셀 파일 경로 및 워크북 전역 변수 선언
excel_file_path = None
wb = None
ws = None
대상년월 = 202408

def 이미지_불러오기(이미지_파일명):
    img = Image.open(이미지_파일명)  # 이미지를 Pillow 형식으로 불러옴
    return np.array(img)  # 이미지를 numpy 배열로 변환하여 반환

def 이미지서치(이미지_데이터, 클릭, 유사도):
    global sc
    try:
        # 1920x1080 해상도 내에서만 검색하도록 영역을 제한
        화면_영역 = (0, 0, 1920, 1080)  # (left, top, width, height)
        위치 = pyautogui.locateCenterOnScreen(이미지_데이터, confidence=유사도, region=화면_영역)  # 제한된 영역에서 이미지 검색
        if 위치 is not None:
            sc = 위치
            pyautogui.moveTo(위치)  # 이미지 위치로 마우스 이동
            if 클릭 == 1:
                pyautogui.click()
        else:
            sc = None
            print("이미지를 찾을 수 없습니다.")
    except Exception as e:
        sc = None

def 수정클릭():
    이미지서치(emptybox, 1, 0.99)
    if sc is not None:
        pyautogui.move(1460, 0)
        pyautogui.click()
        pytime.sleep(2)
    else:
        이미지서치(page, 0, 0.99)
        if sc is not None:
            pyautogui.move(10, 0)
            pyautogui.click()
            wb.save(excel_file_path)
            pytime.sleep(2)
        else:
            이미지서치(end, 0, 0.99)
            if sc is not None:
                print('모든 강사 정리 완료')
                강사추가()
            else:
                pyautogui.scroll(-1000)

        수정클릭()

def 시간_추출_및_포맷(텍스트):
    숫자만 = (re.sub(r'\D', '', 텍스트)[2:]).zfill(4)
    return f"{숫자만[:2]}:{숫자만[2:]}" if len(숫자만) >= 4 else 숫자만

def 시작_종료_시간_가져오기():
    pyautogui.move(486, 0)
    pyautogui.click()
    pyautogui.click()
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'c')
    날짜 = pyperclip.paste()
    pyautogui.move(100, 0)
    pyautogui.dragRel(120, 0, duration=0.2)
    pyautogui.hotkey('ctrl', 'c')
    시작시간 = pyperclip.paste()
    pyautogui.move(60, 0)
    pyautogui.dragRel(120, 0, duration=0.2)
    pyautogui.hotkey('ctrl', 'c')
    종료시간 = pyperclip.paste()
    return 날짜, 시작시간, 종료시간

def 교육시간표():
    시간표_리스트 = []  # 모든 날짜, 시작 시간, 종료 시간을 저장할 리스트
    
    for i in range(1, 8): 
        이미지서치(globals()[f'day{i}'], 0, 0.98)  
        if sc is not None:
            날짜, 시작, 종료 = 시작_종료_시간_가져오기()  # 날짜도 포함하여 반환
            시작_시간 = 시간_추출_및_포맷(시작)
            종료_시간 = 시간_추출_및_포맷(종료)
            
            # 날짜, 시작 시간, 종료 시간을 리스트에 추가
            시간표_리스트.extend([날짜, 시작_시간, 종료_시간])
    
    return 시간표_리스트

def 교육장_강사명():
    이미지서치(place,0,0.98)
    pyautogui.move(600, 50)
    pyautogui.click()
    pyautogui.click()
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'c')
    주소1 = pyperclip.paste() 
    pyautogui.move(400, 0)
    pyautogui.click()
    pyautogui.click()
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'c')
    주소2 = pyperclip.paste() 
    주소 = 주소1 + '' + 주소2
    이미지서치(name,0,0.98)
    pyautogui.move(200, 0)
    pyautogui.click()
    pyautogui.click()
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'c')
    주강사 = pyperclip.paste() 
    pyautogui.move(760, 0)
    pyautogui.click()
    pyautogui.click()
    pyautogui.click()
    pyautogui.hotkey('ctrl', 'c')
    보조강사 = pyperclip.paste()
    if 주강사 == 보조강사:
        보조강사 = None
    pyautogui.hotkey('alt', 'left')
    pytime.sleep(2)
    return 주소, 주강사, 보조강사


def 엑셀저장(시간표_리스트, 관리지역, 주소, 주강사, 보조강사):
    global current_row, ws  # 전역 변수 사용
    for index in range(0, len(시간표_리스트), 3):
        날짜 = 시간표_리스트[index]
        시작_시간 = 시간표_리스트[index + 1]
        종료_시간 = 시간표_리스트[index + 2]

        # 종료 시간에서 시작 시간을 빼기 위해 문자열을 시간으로 변환
        종료_시간_dt = datetime.strptime(종료_시간, '%H:%M')
        시작_시간_dt = datetime.strptime(시작_시간, '%H:%M')
        차이 = 종료_시간_dt - 시작_시간_dt

        # 시간 차이를 "hh:mm" 형식으로 변환
        차이_시간 = 차이.seconds // 3600
        차이_분 = (차이.seconds % 3600) // 60
        차이_형식 = f"{차이_시간:02}:{차이_분:02}"

        # 주강사에 대해 랜덤으로 10분에서 1시간 30분 더하기
        주강사_랜덤_분 = random.randint(10, 90)  # 10분에서 90분 (1시간 30분)
        주강사_종료_시간_랜덤_추가 = 종료_시간_dt + timedelta(minutes=주강사_랜덤_분)

        # A열에 날짜와 주강사 수정된 종료 시간 저장 ("날짜 종료시간" 형식으로)
        ws[f'A{current_row}'] = f"{날짜} {주강사_종료_시간_랜덤_추가.strftime('%H:%M')}"
        ws[f'B{current_row}'] = 관리지역
        ws[f'D{current_row}'] = 주소
        ws[f'E{current_row}'] = 주강사
        ws[f'F{current_row}'] = 시작_시간
        ws[f'G{current_row}'] = 종료_시간
        ws[f'H{current_row}'] = 차이_형식

        current_row += 1  # 다음 행으로 이동

        # 보조강사가 있으면 보조강사 정보 저장
        if 보조강사:
            # 보조강사에 대해 별도의 랜덤으로 10분에서 1시간 30분 더하기
            보조강사_랜덤_분 = random.randint(10, 90)  # 10분에서 90분 (1시간 30분)
            보조강사_종료_시간_랜덤_추가 = 종료_시간_dt + timedelta(minutes=보조강사_랜덤_분)

            ws[f'A{current_row}'] = f"{날짜} {보조강사_종료_시간_랜덤_추가.strftime('%H:%M')}"
            ws[f'B{current_row}'] = 관리지역
            ws[f'D{current_row}'] = 주소
            ws[f'E{current_row}'] = 보조강사
            ws[f'F{current_row}'] = 시작_시간
            ws[f'G{current_row}'] = 종료_시간
            ws[f'H{current_row}'] = 차이_형식

            current_row += 1  # 다음 행으로 이동
    

def 근태관리():
    수정클릭()
    이미지서치(area, 0,0.98)
    pyautogui.move(280, 0)
    pyautogui.dragRel(40, 0, duration=0.2)
    pyautogui.hotkey('ctrl', 'c')
    관리지역 = pyperclip.paste() 
    관리지역 = re.sub(r'\s+', '', 관리지역)
    pyautogui.scroll(-700)
    시간표_리스트 = 교육시간표()
    주소, 주강사, 보조강사 = 교육장_강사명()
    엑셀저장(시간표_리스트, 관리지역, 주소, 주강사, 보조강사)


def 주말삭제(년월):
    # 입력된 년월에서 연도와 월을 추출
    year = int(str(년월)[:4])
    month = int(str(년월)[4:])

    weekdays = []  # 주말을 제외한 날짜를 저장할 리스트

    # 해당 연도와 월의 월간 달력을 주단위로 가져옴 (각 주는 일요일부터 시작)
    cal = calendar.monthcalendar(year, month)

    for week in cal:
        for day in week:
            if day != 0:  # 0은 해당 주의 빈 날짜를 의미
                weekday = datetime(year, month, day).weekday()  # 요일을 구함 (월요일=0, 일요일=6)
                if weekday < 5:  # 0~4는 월요일부터 금요일
                    weekdays.append(f"{year}-{month:02d}-{day:02d}")  # 날짜를 YYYY-MM-DD 형식으로 추가

    return weekdays


def 강사추가():
    global ps, ws, current_row  # 'ps', 'ws', 'current_row' 시트를 전역 변수로 사용

    # '강사유형'이라는 시트를 선택합니다.
    ps = wb['강사유형']

    # 조건에 맞는 데이터를 저장할 리스트 초기화
    필터링된_데이터 = []

    # '상설배움터' 또는 '체험존' 키워드를 포함하는 행을 찾습니다.
    for row in ps.iter_rows(min_row=2, max_row=ps.max_row, min_col=1, max_col=ps.max_column):
        # F열이 6번째 열이므로, row[5]로 접근합니다.
        if '상설배움터' in str(row[5].value) or '체험존' in str(row[5].value):
            # 조건에 맞는 행이 발견되면 해당 데이터를 리스트에 추가합니다.
            데이터_행 = [cell.value for cell in row]  # 해당 행의 모든 셀 값을 가져옵니다.
            필터링된_데이터.append(데이터_행)  # 리스트에 추가

    # 주말 삭제 후 평일 날짜 리스트를 가져옴
    평일_날짜_리스트 = 주말삭제(대상년월)

    # ws 시트의 E열에 대해 ps의 D열과 동일한 값을 가진 행을 삭제합니다.
    ps_values = [ps_row[0].value for ps_row in ps.iter_rows(min_row=2, max_row=ps.max_row, min_col=4, max_col=4)]
    ws_rows_to_delete = []

    # 삭제할 행을 수집
    for ws_row in range(2, ws.max_row + 1):
        ws_value = ws.cell(row=ws_row, column=5).value  # ws의 E열 값
        if ws_value in ps_values:
            ws_rows_to_delete.append(ws_row)

    # 역순으로 행 삭제
    for row in sorted(ws_rows_to_delete, reverse=True):
        ws.delete_rows(row)

    # 행 삭제 후 current_row 재조정
    current_row = ws.max_row + 1

    # 필터링된 데이터를 각 평일 날짜에 대해 반복하여 ws 시트에 추가
    for 날짜 in 평일_날짜_리스트:
        for row in 필터링된_데이터:
            # ps의 열 값을 가져오기
            C_value = row[2]  # C열 값 (인덱스 2)
            D_value = row[3]  # D열 값 (인덱스 3)
            G_value = row[6]  # G열 값 (인덱스 6)
            I_value = row[8]  # I열 값 (인덱스 8)
            J_value = row[9]  # J열 값 (인덱스 9)

            # J_value가 time 형식인지 확인
            if isinstance(J_value, time):
                J_시간 = datetime.combine(datetime.today(), J_value)  # 현재 날짜와 결합하여 datetime 객체로 변환
            else:
                J_시간 = datetime.strptime(J_value, '%H:%M')  # 문자열인 경우 변환

            # 랜덤 시간 추가
            랜덤_분 = random.randint(10, 90)
            수정_시간 = J_시간 + timedelta(minutes=랜덤_분)

            # H열에 J - I 차이 시간 계산
            if isinstance(I_value, time):
                I_시간 = datetime.combine(datetime.today(), I_value)
            else:
                I_시간 = datetime.strptime(I_value, '%H:%M')

            차이 = J_시간 - I_시간
            차이_시간 = 차이.seconds // 3600
            차이_분 = (차이.seconds % 3600) // 60
            차이_형식 = f"{차이_시간:02}:{차이_분:02}"

            # A열에 주말 삭제한 날짜 + 수정된 J열 시간 추가
            ws[f'A{current_row}'] = f"{날짜} {수정_시간.strftime('%H:%M')}"
            ws[f'B{current_row}'] = C_value
            ws[f'D{current_row}'] = G_value
            ws[f'E{current_row}'] = D_value
            ws[f'F{current_row}'] = I_value
            ws[f'G{current_row}'] = J_value
            ws[f'H{current_row}'] = 차이_형식

            current_row += 1  # 다음 행으로 이동

    wb.save(excel_file_path)
    wb.close()
    print("엑셀 파일이 저장되고 닫혔습니다.")
    sys.exit()






    
if __name__ == "__main__":
    sc = 0
    현재_디렉토리 = os.path.dirname(os.path.abspath(__file__))
    excel_file_path = os.path.join(현재_디렉토리, '근태관리결과.xlsx')  # 엑셀 파일 경로
    
    # 엑셀 파일 열기
    wb = openpyxl.load_workbook(excel_file_path)
    
    ws = wb['raw data_근태관리']  # '매크로'라는 이름의 시트가 있으면 선택

    # 현재 시트의 마지막 사용된 행을 찾고, 그 다음 행부터 작성하도록 설정
    current_row = ws.max_row + 1

    사진경로 = os.path.abspath(os.path.join(현재_디렉토리, 'pic'))
    emptybox = 이미지_불러오기(os.path.join(사진경로, 'emptybox.png'))
    page = 이미지_불러오기(os.path.join(사진경로, 'page.png'))
    area = 이미지_불러오기(os.path.join(사진경로, 'area.png'))
    day1 = 이미지_불러오기(os.path.join(사진경로, 'day1.png'))
    day2 = 이미지_불러오기(os.path.join(사진경로, 'day2.png'))
    day3 = 이미지_불러오기(os.path.join(사진경로, 'day3.png'))
    day4 = 이미지_불러오기(os.path.join(사진경로, 'day4.png'))
    day5 = 이미지_불러오기(os.path.join(사진경로, 'day5.png'))
    day6 = 이미지_불러오기(os.path.join(사진경로, 'day6.png'))
    day7 = 이미지_불러오기(os.path.join(사진경로, 'day7.png'))
    place = 이미지_불러오기(os.path.join(사진경로, 'place.png'))
    name = 이미지_불러오기(os.path.join(사진경로, 'name.png'))
    end = 이미지_불러오기(os.path.join(사진경로, 'end.png'))
    
    print("프로그램을 종료하려면 'q' 키를 누르세요.")

    while True:
    
        근태관리()
        
        if keyboard.is_pressed('q'):  # 'q' 키가 눌리면
            print("프로그램을 종료합니다.")
            wb.save(excel_file_path)
            wb.close()
            break  # 루프 탈출 및 프로그램 종료