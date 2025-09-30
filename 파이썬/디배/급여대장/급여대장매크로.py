# pip install pandas openpyxl

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os

def process_excel(filename):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    input_path = os.path.join(base_dir, filename)

    # ---------------------------
    # 1. 급여대장 시트 읽기
    # ---------------------------
    df_wage = pd.read_excel(input_path, sheet_name="급여대장", header=5)

    # 이름 컬럼 자동 탐색
    name_col = next((col for col in df_wage.columns if "성" in str(col)), None)

    # 필요한 컬럼만
    columns_to_keep = [name_col, "국민연금", "건강보험", "고용보험", "장기요양보험료", "소득세", "지방소득세"]
    keep = [col for col in columns_to_keep if col in df_wage.columns]
    filtered_df = df_wage[keep]

    # 이름이 문자열인 행만 (합계 숫자 제거)
    if name_col:
        filtered_df = filtered_df[filtered_df[name_col].apply(lambda x: isinstance(x, str))]

    # ---------------------------
    # 2. 급여산정 시트 읽기
    # ---------------------------
    df_calc = pd.read_excel(input_path, sheet_name="급여산정", header=1)

    if "4대보험" not in df_calc.columns:
        raise ValueError("급여산정 시트에서 '4대보험' 컬럼을 찾지 못했습니다. header 값을 조정해보세요.")

    df_calc_filtered = df_calc[df_calc["4대보험"] == "O"]

    # 이름(C열), 아이디(B열)
    name_col_calc = df_calc.columns[2]  # C열
    id_col_calc = df_calc.columns[1]    # B열

    # 동명이인 체크
    name_counts = df_calc_filtered[name_col_calc].value_counts()
    name_to_id = {}
    for _, row in df_calc_filtered.iterrows():
        name = row[name_col_calc]
        emp_id = row[id_col_calc]
        if name_counts[name] > 1:
            name_to_id[name] = "동명이인"
        else:
            name_to_id[name] = emp_id

    # ---------------------------
    # 3. 결과 저장 (아이디 열 추가)
    # ---------------------------
    output_path = os.path.join(base_dir, filename.replace(".xlsx", "_결과.xlsx"))
    filtered_df.to_excel(output_path, index=False)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    ws.insert_cols(1)
    ws.cell(row=1, column=1, value="아이디")

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=2).value  # 이름은 B열
        if name in name_to_id:
            if name_to_id[name] == "동명이인":
                ws.cell(row=row, column=1, value="동명이인")
                ws.cell(row=row, column=1).fill = red_fill
            else:
                ws.cell(row=row, column=1, value=name_to_id[name])

    wb.save(output_path)

    # ---------------------------
    # 4. 엑셀 파일 자동 실행 (Windows 전용)
    # ---------------------------
    os.startfile(output_path)


if __name__ == "__main__":
    process_excel("전체급여대장.xlsx")
