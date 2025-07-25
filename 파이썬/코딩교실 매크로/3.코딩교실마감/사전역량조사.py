import openpyxl
import os
from playwright.sync_api import sync_playwright
import statistics
import time
import random
import pandas as pd

result = 0
page = None
browser = None
col_a = []  # 데이터를 저장할 전역 변수

def download_template_sajeon():
    """사전역량평가 양식 다운로드 후 파일 자동으로 열기."""
    from tkinter import filedialog, messagebox
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="사전역량향상평가 양식.xlsx"
    )
    if filepath:
        df = pd.DataFrame(columns=['학년', '응답1', '응답2', '응답3', '응답4', '응답5', '응답6', '응답7'])
        with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            worksheet.write('K1', 'A~H에 학년,응답을 복붙')
            worksheet.write('K2', '응답이 5일경우 G,H에 1 작성')
            worksheet.write('K3', 'O에 해당하는 것을 P에 작성')
            worksheet.write('O1', '사전역량조사링크')
            worksheet.write('O2', '목표 4학년 수')
            worksheet.write('O3', '목표 5학년 수')
            worksheet.write('O4', '목표 6학년 수')
        messagebox.showinfo("저장 완료", f"사전역량향상평가 양식 파일이 저장되었습니다 :\n" + filepath)
        os.startfile(filepath)  # 파일을 자동으로 열기

def 엑셀(filepath):
    """사용자가 선택한 엑셀 파일을 불러와 데이터 처리."""
    global wb, col_a, 링크
    wb = openpyxl.load_workbook(filepath, data_only=True)  # 사용자가 선택한 파일 열기
    sheet = wb.active

    # 학년 및 응답 값 처리
    col_a = []
    링크 = sheet["P1"].value
    목표수 = {
        4: sheet["P2"].value,
        5: sheet["P3"].value,
        6: sheet["P4"].value,
    }

    for i in range(2, sheet.max_row + 1):
        학년 = sheet[f"A{i}"].value  # 학년 값은 숫자로 처리
        if 학년 is None:
            continue
        응답 = [sheet.cell(row=i, column=j).value for j in range(2, 9)]
        
        # 응답 값 변환
        변환응답 = [
            5 if 값 == "매우 그렇다" else
            4 if 값 == "그렇다" else
            3 if 값 == "보통이다" else
            2 if 값 == "그렇지 않다" else
            1 if 값 == "전혀 그렇지 않다" else 0
            for 값 in 응답
        ]
        
        col_a.append((학년, 변환응답))

    # 학년별 모자란 데이터 추가
    for 학년, 목표 in 목표수.items():
        현재_인원 = sum(1 for h, _ in col_a if h == 학년)
        부족한수 = 목표 - 현재_인원
        for _ in range(부족한수):
            col_a.append((학년, [1] * 7))  # 응답 값은 모두 1로 추가

    wb.close()  # 엑셀 닫기

def 자동화():
    global page, col_a
    for 학년, 응답 in col_a:  # col_a에 저장된 데이터 처리
        page.goto(링크)  # 링크로 이동
        page.wait_for_selector('span.NPEfkd.RveJvd.snByac', timeout=10000)  # 제출 버튼 대기

        # 응답 클릭
        elements = page.query_selector_all('.AB7Lab.Id5V1')  # 현재 페이지에 있는 모든 클릭 항목을 찾음
        for k, 답변 in enumerate(응답):
            if 답변 > 0:
                elements[답변 - 1 + k * 5].click()  # 정답에 해당하는 항목을 클릭함

        page.click('span.NPEfkd.RveJvd.snByac')  # 제출 클릭

def 결과저장(저장_경로):
    """결과를 엑셀 파일로 저장."""
    result_file = 저장_경로

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "결과"

    current_row = 1  # 엑셀의 현재 행 위치

    # 학년별 인원수 계산
    학년별인원수 = {}
    for h, _ in col_a:
        학년별인원수[h] = 학년별인원수.get(h, 0) + 1

    for 학년, 인원수 in 학년별인원수.items():
        # 학년 및 인원수 작성
        sheet.cell(row=current_row, column=1, value=f"{학년}학년")
        sheet.cell(row=current_row, column=2, value="인원수")
        sheet.cell(row=current_row, column=3, value=인원수)
        current_row += 1

        # 응답 수 헤더 작성 (문항이 가로로)
        headers = ["응답수"] + [f"항목{q}" for q in range(1, 8)] + ["총합"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=current_row, column=col, value=header)
        current_row += 1

        # 응답 1부터 5까지 세로로 작성
        for 응답_num in range(1, 6):
            sheet.cell(row=current_row, column=1, value=f"응답 {응답_num}")
            for q in range(7):  # 항목1부터 항목7까지
                count = sum(1 for h, 응답 in col_a if h == 학년 and 응답[q] == 응답_num)
                sheet.cell(row=current_row, column=q+2, value=count)
            # 총합 응답 수 계산 (해당 응답 번호의 전체 합)
            total = sum(1 for h, 응답 in col_a for question_idx in range(7) if h == 학년 and 응답[question_idx] == 응답_num)
            sheet.cell(row=current_row, column=9, value=total)
            current_row += 1

        # 평균 점수 계산 및 작성
        sheet.cell(row=current_row, column=1, value="평균")
        for q in range(7):
            학년_데이터 = [응답[q] for h, 응답 in col_a if h == 학년 and 응답[q] > 0]
            if 학년_데이터:
                평균 = round(statistics.mean(학년_데이터), 2)
            else:
                평균 = 0
            sheet.cell(row=current_row, column=q+2, value=평균)
        # 총 평균 (전체 문항의 평균)
        전체_데이터 = [점수 for h, 응답 in col_a if h == 학년 for 점수 in 응답]
        if 전체_데이터:
            총평균 = round(statistics.mean(전체_데이터), 2)
        else:
            총평균 = 0
        sheet.cell(row=current_row, column=9, value=총평균)
        current_row += 2  # 다음 학년을 위해 두 줄 띄우기

    wb.save(result_file)
    os.startfile(result_file)

def 동작(filepath, 저장_경로):
    """동작 함수, 선택한 파일으로 작업 실행."""
    global browser, page
    엑셀(filepath)
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        page = browser.new_page()
        자동화()
        browser.close()
    결과저장(저장_경로)

