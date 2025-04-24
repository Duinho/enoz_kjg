from openpyxl import load_workbook
import os
import pandas as pd

def process_names_in_excel(filepath,저장_경로):
    """이름의 두 번째 글자를 'O'로 변경하고 엑셀 파일을 저장합니다."""
    wb = load_workbook(filepath)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        name = cell.value
        if name and isinstance(name, str) and len(name) > 1:
            cell.value = name[0] + 'O' + name[2:]
    wb.save(저장_경로)
    os.startfile(저장_경로)

def download_template_noc():
    """사용자가 저장할 위치를 선택하고 기본 양식의 엑셀 파일을 저장한 후 자동으로 엽니다."""
    from tkinter import filedialog, messagebox
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="이름 O 처리 양식.xlsx"
    )
    if filepath:
        # 데이터프레임 생성
        df = pd.DataFrame({'No.': [1, 2, 3], 'Name': ['홍길동', '김철수', '이영희']})
        # ExcelWriter를 사용하여 엑셀 파일 저장 및 텍스트 추가
        with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            worksheet.write('K1', 'B열에 학생 이름 나열')
            worksheet.write('K2', 'B열의 두 번째 글자를 O로 변경')
        
        messagebox.showinfo("저장 완료", f"양식 파일이 저장되었습니다:\n {filepath}")
        os.startfile(filepath)  # 저장된 파일을 자동으로 엽니다
