import openpyxl
import os
from playwright.sync_api import sync_playwright
import time
import random
from tkinter import filedialog, messagebox, simpledialog
import pandas as pd

def download_template_yl():
    """역량향상도평가 양식 다운로드 후 파일 자동으로 열기."""
    from tkinter import filedialog, messagebox
    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="역량향상도평가 양식.xlsx"
    )
    if filepath:
        df = pd.DataFrame(columns=['이름', '학교', '학년'])
        with pd.ExcelWriter(filepath, engine='xlsxwriter') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            worksheet.write('I1', 'A에 학교 이름(역량향상도평가에 있는 이름과 동일해야함)')
            worksheet.write('I2', 'B에 학생 이름, C에 학년(숫자만 입력)')
            worksheet.write('I3', 'P1~3 구글폼링크, P4~6 정답(숫자만 나열)')
            worksheet.write('O1', '초저링크')
            worksheet.write('O2', '초고링크')
            worksheet.write('O3', '중등링크')
            worksheet.write('O4', '초저정답')
            worksheet.write('O5', '초고정답')
            worksheet.write('O6', '중등정답')
        messagebox.showinfo("저장 완료", f"역량향상도평가 양식 파일이 저장되었습니다 :\n" + filepath)
        os.startfile(filepath)  # 파일을 자동으로 열기

def 엑셀(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet = wb.active  # sheet에 엑셀 값 저장

    col_n = [sheet[f"A{i}"].value for i in range(2, sheet.max_row + 1)]
    col_s = [sheet[f"B{i}"].value for i in range(2, sheet.max_row + 1)]
    col_g = [sheet[f"C{i}"].value for i in range(2, sheet.max_row + 1)]
    col_l = []
    col_a = []
    
    for i in range(2, sheet.max_row + 1):
        grade = sheet[f"C{i}"].value
        if grade is not None:  # grade가 None이 아닐 때만 작업을 수행
            answer_prefix = str(grade - 1)
            if grade == 4 :
                col_l.append(sheet[f"P1"].value)
                answer_value = sheet[f"P4"].value
            elif grade == 5 or grade == 6:
                col_l.append(sheet[f"P2"].value)
                answer_value = sheet[f"P5"].value
            if answer_value is not None:
                col_a.append(answer_prefix + str(answer_value))  # 접두사를 붙여서 col_a에 저장
            else:
                col_a.append(None)  # 값이 없는 경우 None을 할당
        else:
            col_l.append(None)  # grade가 None인 경우
            col_a.append(None)

    wb.close()  # 엑셀 닫음
    return col_n, col_s, col_g, col_l, col_a

def 자동화(col_n, col_s, col_g, col_l, col_a, error_rate):
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        page = browser.new_page()
        
        for i in range(len(col_n)):
            page.goto(col_l[i])
            page.wait_for_selector('span.NPEfkd.RveJvd.snByac', timeout=10000)
            page.fill('input[aria-labelledby="i1 i4"]', str(col_n[i]))
            page.click('div.MocG8c.HZ3kWc.mhLiyf.LMgvRb.KKjvXb.DEh1R')
            time.sleep(0.5)
            
            option_selector = f'//span[text()="{col_s[i]}" and ancestor::div[@role="option"]]'
            page.locator(option_selector).scroll_into_view_if_needed()
            page.locator(option_selector).click()
            
            if col_g[i] > 6:
                col_g[i] = col_g[i] - 6
            col_a_list = []
            for char in str(col_a[i]):
                correct_answer = int(char)
                if random.random() <= error_rate:
                    wrong_answers = [x for x in range(1, 5) if x != correct_answer]
                    if wrong_answers:
                        correct_answer = random.choice(wrong_answers)
                col_a_list.append(correct_answer)
            for j in range(1, len(col_a_list)):  # col_a_list의 길이만큼 반복
                col_a_list[j] += 4 * j  # col_a_list에 0 4 8 12 16로 순차적으로 더함(4개 항목)            
            elements = page.query_selector_all('.AB7Lab.Id5V1')
            for k in range(16):
                elements[col_a_list[k] - 1].click()
            print(f'{col_n[i]} 제출 완료')
            page.click('span.NPEfkd.RveJvd.snByac')
            time.sleep(0.5)

        browser.close()

def 동작(filepath, error_rate):
    col_n, col_s, col_g, col_l, col_a = 엑셀(filepath)
    자동화(col_n, col_s, col_g, col_l, col_a, error_rate)

if __name__ == "__main__":
    동작()
