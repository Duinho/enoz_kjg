import os
import openpyxl
import re, datetime
import time
import shutil
from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup

# ——————————————————————————————
# 설정
# ——————————————————————————————
BASE_DIR            = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH          = os.path.join(BASE_DIR, '보고서수정.xlsx')
REPORT_DIR          = os.path.join(BASE_DIR, '보고서')

# 전역 매핑
name_change_map      = {}
cancellation_map     = {}
additional_map       = {}
session_dates_map    = {}
session_descriptions = []


# ——————————————————————————————
# HTML 후처리 통합
# ——————————————————————————————
def process_html(path,
                 suffix_change=None,
                 attendance=None,
                 clear_fields=(),
                 education_idx=None,
                 additions=None):
    with open(path, 'r', encoding='cp949', errors='ignore') as f:
        soup = BeautifulSoup(f, 'html.parser')

    # 1) suffix 교체
    if suffix_change:
        o_num, n_num = suffix_change
        for lbl in ('Group No', '담당강사명'):
            td0 = soup.find('td', string=lbl)
            if td0:
                td1 = td0.find_next_sibling('td')
                txt = td1.get_text()
                td1.string = txt.rsplit('_', 1)[0] + '_' + n_num

    # 2) 출결현황 + 취소자
    if attendance:
        idx, cancels = attendance
        td0 = soup.find('td', string='출결현황')
        if td0:
            td1 = td0.find_next_sibling('td')
            students = [seg.split('[')[0].strip()
                        for seg in td1.get_text().split(',') if seg.strip()]
            new_p = soup.new_tag('p', **{'class':'MsoNormal'})
            for i, stu in enumerate(students, 1):
                cf = next((n for nm, n in cancels if nm == stu), None)
                status, color = ('결','#EE0000') if cf and idx>=cf else ('출','blue')
                sp0 = soup.new_tag('span', **{'class':'SpellE'}); sp0.string = stu
                sp1 = soup.new_tag('span', lang='EN-US', style=f'color:{color}'); sp1.string = f'[{status}]'
                new_p.append(sp0); new_p.append(sp1)
                if i < len(students): new_p.append(', ')
            # 누락 취소자 → 이전 차시는 [출], 이후는 [결]
            for nm, cf in cancels:
                if nm not in students:
                    if new_p.contents and not str(new_p.contents[-1]).endswith(', '):
                        new_p.append(', ')
                    sp0 = soup.new_tag('span', **{'class':'SpellE'}); sp0.string = nm
                    if idx >= cf:
                        sp1 = soup.new_tag('span', lang='EN-US', style='color:#EE0000'); sp1.string = '[결]'
                    else:
                        sp1 = soup.new_tag('span', lang='EN-US', style='color:blue'); sp1.string = '[출]'
                    new_p.append(sp0); new_p.append(sp1)
            td1.clear()
            td1.append(new_p)

    # 3) clear_fields
    for lbl in clear_fields:
        td0 = soup.find('td', string=lbl)
        if td0:
            td0.find_next_sibling('td').clear()

    # 4) 교육내용 삽입
    if education_idx:
        td0 = soup.find('td', string='교육내용')
        if td0:
            td1 = td0.find_next_sibling('td')
            p = soup.new_tag('p', **{'class':'MsoNormal'})
            p.string = session_descriptions[education_idx-1]
            td1.clear()
            td1.append(p)

    # 5) 추가인원 [출] 삽입
    if additions:
        td0 = soup.find('td', string='출결현황')
        if td0:
            td1 = td0.find_next_sibling('td')
            p = td1.find('p', class_='MsoNormal')
            if not p:
                p = soup.new_tag('p', **{'class':'MsoNormal'})
                td1.clear()
                td1.append(p)

            existing = [span.get_text() for span in p.find_all('span', class_='SpellE')]
            for nm in additions:
                if nm not in existing:
                    if p.contents:
                        p.append(', ')
                    sp0 = soup.new_tag('span', **{'class':'SpellE'})
                    sp0.string = nm
                    sp1 = soup.new_tag('span', lang='EN-US', style='color:blue')
                    sp1.string = '[출]'
                    p.append(sp0)
                    p.append(sp1)

    with open(path, 'w', encoding='cp949', errors='ignore') as f:
        f.write(str(soup))


# ——————————————————————————————
# 설정 불러오기
# ——————————————————————————————
def 초기화():
    global 로그인사이트, 스케쥴관리링크, adminID, adminPW
    global 지역, 대상년월, 월수반, 화목반, 목표월수, 목표화목
    global session_descriptions, additional_map

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['메인']
    로그인사이트   = ws.cell(1,17).value
    스케쥴관리링크  = ws.cell(2,17).value
    adminID        = ws.cell(3,17).value
    adminPW        = ws.cell(4,17).value
    지역           = f"{ws.cell(5,17).value}_"
    대상년월       = ws.cell(6,17).value
    월수반         = ws.cell(7,17).value
    화목반         = ws.cell(8,17).value
    목표월수       = ws.cell(9,17).value
    목표화목       = ws.cell(10,17).value

    session_descriptions[:] = [ws.cell(i,19).value or '' for i in range(1,9)]

    # 반이름 변경
    name_change_map.clear()
    for o,n in wb['반이름변경'].iter_rows(min_row=1, max_col=2, values_only=True):
        if o and n:
            name_change_map[str(o).strip()] = str(n).strip()

    # 취소자 (변경된 반이름 기준으로 저장)
    cancellation_map.clear()
    ws = wb['취소자']
    for row in ws.iter_rows(min_row=1, values_only=True):
        orig = row[0]
        if not orig:
            continue
        mapped = name_change_map.get(str(orig).strip(), str(orig).strip())
        if mapped not in cancellation_map:
            cancellation_map[mapped] = []
        for c in row[1:]:
            if c:
                nm, num = c.split(',')
                cancellation_map[mapped].append((nm.strip(), int(num)))

    # 추가인원 (변경된 반이름 기준으로 저장, 열 구조)
    additional_map.clear()
    ws = wb['추가인원']
    headers = [cell.value for cell in ws[1] if cell.value]
    for col_idx, group in enumerate(headers, start=1):
        if not group:
            continue
        mapped = name_change_map.get(str(group).strip(), str(group).strip())
        additional_map[mapped] = []
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            if row[0]:
                additional_map[mapped].append(str(row[0]).strip())

    # 날짜 맵 초기화
    session_dates_map.clear()


# ——————————————————————————————
# 로그인
# ——————————————————————————————
def 로그인(page):
    page.goto(로그인사이트)
    page.fill('input[name="tbAdminId"]', adminID)
    page.fill('input[name="tbAdminPass"]', adminPW)
    page.press('input[name="tbAdminPass"]', 'Enter')
    page.goto(스케쥴관리링크)
    page.wait_for_selector('select[name="ddlTargetDate"]')
    page.select_option('select[name="ddlTargetDate"]', str(대상년월))
    page.wait_for_selector('select[name="ddlKeyField"]')
    page.select_option('select[name="ddlKeyField"]','a.tutor_id')


# ——————————————————————————————
# 보고서다운
# ——————————————————————————————
def 보고서다운(page):
    os.makedirs(REPORT_DIR, exist_ok=True)
    inp = 'input[name="tbKeyWord"][size="25"]'
    btn = 'a.button_gray_small:has-text("보고서"),a.button_red_small:has-text("보고서")'

    for prefix, count in [('MW',월수반), ('TT',화목반)]:
        for i in range(1, count+1):
            orig   = f"{prefix}_{i:02d}"
            mapped = name_change_map.get(orig, orig)
            code   = f"{지역}{orig}"
            grp    = os.path.join(REPORT_DIR, f"{지역}{mapped}")
            os.makedirs(grp, exist_ok=True)

            page.fill(inp, code); page.press(inp,'Enter')
            loc = page.locator(btn); loc.nth(7).wait_for(timeout=30000)
            bts = loc.all()

            dates = []
            for b in bts[::-1]:
                href = b.get_attribute('href') or ''
                m = re.findall(r"'(.*?)'", href)
                if len(m)>1:
                    dates.append(m[1])
            session_dates_map[orig] = dates

            missing = []
            for idx, b in enumerate(bts[::-1],1):
                cls = b.get_attribute('class') or ''
                date = dates[idx-1] if idx-1 < len(dates) else 'Unknown'
                dst  = os.path.join(grp, f"{지역}{mapped}_{date}.doc")

                if 'button_red_small' in cls:
                    missing.append(date)
                    continue

                with page.expect_popup() as pp: b.click()
                pop = pp.value; pop.wait_for_selector('a.button_yellow')
                with pop.expect_download() as dl: pop.click('a.button_yellow')
                dl.value.save_as(dst); pop.close()

                if orig in name_change_map:
                    o_n, n_n = orig.split('_')[1], mapped.split('_')[1]
                    process_html(dst, suffix_change=(o_n,n_n))

                process_html(dst, attendance=(idx, cancellation_map.get(mapped, [])))

            if missing:
                print(f"{code}의 {', '.join(f'[{d}]' for d in missing)}인 보고서가 작성되지 않았습니다.")


# ——————————————————————————————
# stub 생성
# ——————————————————————————————
def create_stubs():
    for prefix, total in [('MW', 목표월수), ('TT', 목표화목)]:
        template_orig = None
        for i in range(1, total+1):
            o2 = f"{prefix}_{i:02d}"
            m2 = name_change_map.get(o2, o2)
            if os.path.isdir(os.path.join(REPORT_DIR, f"{지역}{m2}")):
                template_orig = o2
                break
        if not template_orig:
            print(f"[WARN] {prefix}용 템플릿 그룹이 없습니다.")
            continue

        dates           = session_dates_map.get(template_orig, [])
        template_mapped = name_change_map.get(template_orig, template_orig)
        src_template    = os.path.join(REPORT_DIR, f"{지역}{template_mapped}", f"{지역}{template_mapped}_{dates[0]}.doc")
        if not os.path.exists(src_template):
            print(f"[ERROR] 템플릿 파일 없음: {src_template}")
            continue

        for i in range(1, total+1):
            orig    = f"{prefix}_{i:02d}"
            mapped  = name_change_map.get(orig, orig)
            grp_dir = os.path.join(REPORT_DIR, f"{지역}{mapped}")
            if os.path.isdir(grp_dir):
                continue

            os.makedirs(grp_dir, exist_ok=True)
            for idx, date in enumerate(dates, start=1):
                dst = os.path.join(grp_dir, f"{지역}{mapped}_{date}.doc")
                if os.path.exists(dst):
                    continue
                shutil.copy(src_template, dst)

                with open(dst, 'r', encoding='cp949', errors='ignore') as f:
                    soup = BeautifulSoup(f, 'html.parser')
                td0 = soup.find('td', string='수업일시')
                if td0:
                    td1 = td0.find_next_sibling('td')
                    raw = td1.get_text()
                    dt = datetime.datetime.strptime(date, "%Y-%m-%d")
                    date_str = dt.strftime("%Y.%m.%d")
                    dow      = dt.strftime("%a").upper()
                    new_prefix = f"{date_str}({dow})"
                    new_text = re.sub(r"\d{4}\.\d{2}\.\d{2}\([A-Z]{3}\)", new_prefix, raw)
                    td1.clear()
                    td1.append(new_text)
                with open(dst, 'w', encoding='cp949', errors='ignore') as f:
                    f.write(str(soup))

                process_html(dst,
                             suffix_change=(template_orig.split('_')[1], orig.split('_')[1]),
                             clear_fields=('출결현황', '특이사항'),
                             education_idx=idx)

    print("▶︎ 누락된 반 스텁 생성 완료")


# ——————————————————————————————
# 추가인원 삽입
# ——————————————————————————————
def add_additional_members():
    for mapped, names in additional_map.items():
        grp_dir = os.path.join(REPORT_DIR, f"{지역}{mapped}")
        if not os.path.isdir(grp_dir):
            continue
        for fn in os.listdir(grp_dir):
            if fn.endswith('.doc'):
                path = os.path.join(grp_dir, fn)
                process_html(path, additions=names)


# ——————————————————————————————
# 메인
# ——————————————————————————————
if __name__ == '__main__':
    초기화()
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page    = browser.new_context(accept_downloads=True).new_page()
        로그인(page)
        보고서다운(page)
        browser.close()
    create_stubs()
    add_additional_members()
