import os
import openpyxl
import random
from playwright.sync_api import sync_playwright
import time

# 숫자 → 옵션 인덱스 매핑 (0부터 시작)
DIGIT_TO_INDEX = {
    '1': 0,
    '2': 1,
    '3': 2,
    '4': 3
}

# "1.", "2.", ..., "15." 형태의 문항 제목인지 확인
def is_problem_heading(text: str) -> bool:
    text = text.strip()
    return any(text.startswith(f"{n}.") for n in range(1, 16))

def main():
    # 원하는 평균 점수(target_score)를 세팅
    target_score = 97
    # 오답 확률 계산
    error_rate = (100 - target_score) / 100  # 0.02

    # 1) 엑셀 로드
    base_dir = os.path.dirname(os.path.abspath(__file__))
    wb_path = os.path.join(base_dir, '학습성취도평가.xlsx')
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    ws = wb.active

    # 2) 링크 매핑
    level_map = {'초저': '초저링크', '초고': '초고링크', '중등': '중등링크'}
    link_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header in level_map.values():
            url = ws.cell(row=1, column=col + 1).value
            lvl = next(k for k, v in level_map.items() if v == header)
            link_map[lvl] = url

    # 3) 정답 행 매핑
    ans_row_for = {'초저': 4, '초고': 5, '중등': 6}

    # 4) Playwright 브라우저 한 번만 띄우기
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        page = browser.new_page()

        # 5) 학생별 반복
        for i, (name, school, grade) in enumerate(ws.iter_rows(min_row=2, max_col=3, values_only=True)):
            # 5-1) 레벨 결정
            if school and '초등학교' in school:
                if grade in (3, 4):
                    lvl = '초저'
                elif grade in (5, 6):
                    lvl = '초고'
                else:
                    lvl = None
            elif school and '중학교' in school:
                lvl = '중등'
            else:
                lvl = None
            if not lvl or lvl not in link_map:
                continue

            url = link_map[lvl]
            raw = ws.cell(row=ans_row_for[lvl], column=16).value
            answer_str = str(raw).strip() if raw else ''

            # 5-2) 페이지 열기 (기존 페이지에서 네비게이트)
            page.goto(url)

            # 5-3) 이름/학교/학년 입력
            page.fill('input[jsname="YPqjbf"]', str(name))
            page.click(f'//div[@role="radio" and @data-value="{school}"]')
            page.click(f'//div[@role="radio" and @data-value="{grade}"]')

            # 5-4) 문제 그룹 필터링
            containers = page.locator('div[jsmodel="CP1oW"]')
            problem_groups = []
            for idx in range(containers.count()):
                heading = containers.nth(idx).locator('.M7eMe').inner_text()
                if is_problem_heading(heading):
                    problem_groups.append(containers.nth(idx).locator('div[role="radiogroup"]'))

            # 5-5) 답안 클릭 (확률적으로 오답)
            for q_idx, group in enumerate(problem_groups):
                if q_idx >= len(answer_str):
                    break
                ch = answer_str[q_idx]
                correct = DIGIT_TO_INDEX.get(ch)
                if correct is None:
                    print(f"  [경고] 문항{q_idx+1} 답 '{ch}' 인식 불가")
                    continue
                # 틀릴 확률 적용
                if random.random() < error_rate:
                    choices = [0, 1, 2, 3]
                    choices.remove(correct)
                    pick = random.choice(choices)
                else:
                    pick = correct
                group.locator('div[role="radio"]').nth(pick).click()

            # 5-6) 제출
            page.click('div[jsname="M2UYVd"][aria-label="Submit"]')
            time.sleep(0.5)

            print(f"{name} 제출 완료")

        browser.close()
    wb.close()

if __name__ == '__main__':
    main()
