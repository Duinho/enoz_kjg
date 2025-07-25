import os
import time
import random
import openpyxl
from playwright.sync_api import sync_playwright

def choose_by_prob(probs):
    """
    probs: [p1, p2, p3, p4, p5] (합이 1이 아니어도 내부에서 정규화)
    1~5 중 하나를 확률에 따라 반환
    """
    total = sum(probs)
    if total == 0:
        # 모두 0 이면 무작위 동등 선택
        return random.randint(1, 5)
    # 정규화
    norm = [p / total for p in probs]
    r = random.random()
    cum = 0.0
    for idx, p in enumerate(norm, start=1):
        cum += p
        if r < cum:
            return idx
    return 5  # 혹시 r 가 극단적으로 1.0 에 가까울 때

def main():
    # 1) 엑셀 로드
    base_dir = os.path.dirname(os.path.abspath(__file__))
    wb_path  = os.path.join(base_dir, '만족도조사.xlsx')
    wb       = openpyxl.load_workbook(wb_path, data_only=True)
    ws       = wb.active

    # 2) P1~P4 읽기
    form_url    = ws['P1'].value
    repeat      = int(ws['P2'].value or 1)
    session1_q  = int(ws['P3'].value or 0)
    session2_q  = int(ws['P4'].value or 0)

    # 3) P5~P9 읽어서 확률 리스트 구성
    raw_probs = [
        float(ws['P5'].value or 0),
        float(ws['P6'].value or 0),
        float(ws['P7'].value or 0),
        float(ws['P8'].value or 0),
        float(ws['P9'].value or 0),
    ]

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False)
        page    = browser.new_page()

        for _ in range(repeat):
            # ─── (A) 폼 열고 첫 번째 “다음” 클릭
            page.goto(form_url)
            page.get_by_role("button", name="다음").click()
            time.sleep(0.1)

            # ─── (B) 1세션 문항 선택
            page.wait_for_selector('div[jsname="skjTt"]', timeout=10000)
            q1 = page.locator('div[jsname="skjTt"]')
            for i in range(session1_q):
                choice = choose_by_prob(raw_probs)
                q1.nth(i).locator(f'div[role="radio"][data-value="{choice}"]').click()
                time.sleep(0.1)

            # ─── (C) 다시 “다음” 클릭
            page.get_by_role("button", name="다음").click()
            time.sleep(0.1)

            # ─── (D) 2세션 문항 선택
            page.wait_for_selector('div[jscontroller="FYWcYb"]', timeout=10000)
            q2 = page.locator('div[jscontroller="FYWcYb"] div[jsname="cnAzRb"]')
            for i in range(session2_q):
                choice = choose_by_prob(raw_probs)
                q2.nth(i).locator(f'div[role="radio"][data-value="{choice}"]').click()
                time.sleep(0.1)

            # ─── (E) “제출” 클릭
            # XPath 로 “제출” 텍스트가 있는 span 의 상위 버튼을 클릭
            submit_button = '//span[text()="제출"]/ancestor::div[@role="button"]'
            page.wait_for_selector(submit_button, timeout=10000)
            page.click(submit_button)
            time.sleep(0.1)

        browser.close()
    wb.close()

if __name__ == '__main__':
    main()
