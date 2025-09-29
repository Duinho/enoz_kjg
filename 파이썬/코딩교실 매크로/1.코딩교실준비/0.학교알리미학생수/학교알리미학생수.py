# 라이브러리 설치를 위해 터미널에 아래의 pip 명령어 입력
# 먼저 pip install openpyxl playwright 입력
# 그 다음 playwright install 입력
# 위 터미널 창에 붙여넣기하여 라이브러리 설치
import os
import openpyxl
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
import time

폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 폴더경로를 코드가 있는 디렉토리로 저장
excel_file_path = os.path.join(폴더경로, '학교알리미학생수.xlsx')  # 엑셀 파일 경로

def 엑셀(): # 엑셀을 열어 정보를 가져오는 함수
    global 학교알리미주소, 학교, 시도, 시군구, 학년 
    wb = openpyxl.load_workbook(excel_file_path)    # 엑셀파일 로드
    ws = wb['정보']                                 # 엑셀파일 정보를 ws에 저장
    학교알리미주소 = ws.cell(row=1, column=17).value # 각 행과열에서 필요한 정보 변수에 저장
    학교 = ws.cell(row=2, column=17).value
    시도 = ws.cell(row=3, column=17).value
    시군구 = ws.cell(row=4, column=17).value
    if "초등학교" in 학교:                           # 분류가 초등학교면 1~6학년 그게 아니면 (중,고) 1~3학년
        학년 = ['1', '2', '3', '4', '5', '6']
    else:
        학년 = ['1', '2', '3']
    wb.close() # 엑셀파일 닫기
    return ws  # ws값 리턴

def 헤더추가(ws): # 엑셀에 제일 위에 헤더를 미리 추가하는 함수
    headers = ['학교명', '1학년', '2학년', '3학년', '4학년', '5학년', '6학년']
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

def 학교받아오기(page):       
    # 엑셀에 저장한 학교알리미 주소로 접속
    page.goto(학교알리미주소)

    # level1: 학교급 (초/중/고)
    radio_buttons = page.query_selector_all('input[name="level1"]')
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 학교 in label.inner_text():
            label.click()
            page.wait_for_load_state('load')
            break

    # level2: 시·도
    radio_buttons = page.query_selector_all('input[name="level2"]')
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 시도 in label.inner_text():
            label.click()
            page.wait_for_load_state('load')
            break

    # level3: 시·군·구
    radio_buttons = page.query_selector_all('input[name="level3"]')
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 시군구 in label.inner_text():
            label.click()
            page.wait_for_load_state('load')
            break

    # 학교 목록 수집
    학교명 = []
    labels = page.query_selector_all('label[for^="shlCd"]')
    for label in labels:
        학교명.append(label.inner_text())

    return 학교명

def 학교정보검색(page, 학교명, 학년, ws):  
    row_index = 2  # 첫 번째 줄은 헤더이므로 두 번째 줄부터 시작

    for 학교 in 학교명:  
        label = page.query_selector(f'label:has-text("{학교}")')  
        if label:
            # 학교 클릭 및 이동
            label.click()
            page.wait_for_load_state('load') 
            page.click('a[data-tab-id="tabSel"]')
            page.wait_for_load_state('load')  
            page.click('a.accordian_title:has-text("학생현황")') 
            page.wait_for_load_state('load')                    
            page.click('label[for="hangmok01"]')   # 학생현황 선택
            page.wait_for_load_state('load')   

            # 검색 버튼 클릭
            page.click('#webSearchButton')                  
            page.wait_for_load_state('load')

            # "입력된 데이터가 없습니다." 체크
            no_data = page.query_selector('p:has-text("입력된 데이터가 없습니다.")')
            if no_data:
                print(f'{학교} → 데이터 없음, 건너뜀')
                # 다음 학교로 이동
                page.click('a.slidedown[href="javascript:research();"]') 
                page.wait_for_load_state('load')
                continue  # 이 학교는 패스

            # 데이터가 있는 경우에만 처리
            page.wait_for_selector('xpath=//tr[th[normalize-space(.)="학급수"]]', timeout=10000)
            page.wait_for_selector('xpath=//tr[th[normalize-space(.)="학생수"]]', timeout=10000)

            # 학급수 행
            class_counts = []
            class_row = page.query_selector('xpath=//tr[th[normalize-space(.)="학급수"]]')
            if class_row:
                tds = class_row.query_selector_all('td')
                for i, grade in enumerate(학년):
                    if i < len(tds):
                        raw_value = tds[i].inner_text().strip()
                        if raw_value in ('', '-'):
                            class_counts.append(0)
                        else:
                            class_counts.append(int(raw_value.split('(')[0].replace(',', '')))
                    else:
                        class_counts.append(0)

            # 학생수 행
            values = []
            student_row = page.query_selector('xpath=//tr[th[normalize-space(.)="학생수"]]')
            if student_row:
                tds = student_row.query_selector_all('td')
                for i, grade in enumerate(학년):
                    if i < len(tds):
                        raw_value = tds[i].inner_text().strip()
                        if raw_value in ('', '-'):
                            values.append(0)
                        else:
                            values.append(int(raw_value.replace(',', '').split('(')[0]))
                    else:
                        values.append(0)

            # ----------------------------
            # 엑셀 저장
            # ----------------------------
            ws.cell(row=row_index, column=1, value=학교) 
            for col_index, value in enumerate(values, start=2):
                ws.cell(row=row_index, column=col_index, value=value)  # 학생 수
            for col_index, class_count in enumerate(class_counts, start=8):  
                ws.cell(row=row_index, column=col_index, value=class_count)  # 반 수

            row_index += 1  
            print(f'{학교} 학생수 및 반 수 저장 완료')

            # 다음 학교로 이동
            page.click('a.slidedown[href="javascript:research();"]') 
            page.wait_for_load_state('load')



def 동작():
    global browser, page  # 전역 변수로 사용
    with sync_playwright() as p:  # 크로미움 브라우저 열기
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()           # 새로운 창이 열리면 page에 저장
        ws = 엑셀()                         # 엑셀을 통해 ws에 엑셀값 저장
        헤더추가(ws)                        # ws에 헤더추가하여 저장
        학교명 = 학교받아오기(page)          # 학교명에 학교받아오기로 받아온 학교 정보를 저장
        학교정보검색(page, 학교명, 학년, ws) # 각 정보를 통해 
        ws.parent.save(excel_file_path)    # 동작이 끝나면 엑셀 파일 저장
        print('모든 학교 저장 완료')
        browser.close()  # 브라우저 닫기

동작()