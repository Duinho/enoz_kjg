# 라이브러리 설치 메모(터미널):
# pip install openpyxl playwright
# playwright install

import os
import openpyxl
from openpyxl import Workbook
from playwright.sync_api import sync_playwright
import random
import time
from datetime import datetime
from collections import deque

# --- 경로 설정 --------------------------------------------------------------
폴더경로 = os.path.dirname(os.path.abspath(__file__))
사용자_데이터_디렉토리 = r"C:\play\UserData"   # <= 이렇게 바꿔주세요 (직접 생성 안 해도 됩니다)
excel_file_path = os.path.join(폴더경로, '추가인원 가입 수강신청.xlsx')

# --- 전역 상태 --------------------------------------------------------------
클릭 = 0
학교_리스트    = []
아파트_리스트  = []
동_리스트      = []
아이디1_리스트 = []
아이디2_리스트 = []
아이디3_리스트 = []
남아_리스트    = []
여아_리스트    = []
부_리스트      = []   # 아버지 이름 풀(‘이름’ 부분만)
모_리스트      = []   # 어머니 이름 풀(‘이름’ 부분만)

# 이름 분산용 큐/카운트(히스토리 포함)
남아_큐 = deque()
여아_큐 = deque()
부_큐  = deque()
모_큐  = deque()

이름_카운트 = {}      # 자녀 이름 사용 누적(성 제외)
부모_이름_카운트 = {}  # 부모 이름 사용 누적(성 제외, 부/모 통합 카운트)

# --- 성씨 및 확률 ----------------------------------------------------------
성 = [
    ("김", 0.21782), ("이", 0.14882), ("박", 0.08536), ("최", 0.04754), ("정", 0.04402),
    ("강", 0.02355), ("조", 0.02161), ("윤", 0.02077), ("장", 0.02017), ("임", 0.01674),
    ("한", 0.01560), ("오", 0.01548), ("서", 0.01524), ("신", 0.01526), ("권", 0.01427),
    ("황", 0.01412), ("안", 0.01393), ("송", 0.01389), ("류", 0.01300), ("전", 0.01098),
    ("홍", 0.01137), ("고", 0.00957), ("문", 0.00938), ("양", 0.00884), ("손", 0.00919),
    ("배", 0.00812), ("백", 0.00773), ("허", 0.00660), ("유", 0.00566), ("남", 0.00559),
    ("심", 0.00551), ("노", 0.00498), ("정", 0.00442), ("하", 0.00463), ("곽", 0.00410),
    ("성", 0.00405), ("차", 0.00396), ("주", 0.00389), ("우", 0.00390), ("구", 0.00392),
    ("신", 0.00369), ("임", 0.00379), ("라", 0.00377), ("전", 0.00386), ("민", 0.00350),
    ("유", 0.00381), ("진", 0.00315), ("지", 0.00311), ("엄", 0.00291), ("채", 0.00254),
    ("원", 0.00262), ("천", 0.00239), ("방", 0.00191), ("공", 0.00183), ("강", 0.00201),
    ("현", 0.00180), ("함", 0.00164), ("변", 0.00164), ("염", 0.00139), ("양", 0.00181),
    ("변", 0.00121), ("여", 0.00123), ("추", 0.00121), ("노", 0.00134), ("도", 0.00115),
    ("소", 0.00098), ("신", 0.00103), ("석", 0.00100), ("선", 0.00085), ("설", 0.00086),
    ("마", 0.00078), ("길", 0.00075), ("주", 0.00081), ("연", 0.00068), ("방", 0.00068),
    ("위", 0.00062), ("표", 0.00062), ("명", 0.00059), ("기", 0.00056), ("반", 0.00054),
    ("왕", 0.00051), ("금", 0.00051), ("옥", 0.00051), ("육", 0.00047), ("인", 0.00045),
    ("맹", 0.00044), ("제", 0.00044), ("모", 0.00042), ("장", 0.00041), ("남", 0.00042),
    ("탁", 0.00043), ("국", 0.00039), ("여", 0.00039), ("진", 0.00042), ("어", 0.00038),
    ("은", 0.00034), ("편", 0.00033), ("구", 0.00029), ("용", 0.00016), ("남궁", 0.00015)
]
성씨, weights = zip(*성)

# --- 고정 비밀번호 ----------------------------------------------------------
비밀번호 = 'enoz7223!'

# --- 정보 시트에서 불러올 변수들 -------------------------------------------
BG회원가입   = ""
BG수강신청   = ""
PH회원가입   = ""
PH수강신청   = ""
반복횟수     = 0
랜덤최소     = 0
랜덤최대     = 0

# --- 유틸: 복성 처리 --------------------------------------------------------
복성_목록 = {'남궁','황보','제갈','선우','서문','독고','동방','사공','사마','소봉','어금','장곡','탁안'}

def 이름부분만(전체이름: str) -> str:
    """성(1글자 또는 복성 2글자)을 제거한 '이름'만 반환."""
    if not 전체이름:
        return ""
    if len(전체이름) >= 3 and 전체이름[:2] in 복성_목록:
        return 전체이름[2:]
    return 전체이름[1:]

# --- 히스토리 로딩: 이름/부모이름 카운트 -----------------------------------
def 이름_카운트_로드():
    """가입 현황 파일에서 자녀/부모 이름(성 제외) 사용 누적 카운트를 읽어옴."""
    child_counts = {}
    parent_counts = {}

    현황_경로 = os.path.join(폴더경로, '추가인원 가입 현황.xlsx')
    if not os.path.exists(현황_경로):
        return child_counts, parent_counts

    wb = openpyxl.load_workbook(현황_경로, data_only=True)
    for 시트 in wb.sheetnames:
        ws = wb[시트]
        # 예상 열: 이름, 아이디, 나이, 요일, 학교, 주소, (부모이름 - 있을 수 있음), 가입일시
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            # 자녀 이름
            전체이름 = row[0]
            if 전체이름:
                nm = 이름부분만(전체이름)
                if nm:
                    child_counts[nm] = child_counts.get(nm, 0) + 1
            # 부모이름(있을 수도, 없을 수도)
            if len(row) >= 8:
                부모이름 = row[-2]  # 가입일시가 마지막이라 가정 -> 바로 앞이 부모이름
                if 부모이름:
                    pnm = 이름부분만(부모이름)
                    if pnm:
                        parent_counts[pnm] = parent_counts.get(pnm, 0) + 1
    wb.close()
    return child_counts, parent_counts

# --- 큐 초기화 --------------------------------------------------------------
def 큐_초기화():
    """남/여/부/모 이름 큐 섞어서 준비(1회전 중복 없음)."""
    global 남아_큐, 여아_큐, 부_큐, 모_큐
    남 = [x for x in 남아_리스트 if x]
    여 = [x for x in 여아_리스트 if x]
    부 = [x for x in 부_리스트 if x]
    모 = [x for x in 모_리스트 if x]
    random.shuffle(남)
    random.shuffle(여)
    random.shuffle(부)
    random.shuffle(모)
    남아_큐 = deque(남)
    여아_큐 = deque(여)
    부_큐  = deque(부)
    모_큐  = deque(모)

# --- 분산 추출 --------------------------------------------------------------
def 분산_선택(데크: deque, 카운트_딕트: dict) -> str:
    """
    데크에 들어있는 후보들 중 '현재 사용 카운트가 최소'인 집합에서 랜덤 1개 선택.
    선택된 값은 데크에서 1회 제거, 카운트 +1.
    데크가 비면 재초기화는 호출부에서 책임짐.
    """
    if not 데크:
        return ""  # 호출부에서 재초기화하도록
    풀 = list(데크)
    최소 = min(카운트_딕트.get(n, 0) for n in 풀)
    후보 = [n for n in 풀 if 카운트_딕트.get(n, 0) == 최소]
    선택 = random.choice(후보)
    데크.remove(선택)
    카운트_딕트[선택] = 카운트_딕트.get(선택, 0) + 1
    return 선택

# --- 초기화: 정보 시트 읽기 -------------------------------------------------
def 초기화():
    """'정보' 시트에서 반복/랜덤값과 이름 풀을 읽고, 카운트/큐 준비."""
    global BG회원가입, BG수강신청, PH회원가입, PH수강신청
    global 반복횟수, 랜덤최소, 랜덤최대
    global 아이디1_리스트, 아이디2_리스트, 아이디3_리스트
    global 남아_리스트, 여아_리스트, 부_리스트, 모_리스트
    global 이름_카운트, 부모_이름_카운트

    아이디1_리스트.clear()
    아이디2_리스트.clear()
    아이디3_리스트.clear()
    남아_리스트.clear()
    여아_리스트.clear()
    부_리스트.clear()
    모_리스트.clear()

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb["정보"]

    반복횟수   = ws.cell(row=1, column=20).value
    랜덤최소   = ws.cell(row=2, column=20).value
    랜덤최대   = ws.cell(row=3, column=20).value

    아이디마지막행 = 92
    자녀마지막행   = 501
    부모마지막행   = 78

    for row in ws.iter_rows(min_row=2, max_row=아이디마지막행, values_only=True):
        아이디1_리스트.append(row[5])
        아이디2_리스트.append(row[6])
        아이디3_리스트.append(row[7])

    for row in ws.iter_rows(min_row=2, max_row=자녀마지막행, values_only=True):
        남아_리스트.append(row[10])
        여아_리스트.append(row[11])

    for row in ws.iter_rows(min_row=2, max_row=부모마지막행, values_only=True):
        부_리스트.append(row[14])
        모_리스트.append(row[15])
    wb.close()

    # 히스토리 카운트
    이름_카운트, 부모_이름_카운트 = 이름_카운트_로드()
    # 큐 준비
    큐_초기화()

# --- 시트별 주소 목록 로더 ---------------------------------------------------
def load_lists(sheet_name):
    global 학교_리스트, 아파트_리스트, 동_리스트
    global 회원가입사이트, 수강신청사이트

    학교_리스트.clear()
    아파트_리스트.clear()
    동_리스트.clear()

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb[sheet_name]

    학교마지막행 = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=학교마지막행, values_only=True):
        학교_리스트.append(row[0])
        아파트_리스트.append(row[1])
        동_리스트.append(row[2])

    회원가입사이트 = ws.cell(row=1, column=20).value
    수강신청사이트 = ws.cell(row=2, column=20).value
    wb.close()

# --- 현황 파일 저장 ---------------------------------------------------------
def 엑셀_초기화_및_데이터_저장(sheet_name, 이름, 아이디, 나이, 요일, 학교, 주소, 부모이름):
    """
    '추가인원 가입 현황.xlsx'에 sheet_name(PH 또는 BG) 시트를 만들어(또는 기존 시트 사용)
    가입 기록 저장. '부모이름' 컬럼 추가. 마지막 컬럼은 가입일시 유지.
    """
    현황_경로 = os.path.join(폴더경로, '추가인원 가입 현황.xlsx')
    if not os.path.exists(현황_경로):
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = sheet_name
        ws2.append(["이름","아이디","나이","요일","학교","주소","부모이름","가입일시"])
    else:
        wb2 = openpyxl.load_workbook(현황_경로)
        if sheet_name in wb2.sheetnames:
            ws2 = wb2[sheet_name]
        else:
            ws2 = wb2.create_sheet(title=sheet_name)
            ws2.append(["이름","아이디","나이","요일","학교","주소","부모이름","가입일시"])

    ws2.append([
        이름,
        아이디,
        나이,
        요일,
        학교,
        주소,
        부모이름,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ])
    wb2.save(현황_경로)
    wb2.close()

# --- 팝업 처리 --------------------------------------------------------------
def handle_alert(dialog):
    global 클릭
    time.sleep(1)
    클릭 = 1
    dialog.accept()

# --- 회원가입 ---------------------------------------------------------------
def 회원가입(page, sheet_name):
    global 이름, 요일

    load_lists(sheet_name)
    page.goto(회원가입사이트)
    time.sleep(2)
    page.keyboard.press('F5')
    time.sleep(3)
    if page.locator('a[href="/Account/LoginProc"]:has-text("LOGOUT")').count() > 0:
        page.locator('a[href="/Account/LoginProc"]').click()
        time.sleep(1)
        page.goto(회원가입사이트)

    page.wait_for_load_state('networkidle')
    page.locator("#cbAgree1").click(force=True)
    page.keyboard.press("End")
    page.locator("#cbAgree2").click(force=True)
    page.click(".btn_type4.c1:has-text('다음')")
    page.wait_for_selector("input[name='tbMemName'].type2.w1.IsKor")

    # --- 이름 & 성별(분산 추출) ---
    성한자 = random.choices(성씨, weights=weights, k=1)[0]
    아이성 = 성한자
    성별값 = random.choice(['남자','여자'])
    page.select_option('select[name="ddlSex"]', value=성별값)

    # 자녀 이름 선택(히스토리 기반 분산)
    global 남아_큐, 여아_큐, 이름_카운트
    if 성별값 == '남자':
        if not 남아_큐:
            큐_초기화()
        이름후 = 분산_선택(남아_큐, 이름_카운트)
        if not 이름후:  # 안전장치
            이름후 = random.choice([x for x in 남아_리스트 if x])
    else:
        if not 여아_큐:
            큐_초기화()
        이름후 = 분산_선택(여아_큐, 이름_카운트)
        if not 이름후:
            이름후 = random.choice([x for x in 여아_리스트 if x])

    이름 = 성한자 + 이름후
    page.fill('input[name="tbMemName"]', 이름)

    # --- 아이디 생성 ---
    랜덤 = random.randint(0, 2)
    if 랜덤 == 0:
        아이디 = 아이디1_리스트[random.randint(0, len(아이디1_리스트)-1)]
    elif 랜덤 == 1:
        아이디 = 아이디2_리스트[random.randint(0, len(아이디2_리스트)-1)]
    else:
        아이디 = 아이디3_리스트[random.randint(0, len(아이디3_리스트)-1)]
    랜덤 = random.randint(0, 11)
    if 랜덤 < 3:
        아이디 += 아이디1_리스트[random.randint(0, len(아이디1_리스트)-1)]
    elif 랜덤 < 6:
        아이디 += 아이디2_리스트[random.randint(0, len(아이디2_리스트)-1)]
    elif 랜덤 < 9:
        아이디 += 아이디3_리스트[random.randint(0, len(아이디3_리스트)-1)]
    else:
        아이디 += str(random.randint(0, 9))
    if random.choice([True, False]):
        아이디 += str(random.randint(0, 9))
    if len(아이디) > 15:
        아이디 = 아이디[:15]
    page.fill('input[name="tbMemID"]', 아이디)
    page.locator("text=중복확인").click()
    page.fill('input[name="tbMemPass"]', 비밀번호)
    page.fill('input[name="tbMemPass2"]', 비밀번호)

    # --- 전화번호 ---
    page.fill('input[name="tbMobile1"]', '010')
    중간번호 = ''.join(str(random.randint(0,9)) for _ in range(4))
    끝번호   = ''.join(str(random.randint(0,9)) for _ in range(4))
    page.fill('input[name="tbMobile2"]', 중간번호)
    page.fill('input[name="tbMobile3"]', 끝번호)

    # --- 생년월일/나이 ---
    나이     = random.randint(11, 13)
    출생년도 = 2026 - 나이
    page.select_option('select[name="ddlBirthDay1"]', value=str(출생년도))
    page.select_option('select[name="ddlBirthDay2"]', value=f"{random.randint(1,12):02}")
    page.select_option('select[name="ddlBirthDay3"]', value=f"{random.randint(1,28):02}")

    # --- 학교 & 주소 ---
    학교선택 = random.randint(0, len(학교_리스트)-1)
    학교     = 학교_리스트[학교선택]
    page.select_option('select[name="ddlAcaName1"]', value=학교)
    page.fill('input[name="tbGrade1"]', str(나이 - 7))
    page.fill('input[name="tbClass1"]', str(random.randint(1,5)))
    page.click("span:text('우편번호 찾기')")
    time.sleep(2)
    page.keyboard.type(아파트_리스트[학교선택])
    time.sleep(1)
    page.keyboard.press('Enter')
    time.sleep(1)
    for _ in range(2):
        page.keyboard.press('Tab')
    page.keyboard.press('Enter')
    time.sleep(1)
    호   = f"{random.randint(1,5)}0{random.randint(1,9)}"
    주소 = f"{동_리스트[학교선택]} {호}호"
    page.fill('input[name="tbAddr2"]', 주소)
    주소 = f"{아파트_리스트[학교선택]} {주소}"

    # --- 부모 정보(분산 추출) ---
    global 부_큐, 모_큐, 부모_이름_카운트
    부모성별 = random.choice(['부','모'])
    if 부모성별 == '부':
        # 아버지는 보통 자녀 성과 동일하게 구성
        if not 부_큐:
            큐_초기화()
        부이름후 = 분산_선택(부_큐, 부모_이름_카운트)
        if not 부이름후:
            부이름후 = random.choice([x for x in 부_리스트 if x])
        부모이름 = f"{아이성}{부이름후}"
    else:
        # 어머니는 무작위 성씨 + 모 이름
        if not 모_큐:
            큐_초기화()
        모이름후 = 분산_선택(모_큐, 부모_이름_카운트)
        if not 모이름후:
            모이름후 = random.choice([x for x in 모_리스트 if x])
        부모이름 = f"{random.choices(성씨, weights=weights, k=1)[0]}{모이름후}"

    page.fill('input[name="tbPName"]', 부모이름)
    page.fill('input[name="tbPMobile1"]', '010')
    page.fill('input[name="tbPMobile2"]', 중간번호)
    page.fill('input[name="tbPMobile3"]', 끝번호)

    # --- 이메일 ---
    포털 = ["@gmail.com","@naver.com","@daum.net","@nate.com"]
    page.fill('input[name="tbMemEmail"]', f"{아이디}{random.choice(포털)}")

    page.click("#rdRoute1")
    time.sleep(1)
    page.wait_for_selector("span:text('회원가입')", state="visible")
    page.once('dialog', handle_alert)
    page.click("span:text('회원가입')")
    for _ in range(3):
        time.sleep(1)
        page.keyboard.press('Enter')

    # 수강 요일 레이블
    wc = random.randint(0,1)
    요일 = "월수" if wc == 0 else "화목"

    # 기록 저장(부모이름 포함)
    엑셀_초기화_및_데이터_저장(sheet_name, 이름, 아이디, 나이, 요일, 학교, 주소, 부모이름)

# --- 수강신청 ---------------------------------------------------------------
def 신청(page, sheet_name):
    page.goto(수강신청사이트)
    time.sleep(2)
    if page.locator('input#q1_2').count() > 0:
        page.locator('input#q1_2').click()
        time.sleep(1)

    if sheet_name == "BG":
        if 요일 == "월수":
            page.locator('label.cc-cc.check[for="3_23"]').click()
        else:
            page.locator('label.cc-cc.check[for="4_24"]').click()
    elif sheet_name == "PH":
        if 요일 == "월수":
            page.locator('button.btn_week[data-week="2,4"]').click()
            time.sleep(1)
            page.locator('label.cc-cc.check[for="1_53"]').click()
        else:
            page.locator('button.btn_week[data-week="3,5"]').click()
            time.sleep(1)
            page.locator('label.cc-cc.check[for="2_54"]').click()
        time.sleep(1)

    page.once('dialog', handle_alert)
    page.locator('a.btn_type5.mb30:has-text("수강신청")').click()
    time.sleep(1)
    page.keyboard.press('Enter')
    time.sleep(1)

    page.wait_for_selector('input#btn_poll', timeout=10000)
    poll_buttons = page.locator('input[id^="q"][id$="_1"]')
    total = poll_buttons.count()
    for idx in range(total):
        poll_buttons.nth(idx).click()

    page.click('input#btn_poll')
    time.sleep(1)
    page.click('input#cbAddress')
    time.sleep(1)
    page.click('input#cbPc')
    time.sleep(1)
    page.once('dialog', handle_alert)
    page.click('a:has(span:text("신청"))')
    time.sleep(1)
    page.keyboard.press('Enter')
    time.sleep(1)

    현황_경로 = os.path.join(폴더경로, '추가인원 가입 현황.xlsx')
    wb2 = openpyxl.load_workbook(현황_경로, data_only=True)
    ws2 = wb2[sheet_name]
    total = ws2.max_row - 1
    today_str = datetime.now().strftime('%Y-%m-%d')
    today_count = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if str(row[-1]).startswith(today_str):  # 가입일시가 마지막 열
            today_count += 1
    wb2.close()
    print(f"{이름} 신청 완료 / 총 {total}명, 오늘 {today_count}명 추가되었습니다.")
    time.sleep(random.randint(랜덤최소, 랜덤최대))

# --- 전체 동작 --------------------------------------------------------------
def 동작():
    if not os.path.exists(사용자_데이터_디렉토리):
        os.makedirs(사용자_데이터_디렉토리)

    초기화()
    with sync_playwright() as playwright:
        context = playwright.firefox.launch_persistent_context(
            user_data_dir=사용자_데이터_디렉토리,
            headless=False,
            args=['--disable-popup-blocking']
        )
        page = context.new_page()

        # BG 시트 반복 처리(필요 시 PH로 변경/추가)
        for _ in range(반복횟수):
            회원가입(page, "BG")
            신청(page, "BG")

        time.sleep(5)
        context.close()

if __name__ == "__main__":
    동작()
