# 라이브러리 설치를 위해 터미널에 아래의 pip 명령어 입력
# 먼저 playwright install 입력
# 그 다음 pip install openpyxl playwright 입력
# 위 터미널 창에 붙여넣기하여 라이브러리 설치

import os
import openpyxl
from playwright.sync_api import sync_playwright
import random
import time

폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '기촌 추가인원 수강신청.xlsx')  # 엑셀 파일 경로
사용자_데이터_디렉토리 = os.path.join(폴더경로, 'UserData')

클릭 = 0
아이디_리스트 = []
비밀번호 = 'enoz7223'

"""엑셀 파일 값 받아오기 함수 시작 """

def 초기화():
    global wb, ws, 사이트, 로그인사이트, 수강신청사이트, 랜덤최소, 랜덤최대, 아이디_리스트, 사전역량조사
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    사이트 = ws.cell(row=1, column=20).value
    로그인사이트 = ws.cell(row=2, column=20).value
    수강신청사이트 = ws.cell(row=3, column=20).value
    랜덤최소 = ws.cell(row=4, column=20).value
    랜덤최대 = ws.cell(row=5, column=20).value
    사전역량조사 = ws.cell(row=6, column=20).value
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5, values_only=True):
        아이디_리스트.append({'아이디': row[0], '이름': row[1], '요일': row[2], '학년': row[4]})

"""엑셀 파일 값 받아오기 함수 끝"""

"""수강신청 함수 시작"""
def 신청(page, 아이디, 이름, 요일, 학년):
    page.goto(사이트, wait_until='networkidle')
    logout_link = page.locator('a[href="/Account/LoginProc"]:has-text("LOGOUT")') # 로그인이 되어있는 상태라면
    if logout_link.count() > 0:                                                          
        logout_link.click()                                                       # 로그아웃 클릭
        time.sleep(1)
    page.goto(로그인사이트, wait_until='networkidle')                               # 로그인
    page.fill('input[name="tbID"]', 아이디)                                        
    page.fill('input[name="tbPass"]', 비밀번호)                                    
    time.sleep(0.3)
    page.locator('a.btn_type2').click()
    page.goto(수강신청사이트, wait_until='networkidle')
    #page.locator('label[for="1"]').click()
    # 요일에 따른 동작
    if 요일 == '화목':                                                             # 학생이 화목이면 화목클릭 
        page.locator('button.btn_week[data-week="3,5"]').click()                  # 이 부분과 아래의 선택자가 다를 수 있음. 사이트에서 관리자 모드로 화목 선택자와 버튼을 활성화 하는 선택자로 수정해야함. 챗지피티 참고
        time.sleep(2)
        page.locator('label.cc-cc.check[for="2_48"]').click()
    else:
        page.locator('button.btn_week[data-week="2,4"]').click()                   # 확생이 월수면 월수 클릭
        time.sleep(2)
        page.locator('label.cc-cc.check[for="1_47"]').click()
    
    time.sleep(2)
    page.once('dialog', handle_alert)
    page.locator('a.btn_type5.mb30:has-text("수강신청")').click()
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    page.wait_for_selector('input#btn_poll', timeout=10000) 
    page.locator('input#q1_1').click()                               
    page.locator('input#q2_1').click()
    page.locator('input#q3_1').click()
    page.locator('input#q4_1').click()
    page.locator('input#q5_1').click()
    if 사전역량조사 > 5 :
        page.locator('input#q6_1').click()
        page.locator('input#q7_1').click()
    page.click('input#btn_poll')
    time.sleep(2)
    page.click('input#cbAddress')
    time.sleep(2)
    page.click('input#cbPc')
    time.sleep(2)
    page.once('dialog', handle_alert)
    page.click('a:has(span:text("신청"))')
    time.sleep(2)
    page.keyboard.press('Enter')
    time.sleep(2)
    page.keyboard.press('Enter')
    print(f"{이름} 신청 완료")
    엑셀_초기화_및_데이터_저장(이름, 아이디, 학년)
    time.sleep(random.randint(랜덤최소, 랜덤최대))   # 엑셀에 적어둔 랜덤최소와 랜덤최대에 맞춰 다음 가입 및 신청 대기
"""수강신청 함수 끝"""


def handle_alert(dialog):
    global 클릭
    time.sleep(3)
    클릭 = 1
    dialog.accept()


def 엑셀_초기화_및_데이터_저장(이름, 아이디, 학년):   # 추가인원을 수강 신청한 현황을 저장하는 함수
    파일_경로 = os.path.join(폴더경로, '추가인원 신청 현황.xlsx')
    if not os.path.exists(파일_경로):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "신청 정보"
        ws['A1'] = "이름"
        ws['B1'] = "아이디"
        ws['C1'] = "학년"
    else:
        wb = openpyxl.load_workbook(파일_경로)
        ws = wb.active
    
    new_row = ws.max_row + 1  # 마지막행의 한 칸 아래에 각 정보를 추가한 후 저장
    ws[f'A{new_row}'] = 이름
    ws[f'B{new_row}'] = 아이디
    ws[f'C{new_row}'] = 학년
    wb.save(파일_경로)
    wb.close()

def 동작():
    if not os.path.exists(사용자_데이터_디렉토리):
        os.makedirs(사용자_데이터_디렉토리)
    with sync_playwright() as playwright:
        # 파이어폭스를 사용하여 launch_persistent_context 호출
        context = playwright.firefox.launch_persistent_context(
            user_data_dir=사용자_데이터_디렉토리,
            headless=False,
            args=['--disable-popup-blocking']
        )
        page = context.new_page()
        초기화()
        for 사용자 in 아이디_리스트:
            신청(page, 사용자['아이디'], 사용자['이름'], 사용자['요일'], 사용자['학년'])
        time.sleep(5)
        context.close()

동작()
