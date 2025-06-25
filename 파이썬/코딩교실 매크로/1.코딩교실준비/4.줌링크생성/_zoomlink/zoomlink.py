import os
import openpyxl
import time
from playwright.sync_api import sync_playwright
import pyautogui
from pyautogui import PyAutoGUIException
# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '줌링크생성.xlsx')  # 엑셀 파일 경로

def 초기화():
    global wb, ws
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

def 엑셀():
    global 로그인사이트,ID,PW,사용자관리링크,사업명,이메일,호스트키,줌링크
    로그인사이트 = ws.cell(row=1, column=17).value
    사용자관리링크 = ws.cell(row=2, column=17).value
    ID = ws.cell(row=3, column=17).value
    PW = ws.cell(row=4, column=17).value
    사업명 = ws.cell(row=5, column=17).value
    이메일 = []
    호스트키 = []
    줌링크 = []
    row = 2  # 1행은 헤더이므로 2행부터 시작
    while True:
        email = ws.cell(row=row, column=3).value  # C열은 column=3
        if email is None:
            break
        이메일.append(email)
        row += 1

def 로그인(page):
    # 로그인 프로세스
    page.goto(로그인사이트)
    page.fill('input#email', ID)
    page.click('span.zm-button__slot:text("Next")')
    page.wait_for_selector('input#password')
    page.fill('input#password', PW)
    page.keyboard.press("Enter")
    time.sleep(10)

def click_image_with_offset(
    image_name: str,
    offset: tuple[int, int] = (0, 0),
    timeout: float = 10,
    confidence: float = 0.8
) -> bool:
    # 파일 경로 구하고
    if "__file__" in globals():
        script_dir = os.path.dirname(os.path.abspath(__file__))
    else:
        script_dir = os.getcwd()
    image_path = os.path.join(script_dir, "pic", image_name)
    print(f"Searching for image at: {image_path}")

    # 파일 자체 없으면 바로 False
    if not os.path.exists(image_path):
        print(f"❌ 이미지 파일이 없습니다: {image_path}")
        return False

    end_time = time.time() + timeout
    while time.time() < end_time:
        try:
            box = pyautogui.locateOnScreen(image_path, confidence=confidence)
        except PyAutoGUIException as e:
            # pyscreeze·Pillow 관련 에러는 여기서 잡아서 False 반환
            print(f"[warn] locateOnScreen 에러: {e}")
            return False
        except Exception as e:
            # 그 외 예기치 않은 에러도 안전하게 잡아서 False
            print(f"[warn] locateOnScreen 예외: {e}")
            return False

        if box:
            center_x, center_y = pyautogui.center(box)
            pyautogui.click(center_x + offset[0], center_y + offset[1])
            return True

        time.sleep(0.5)

    # 시간 초과해서 찾지 못해도 False
    return False

def 이메일작업(page, 이메일주소, 순번):
    try:
        page.goto(사용자관리링크)
        page.wait_for_selector('input.zoom-input__inner[placeholder="검색"]')
        page.locator('input.zoom-input__inner[placeholder="검색"]').first.fill(이메일주소)
        page.keyboard.press("Enter")
        page.wait_for_selector('mark[data-markjs="true"]')  # 검색 결과 하이라이트 등장까지 대기
        page.click('mark[data-markjs="true"]')

        page.wait_for_selector('button[data-testid="host-key:show-btn"]')
        page.click('button[data-testid="host-key:show-btn"]')
        page.wait_for_selector('#hostKey-subShowLabel')
        time.sleep(1)
        호스트키.append(page.text_content('#hostKey-subShowLabel').strip())
        page.locator('a.dropdown-toggle > b.caret').nth(1).click()
        page.wait_for_selector('a:has-text("향후 회의")')
        page.click('a:has-text("향후 회의")')
        page.wait_for_selector('a:has-text("회의 예약")')
        page.click('a:has-text("회의 예약")')
        주제 = f"{사업명} {순번 + 1:02}반"
        page.wait_for_selector('input#topic')
        pyautogui.click(960, 300)
        pyautogui.scroll(-1000)
        time.sleep(1)
        click_image_with_offset("1.png", offset=(-30,0), timeout=15, confidence=0.9)
        pyautogui.scroll(-500)
        print("1")
        time.sleep(1)
        click_image_with_offset("2.png", offset=(0,0), timeout=15, confidence=0.9)
        print("1")
        time.sleep(1)
        click_image_with_offset("3.png", offset=(-10,0), timeout=15, confidence=0.9)
        print("1")

        page.fill('input#topic', 주제)
        page.click('span:has-text("되풀이 회의")')
        page.click('span#recurringType')
        page.click('dd#select-item-recurringType-3')       
        page.fill('input.zm-select__input[role="combobox"]', '운영사무국')
        time.sleep(1)
        page.keyboard.press("Enter")
        time.sleep(1)
        page.wait_for_selector('button.save-btn')
        page.click('button.save-btn')

        page.wait_for_selector('a[aria-labelledby="view-registration"]')
        time.sleep(3)
        줌링크.append(page.get_attribute('a[aria-labelledby="view-registration"]', 'href'))

    except Exception as e:
        print(f"[실패] {이메일주소} 작업 중 오류 발생: {e}")
        호스트키.append("오류")
        줌링크.append("오류")


def 엑셀저장():
    for i, (키, 링크) in enumerate(zip(호스트키, 줌링크)):
        ws.cell(row=i+2, column=4).value = 링크   # D열: 줌링크
        ws.cell(row=i+2, column=5).value = 키     # E열: 호스트키
    wb.save(excel_file_path)

def 동작():
    global browser
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()
        초기화()
        엑셀()
        로그인(page)
        for i, 주소 in enumerate(이메일):
            이메일작업(page, 주소, i)
        엑셀저장()
        browser.close()
        print("작업이 완료되었습니다. 엑셀 파일에 호스트키 및 줌링크가 저장되었습니다.")

동작()
