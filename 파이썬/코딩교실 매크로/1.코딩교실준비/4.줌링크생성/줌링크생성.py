import os
import openpyxl
import time
from playwright.sync_api import sync_playwright

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '줌링크생성.xlsx')  # 엑셀 파일 경로

def 초기화():
    global wb, ws
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

def 엑셀():
    global 로그인사이트,ID,PW,사용자관리링크,사업명,이메일,호스트키,줌링크
    로그인사이트 = ws.cell(row=1, column=17).value
    사용자관리링크 = ws.cell(row=2, column=17).value
    ID = ws.cell(row=3, column=17).value
    PW = ws.cell(row=4, column=17).value
    사업명 = ws.cell(row=5, column=17).value
    이메일 = []
    호스트키 = []
    줌링크 = []
    row = 2  # 1행은 헤더이므로 2행부터 시작
    while True:
        email = ws.cell(row=row, column=3).value  # C열은 column=3
        if email is None:
            break
        이메일.append(email)
        row += 1

def 로그인(page):
    # 로그인 프로세스
    page.goto(로그인사이트)
    page.fill('input#email', ID)
    page.click('span.zm-button__slot:text("Next")')
    page.wait_for_selector('input#password')
    page.fill('input#password', PW)
    page.keyboard.press("Enter")
    time.sleep(10)

def scroll_down(page, pixels):
    page.evaluate(f"window.scrollBy(0, {pixels});")
    time.sleep(0.5)

def 이메일작업(page, 이메일주소, 순번):
    try:
        page.goto(사용자관리링크)
        page.wait_for_selector('input.zoom-input__inner[placeholder="검색"]')
        page.locator('input.zoom-input__inner[placeholder="검색"]').first.fill(이메일주소)
        page.keyboard.press("Enter")
        page.wait_for_selector('mark[data-markjs="true"]')  # 검색 결과 하이라이트 등장까지 대기
        page.click('mark[data-markjs="true"]')

        page.wait_for_selector('button[data-testid="host-key:show-btn"]')
        page.click('button[data-testid="host-key:show-btn"]')
        page.wait_for_selector('#hostKey-subShowLabel')
        time.sleep(1)
        호스트키.append(page.text_content('#hostKey-subShowLabel').strip())

        page.locator('a.dropdown-toggle > b.caret').nth(1).click()
        page.wait_for_selector('a:has-text("향후 회의")')
        page.click('a:has-text("향후 회의")')
        page.wait_for_selector('a:has-text("회의 예약")')
        page.click('a:has-text("회의 예약")')
        time.sleep(10)
        주제 = f"{사업명} {순번 + 1:02}반"
        page.wait_for_selector('input#topic')
        page.fill('input#topic', 주제)
        page.click('span:has-text("되풀이 회의")')
        page.click('span#recurringType')
        page.click('dd#select-item-recurringType-3')
        
        scroll_down(page, 500)
        # ── 대기실 옵션 비활성화 처리 ──
        # 1) attached 상태(=DOM에만 올라와 있으면 OK)로 대기
        page.wait_for_selector('input#checkbox_2', state='attached')

        # 2) locator 가져오기
        waitroom_cb = page.locator('input#checkbox_2')

        # 3) aria-checked 값 확인 (동기 API)
        aria_val = waitroom_cb.get_attribute('aria-checked')
        print(f"[DEBUG] 대기실 aria-checked = {aria_val!r}")  # 확인용 로그

        # 4) true 일 때만 토글(uncheck) – force=True 로 숨겨진 상태에서도 동작
        if aria_val == 'true':
            waitroom_cb.uncheck(force=True)
            # 속성 변경 반영 대기
            page.wait_for_function(
                "() => document.getElementById('checkbox_2').getAttribute('aria-checked') === 'false'",
                timeout=5000
            )
            print("[DEBUG] 대기실 옵션 OFF 완료")
        else:
            print("[DEBUG] 이미 OFF 상태라 건너뜀")
        # ────────────────────────────────────────
        
        
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1)
        page.click('button[aria-label*="Show More Options"]')
        time.sleep(1)
        # "참가자가 언제든지 참가하도록 허용" 체크박스가 체크되어 있지 않으면 클릭
        checkbox_label = page.locator('label:has-text("참가자가 언제든지 참가하도록 허용")')
        checkbox = checkbox_label.locator('input[type="checkbox"]')
        if checkbox.get_attribute('aria-checked') == 'false':
            checkbox.check()

        page.fill('input.zm-select__input[role="combobox"]', '운영사무국')
        time.sleep(1)
        page.keyboard.press("Enter")
        time.sleep(1)
        page.wait_for_selector('button.save-btn')
        page.click('button.save-btn')

        page.wait_for_selector('a[aria-labelledby="view-registration"]')
        time.sleep(3)
        줌링크.append(page.get_attribute('a[aria-labelledby="view-registration"]', 'href'))

    except Exception as e:
        print(f"[실패] {이메일주소} 작업 중 오류 발생: {e}")
        호스트키.append("오류")
        줌링크.append("오류")

        
    except Exception as e:
        print(f"[실패] {이메일주소} 작업 중 오류 발생: {e}")
        호스트키.append("오류")
        줌링크.append("오류")    

def 엑셀저장():
    for i, (키, 링크) in enumerate(zip(호스트키, 줌링크)):
        ws.cell(row=i+2, column=4).value = 링크   # D열: 줌링크
        ws.cell(row=i+2, column=5).value = 키     # E열: 호스트키
    wb.save(excel_file_path)

def 동작():
    global browser
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()
        초기화()
        엑셀()
        로그인(page)
        for i, 주소 in enumerate(이메일):
            이메일작업(page, 주소, i)
        엑셀저장()
        browser.close()
        print("작업이 완료되었습니다. 엑셀 파일에 호스트키 및 줌링크가 저장되었습니다.")

동작()
