import os
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# 경로
폴더경로 = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(폴더경로, 'LMS전산반배정.xlsx')

def 초기화():
    """엑셀 로드 및 메인 시트 바인딩"""
    global wb, ws
    wb = load_workbook(excel_file_path, data_only=True)
    try:
        ws = wb["메인"]
    except KeyError:
        ws = wb.active  # 메인 시트를 못 찾으면 활성 시트 사용

def 메인_A_C_자동채우기():
    """
    - '월수'와 '화목'시트를 읽어 '메인'!A:C(반/이름/아이디)를 자동 구성합니다.
    - 월수/화목 규칙:
        * E=이름, F=아이디
        * A열(병합)이 분반 번호(1→01)로 간주
        * 시트의 Q7을 prefix로 사용(없으면 'BG')
        * 반코드 = {prefix}_{MW|TT}_{분반번호:02d}
    - 메인의 F~M은 수정하지 않음 (A~C만 갱신)
    """
    def _collect(ws_sheet, sheet_code, prefix_cell="Q7", default_prefix="BG"):
        prefix = ws_sheet[prefix_cell].value
        if prefix is None:
            prefix = default_prefix
        prefix = str(prefix).strip().strip("_")

        rows = []
        last_group = None  # 병합된 A열 값(분반 번호) carry
        # 보통 1~2행은 머리글이라 3행부터 훑음 (업로드된 샘플 기준)
        for r in range(3, ws_sheet.max_row + 1):
            a_val = ws_sheet.cell(row=r, column=1).value  # A(분반, 병합)
            if a_val not in (None, ""):
                s = str(a_val).strip()
                last_group = int(s) if s.isdigit() else s

            name = ws_sheet.cell(row=r, column=5).value  # E: 이름
            uid  = ws_sheet.cell(row=r, column=6).value  # F: 아이디
            # 아이디가 없으면 배정 함수가 멈추므로 스킵
            if (name is None and uid is None) or (uid is None or str(uid).strip() == ""):
                continue

            try:
                num_str = f"{int(last_group):02d}"
            except Exception:
                num_str = str(last_group)

            group_code = f"{prefix}_{sheet_code}_{num_str}"
            rows.append((group_code, str(name).strip(), str(uid).strip()))
        return rows

    ws_main = wb["메인"]
    ws_mw   = wb["월수"]
    ws_tt   = wb["화목"]

    # 월수(MW) → 화목(TT) 순으로 합치기
    rows = _collect(ws_mw, "MW") + _collect(ws_tt, "TT")

    # 메인!A:C 초기화(헤더 유지), F~M은 건드리지 않음
    for r in range(2, ws_main.max_row + 1):
        ws_main.cell(row=r, column=1).value = None  # A
        ws_main.cell(row=r, column=2).value = None  # B
        ws_main.cell(row=r, column=3).value = None  # C

    # 메인!A:C 쓰기
    write_row = 2
    for klass, name, uid in rows:
        ws_main.cell(row=write_row, column=1, value=klass)  # A: 반
        ws_main.cell(row=write_row, column=2, value=name)   # B: 이름
        ws_main.cell(row=write_row, column=3, value=uid)    # C: 아이디
        write_row += 1

def 로그인(page):
    """메인(Q열)에 적힌 접속정보로 로그인 후 검색 조건 세팅"""
    global 강의날짜
    로그인사이트   = ws.cell(row=1, column=17).value  # Q1
    수강관리링크   = ws.cell(row=2, column=17).value  # Q2
    adminID      = ws.cell(row=3, column=17).value    # Q3
    adminPW      = ws.cell(row=4, column=17).value    # Q4
    강의날짜        = ws.cell(row=5, column=17).value    # Q5

    page.goto(로그인사이트)
    page.fill('input[name="tbAdminId"]', str(adminID))
    page.fill('input[name="tbAdminPass"]', str(adminPW))
    page.press('input[name="tbAdminPass"]', 'Enter')

    page.goto(수강관리링크)
    page.select_option('select[name="ddlTargetDate"]', value=str(강의날짜))
    page.select_option('select[name="ddlKeyField"]', value='b.m_id')

def 반배정(page, 대상):
    """대상 칼럼 범위 선택 후 배정 실행"""
    if 대상 == '학생배정':
        col_range = slice(0, 3)   # 메인 A~C
    elif 대상 == '망령출동':
        col_range = slice(5, 8)   # 메인 F~H
    elif 대상 == '망령퇴장':
        col_range = slice(10, 13) # 메인 K~M
    else:
        print(f"알 수 없는 대상: {대상}")
        return

    마지막_행 = ws.max_row
    for row in ws.iter_rows(min_row=2, max_row=마지막_행, values_only=True):
        반이름, 학생이름, 아이디 = row[col_range]
        아이디 = f'{아이디}' if 아이디 is not None else None

        if 아이디 is None:
            print(f"{대상}에서 처리할 아이디가 없습니다. 다음으로 넘어갑니다.")
            return

        page.fill('input[name="tbKeyWord"].font_blue', 아이디)
        page.press('input[name="tbKeyWord"].font_blue', 'Enter')

        # 해당 날짜 링크 클릭 → 새 탭 열림
        page.click(f'a[href*="{강의날짜}"].button_red_small')

        # 새 탭 대기
        context = page.context
        new_page = context.wait_for_event("page")
        new_page.wait_for_load_state("load")

        # 새 탭에서 반이름 선택
        new_page.evaluate(
            '''(name) => { 
                const selectElement = document.querySelector('select[name="ddlTargetGroupNo"]');
                if (!selectElement) return;
                for (const option of selectElement.options) {
                    const text = (option.textContent || "").trim();
                    if (text.includes(name)) {
                        selectElement.value = option.value;
                        selectElement.dispatchEvent(new Event('change', { bubbles: true }));
                        break;
                    }
                }
            }''',
            반이름
        )

        # 팝업 자동 승인
        def _on_dialog(dialog):
            try:
                dialog.accept()
            except Exception:
                pass
        new_page.once('dialog', _on_dialog)

        # 수강 변경(or 정원초과 해제) 버튼 클릭
        new_page.locator(
            'a.button_yellow.bold:has-text("수강 변경"), '
            'a.button_red.bold:has-text("수강 인원이 모두 찼습니다. (변경불가 => 가능)")'
        ).click()

        new_page.close()
        page.bring_to_front()
        print(f"{학생이름}({아이디})이(가) {반이름}으로 배정 완료")

def 동작():
    global browser
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()

        # 1) 엑셀 로드
        초기화()

        # 2) 월수/화목 → 메인 A:C 자동 채우기 (F~M은 그대로)
        메인_A_C_자동채우기()
        wb.save(excel_file_path)  # 메인 A:C 갱신사항 저장

        # 3) 로그인 및 배정
        로그인(page)
        반배정(page, '망령출동')
        반배정(page, '학생배정')
        반배정(page, '망령퇴장')

        browser.close()

if __name__ == "__main__":
    동작()
