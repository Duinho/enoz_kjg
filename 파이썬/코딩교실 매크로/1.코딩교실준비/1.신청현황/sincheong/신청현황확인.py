import os
import pandas as pd
import openpyxl
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
import time
import subprocess
import pyautogui as pag
import pyperclip
import mss  # mss 모듈 추가
import cv2  # OpenCV 모듈 추가
import numpy as np

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 현재 작업 디렉토리로 설정
체크 = 0

"""이미지 캡쳐 및 인식 전용 함수 시작"""

def 전체화면캡쳐():
    with mss.mss() as sct:
        monitor = sct.monitors[0]  # 모든 모니터 캡처
        screenshot = sct.grab(monitor)
        img = np.array(screenshot)
        return cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)  # OpenCV 형식으로 변환

def 이미지찾기(이미지_이름, confidence=0.9):
    이미지경로 = os.path.join(폴더경로, 'pic', f'{이미지_이름}.png')

    target_image = cv2.imread(이미지경로)
    screen_image = 전체화면캡쳐()

    res = cv2.matchTemplate(screen_image, target_image, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
    if max_val >= confidence:
        pag.click(max_loc[0] + target_image.shape[1] // 2, max_loc[1] + target_image.shape[0] // 2)
        return True
    else:
        return False

def 인식대기(이미지_이름, confidence=0.9):
    while True:
        if 이미지찾기(이미지_이름, confidence=confidence):
            break
        time.sleep(0.5) 

"""이미지 캡쳐 및 인식 전용 함수 끝"""


""" 엑셀 값 받아오기 함수 시작"""

def 엑셀():
    global wb, sheet, 아이디, 비번, 포항로그인, 포항결제관리, 북구로그인, 구미결제관리, 구미로그인, 북구결제관리, 포항년월, 북구년월, 구미년월
    excel_file_path = os.path.join(폴더경로, '신청현황확인.xlsx')  # 엑셀 파일 경로
    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    sheet = wb.active
    아이디 = sheet.cell(row=1, column=17).value
    비번 = sheet.cell(row=2, column=17).value
    포항로그인 = sheet.cell(row=3, column=17).value
    포항결제관리 = sheet.cell(row=4, column=17).value
    북구로그인 = sheet.cell(row=5, column=17).value
    북구결제관리 = sheet.cell(row=6, column=17).value
    구미로그인 = sheet.cell(row=7, column=17).value
    구미결제관리 = sheet.cell(row=8, column=17).value
    포항년월 = [202406,202409, 202411]
    북구년월 = [202409, 202411]
    구미년월 = [202409]
    wb.close()

""" 엑셀 값 받아오기 함수 끝"""

""" LMS 로그인 및 초기 설정 함수 시작 """

def 로그인결제관리(page, 위치, 체크):
    global 로그인, 결제관리, 년월
    if 위치 == 'pohang':
        로그인 = 포항로그인
        결제관리 = 포항결제관리
        년월 = 포항년월
    elif 위치 == 'bukgu':
        로그인 = 북구로그인
        결제관리 = 북구결제관리
        년월 = 북구년월
    elif 위치 == 'gumi' :
        로그인 = 구미로그인
        결제관리 = 구미결제관리
        년월 = 구미년월

    page.goto(로그인)
    page.wait_for_selector('input[name="tbAdminId"]')
    page.fill('input[name="tbAdminId"]', 아이디)
    page.fill('input[name="tbAdminPass"]', 비번)
    page.press('input[name="tbAdminPass"]', 'Enter')
    page.wait_for_load_state('networkidle')

    page.goto(결제관리)
    page.wait_for_selector('select[name="ddlTargetDate"]')
    page.select_option('select[name="ddlTargetDate"]', value=str(년월[체크]))
    page.select_option('select[name="ddlOrderType"]', value='결제완료')
    page.press('input[name="tbKeyWord"].font_blue', 'Enter')
    page.wait_for_load_state('networkidle')

""" LMS 로그인 및 초기 설정 함수 시작 """

""" LMS에서 엑셀 다운로드 및 변경 후 저장 함수 시작 """

def 엑셀다운(page, 위치, 체크):
    # 팝업을 처리하는 이벤트 핸들러 추가
    def handle_dialog(dialog):
        try:
            dialog.accept()
        except Exception as e:
            print(f"Dialog already handled: {e}")

    page.on("dialog", handle_dialog)
    
    with page.expect_download() as download_info:
        page.click('a.button_yellow_small[title="일반 엑셀 출력"]')
    download = download_info.value
    download_path = os.path.join(폴더경로, 'excel', f'{위치}{체크}.xls')  # 확장자는 xls로 유지
    download.save_as(download_path)
    time.sleep(2)
    page.click('a.button_gray_small:has-text("LOGOUT")')

    # 기존에 xlsx 파일이 있으면 삭제
    xlsx_path = os.path.join(폴더경로, 'excel', f'{위치}{체크}.xlsx')
    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)

    # 엑셀 파일을 기본 프로그램으로 열기
    subprocess.Popen(["start", "excel", download_path], shell=True)

    # .xls 파일을 .xlsx로 변환하는 동작 수행
    인식대기('yes', confidence=0.9)
    pag.hotkey('alt', 'f2')
    인식대기('html', confidence=0.9)
    인식대기('xlsx', confidence=0.9)
    인식대기('file', confidence=0.9)
    pag.move(100, 0)
    pag.click()
    pag.write(f'{위치}{체크}')
    pag.press('enter')
    time.sleep(0.5)
    pag.hotkey('alt', 'f4')

    # 변환된 .xlsx 파일 경로 설정
    xlsx_path = os.path.join(폴더경로, 'excel', f'{위치}{체크}.xlsx')

    # .xlsx 파일을 편집하여 저장
    엑셀파일_편집(xlsx_path, 위치, 체크)


def 엑셀파일_편집(xlsx_path, 위치, 체크):
    # 변환된 .xlsx 파일 열기
    wb = openpyxl.load_workbook(xlsx_path)
    source_sheet = wb.active

    # Z열의 3번 행부터 "결제완료"가 아닌 행 삭제
    for row in range(source_sheet.max_row, 2, -1):  # 3번 행부터 시작하므로 인덱스 2부터
        cell_value = source_sheet[f"Z{row}"].value
        if cell_value != "결제완료":
            source_sheet.delete_rows(row)

    # 병합된 셀 모두 해제
    if source_sheet.merged_cells.ranges:
        for merged_cell in list(source_sheet.merged_cells):
            source_sheet.unmerge_cells(str(merged_cell))

    # 'H' 열의 특정 문자열을 포함하는 경우 수정
    for row in range(1, source_sheet.max_row + 1):
        cell_value = source_sheet[f"H{row}"].value
        if cell_value and '화목' in cell_value:
            source_sheet[f"H{row}"].value = '화목'
        elif cell_value and '월수' in cell_value:
            source_sheet[f"H{row}"].value = '월수'

    # 코딩교실 신청 현황.xlsx 파일 열기 또는 생성
    main_file_path = os.path.join(폴더경로, '코딩교실 신청 현황.xlsx')
    if os.path.exists(main_file_path):
        main_wb = openpyxl.load_workbook(main_file_path)
    else:
        main_wb = openpyxl.Workbook()
    
    # 해당 위치의 시트 가져오기 또는 새로 생성
    target_sheet_name = f"{위치}{체크}"
    if target_sheet_name in main_wb.sheetnames:
        target_sheet = main_wb[target_sheet_name]
    else:
        target_sheet = main_wb.create_sheet(title=target_sheet_name)

    # 'AS3' 셀이 '사배자정보'인지 확인
    if source_sheet['AS2'].value == '사배자정보':
        # 아래 열 매핑 사용
        열_매핑 = {
            'A': 'A', 'H': 'B', 'B': 'C','L': 'D', 'M': 'E', 'N': 'F', 'AL': 'G', 'AM': 'H',
            'AJ': 'I', 'K': 'J', 'AR': 'K', 'AH': 'L', 'AT': 'M', 'AU': 'N',
            'AV': 'O', 'AW': 'P', 'AX': 'Q', 'AY': 'R'
        }
    elif source_sheet['AZ2'].value == '7번':
        열_매핑 = {
            'A': 'A', 'H': 'B', 'B': 'C','L': 'D', 'M': 'E', 'N': 'F', 'AL': 'G', 'AM': 'H',
            'AJ': 'I', 'K': 'J', 'AR': 'K', 'AH': 'L', 'AS': 'M', 'AT': 'N',
            'AU': 'O', 'AV': 'P', 'AW': 'Q', 'AX': 'R', 'AY':'S', 'AZ': 'T'
        }
    else:
        열_매핑 = {
            'A': 'A', 'H': 'B', 'B': 'C','L': 'D', 'M': 'E', 'N': 'F', 'AL': 'G', 'AM': 'H',
            'AJ': 'I', 'K': 'J', 'AR': 'K', 'AH': 'L', 'AS': 'M', 'AT': 'N',
            'AU': 'O', 'AV': 'P', 'AW': 'Q', 'AX': 'R'
        }

    # 열 재배치 작업 수행
    for source_col, target_col in 열_매핑.items():
        for row in range(1, source_sheet.max_row + 1):
            source_cell_value = source_sheet[f"{source_col}{row}"].value
            target_sheet[f"{target_col}{row}"] = source_cell_value

    # 수정된 데이터를 기존 '코딩교실 신청 현황.xlsx' 파일에 저장
    main_wb.save(main_file_path)
    main_wb.close()
    wb.close()

""" LMS에서 엑셀 다운로드 및 변경 후 저장 함수 끝 """


def 동작():
    global browser
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        엑셀()
        로그인결제관리(page, 'pohang', 0) # 'pohang'의 0번째 날짜를 지정 ex) 엑셀() 함수에 포항년월 = [202406,202409, 202411]이라고 작성했다면, pohang과 1이라면 202409이 대상
        엑셀다운(page, 'pohang', 0)
        로그인결제관리(page, 'pohang', 1)
        엑셀다운(page, 'pohang', 1)
        로그인결제관리(page, 'pohang', 2)
        엑셀다운(page, 'pohang', 2)
        
        main_file_path = os.path.join(폴더경로, '코딩교실 신청 현황.xlsx')  # 코드가 끝나면 코딩교실 신청 현황 엑셀에 저장
        subprocess.Popen(["start", "excel", main_file_path], shell=True)  # 저장한 엑셀파일을 엶
        browser.close()

        
동작()
