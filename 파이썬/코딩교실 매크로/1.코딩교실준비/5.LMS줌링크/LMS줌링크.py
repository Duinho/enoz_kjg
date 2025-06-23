import os
import openpyxl
from playwright.sync_api import sync_playwright

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, 'LMS줌링크.xlsx')  # 엑셀 파일 경로

# 전역 리스트
instructor_ids = []
instructor_links = []

def 초기화():
    global wb, ws
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active

def 로그인(page):
    # 로그인 사이트 정보를 읽어옴
    로그인사이트   = ws.cell(row=1, column=17).value
    강사관리링크    = ws.cell(row=2, column=17).value
    adminID        = ws.cell(row=3, column=17).value
    adminPW        = ws.cell(row=4, column=17).value

    # 로그인 프로세스
    page.goto(로그인사이트)
    page.fill('input[name="tbAdminId"]', adminID)
    page.fill('input[name="tbAdminPass"]', adminPW)
    page.press('input[name="tbAdminPass"]', 'Enter')
    page.goto(강사관리링크)

    # 로그인 후에 엑셀의 강사ID/링크를 리스트에 담기
    _엑셀_링크목록_불러오기()

def _엑셀_링크목록_불러오기():
    """엑셀의 A열(강사ID)과 B열(링크)를 읽어서 전역 리스트에 저장"""
    global instructor_ids, instructor_links
    instructor_ids.clear()
    instructor_links.clear()

    # 헤더가 1행이므로 2행부터 시작, 빈 ID를 만나면 종료
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        강사ID, 링크 = row
        if not 강사ID:
            break
        instructor_ids.append(강사ID)
        instructor_links.append(링크)

def 링크부여(page):
    for idx, 강사ID in enumerate(instructor_ids):
        링크 = instructor_links[idx]
        page.click("a.button:has-text(\"검색 보이기 / 감추기\")")
        # 1) 강사 검색
        page.fill('input[name="teacher_name"]', 강사ID)
        page.press('input[name="teacher_name"]', 'Enter')
        page.wait_for_load_state('networkidle')
        # 2) 팝업 열기 (강사 클릭 → 팝업)
        with page.expect_popup() as popup_info:
            page.click(f'span.font_b:has-text("{강사ID}")')
        popup_page = popup_info.value
        popup_page.wait_for_load_state('networkidle')
        # 3) 팝업에 줌링크 입력
        popup_page.fill('input[name="video_link"]', 링크)
        # 4) 저장 버튼 클릭 (onclick="ChkFrm()" 실행 → 팝업 닫힘)
        popup_page.click('input[name="Submit"][value="저 장"]')

        print(f"{강사ID}에 줌링크 저장 완료")

def 동작():
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()
        초기화()
        로그인(page)
        링크부여(page)

        browser.close()

if __name__ == "__main__":
    동작()
