# 라이브러리 설치를 위해 터미널에 아래의 pip 명령어 입력
# 먼저 playwright install 입력
# 그 다음 pip install openpyxl playwright 입력
# 위 터미널 창에 붙여넣기하여 라이브러리 설치

import asyncio
import re
import os
import sys
import openpyxl
import shutil
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError

fm = None
err = None
강의이름 = None
강의날짜 = None
폴더이름 = None
폴더정보 = {}
download = None
item = None
page = None
browser = None
new_handle = None
다운로드_작업들 = []

폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '영상보내기.xlsx')  # 엑셀 파일 경로

"""엑셀 정보 가져오기 함수 시작"""

def 엑셀():
    global 줌로그인, 줌링크, 줌아이디, 줌비번, 알람받을번호, 문자박스링크, 문자박스아이디, 문자박스비번, 한영 
    wb = openpyxl.load_workbook(excel_file_path)    # 엑셀 파일을 wb에 저장
    ws = wb.active                                  # 엑셀 정보를 ws에 저장
    줌로그인 = ws.cell(row=1, column=15).value      # 각 행과 열의 값을 저장
    줌링크 = ws.cell(row=2, column=15).value
    줌아이디 = ws.cell(row=3, column=15).value
    줌비번 = ws.cell(row=4, column=15).value
    알람받을번호 = ws.cell(row=5, column=15).value
    문자박스링크 = ws.cell(row=8, column=15).value
    문자박스아이디 = ws.cell(row=9, column=15).value
    문자박스비번 = ws.cell(row=10, column=15).value
    한영 = ws.cell(row=11, column=15).value

"""엑셀 정보 가져오기 함수 종료"""

"""문자박스 함수 시작"""

async def 문자박스():
    await page.goto(문자박스링크)                           # 문자박스에 들어가서 아이디 비번치고 로그인
    await page.fill('input[name="id"]', 문자박스아이디)
    await page.fill('input[name="pwd"]', 문자박스비번)
    await page.press('input[name="pwd"]', 'Enter')
    await asyncio.sleep(3)    
    try:
        await page.click("button[onclick*='contentsLayerClose']")  # 닫기 버튼 클릭  # 어떤 알림 나오는데 그거 나오면 닫기 버튼 클릭
    except:
        pass
    await asyncio.sleep(3)
    await page.fill('textarea#recvList', 알람받을번호)
    await page.fill('textarea#msg', '영상 다운로드 완료')
    await page.locator("a.hand.openLayer", has_text="선택").click()
    frame = page.frame(name="callbackFrame")
    await frame.wait_for_selector("span:has-text('전체 회신번호')", timeout=5000)
    await frame.locator("span:has-text('전체 회신번호')").click()
    await frame.wait_for_selector('#bulkCallbackNum', timeout=3000)
    await frame.fill('#bulkCallbackNum', "01053006552")
    await frame.click("button[onclick*='callbackCheckForm']")
    await asyncio.sleep(0.3)
    await page.locator("a.hand.openLayer", has_text="전송하기").click()
    await asyncio.sleep(0.3)
    await page.keyboard.press('Enter')
    await asyncio.sleep(0.3)
    await page.keyboard.press('Enter')
    
"""문자박스 함수 종료"""

"""줌 로그인 후 녹화 및 대화 내용으로 이동 함수 시작"""

async def 로그인():
    global page                 # 전역 변수로 사용
    await page.goto(줌로그인)   # Zoom 로그인 페이지로 이동
    await page.wait_for_load_state()
    await page.wait_for_selector('button#onetrust-accept-btn-handler', timeout=10000)  # 쿠키 허용 버튼 클릭
    await page.click('button#onetrust-accept-btn-handler')
    await asyncio.sleep(3)
    await page.fill('input#email', 줌아이디)  # 이메일 및 비밀번호 입력하고 로그인
    await page.click('span.zm-button__slot:has-text("Next")')
    await page.fill('input#password', 줌비번)
    await asyncio.sleep(1)
    #await page.press('#password', 'Enter')
    await page.click('span.zm-button__slot:has-text("Sign In")')
    await page.wait_for_selector('h3._user__name_66l5p_119:has-text("운영사무국")', timeout=30000)
    await page.goto(줌링크)  # 기록관리 페이지로 이동
    await asyncio.sleep(1)
    if 한영 == "한국어":
        await page.wait_for_selector('//h1[contains(text(), "녹화 및 대화 내용") and contains(@style, "font-size: 24px;")]', timeout=30000)
    else:
        await page.wait_for_selector('//h1[contains(text(), "Recording and Transcript Management") and contains(@style, "font-size: 24px;")]', timeout=30000)
    await asyncio.sleep(5)
    
"""줌 로그인 후 녹화 및 대화 내용으로 이동 함수 종료"""

"""영상 목록에서 영상을 클릭하고 영상 정보 가져오기 함수 시작"""

async def 정보가져오기():
    global page, 강의이름, 강의날짜, 폴더정보, fm, err
    try: 
        if err == 1:
            pass
        else:
            fm = await page.query_selector('div.zoom-empty__title:has-text("검색과 일치하는 결과를 찾을 수 없습니다")')
            if fm:                                      # 첫 번째 회의 링크가 없으면 종료(다운로드 받을 회의가 없으면)
                await asyncio.gather(*다운로드_작업들)   # 모든 다운로드가 완료될 때까지 대기
                await 문자박스()
                print("모든 영상 다운로드 완료")
                await browser.close()
                sys.exit()
            link_element = await page.wait_for_selector('td.zoom-virtual-table_1_column_3 a.topic-actived', timeout=30000)  # 첫 번째 회의 링크 찾기
            await link_element.click()
            await asyncio.sleep(1)
            error_count = await page.locator('div.error-wrap').count()
            if error_count > 0 and await page.locator('div.error-wrap').text_content() == "녹화가 너무 짧아 저장되지 않았습니다.":
                await page.goto(줌링크)
                await asyncio.sleep(1)
                first_button = await page.locator('button.grey.zm-button--icon.zm-button--plain.zm-button--icon-ghost.zm-button--icon.zm-button--mini.is-ghost.zm-button.zm-dropdown-selfdefine').first
                await first_button.click()
                await page.locator('div.textRed', has_text="삭제").click()
                await asyncio.sleep(3)
                await page.goto(줌링크)
                link_element = await page.query_selector('a.mgb-0.cursor-pointer.topic-actived')  # 첫 번째 회의 링크 찾기        
                await link_element.click()
                await asyncio.sleep(1)
            await page.wait_for_selector('span.meeting-topic', timeout=30000)
            await asyncio.sleep(1)
            topic_element = await page.query_selector('span.meeting-topic')                         # 회의 주제(타이틀) 가져오기
            if topic_element:
                강의이름 = await topic_element.inner_text()                                         # 강의 이름 가져오기
            date_element = await page.query_selector('span.mgr-md')                                 # 날짜 값 가져오기
            if date_element:
                date_text = await date_element.inner_text()                                         # 정규식을 사용하여 날짜 값을 추출
                pattern = r'(\d+년 \d+월 \d+일)'                                                    # @@년 @월 @@일로 수정 ex) 2025년 10월 24일 
                matches = re.search(pattern, date_text)
                강의날짜 = matches.group(1)
                폴더이름 = f"{강의이름}_{강의날짜}"                                                   # 폴더이름을 강의이름과 강의날짜로 지정
                if "경산" in 강의이름:                                                               # 경산이면 경산, 포항이면 포항 폴더에 저장되도록 설정 북구는 지역이름이 없어서 그냥 저장됩니다. 참고바랍니다.
                    폴더정보[강의이름] = os.path.join(폴더경로, '녹화영상', '경산', 강의날짜, 폴더이름)
                elif "포항" in 강의이름:
                    폴더정보[강의이름] = os.path.join(폴더경로, '녹화영상', '포항', 강의날짜, 폴더이름)
                elif "구미" in 강의이름:
                    폴더정보[강의이름] = os.path.join(폴더경로, '녹화영상', '구미', 강의날짜, 폴더이름)
                elif "온라인SW" in 강의이름:
                    폴더정보[강의이름] = os.path.join(폴더경로, '녹화영상', '대구', 강의날짜, 폴더이름)
                else:
                    폴더정보[강의이름] = os.path.join(폴더경로, '녹화영상',  강의날짜, 폴더이름)                
                os.makedirs(폴더정보[강의이름], exist_ok=True)
    except Exception as e:
        print(f"{강의이름} 정보가져오기에서 오류 발생: {e}")  # 디버깅을 위한 오류 위치 출력    
        err = 1

"""영상 목록에서 영상을 클릭하고 영상 정보 가져오기 함수 종료"""

"""가져온 정보의 이름과 경로로 영상 다운로드 함수 시작"""

async def 영상다운(강의이름, 이름):
    global item, browser, 폴더정보, 강의날짜, new_handle, err  # 전역 변수로 사용
    try:
        await item.click()              # 다운로드 클릭
        new_handle = None               # 새로 열린 창 핸들 얻기
        while not new_handle:
            for handle in browser.contexts[0].pages:
                if handle != page:
                    new_handle = handle
                    break
            await asyncio.sleep(1)    
        await new_handle.wait_for_selector('a.download-btn', timeout=90000)                               # 새로 열린 창에서 다운로드 버튼 찾기
        dl_element = await new_handle.query_selector('a.download-btn')      
        if dl_element:
            async with new_handle.expect_download() as download_info:
                await dl_element.click()                                                                  # 다운로드 클릭
                다운로드_작업들.append(asyncio.create_task(handle_download(download_info, 강의이름, 이름))) # 영상을 다운로드 경로와 설정한 아름으로 저장합니다.
            await new_handle.close()                                                                      # 다운로드 클릭하면 새로 열린 창을 닫습니다.
            new_handle = None
            await page.wait_for_selector('span.meeting-topic', timeout=30000)
            await asyncio.sleep(1)
    except PlaywrightTimeoutError as e:
        print(f"{강의이름}의 {이름} 다운로드 대기열이 가득 찼습니다.")
        if new_handle:
            await new_handle.close()
            await asyncio.sleep(25)
            err = 1
    except Exception as e:
        print(f"{강의이름}의 {이름} 다운로드에서 예외 발생: {e}")
        if new_handle:
            await new_handle.close()
            err = 1

"""가져온 정보의 이름과 경로로 영상 다운로드 함수 종료"""

"""다운로드 팝업 관리 함수 시작"""
            
async def handle_download(download_info, 강의이름, 이름):
    try:
        download = await download_info.value                                      # 다운로드 중인 정보 저장
        download_path = await download.path()           
        target_path = os.path.join(폴더정보[강의이름], f"{강의이름}_{강의날짜}.mp4") # 다운로드 파일명을 지정
        shutil.move(download_path, target_path)                                   # 파일명으로 변경하여 저장
    except Exception as e:
        print(f"{강의이름}의 {이름} 다운로드 이동 중 오류 발생: {e}")
        err = 1

"""다운로드 팝업 관리 함수 종료"""

"""다운로드 후 영상 삭제 함수 시작"""

async def 삭제():
    global page, err  # 전역 변수로 사용
    try:
        if err:
            await page.goto(줌링크)  # 기록관리 페이지로 이동
            await asyncio.sleep(5)
            fm = None
            fm = await page.query_selector('span[role="alert"]:has-text("검색과 일치하는 결과를 찾을 수 없습니다")')  # 만약 검색과 일치하는 결과를 찾을 수 없습니다. 라고 나오지 않는다면
            if fm is not None:
                await asyncio.sleep(1)
                if 한영 == "한국어":        # 한영 부분이 한국어 라고 적혀있으면 녹화 및 대화 내용을 찾기 아니면 Recording and Transcript Management을 찾기
                    await page.wait_for_selector('//h1[contains(text(), "녹화 및 대화 내용") and contains(@style, "font-size: 24px;")]', timeout=30000)
                else:
                    await page.wait_for_selector('//h1[contains(text(), "Recording and Transcript Management") and contains(@style, "font-size: 24px;")]', timeout=30000)
                print(f"설치 재시작")              
                await asyncio.sleep(1)
            err = None
        else:
            delete_element = await page.query_selector('button:nth-child(4)')  # 삭제 버튼 누르기
            if delete_element:
                await delete_element.click()                                   # 버튼 누르기
                await asyncio.sleep(1)
                delete1_element = await page.query_selector('button.zm-button--primary.zm-button--small.zm-button > span')  # 휴지통으로 이동 찾기       
                await asyncio.sleep(1)
                if delete1_element: 
                    await page.get_by_role("button", name="휴지통으로 이동").click()  # 버튼 누르기
                    await asyncio.sleep(1)
                    await page.wait_for_load_state()            
            await asyncio.sleep(1)
            await page.reload() 
            await page.wait_for_selector('text=휴지통', timeout=30000)              # 휴지통이라는 말이 나올때까지 대기 (정상적으로 삭제되면 뜨는 것)
            await asyncio.sleep(3)
    except Exception as e:
        print(f"{강의이름} 삭제에서 오류 발생: {e}")  # 디버깅을 위한 오류 위치 출력
        await page.goto(줌링크)                     # 기록관리 페이지로 이동
        await asyncio.sleep(1)
        if 한영 == "한국어":
            await page.wait_for_selector('//h1[contains(text(), "녹화 및 대화 내용") and contains(@style, "font-size: 24px;")]', timeout=30000)
        else:
            await page.wait_for_selector('//h1[contains(text(), "Recording and Transcript Management") and contains(@style, "font-size: 24px;")]', timeout=30000)
        await asyncio.sleep(3)

"""다운로드 후 영상 삭제 함수 종료"""

async def 동작():
    global browser, page, item, err, 다운로드_작업들            # 전역 변수로 사용
    async with async_playwright() as playwright:               # 크로미움 브라우저 열기
        browser = await playwright.chromium.launch(
            headless=False,
        )
        page = await browser.new_page(accept_downloads=True)   # 새로운 창이 열리면 page에 저장
        엑셀()
        await 로그인()
        while True:
            await 정보가져오기()
            downloadlist = await page.query_selector_all('div.item_list_header.relative > a')  # 영상 다운로드 하기 위해 각 항목으로 접속
            for item in downloadlist:
                inner_text = await item.inner_text()                                # 영상 리스트를 가져옴
                if inner_text == "갤러리 보기가 포함된 공유 화면" or inner_text == "발표자 보기가 포함된 공유 화면": # 갤러리 보기가 포함된 공유 화면이거나 발표자 보기가 포함된 공유 화면을 선택
                    await 영상다운(강의이름, '강의 영상')
            await 삭제()

asyncio.run(동작())