import os
import openpyxl
import time
from playwright.sync_api import sync_playwright
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import re
from datetime import datetime, timedelta
import pytz

# 변수 선언
폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 지금 코드 파일이 있는 위치를 저장
excel_file_path = os.path.join(폴더경로, '영상보내기.xlsx')  # 엑셀 파일 경로

def find_folder_id(드라이브, folder_name, parent_folder_id=None):
    query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder'"
    if parent_folder_id:
        query += f" and '{parent_folder_id}' in parents"
    results = 드라이브.files().list(q=query, fields="files(id, name, webViewLink)").execute()
    items = results.get('files', [])
    if not items:
        raise ValueError(f"No folder named '{folder_name}' found.")
    return items[0]['id'], items[0].get('webViewLink')

def get_subfolders(드라이브, parent_folder_id):
    query = f"'{parent_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
    results = 드라이브.files().list(q=query, fields="files(id, name)").execute()
    return results.get('files', [])

def format_date(date_str):
    match = re.findall(r'\d+', date_str)
    if not match or len(match) < 2:
        raise ValueError("Invalid date format")
    return ''.join(match[-2:])

def extract_spreadsheet_id(url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url)
    if not match:
        raise ValueError("Invalid Google Sheets URL")
    return match.group(1)

def 드라이브링크복사():
    global 영상링크, 문자박스내용
    try:
        드라이브 = build('drive', 'v3', developerKey=api키)
        FOLDER_ID = re.search(r"/folders/([a-zA-Z0-9-_]+)", 드라이브링크).group(1)
        날짜가공 = format_date(날짜)
        반이름 = f' {반번호}'  # 앞에 공백 하나만 추가하여 ' 02반' 형식으로 맞춤

        folder_id, _ = find_folder_id(드라이브, 요일, parent_folder_id=FOLDER_ID)

        subfolders = get_subfolders(드라이브, folder_id)
        target_folder_id = None
        for subfolder in subfolders:
            try:
                if format_date(subfolder['name']) == 날짜가공:
                    target_folder_id = subfolder['id']
                    break
            except ValueError:
                print("Invalid date format in subfolder name")  # 디버깅 메시지
                continue

        if target_folder_id:
            반_subfolders = get_subfolders(드라이브, target_folder_id)
            for subfolder in 반_subfolders:
                # 반이름을 폴더 이름에서 찾기
                if re.search(rf'\b{반이름}\b', subfolder['name']):
                    _, 영상링크 = find_folder_id(드라이브, subfolder['name'], parent_folder_id=target_folder_id)
                    break

            if 영상링크:
                문자박스내용 = f"{멘트1}\n영상 링크 : {영상링크}\n{멘트2}"
            else:
                print("No matching 반 folder found.")
        else:
            print("No matching 날짜 folder found.")

    except ValueError as e:
        print(f"Error during 드라이브링크복사: {e}")  # 디버깅 메시지


def 엑셀(시트세트):
    global 알람받을번호, 날짜, api키, 시트링크, 시트이름, 드라이브링크, wb, ws, 요일, 멘트1, 멘트2, 문자박스링크, 문자박스아이디, 문자박스비번, 회신번호
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    알람받을번호 = ws.cell(row=5, column=15).value
    날짜 = ws.cell(row=1, column=17).value    

    # 첫 번째 세트와 두 번째 세트를 구분하여 링크와 드라이브 링크 설정
    if 시트세트 == 1:
        시트링크_월수 = ws.cell(row=3, column=17).value
        시트링크_화목 = ws.cell(row=4, column=17).value
        드라이브링크 = ws.cell(row=5, column=17).value
        회신번호 = ws.cell(row=6, column=17).value
    elif 시트세트 == 2:
        시트링크_월수 = ws.cell(row=7, column=17).value
        시트링크_화목 = ws.cell(row=8, column=17).value
        드라이브링크 = ws.cell(row=9, column=17).value
        회신번호 = ws.cell(row=10, column=17).value
    회신번호 = f'010{회신번호}'

    # 요일과 시트 이름 설정
    if '월' in 날짜 or '수' in 날짜:
        시트이름 = '출석부(월/수)'
        요일 = '월수'
        시트링크 = 시트링크_월수
    elif '화' in 날짜 or '목' in 날짜:
        시트이름 = '출석부(화/목)'
        요일 = '화목'
        시트링크 = 시트링크_화목

    api키 = ws.cell(row=2, column=17).value
    멘트1 = ws.cell(row=6, column=15).value
    멘트2 = ws.cell(row=7, column=15).value
    문자박스링크 = ws.cell(row=8, column=15).value
    문자박스아이디 = ws.cell(row=9, column=15).value
    문자박스비번 = ws.cell(row=10, column=15).value
    kst = pytz.timezone('Asia/Seoul')
    now = datetime.now(kst)
    yesterday = now - timedelta(1)
    month = yesterday.month  # 월 (앞에 0을 제거한 형태)
    day = yesterday.day  # 일
    weekday_dict = {
        'Monday': '월',
        'Tuesday': '화',
        'Wednesday': '수',
        'Thursday': '목',
        'Friday': '금',
        'Saturday': '토',
        'Sunday': '일'
    }
    weekday_eng = yesterday.strftime('%A')
    weekday_kor = weekday_dict[weekday_eng]
    formatted_date = f"{month}월 {day}일"
    어제날짜 = f"{formatted_date}({weekday_kor})"
    if 날짜 == '자동' :
        날짜 = 어제날짜
        if weekday_kor == '월' or  weekday_kor == '수' :
            시트이름 = '출석부(월/수)'
            요일 = '월수'
            시트링크 = 시트링크_월수
        elif weekday_kor == '화' or  weekday_kor == '목' :
            시트이름 = '출석부(화/목)'
            요일 = '화목'
            시트링크 = 시트링크_화목    


def 날짜검색(시트데이터, search_value):
    for col_index, value in enumerate(시트데이터[1], start=1):
        if value == search_value:
            return col_index
    return None  

def 영상검색(시트데이터, column_index, search_value):
    rows = []
    for row_index, row in enumerate(시트데이터[1:], start=2):
        if len(row) >= column_index and row[column_index - 1] == search_value:
            rows.append(row_index)
    return rows

def 특정열값검색(시트데이터, rows, page):
    global 반번호, 학생번호, 학부모번호, 학생이름
    for row in rows:
        # A 열에서 최초로 나오는 값을 찾기 위한 로직 추가
        반번호 = None
        current_row = row
        while current_row > 0:
            if len(시트데이터[current_row - 1]) > 0:
                반번호 = 시트데이터[current_row - 1][0]
                if 반번호:
                    # 반번호에서 숫자와 반을 분리하여 숫자는 두 자리로 만들기
                    match = re.match(r'(\d+)(반)', str(반번호))
                    if match:
                        number_part = match.group(1).zfill(2)  # 숫자 부분을 두 자리로 변환
                        반번호 = f'{number_part}{match.group(2)}'  # 두 자리 숫자와 '반'을 결합
                    break
            current_row -= 1

        # j_col_index, l_col_index, e_col_index 위치의 값을 학생번호, 학부모번호, 학생이름에 저장
        학생이름 = 시트데이터[row - 1][e_col_index - 1] if len(시트데이터[row - 1]) >= e_col_index else None
        학생번호 = 시트데이터[row - 1][j_col_index - 1] if len(시트데이터[row - 1]) >= j_col_index else None
        학부모번호 = 시트데이터[row - 1][l_col_index - 1] if len(시트데이터[row - 1]) >= l_col_index else None

        드라이브링크복사()
        page.fill('textarea#recvList', 학생번호)
        if 학생번호 == 학부모번호:
            page.fill('textarea#recvList', 학생번호)
        else:
            page.fill('textarea#recvList', f'{학생번호}\n{학부모번호}')
        page.fill('textarea#msg', 문자박스내용)
        page.click('div.num_select')
        time.sleep(0.3)
        frame = page.frame(name='callbackFrame')
        frame.click(f"a:has-text('{str(회신번호)}')")
        page.click('a:has(img[src*="send_btn.gif"])')
        time.sleep(0.3)
        page.keyboard.press('Enter')
        time.sleep(0.3)
        page.keyboard.press('Enter')
        print(f'{반번호} {학생이름} 학생 영상 발송 완료')
    page.fill('textarea#recvList', 알람받을번호)
    page.fill('textarea#msg', '영상 발송 완료')
    page.click('div.num_select')
    time.sleep(0.3)
    frame = page.frame(name='callbackFrame')
    frame.wait_for_selector(f"a:has-text('{str(회신번호)}')", timeout=30000)
    frame.click(f"a:has-text('{str(회신번호)}')")
    page.click('a:has(img[src*="send_btn.gif"])')
    time.sleep(0.3)
    page.keyboard.press('Enter')
    time.sleep(0.3)
    page.keyboard.press('Enter')


def 시트확인(page):
    global 날짜위치, 영상_행목록, j_col_index, l_col_index, e_col_index
    시트아이디 = extract_spreadsheet_id(시트링크)   
    시트 = build('sheets', 'v4', developerKey=api키)
    
    # 시트 데이터 한 번에 가져오기
    range_name = f'{시트이름}'
    result = 시트.spreadsheets().values().get(spreadsheetId=시트아이디, range=range_name).execute()
    시트데이터 = result.get('values', [])
    if not 시트데이터:
        print("No data found.")
        return

    # "전화번호"라는 텍스트를 가진 열의 인덱스를 찾음
    header_row = 시트데이터[1]  # 두 번째 행 (1번 인덱스)
    j_col_index = None
    l_col_index = None
    e_col_index = None
    for col_index, value in enumerate(header_row, start=1):
        if value == "전화번호":
            if j_col_index is None:
                j_col_index = col_index  # 첫 번째 "전화번호"
            else:
                l_col_index = col_index  # 두 번째 "전화번호"
        elif value == "이름" and e_col_index is None:
            e_col_index = col_index  # "이름" 열의 인덱스 저장

    if j_col_index is None or l_col_index is None:
        print("전화번호 열이 두 개 발견되지 않았습니다.")
        return

    if e_col_index is None:
        print("이름 열이 발견되지 않았습니다.")
        return

    print(f"열 위치: j_col_index={j_col_index}, l_col_index={l_col_index}, e_col_index={e_col_index}")  # 디버깅 메시지

    날짜위치 = 날짜검색(시트데이터, 날짜)  # 셀 주소 찾기
    if 날짜위치:
        영상_행목록 = 영상검색(시트데이터, 날짜위치, '영상')
        특정열값검색(시트데이터, 영상_행목록, page)
    else:
        print(f'날짜 "{날짜}" not found in the second row')

def 시트수정():
    global 날짜위치, 영상_행목록, 시트이름, 폴더경로, 시트링크

    # JSON 파일 경로 설정
    json_file_path = os.path.join(폴더경로, 'auto-send-link-b97c1597bb00.json')

    # OAuth2 인증 설정
    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
    creds = Credentials.from_service_account_file(json_file_path, scopes=SCOPES)
    service = build('sheets', 'v4', credentials=creds)

    # 스프레드시트 ID 추출
    시트아이디 = extract_spreadsheet_id(시트링크)

    # 시트 ID 가져오기
    spreadsheet = service.spreadsheets().get(spreadsheetId=시트아이디).execute()
    sheet_id = None
    for sheet in spreadsheet['sheets']:
        if sheet['properties']['title'] == 시트이름:
            sheet_id = sheet['properties']['sheetId']
            break

    if sheet_id is None:
        raise ValueError(f"Sheet name '{시트이름}' not found in the spreadsheet")

    # 수정할 범위 및 값을 준비
    requests = []
    for i in range(len(영상_행목록)):
        cell_range = f'{시트이름}!{chr(64 + 날짜위치)}{영상_행목록[i]}'  # 날짜위치는 열, 영상_행목록[i]는 행
        requests.append({
            'range': cell_range,
            'values': [['O']]
        })

    # requests가 비어있지 않을 때만 batchUpdate 실행
    if requests:
        body = {
            'valueInputOption': 'RAW',
            'data': requests
        }
        response = service.spreadsheets().values().batchUpdate(
            spreadsheetId=시트아이디,
            body=body
        ).execute()

        # 메모 추가
        requests = []
        for i in range(len(영상_행목록)):
            requests.append({
                "updateCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": 영상_행목록[i] - 1,
                        "endRowIndex": 영상_행목록[i],
                        "startColumnIndex": 날짜위치 - 1,
                        "endColumnIndex": 날짜위치
                    },
                    "rows": [
                        {
                            "values": [
                                {
                                    "userEnteredValue": {
                                        "stringValue": "O"
                                    },
                                    "note": "영상 발송 완료"
                                }
                            ]
                        }
                    ],
                    "fields": "note,userEnteredValue"
                }
            })

        # requests가 비어있지 않을 때만 batchUpdate 실행
        if requests:
            body = {
                'requests': requests
            }
            response = service.spreadsheets().batchUpdate(
                spreadsheetId=시트아이디,
                body=body
            ).execute()
    else:
        print("No updates to send to Google Sheets.")

        

def 로그인(page):
    page.goto(문자박스링크)
    page.fill('input[name="id"]', 문자박스아이디)
    page.fill('input[name="pwd"]', 문자박스비번)
    page.press('input[name="pwd"]', 'Enter')
    time.sleep(1) 
    try:
        page.click("a[onclick*='contentsLayerClose']")  # 닫기 버튼 클릭
    except:
        pass 
    
def 영상발송(page):
    시트확인(page)
    시트수정()    
    
def 동작():
    global browser
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()
        # 첫 번째 세트 동작
        엑셀(시트세트=1)
        로그인(page)
        영상발송(page)

        # 두 번째 세트 동작
        엑셀(시트세트=2)
        영상발송(page)

        print('모든 영상 링크 발송 완료')
        browser.close()

동작()
