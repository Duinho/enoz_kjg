import os
import json
import requests
import shutil
import openpyxl
import time
from datetime import datetime
from requests.auth import HTTPBasicAuth
from concurrent.futures import ThreadPoolExecutor, as_completed
from playwright.sync_api import sync_playwright

# ✅ 현재 코드 경로 기준으로 설정
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, 'config.json')  # 기존 config.json 사용 가능
excel_file_path = os.path.join(script_dir, '영상보내기.xlsx')
BASE_DIR = os.path.join(script_dir, "녹화영상")  # 영상 폴더 고정


# ✅ config.json (Zoom 관련 설정만 여기서 가져옴)
with open(config_path, 'r', encoding='utf-8') as f:
    config = json.load(f)

CLIENT_ID = config['Zoom']['client_id']
CLIENT_SECRET = config['Zoom']['client_secret']
ACCOUNT_ID = config['Zoom']['account_id']
START = config['Recordings']['start_date']
END = config['Recordings']['end_date']


# ✅ Zoom Access Token 발급
def get_access_token():
    url = 'https://zoom.us/oauth/token'
    params = {
        'grant_type': 'account_credentials',
        'account_id': ACCOUNT_ID
    }
    res = requests.post(url, params=params, auth=HTTPBasicAuth(CLIENT_ID, CLIENT_SECRET))
    res.raise_for_status()
    return res.json()['access_token']


# ✅ 사용자 목록 가져오기
def get_all_users(token):
    url = 'https://api.zoom.us/v2/users'
    headers = {'Authorization': f'Bearer {token}'}
    params = {'page_size': 100}
    users = []

    while True:
        res = requests.get(url, headers=headers, params=params)
        res.raise_for_status()
        data = res.json()
        users.extend(data.get('users', []))
        if not data.get('next_page_token'):
            break
        params['next_page_token'] = data['next_page_token']

    return users


# ✅ 회의 녹화 목록
def list_recordings(token, user_id):
    url = f'https://api.zoom.us/v2/users/{user_id}/recordings'
    headers = {'Authorization': f'Bearer {token}'}
    params = {'from': START, 'to': END, 'page_size': 50}

    meetings = []
    while True:
        res = requests.get(url, headers=headers, params=params)
        res.raise_for_status()
        data = res.json()
        meetings.extend(data.get('meetings', []))
        if not data.get('next_page_token'):
            break
        params['next_page_token'] = data['next_page_token']
    return meetings

# 한 번 처리한 meeting_id 스킵용
# 한 번 처리한 meeting_id 스킵용
processed_meetings = set()

# MP4 우선순위 타입 정의
priority_types = [
    "shared_screen_with_gallery_view",
    "shared_screen_with_speaker_view",
    "shared_screen",
    "speaker_view",
    "gallery_view"
]

def download_and_delete(meeting, token, user_email):
    """
    1) 중복 meeting_id 스킵
    2) 날짜는 'YYYY년 M월 D일' (앞자리 0 없음)
    3) priority_types 에서 첫 매칭 MP4 한 개만 다운로드
    4) 파일명 충돌 시에만 _1, _2… 추가
    """
    meeting_id = meeting['id']
    if meeting_id in processed_meetings:
        print(f"⏭️ 이미 처리된 회의 스킵: {meeting_id}")
        return
    processed_meetings.add(meeting_id)

    topic = meeting.get('topic', 'NoTopic')
    dt = datetime.fromisoformat(meeting['start_time'].rstrip('Z'))
    date_str = f"{dt.year}년 {dt.month}월 {dt.day}일"

    # 지역 분류
    if "경산" in topic:
        region = "경산"
    elif "포항" in topic:
        region = "포항"
    elif "구미" in topic:
        region = "구미"
    elif "온라인SW" in topic:
        region = "대구"
    else:
        region = "기타"

    # 폴더: BASE_DIR/region/YYYY년 M월 D일/토픽_YYYY년 M월 D일
    folder_path = os.path.join(BASE_DIR, region, date_str, f"{topic}_{date_str}")
    os.makedirs(folder_path, exist_ok=True)

    # 녹화 파일 목록 조회
    headers = {'Authorization': f'Bearer {token}'}
    rec = requests.get(
        f'https://api.zoom.us/v2/meetings/{meeting_id}/recordings',
        headers=headers
    )
    rec.raise_for_status()
    files = rec.json().get('recording_files', [])

    # 우선순위 MP4 한 개만 선택
    selected = None
    for p_type in priority_types:
        for f in files:
            if f.get('recording_type') == p_type and f.get('file_type') == 'MP4':
                selected = f
                break
        if selected:
            break

    if not selected:
        print(f"⚠️ MP4 파일 없음: {topic} ({meeting_id})")
        return

    # 다운로드 URL & 기본 파일명
    download_url = selected['download_url'] + f"?access_token={token}"
    base_filename = f"{topic} _{date_str}.mp4"
    file_path = os.path.join(folder_path, base_filename)

    # 파일명 충돌 시에만 _1, _2… 추가
    counter = 1
    name, ext = os.path.splitext(base_filename)
    while os.path.exists(file_path):
        file_path = os.path.join(folder_path, f"{name}_{counter}{ext}")
        counter += 1

    print(f"🔽 {user_email} 회의 → 다운로드: {os.path.basename(file_path)}")
    with requests.get(download_url, stream=True) as r:
        r.raise_for_status()
        with open(file_path, 'wb') as out:
            shutil.copyfileobj(r.raw, out)

    # 녹화 전체 삭제
    res_del = requests.delete(
        f'https://api.zoom.us/v2/meetings/{meeting_id}/recordings',
        headers=headers
    )
    if res_del.status_code == 204:
        print(f"🧹 삭제 완료: {topic} ({meeting_id})")
    else:
        print(f"❌ 삭제 실패 ({res_del.status_code}): {res_del.text}")


# ✅ 문자 정보 Excel에서 가져오기 (시트세트 1 고정)
def get_sms_config_from_excel():
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    문자박스링크 = ws.cell(row=8, column=15).value
    문자박스아이디 = ws.cell(row=9, column=15).value
    문자박스비번 = ws.cell(row=10, column=15).value
    회신번호 = '010' + str(ws.cell(row=6, column=17).value)
    알람받을번호 = ws.cell(row=5, column=15).value
    return 문자박스링크, 문자박스아이디, 문자박스비번, 회신번호, 알람받을번호


# ✅ 문자 발송 (Playwright)
def send_sms_via_playwright(link, user_id, user_pwd, recv_number, callback_number):
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        page = browser.new_page()
        page.goto(link)
        page.fill('input[name="id"]', user_id)
        page.fill('input[name="pwd"]', user_pwd)
        page.press('input[name="pwd"]', 'Enter')
        time.sleep(1)
        try:
            page.click("button[onclick*='contentsLayerClose']")
        except:
            pass
        page.fill('textarea#recvList', recv_number)
        page.fill('textarea#msg', '영상 발송 완료')
        page.locator("a.hand.openLayer", has_text="선택").click()
        frame = page.frame(name="callbackFrame")
        frame.wait_for_selector("span:has-text('전체 회신번호')", timeout=5000)
        frame.locator("span:has-text('전체 회신번호')").click()
        frame.wait_for_selector('#bulkCallbackNum', timeout=3000)
        frame.fill('#bulkCallbackNum', callback_number)
        frame.click("button[onclick*='callbackCheckForm']")
        time.sleep(0.3)
        page.locator("a.hand.openLayer", has_text="전송하기").click()
        time.sleep(0.3)
        page.keyboard.press('Enter')
        time.sleep(0.3)
        page.keyboard.press('Enter')

        print("📨 영상 발송 완료 메시지 전송 완료")
        browser.close()


# ✅ 메인 로직
def main():
    print("🔐 Access Token 발급 중...")
    token = get_access_token()

    print("👥 사용자 목록 조회 중...")
    users = get_all_users(token)
    print(f"총 사용자 수: {len(users)}")

    download_tasks = []
    with ThreadPoolExecutor(max_workers=4) as executor:
        for user in users:
            user_id = user['id']
            user_email = user['email']
            print(f"\n📁 [{user_email}] 회의 목록 조회 중...")

            try:
                meetings = list_recordings(token, user_id)
                print(f"📋 회의 수: {len(meetings)}")

                for meeting in meetings:
                    task = executor.submit(download_and_delete, meeting, token, user_email)
                    download_tasks.append(task)

            except Exception as e:
                print(f"⚠️ 사용자 {user_email} 오류: {e}")

        for future in as_completed(download_tasks):
            try:
                future.result()
            except Exception as e:
                print(f"❌ 병렬 다운로드 오류: {e}")

    print("\n🎉 모든 병렬 다운로드 완료")

    # 문자 발송
    문자박스링크, 문자박스아이디, 문자박스비번, 회신번호, 알람받을번호 = get_sms_config_from_excel()
    send_sms_via_playwright(
        문자박스링크,
        문자박스아이디,
        문자박스비번,
        알람받을번호,
        회신번호
    )


# ✅ 실행
if __name__ == "__main__":
    main()
