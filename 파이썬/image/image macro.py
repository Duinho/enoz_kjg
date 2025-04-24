import cv2
import numpy as np
import pyautogui
import pytesseract
import os
import keyboard
from pykospacing import Spacing
from openpyxl import load_workbook

# Tesseract 실행 파일 경로 설정
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# 현재 디렉토리의 'pic' 폴더로 설정
사진폴더 = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'pic')
엑셀파일경로 = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'image.xlsx')

# 캡처 폴더 경로 설정
캡처폴더 = os.path.join(사진폴더, 'cap')
if not os.path.exists(캡처폴더):
    os.makedirs(캡처폴더)

# 새로 캡처하기 변수 초기화
새로캡쳐하기 = 0
current_screenshot = None

def 화면캡쳐(search_area):
    global 새로캡쳐하기, current_screenshot
    if 새로캡쳐하기 == 1 or current_screenshot is None:
        current_screenshot = pyautogui.screenshot(region=search_area)
        current_screenshot = cv2.cvtColor(np.array(current_screenshot), cv2.COLOR_RGB2BGR)
        screenshot_path = os.path.join(사진폴더, 'screenshot.png')
        cv2.imwrite(screenshot_path, current_screenshot)
        새로캡쳐하기 = 0
        print("Screenshot captured.")
    return current_screenshot

def 템플릿매칭(screenshot, target_image, threshold):
    result = cv2.matchTemplate(screenshot, target_image, cv2.TM_CCOEFF_NORMED)
    _, max_val, _, max_loc = cv2.minMaxLoc(result)
    if max_val >= threshold:
        target_height, target_width = target_image.shape[:2]
        target_center = (max_loc[0] + target_width // 2, max_loc[1] + target_height // 2)
        return target_center
    else:
        return None

def 이미지전처리(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)  # 그레이스케일로 변환
    gray = cv2.resize(gray, None, fx=3, fy=3, interpolation=cv2.INTER_CUBIC)  # 이미지 확대
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)  # 이진화

    # 글자 영역만 남기고 나머지는 흰색으로 만들기
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    if contours:
        x, y, w, h = cv2.boundingRect(contours[0])
        for contour in contours:
            x_, y_, w_, h_ = cv2.boundingRect(contour)
            x = min(x, x_)
            y = min(y, y_)
            w = max(w, x_ + w_ - x)
            h = max(h, y_ + h_ - y)
        cropped_image = image[y:y+h, x:x+w]

    return cropped_image

def 띄어쓰기교정(text):
    spacing = Spacing()
    lines = text.split('\n')
    corrected_lines = [spacing(line.replace(" ", "")) for line in lines]
    corrected_text = '\n'.join(corrected_lines)
    return corrected_text

def 텍스트추출(image, custom_oem_psm_config):
    text_area = 이미지전처리(image)  # 전처리
    extracted_text = pytesseract.image_to_string(text_area, lang='kor', config=custom_oem_psm_config)
    corrected_text = 띄어쓰기교정(extracted_text)
    print(f"Extracted text after correction: {corrected_text}")
    return corrected_text, text_area

def 엑셀에저장(텍스트, 행, 열):
    wb = load_workbook(엑셀파일경로)
    ws = wb.active
    ws.cell(row=행, column=열).value = 텍스트
    wb.save(엑셀파일경로)
    print(f"Text saved to Excel at row {행}, column {열}")

def 서칭글자인식하기(image_name, threshold=0.8, search_area=(0, 0, 1920, 1080), text_search_area=(0, 0, 100, 50)):
    global 행, 열, current_screenshot
    target_image_path = os.path.join(사진폴더, image_name)
    target_image = cv2.imread(target_image_path, cv2.IMREAD_COLOR)
    if target_image is not None:
        target_center = 템플릿매칭(current_screenshot, target_image, threshold)
        if target_center:
            text_area_x = target_center[0] + search_area[0] + text_search_area[0]
            text_area_y = target_center[1] + search_area[1] + text_search_area[1]
            text_area_w = text_search_area[2]
            text_area_h = text_search_area[3]
            if (text_area_x >= 0 and text_area_y >= 0 and 
                text_area_x + text_area_w <= current_screenshot.shape[1] and 
                text_area_y + text_area_h <= current_screenshot.shape[0]):
                
                text_area = current_screenshot[text_area_y:text_area_y+text_area_h, text_area_x:text_area_x+text_area_w]
                custom_oem_psm_config = r'--oem 3 --psm 3'
                텍스트, 전처리된_이미지 = 텍스트추출(text_area, custom_oem_psm_config)
                text_area_path = os.path.join(캡처폴더, f'{행}_{열}.png')
                cv2.imwrite(text_area_path, 전처리된_이미지)
                print(f"Text area saved at {text_area_path}")
                엑셀에저장(텍스트, 행, 열)
                열 += 1
                return 텍스트
    return None

def 글자인식하기(search_area, text_search_area):
    global 행, 열, current_screenshot
    text_area_x = text_search_area[0]
    text_area_y = text_search_area[1]
    text_area_w = text_search_area[2]
    text_area_h = text_search_area[3]
    if (text_area_x >= 0 and text_area_y >= 0 and 
        text_area_x + text_area_w <= current_screenshot.shape[1] and 
        text_area_y + text_area_h <= current_screenshot.shape[0]):
        
        text_area = current_screenshot[text_area_y:text_area_y+text_area_h, text_area_x:text_area_x+text_area_w]
        custom_oem_psm_config = r'--oem 3 --psm 3'
        텍스트, 전처리된_이미지 = 텍스트추출(text_area, custom_oem_psm_config)
        text_area_path = os.path.join(캡처폴더, f'{행}_{열}.png')
        cv2.imwrite(text_area_path, 전처리된_이미지)
        print(f"Text area saved at {text_area_path}")
        엑셀에저장(텍스트, 행, 열)
        열 += 1
        return 텍스트
    else:
        print("Error: Text search area is out of screenshot bounds.")
        return None

def main_loop():
    global 행, 열, 새로캡쳐하기  # 초기 셀 위치 설정
    행 = 1
    열 = 1
    새로캡쳐하기 = 1
    while True:
        if keyboard.is_pressed('ctrl+alt+s'):
            print("프로그램이 종료되었습니다.")
            break
        
        pyautogui.click(x=1663, y=968)
        for _ in range(10):
            새로캡쳐하기 = 1
            화면캡쳐((0, 0, 1920, 1080))
            글자인식하기((0, 0, 1920, 1080), (155, 340, 64, 20))
            글자인식하기((0, 0, 1920, 1080), (224, 340, 150, 20))
            글자인식하기((0, 0, 1920, 1080), (377, 340, 36, 20))
            글자인식하기((0, 0, 1920, 1080), (414, 340, 106, 20))
            열 = 1
            행 += 1  # 다음 행으로 이동
            pyautogui.press('down')
        break

if __name__ == "__main__":
    main_loop()
