import argparse
import os
import re
import time
from pathlib import Path
from urllib.parse import urljoin

import openpyxl
import requests
from bs4 import BeautifulSoup
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
CONFIG_XLSX = BASE_DIR / "신청현황확인.xlsx"
OUTPUT_XLSX = BASE_DIR / "코딩교실 신청 현황.xlsx"
DOWNLOAD_DIR = BASE_DIR / "excel"
ADDITIONAL_STATUS_XLSX_CANDIDATES = [
    Path(os.environ["ADDITIONAL_STATUS_XLSX"])
    if os.environ.get("ADDITIONAL_STATUS_XLSX")
    else None,
    Path.home()
    / "Documents"
    / "GitHub"
    / "enoz_kjg"
    / "파이썬"
    / "코딩교실"
    / "1.코딩교실준비"
    / "2.추가인원 수강신청"
    / "1.추가인원 가입 수강신청"
    / "추가인원 가입 현황.xlsx",
]
DEFAULT_TARGETS = [
    ("pohang", 1),
    ("gumi", 1),
]

LOCATION_CONFIG = {
    "pohang": {
        "months": [202605, 202409, 202510],
        "gubun": "PH",
        "headers": [
            "No.",
            "학교레벨",
            "강좌명",
            "재수강",
            "회원명",
            "회원ID",
            "성별",
            "학교",
            "학년",
            "휴대전화",
            "학부모명",
            "학부모연락처",
            "주소2 (교재배송)",
            "1번",
            "2번",
            "3번",
            "4번",
            "5번",
            "6번",
            "7번",
            "회원할인",
        ],
    },
    "bukgu": {
        "months": [202506, 202509, 202510],
        "gubun": "BG",
        "headers": [
            "No.",
            "학교레벨",
            "강좌명",
            "재수강",
            "회원명",
            "회원ID",
            "성별",
            "학교",
            "학년",
            "휴대전화",
            "학부모명",
            "학부모연락처",
            "주소2 (교재배송)",
            "Q1",
            "Q2",
            "Q3",
            "Q4",
            "Q5",
            "2021",
            "2022",
            "2023",
            "2024",
            "2025",
        ],
    },
    "gumi": {
        "months": [202605],
        "gubun": "GM",
        "headers": [
            "No.",
            "학교레벨",
            "강좌명",
            "재수강",
            "회원명",
            "회원ID",
            "성별",
            "학교",
            "학년",
            "휴대전화",
            "학부모명",
            "학부모연락처",
            "주소2 (교재배송)",
            "Q1",
            "Q2",
            "Q3",
            "Q4",
            "Q5",
            "회원할인",
        ],
    },
}

RUNTIME = {}
ADDITIONAL_MEMBER_IDS = None


def log(message: str):
    print(f"[{time.strftime('%H:%M:%S')}] {message}", flush=True)


def elapsed_text(start: float) -> str:
    return f"{time.perf_counter() - start:.1f}s"


def read_html_text(path: Path) -> str:
    raw = path.read_bytes()
    meta_sample = raw[:2048].decode("ascii", errors="ignore").lower()
    encodings = []
    match = re.search(r"charset=([a-z0-9_-]+)", meta_sample)
    if match:
        encodings.append(match.group(1))
    encodings.extend(["utf-8-sig", "euc-kr", "cp949", "utf-8"])

    tried = set()
    for encoding in encodings:
        if encoding in tried:
            continue
        tried.add(encoding)
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw.decode("utf-8", errors="replace")


def expand_cells(row) -> list[str]:
    values = []
    for cell in row.find_all(["td", "th"]):
        value = cell.get_text(" ", strip=True)
        colspan = int(cell.get("colspan", 1) or 1)
        values.extend([value] * colspan)
    return values


def normalize_course_name(value):
    if not value:
        return value
    if "화목" in value:
        return "화목"
    if "월수" in value:
        return "월수"
    return value


def normalize_cell_value(header, value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None

    if header == "강좌명":
        return normalize_course_name(value)

    if header in {"No.", "학년"} and isinstance(value, str):
        text = value.replace(",", "")
        if text.isdigit():
            return int(text)

    return value


def output_header_value(header):
    if header.isdigit():
        return int(header)
    return header


def find_additional_status_xlsx() -> Path | None:
    for candidate in ADDITIONAL_STATUS_XLSX_CANDIDATES:
        if candidate and candidate.exists():
            return candidate
    return None


def load_additional_member_ids() -> set[str]:
    global ADDITIONAL_MEMBER_IDS
    if ADDITIONAL_MEMBER_IDS is not None:
        return ADDITIONAL_MEMBER_IDS

    status_path = find_additional_status_xlsx()
    if status_path is None:
        log("추가인원 현황 파일을 찾지 못했습니다. 추가 열은 빈칸으로 처리합니다.")
        ADDITIONAL_MEMBER_IDS = set()
        return ADDITIONAL_MEMBER_IDS

    workbook = openpyxl.load_workbook(status_path, data_only=True)
    try:
        member_ids = set()
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 2:
                    continue
                member_id = str(row[1] or "").strip()
                if member_id:
                    member_ids.add(member_id)
        ADDITIONAL_MEMBER_IDS = member_ids
        log(f"추가인원 ID {len(member_ids)}개 로드: {status_path}")
        return ADDITIONAL_MEMBER_IDS
    finally:
        workbook.close()


def parse_download_table(download_path: Path, location: str):
    html = read_html_text(download_path)
    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if table is None:
        raise ValueError(f"표를 찾지 못했습니다: {download_path}")

    rows = table.find_all("tr")
    if len(rows) < 2:
        raise ValueError(f"헤더 행을 찾지 못했습니다: {download_path}")

    title_cell = rows[0].find(["td", "th"])
    title = title_cell.get_text(" ", strip=True) if title_cell else download_path.stem

    headers = expand_cells(rows[1])
    required_headers = set(LOCATION_CONFIG[location]["headers"]) | {"결제상태"}
    missing_headers = sorted(required_headers - set(headers))
    if missing_headers:
        raise ValueError(
            f"{location} 다운로드에 필요한 헤더가 없습니다: {', '.join(missing_headers)}"
        )

    output_headers = LOCATION_CONFIG[location]["headers"] + ["추가"]
    additional_member_ids = load_additional_member_ids()
    output_rows = []

    for row in rows[2:]:
        values = expand_cells(row)
        if not values or len(values) == 1:
            continue
        if len(values) < len(headers):
            values.extend([""] * (len(headers) - len(values)))

        record = dict(zip(headers, values[: len(headers)]))
        if not str(record.get("No.", "")).strip():
            continue
        if record.get("결제상태") != "결제완료":
            continue

        row_values = [
            normalize_cell_value(header, record.get(header, ""))
            for header in LOCATION_CONFIG[location]["headers"]
        ]
        member_id = str(record.get("회원ID", "") or "").strip()
        row_values.append("추가" if member_id in additional_member_ids else None)
        output_rows.append(row_values)

    return title, output_headers, output_rows


def clear_sheet(sheet):
    if sheet.merged_cells.ranges:
        for merged_range in list(sheet.merged_cells.ranges):
            sheet.unmerge_cells(str(merged_range))

    if sheet.max_row > 0:
        sheet.delete_rows(1, sheet.max_row)


def ensure_output_file_closed():
    lock_path = OUTPUT_XLSX.with_name(f"~${OUTPUT_XLSX.name}")
    if lock_path.exists():
        raise PermissionError(
            f"결과 파일이 Excel에서 열려 있는 것 같습니다. 닫고 다시 실행하세요: {OUTPUT_XLSX}"
        )


def load_output_workbook():
    if OUTPUT_XLSX.exists():
        return openpyxl.load_workbook(OUTPUT_XLSX)
    return openpyxl.Workbook()


def write_output_sheet_to_workbook(workbook, title, headers, rows, location, check_no):
    default_sheet = None
    if (
        "Sheet" in workbook.sheetnames
        and len(workbook.sheetnames) == 1
        and workbook["Sheet"]["A1"].value is None
    ):
        default_sheet = workbook["Sheet"]

    sheet_name = f"{location}{check_no}"
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        clear_sheet(sheet)
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        if default_sheet is not None:
            del workbook[default_sheet.title]

    for col_idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_idx, value=output_header_value(header))

    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    if headers:
        last_column = get_column_letter(len(headers))
        last_row = max(len(rows) + 1, 1)
        sheet.auto_filter.ref = f"A1:{last_column}{last_row}"
        sheet.freeze_panes = "A2"


def write_output_sheet(title, headers, rows, location, check_no):
    ensure_output_file_closed()
    workbook = load_output_workbook()
    try:
        write_output_sheet_to_workbook(workbook, title, headers, rows, location, check_no)
        workbook.save(OUTPUT_XLSX)
    finally:
        workbook.close()


def load_runtime_config():
    workbook = openpyxl.load_workbook(CONFIG_XLSX, data_only=True)
    sheet = workbook.active

    RUNTIME["admin_id"] = sheet.cell(row=1, column=17).value
    RUNTIME["admin_password"] = sheet.cell(row=2, column=17).value
    RUNTIME["pohang_login_url"] = sheet.cell(row=3, column=17).value
    RUNTIME["pohang_order_url"] = sheet.cell(row=4, column=17).value
    RUNTIME["bukgu_login_url"] = sheet.cell(row=5, column=17).value
    RUNTIME["bukgu_order_url"] = sheet.cell(row=6, column=17).value
    RUNTIME["gumi_login_url"] = sheet.cell(row=7, column=17).value
    RUNTIME["gumi_order_url"] = sheet.cell(row=8, column=17).value

    workbook.close()


def get_location_urls(location: str):
    if location == "pohang":
        return RUNTIME["pohang_login_url"], RUNTIME["pohang_order_url"]
    if location == "bukgu":
        return RUNTIME["bukgu_login_url"], RUNTIME["bukgu_order_url"]
    if location == "gumi":
        return RUNTIME["gumi_login_url"], RUNTIME["gumi_order_url"]
    raise ValueError(f"지원하지 않는 위치입니다: {location}")


def get_target_month(location: str, check_no: int) -> str:
    month_list = LOCATION_CONFIG[location]["months"]
    month_index = check_no - 1
    if not 0 <= month_index < len(month_list):
        raise IndexError(
            f"체크값 {check_no}에 해당하는 년월이 없습니다. (리스트 길이: {len(month_list)})"
        )
    return str(month_list[month_index])


def target_label(location: str, check_no: int) -> str:
    return f"{location}{check_no}({get_target_month(location, check_no)})"


def parse_target(value: str) -> tuple[str, int]:
    match = re.fullmatch(r"([a-z]+)\s*[:=-]?\s*(\d+)", value.strip().lower())
    if not match:
        raise argparse.ArgumentTypeError("대상은 pohang:1, bukgu:3, gumi1 형식으로 입력하세요.")

    location, check_no_text = match.groups()
    if location not in LOCATION_CONFIG:
        raise argparse.ArgumentTypeError(f"지원하지 않는 지역입니다: {location}")

    check_no = int(check_no_text)
    max_check_no = len(LOCATION_CONFIG[location]["months"])
    if not 1 <= check_no <= max_check_no:
        raise argparse.ArgumentTypeError(
            f"{location}은 1~{max_check_no}번만 선택할 수 있습니다."
        )

    return location, check_no


def configured_targets() -> list[tuple[str, int]]:
    return [
        (location, check_no)
        for location, config in LOCATION_CONFIG.items()
        for check_no in range(1, len(config["months"]) + 1)
    ]


def print_available_targets():
    for location, config in LOCATION_CONFIG.items():
        targets = [
            f"{location}{idx}={month}"
            for idx, month in enumerate(config["months"], start=1)
        ]
        print(", ".join(targets))


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="코딩교실 신청 현황을 다운로드/정리합니다.")
    parser.add_argument(
        "-t",
        "--target",
        action="append",
        type=parse_target,
        metavar="LOCATION:CHECK",
        help="처리 대상입니다. 예: --target pohang:1 --target bukgu:3",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="LOCATION_CONFIG에 등록된 모든 대상을 처리합니다.",
    )
    parser.add_argument(
        "--cache",
        action="store_true",
        help="새로 다운로드하지 않고 excel 폴더의 기존 .xls 파일을 사용합니다.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="다운로드/파싱 결과만 확인하고 결과 엑셀은 저장하지 않습니다.",
    )
    parser.add_argument(
        "--no-open",
        action="store_true",
        help="작업 완료 후 결과 엑셀을 자동으로 열지 않습니다.",
    )
    parser.add_argument(
        "--list-targets",
        action="store_true",
        help="선택 가능한 대상과 월을 출력합니다.",
    )

    args = parser.parse_args(argv)
    if args.all and args.target:
        parser.error("--all과 --target은 함께 사용할 수 없습니다.")
    return args


def login_session(login_url: str) -> requests.Session:
    session = requests.Session()
    response = session.post(
        urljoin(login_url, "login.asp?proc=in"),
        data={
            "tbAdminId": RUNTIME["admin_id"],
            "tbAdminPass": RUNTIME["admin_password"],
            "chk_type": "0",
        },
        allow_redirects=True,
        timeout=30,
    )
    response.raise_for_status()
    return session


def get_order_form_fields(session: requests.Session, order_url: str) -> dict[str, str]:
    response = session.get(order_url, timeout=30)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    form = soup.find("form", attrs={"name": "frm"})
    if form is None:
        raise ValueError(f"주문 폼을 찾지 못했습니다: {order_url}")

    fields = {}
    for tag in form.find_all(["input", "select"]):
        field_name = tag.get("name")
        if not field_name or field_name in fields:
            continue

        if tag.name == "select":
            selected = tag.find("option", selected=True) or tag.find("option")
            fields[field_name] = selected.get("value", "") if selected else ""
            continue

        input_type = (tag.get("type") or "").lower()
        if input_type in {"radio", "checkbox"}:
            if tag.has_attr("checked"):
                fields[field_name] = tag.get("value", "on")
            continue

        fields[field_name] = tag.get("value", "")

    return fields


def download_excel_file(session: requests.Session, order_url: str, location: str, check_no: int):
    DOWNLOAD_DIR.mkdir(exist_ok=True)
    download_path = DOWNLOAD_DIR / f"{location}{check_no}.xls"
    if download_path.exists():
        download_path.unlink()

    fields = get_order_form_fields(session, order_url)
    fields.update(
        {
            "ddlTargetDate": get_target_month(location, check_no),
            "ddlKeyGroup": LOCATION_CONFIG[location]["gubun"],
            "ddlOrderType": "3",
        }
    )

    excel_url = urljoin(
        order_url,
        f"OrderListExcel.asp?gubun={LOCATION_CONFIG[location]['gubun']}",
    )
    response = session.post(
        excel_url,
        data=fields,
        headers={"Referer": order_url},
        timeout=60,
    )
    response.raise_for_status()

    if b"<meta http-equiv=\"Content-Type\"" not in response.content[:200]:
        raise ValueError(f"엑셀 응답 형식이 예상과 다릅니다: {excel_url}")

    download_path.write_bytes(response.content)
    return download_path


def get_cached_download_path(location: str, check_no: int) -> Path:
    download_path = DOWNLOAD_DIR / f"{location}{check_no}.xls"
    if not download_path.exists():
        raise FileNotFoundError(f"캐시 파일이 없습니다: {download_path}")
    return download_path


def download_and_merge(
    location: str,
    check_no: int,
    *,
    use_cache: bool = False,
    dry_run: bool = False,
    workbook=None,
    session: requests.Session | None = None,
):
    if location not in LOCATION_CONFIG:
        raise ValueError(f"지원하지 않는 위치입니다: {location}")

    label = target_label(location, check_no)
    if use_cache:
        download_path = get_cached_download_path(location, check_no)
        log(f"{label}: 기존 다운로드 파일 사용")
    else:
        login_url, order_url = get_location_urls(location)
        start = time.perf_counter()
        if session is None:
            log(f"{label}: 로그인/다운로드 시작")
            with login_session(login_url) as own_session:
                download_path = download_excel_file(own_session, order_url, location, check_no)
        else:
            log(f"{label}: 다운로드 시작")
            download_path = download_excel_file(session, order_url, location, check_no)
        log(f"{label}: 다운로드 완료 ({elapsed_text(start)})")

    start = time.perf_counter()
    title, headers, rows = parse_download_table(download_path, location)
    target_month = get_target_month(location, check_no)
    if target_month not in title:
        log(f"{label}: 주의 - 파일 제목에 대상 월 {target_month}이 없습니다. 제목: {title}")
    log(f"{label}: 결제완료 {len(rows)}명 파싱 완료 ({elapsed_text(start)})")

    if dry_run:
        return {"target": label, "rows": len(rows), "written": False}

    if workbook is None:
        write_output_sheet(title, headers, rows, location, check_no)
    else:
        write_output_sheet_to_workbook(workbook, title, headers, rows, location, check_no)

    log(f"{label}: 결과 시트 반영 완료")
    return {"target": label, "rows": len(rows), "written": True}


def process_targets(
    targets: list[tuple[str, int]],
    *,
    use_cache: bool = False,
    dry_run: bool = False,
):
    sessions = {}
    workbook = None
    results = []

    if not dry_run:
        ensure_output_file_closed()
        workbook = load_output_workbook()

    try:
        for location, check_no in targets:
            session = None
            if not use_cache:
                session = sessions.get(location)
                if session is None:
                    login_url, _ = get_location_urls(location)
                    start = time.perf_counter()
                    log(f"{location}: 로그인 시작")
                    session = login_session(login_url)
                    sessions[location] = session
                    log(f"{location}: 로그인 완료 ({elapsed_text(start)})")

            results.append(
                download_and_merge(
                    location,
                    check_no,
                    use_cache=use_cache,
                    dry_run=dry_run,
                    workbook=workbook,
                    session=session,
                )
            )

        if workbook is not None:
            start = time.perf_counter()
            workbook.save(OUTPUT_XLSX)
            log(f"결과 저장 완료: {OUTPUT_XLSX} ({elapsed_text(start)})")
    finally:
        if workbook is not None:
            workbook.close()
        for session in sessions.values():
            session.close()

    return results


def run(argv=None):
    args = parse_args(argv)
    if args.list_targets:
        print_available_targets()
        return

    targets = configured_targets() if args.all else args.target or DEFAULT_TARGETS
    total_start = time.perf_counter()

    log(f"처리 대상: {', '.join(target_label(*target) for target in targets)}")
    if args.cache:
        log("캐시 모드: 새 다운로드 없이 excel 폴더의 기존 파일을 사용합니다.")
    if args.dry_run:
        log("점검 모드: 결과 엑셀 저장을 생략합니다.")

    load_runtime_config()
    results = process_targets(targets, use_cache=args.cache, dry_run=args.dry_run)

    total_rows = sum(result["rows"] for result in results)
    log(f"전체 완료: {len(results)}개 대상, 결제완료 {total_rows}명 ({elapsed_text(total_start)})")

    if not args.dry_run and not args.no_open:
        os.startfile(str(OUTPUT_XLSX))


if __name__ == "__main__":
    run()
