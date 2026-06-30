from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "_sincheong" / "코딩교실 신청 현황.xlsx"
DEFAULT_SPREADSHEET_ID = "1eG8aPnjIbI2UQiAJbii6QnCoW2kgUK4fGRwQnKz6JT4"
DEFAULT_SHEETS = ("pohang1", "gumi1")
DAILY_STATUS_SHEET = "일별현황"
UPDATE_LOG_SHEET = "업데이트로그"
DASHBOARD_SHEET = "대시보드"
TARGET_SNAPSHOTS = (
    {
        "title": "포항 1기",
        "sheet_name": "pohang1",
        "online_cell": "F7",
        "offline_cell": "G5",
        "previous_online_cell": "H5",
        "additional_cell": "I5",
    },
    {
        "title": "구미 1기",
        "sheet_name": "gumi1",
        "online_cell": "F17",
        "offline_cell": "G15",
        "previous_online_cell": "H15",
        "additional_cell": "I15",
    },
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync local application status workbook to Google Sheets.")
    parser.add_argument("--credentials", required=True, help="Path to Google service-account JSON credentials.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Local workbook path.")
    parser.add_argument("--spreadsheet-id", default=DEFAULT_SPREADSHEET_ID, help="Google spreadsheet ID.")
    parser.add_argument(
        "--sheets",
        default=",".join(DEFAULT_SHEETS),
        help="Comma-separated worksheet names to sync from the local workbook.",
    )
    return parser.parse_args()


def import_gspread():
    try:
        import gspread
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: gspread. Run `py -3 -m pip install -r requirements.txt` first."
        ) from exc
    return gspread


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    return value


def trim_grid(rows: list[list[Any]]) -> list[list[Any]]:
    last_row = 0
    last_col = 0

    for row_index, row in enumerate(rows, start=1):
        row_has_value = False
        for col_index, value in enumerate(row, start=1):
            if value != "":
                row_has_value = True
                last_col = max(last_col, col_index)
        if row_has_value:
            last_row = row_index

    if last_row == 0 or last_col == 0:
        return [[]]

    return [row[:last_col] for row in rows[:last_row]]


def read_workbook_sheet(workbook_path: Path, sheet_name: str) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(workbook_path, data_only=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Sheet not found in local workbook: {sheet_name}")

        worksheet = workbook[sheet_name]
        rows = [
            [normalize_cell(cell.value) for cell in row]
            for row in worksheet.iter_rows()
        ]
        return trim_grid(rows)
    finally:
        workbook.close()


def get_or_create_worksheet(spreadsheet, title: str, rows: int, cols: int):
    try:
        return spreadsheet.worksheet(title)
    except Exception:
        return spreadsheet.add_worksheet(title=title, rows=max(rows, 100), cols=max(cols, 20))


def worksheet_update(worksheet, values: list[list[Any]]) -> None:
    row_count = max(len(values), 1)
    col_count = max((len(row) for row in values), default=1)

    if values and values != [[]]:
        # Do not clear first. If the Google update fails after a clear, the sheet is left blank.
        worksheet.resize(
            rows=max(row_count, worksheet.row_count, 1),
            cols=max(col_count, worksheet.col_count, 1),
        )
        try:
            worksheet.update(values=values, range_name="A1", value_input_option="USER_ENTERED")
        except TypeError:
            worksheet.update("A1", values, value_input_option="USER_ENTERED")
        worksheet.resize(rows=max(row_count, 1), cols=max(col_count, 1))


def validate_source_values(sheet_name: str, values: list[list[Any]]) -> None:
    nonempty_rows = [
        row for row in values
        if any(str(cell or "").strip() for cell in row)
    ]
    if len(nonempty_rows) < 2:
        raise RuntimeError(
            f"Local source for {sheet_name} is too small ({len(nonempty_rows)} non-empty rows). "
            "Skipped sync to avoid wiping Google Sheets data."
        )


def count_additional_rows(values: list[list[Any]]) -> int:
    if len(values) < 1:
        return 0

    headers = [str(value or "").strip() for value in values[0]]
    try:
        additional_col_index = headers.index("추가")
    except ValueError:
        return 0

    count = 0
    for row in values[1:]:
        if additional_col_index < len(row) and str(row[additional_col_index] or "").strip() == "추가":
            count += 1
    return count


def update_dashboard_additional_counts(dashboard, additional_counts: dict[str, int]) -> None:
    updates = []
    for target in TARGET_SNAPSHOTS:
        sheet_name = target.get("sheet_name")
        additional_cell = target.get("additional_cell")
        if sheet_name and additional_cell and sheet_name in additional_counts:
            updates.append(
                {
                    "range": additional_cell,
                    "values": [[additional_counts[sheet_name]]],
                }
            )

    if updates:
        dashboard.batch_update(updates, value_input_option="USER_ENTERED")


def clear_basic_filter(spreadsheet, worksheet) -> None:
    try:
        spreadsheet.batch_update(
            {"requests": [{"clearBasicFilter": {"sheetId": worksheet.id}}]}
        )
    except Exception:
        pass


def apply_data_row_format(spreadsheet, worksheet, rows: int, cols: int) -> None:
    if rows <= 2 or cols <= 0:
        return

    clear_basic_filter(spreadsheet, worksheet)
    spreadsheet.batch_update(
        {
            "requests": [
                {
                    "copyPaste": {
                        "source": {
                            "sheetId": worksheet.id,
                            "startRowIndex": 1,
                            "endRowIndex": 2,
                            "startColumnIndex": 0,
                            "endColumnIndex": cols,
                        },
                        "destination": {
                            "sheetId": worksheet.id,
                            "startRowIndex": 2,
                            "endRowIndex": rows,
                            "startColumnIndex": 0,
                            "endColumnIndex": cols,
                        },
                        "pasteType": "PASTE_FORMAT",
                        "pasteOrientation": "NORMAL",
                    }
                }
            ]
        }
    )


def remove_bold_format(spreadsheet, worksheet, rows: int, cols: int) -> None:
    if rows <= 0 or cols <= 0:
        return

    spreadsheet.batch_update(
        {
            "requests": [
                {
                    "repeatCell": {
                        "range": {
                            "sheetId": worksheet.id,
                            "startRowIndex": 0,
                            "endRowIndex": rows,
                            "startColumnIndex": 0,
                            "endColumnIndex": cols,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "textFormat": {
                                    "bold": False,
                                }
                            }
                        },
                        "fields": "userEnteredFormat.textFormat.bold",
                    }
                }
            ]
        }
    )


def reset_basic_filter(spreadsheet, worksheet, rows: int, cols: int) -> None:
    if rows <= 0 or cols <= 0:
        return

    clear_basic_filter(spreadsheet, worksheet)
    spreadsheet.batch_update(
        {
            "requests": [
                {
                    "setBasicFilter": {
                        "filter": {
                            "range": {
                                "sheetId": worksheet.id,
                                "startRowIndex": 0,
                                "endRowIndex": rows,
                                "startColumnIndex": 0,
                                "endColumnIndex": cols,
                            }
                        }
                    }
                }
            ]
        }
    )


def parse_date(value: Any, now: dt.datetime) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value or "").strip()
    if not text:
        return None

    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return dt.date(year, month, day)

    match = re.search(r"(\d{1,2})월\s*(\d{1,2})일", text)
    if match:
        month, day = map(int, match.groups())
        parsed = dt.date(now.year, month, day)
        if parsed > now.date():
            parsed = dt.date(now.year - 1, month, day)
        return parsed

    return None


def get_last_sync_date(spreadsheet, now: dt.datetime) -> dt.date | None:
    try:
        log_sheet = spreadsheet.worksheet(UPDATE_LOG_SHEET)
        values = log_sheet.get("A:A", value_render_option="UNFORMATTED_VALUE")
        for row in reversed(values):
            if not row:
                continue
            parsed = parse_date(row[0], now)
            if parsed:
                return parsed
    except Exception:
        pass

    try:
        dashboard = spreadsheet.worksheet(DASHBOARD_SHEET)
        values = dashboard.get("F1", value_render_option="FORMATTED_VALUE")
        if values and values[0]:
            parsed = parse_date(values[0][0], now)
            if parsed:
                return parsed
    except Exception:
        pass

    try:
        daily_status = spreadsheet.worksheet(DAILY_STATUS_SHEET)
        values = daily_status.get("A:A", value_render_option="UNFORMATTED_VALUE")
        dates = [parse_date(row[0], now) for row in values if row]
        dates = [date for date in dates if date]
        if dates:
            return max(dates)
    except Exception:
        pass

    return None


def parse_number(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)

    text = re.sub(r"[^\d.-]", "", str(value))
    if text in ("", "-", ".", "-."):
        return 0
    return int(float(text))


def read_number_cell(worksheet, cell: str) -> int:
    values = worksheet.get(cell, value_render_option="UNFORMATTED_VALUE")
    if not values or not values[0]:
        return 0
    return parse_number(values[0][0])


def read_dashboard_snapshots(dashboard) -> list[dict[str, Any]]:
    snapshots = []
    for target in TARGET_SNAPSHOTS:
        online = read_number_cell(dashboard, target["online_cell"])
        offline = read_number_cell(dashboard, target["offline_cell"])
        snapshots.append(
            {
                "target": target["title"],
                "online": online,
                "offline": offline,
                "total": online + offline,
                "previous_online_cell": target["previous_online_cell"],
            }
        )
    return snapshots


def update_previous_online_cells(dashboard, snapshots: list[dict[str, Any]]) -> None:
    updates = [
        {
            "range": snapshot["previous_online_cell"],
            "values": [[snapshot["online"]]],
        }
        for snapshot in snapshots
    ]
    dashboard.batch_update(updates, value_input_option="USER_ENTERED")


def ensure_daily_status_header(worksheet) -> None:
    header = ["기준일", "대상", "온라인계", "오프라인", "총인원", "갱신시각", "비고"]
    values = worksheet.get("A1:G1", value_render_option="UNFORMATTED_VALUE")
    if values and values[0][: len(header)] == header:
        return
    worksheet.update(values=[header], range_name="A1:G1", value_input_option="USER_ENTERED")


def upsert_daily_status(
    spreadsheet,
    snapshot_date: dt.date,
    snapshots: list[dict[str, Any]],
    now: dt.datetime,
) -> None:
    worksheet = get_or_create_worksheet(spreadsheet, DAILY_STATUS_SHEET, 500, 8)
    ensure_daily_status_header(worksheet)

    existing_rows = worksheet.get("A:G", value_render_option="UNFORMATTED_VALUE")
    row_by_key = {}
    for row_index, row in enumerate(existing_rows[1:], start=2):
        if len(row) < 2:
            continue
        date_value = parse_date(row[0], now)
        target = str(row[1]).strip()
        if date_value and target:
            row_by_key[(date_value.isoformat(), target)] = row_index

    timestamp = now.strftime("%Y-%m-%d %H:%M:%S")
    note = "자동 전일 기록"
    append_rows = []
    update_rows = []

    for snapshot in snapshots:
        row_values = [
            snapshot_date.isoformat(),
            snapshot["target"],
            snapshot["online"],
            snapshot["offline"],
            snapshot["total"],
            timestamp,
            note,
        ]
        key = (snapshot_date.isoformat(), snapshot["target"])
        row_index = row_by_key.get(key)
        if row_index:
            update_rows.append((row_index, row_values))
        else:
            append_rows.append(row_values)

    for row_index, row_values in update_rows:
        worksheet.update(
            values=[row_values],
            range_name=f"A{row_index}:G{row_index}",
            value_input_option="USER_ENTERED",
        )

    if append_rows:
        worksheet.append_rows(append_rows, value_input_option="USER_ENTERED")


def carry_forward_previous_day(spreadsheet, now: dt.datetime) -> str | None:
    last_sync_date = get_last_sync_date(spreadsheet, now)
    if not last_sync_date or last_sync_date >= now.date():
        return None

    dashboard = spreadsheet.worksheet(DASHBOARD_SHEET)
    snapshots = read_dashboard_snapshots(dashboard)
    update_previous_online_cells(dashboard, snapshots)
    upsert_daily_status(spreadsheet, last_sync_date, snapshots, now)

    summary = ", ".join(
        f"{snapshot['target']} 온라인계 {snapshot['online']}명"
        for snapshot in snapshots
    )
    return f"{last_sync_date.isoformat()} previous-day snapshot saved: {summary}"


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook).resolve()
    credentials_path = Path(args.credentials).resolve()
    sheet_names = [name.strip() for name in args.sheets.split(",") if name.strip()]

    if not workbook_path.exists():
        print(f"Local workbook not found: {workbook_path}", file=sys.stderr)
        return 1
    if not credentials_path.exists():
        print(f"Google credentials not found: {credentials_path}", file=sys.stderr)
        return 1
    if not sheet_names:
        print("No sheets were requested.", file=sys.stderr)
        return 1

    gspread = import_gspread()
    client = gspread.service_account(filename=str(credentials_path))
    spreadsheet = client.open_by_key(args.spreadsheet_id)

    now = dt.datetime.now()
    rollover_message = carry_forward_previous_day(spreadsheet, now)
    if rollover_message:
        print(rollover_message)

    synced = []
    additional_counts = {}
    for sheet_name in sheet_names:
        values = read_workbook_sheet(workbook_path, sheet_name)
        validate_source_values(sheet_name, values)
        additional_counts[sheet_name] = count_additional_rows(values)
        rows = max(len(values), 1)
        cols = max((len(row) for row in values), default=1)
        worksheet = get_or_create_worksheet(spreadsheet, sheet_name, rows, cols)
        clear_basic_filter(spreadsheet, worksheet)
        worksheet_update(worksheet, values)
        apply_data_row_format(spreadsheet, worksheet, rows, cols)
        remove_bold_format(spreadsheet, worksheet, rows, cols)
        reset_basic_filter(spreadsheet, worksheet, rows, cols)
        synced.append(f"{sheet_name}({rows}x{cols})")

    try:
        dashboard = spreadsheet.worksheet(DASHBOARD_SHEET)
        dashboard.update_acell("F1", now.strftime("%m월 %d일 %H시 %M분 현황"))
        update_dashboard_additional_counts(dashboard, additional_counts)
    except Exception as exc:
        print(f"Dashboard timestamp update skipped: {exc}", file=sys.stderr)

    try:
        log_sheet = get_or_create_worksheet(spreadsheet, UPDATE_LOG_SHEET, 100, 4)
        log_sheet.append_row(
            [now.strftime("%Y-%m-%d %H:%M:%S"), "google_sync", ", ".join(synced), "OK"],
            value_input_option="USER_ENTERED",
        )
    except Exception as exc:
        print(f"Update log append skipped: {exc}", file=sys.stderr)

    print("Google Sheets sync completed: " + ", ".join(synced))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
