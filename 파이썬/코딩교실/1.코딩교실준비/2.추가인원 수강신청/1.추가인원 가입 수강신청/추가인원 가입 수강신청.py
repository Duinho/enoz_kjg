# 라이브러리 설치 메모(터미널):
# pip install openpyxl playwright
# playwright install

import os
import json
import re
import openpyxl
from openpyxl import Workbook
from playwright.sync_api import sync_playwright
import random
import time
from datetime import datetime, timedelta
from collections import deque

# --- 경로 설정 --------------------------------------------------------------
폴더경로 = os.path.dirname(os.path.abspath(__file__))
# (참고) 이전 버전의 persistent user data dir는 사용하지 않습니다. 매 반복 새 컨텍스트 전략으로 전환.
excel_file_path = os.path.join(폴더경로, '추가인원 가입 수강신청.xlsx')
status_file_path = os.path.join(폴더경로, '추가인원 가입 현황.xlsx')
debug_log_path = os.path.join(폴더경로, 'signup_debug.log')
run_state_path = os.path.join(폴더경로, 'signup_run_state.json')

DASHBOARD_SHEET_NAME = "대시보드"
TARGET_SHEET_NAME = "BG"
TARGET_GENERATION = "1"
STATUS_SHEET_NAME_CACHE = {}
ADDRESS_REQUIRED_KEYWORDS = {
    "BG": "북구",
    "PH": "포항",
}
TOTAL_TARGET_COUNT = 200
DAILY_TARGET_MIN = 0
DAILY_TARGET_MAX = 120
DAILY_TARGET_BUFFER = 5
TODAY_REMAINING_TARGET_OVERRIDE_DATE = "2026-04-28"
TODAY_REMAINING_TARGET_OVERRIDE_COUNT = 80
TODAY_REMAINING_TARGET_OVERRIDE_KEY = (
    f"remaining_boost_{TODAY_REMAINING_TARGET_OVERRIDE_DATE}_{TODAY_REMAINING_TARGET_OVERRIDE_COUNT}"
)
ACTIVE_START_HOUR = 9
ACTIVE_END_HOUR = 21
DEADLINE_AT = datetime(2026, 6, 17, 21, 0, 0)
SCHEDULE_DEADLINE_BUFFER_MINUTES = 10
SIGNUP_MAX_ATTEMPTS = 4
APPLY_RETRY_DELAY_SECONDS = 1
OUTER_RETRY_DELAY_SECONDS = 0
POST_SUCCESS_SLEEP_MIN_SECONDS = 0
POST_SUCCESS_SLEEP_MAX_SECONDS = 0
MAX_APPLY_ATTEMPTS_PER_MEMBER = 2
RUN_WITHOUT_DELAY = False
URGENT_FORCE_DAY = ""
DASHBOARD_SETTING_CELLS = {
    "target_sheet": "B4",
    "target_generation": "B5",
    "target_total": "B6",
    "deadline_at": "B7",
    "active_start_hour": "B8",
    "active_end_hour": "B9",
    "random_schedule": "B10",
    "daily_target_min": "B11",
    "daily_target_max": "B12",
    "daily_target_buffer": "B13",
    "post_success_sleep_min": "B14",
    "post_success_sleep_max": "B15",
}
DASHBOARD_PROGRESS_CELLS = {
    "status_sheet": "E4",
    "current_total": "E5",
    "today_count": "E6",
    "remaining_total": "E7",
    "today_target": "E8",
    "today_remaining": "E9",
    "schedule_count": "E10",
    "next_schedule": "E11",
    "last_schedule": "E12",
    "last_updated": "E13",
    "run_status": "E14",
    "latest_log": "E15",
}

def debug_log(message):
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line = f"[{timestamp}] {message}"
    print(line, flush=True)
    with open(debug_log_path, "a", encoding="utf-8") as f:
        f.write(line + "\n")

def text_or_empty(value):
    return "" if value is None else str(value).strip()

def parse_int_setting(value, default, *, minimum=None, maximum=None):
    text = text_or_empty(value)
    if not text:
        return default
    try:
        parsed = int(float(text))
    except Exception:
        return default
    if minimum is not None:
        parsed = max(minimum, parsed)
    if maximum is not None:
        parsed = min(maximum, parsed)
    return parsed

def parse_bool_setting(value, default=False):
    if isinstance(value, bool):
        return value
    text = text_or_empty(value).lower()
    if not text:
        return default
    if text in {"1", "true", "yes", "y", "on", "예", "사용", "분산", "랜덤"}:
        return True
    if text in {"0", "false", "no", "n", "off", "아니오", "미사용", "즉시"}:
        return False
    return default

def parse_datetime_setting(value, default):
    if isinstance(value, datetime):
        return value.replace(second=0, microsecond=0)
    text = text_or_empty(value)
    if not text:
        return default
    candidates = [
        text,
        text.replace(".", "-").replace("/", "-"),
    ]
    for candidate in candidates:
        for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(candidate, fmt)
                if fmt == "%Y-%m-%d":
                    parsed = parsed.replace(hour=21, minute=0)
                return parsed
            except ValueError:
                pass
    try:
        return datetime.fromisoformat(text)
    except Exception:
        return default

def load_dashboard_settings():
    if not os.path.exists(excel_file_path):
        return {}

    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    try:
        if DASHBOARD_SHEET_NAME not in workbook.sheetnames:
            return {}
        worksheet = workbook[DASHBOARD_SHEET_NAME]
        return {
            key: worksheet[cell].value
            for key, cell in DASHBOARD_SETTING_CELLS.items()
        }
    finally:
        workbook.close()

def apply_dashboard_settings():
    global TARGET_SHEET_NAME, TARGET_GENERATION, TOTAL_TARGET_COUNT
    global DAILY_TARGET_MIN, DAILY_TARGET_MAX, DAILY_TARGET_BUFFER
    global ACTIVE_START_HOUR, ACTIVE_END_HOUR, DEADLINE_AT
    global POST_SUCCESS_SLEEP_MIN_SECONDS, POST_SUCCESS_SLEEP_MAX_SECONDS
    global RUN_WITHOUT_DELAY

    settings = load_dashboard_settings()
    if not settings:
        return

    previous_target = TARGET_SHEET_NAME
    target_sheet = text_or_empty(settings.get("target_sheet")).upper()
    if target_sheet in {"BG", "PH"}:
        TARGET_SHEET_NAME = target_sheet

    TARGET_GENERATION = parse_generation_number(settings.get("target_generation"))
    TOTAL_TARGET_COUNT = parse_int_setting(
        settings.get("target_total"), TOTAL_TARGET_COUNT, minimum=1
    )
    DEADLINE_AT = parse_datetime_setting(settings.get("deadline_at"), DEADLINE_AT)
    ACTIVE_START_HOUR = parse_int_setting(
        settings.get("active_start_hour"), ACTIVE_START_HOUR, minimum=0, maximum=23
    )
    ACTIVE_END_HOUR = parse_int_setting(
        settings.get("active_end_hour"), ACTIVE_END_HOUR, minimum=1, maximum=24
    )
    if ACTIVE_END_HOUR <= ACTIVE_START_HOUR:
        ACTIVE_END_HOUR = min(24, ACTIVE_START_HOUR + 1)

    random_schedule = parse_bool_setting(settings.get("random_schedule"), not RUN_WITHOUT_DELAY)
    RUN_WITHOUT_DELAY = not random_schedule
    DAILY_TARGET_MIN = parse_int_setting(
        settings.get("daily_target_min"), DAILY_TARGET_MIN, minimum=0
    )
    DAILY_TARGET_MAX = parse_int_setting(
        settings.get("daily_target_max"), DAILY_TARGET_MAX, minimum=1
    )
    DAILY_TARGET_BUFFER = parse_int_setting(
        settings.get("daily_target_buffer"), DAILY_TARGET_BUFFER, minimum=0
    )
    POST_SUCCESS_SLEEP_MIN_SECONDS = parse_int_setting(
        settings.get("post_success_sleep_min"), POST_SUCCESS_SLEEP_MIN_SECONDS, minimum=0
    )
    POST_SUCCESS_SLEEP_MAX_SECONDS = parse_int_setting(
        settings.get("post_success_sleep_max"), POST_SUCCESS_SLEEP_MAX_SECONDS, minimum=0
    )
    if POST_SUCCESS_SLEEP_MAX_SECONDS < POST_SUCCESS_SLEEP_MIN_SECONDS:
        POST_SUCCESS_SLEEP_MAX_SECONDS = POST_SUCCESS_SLEEP_MIN_SECONDS

    if previous_target != TARGET_SHEET_NAME:
        STATUS_SHEET_NAME_CACHE.clear()

# --- 전역 상태 --------------------------------------------------------------
클릭 = 0
학교_리스트    = []
아파트_리스트  = []
동_리스트      = []
아이디1_리스트 = []
아이디2_리스트 = []
아이디3_리스트 = []
남아_리스트    = []
여아_리스트    = []
부_리스트      = []   # 아버지 이름 풀(‘이름’ 부분만)
모_리스트      = []   # 어머니 이름 풀(‘이름’ 부분만)

# 이름 분산용 큐/카운트(히스토리 포함)
남아_큐 = deque()
여아_큐 = deque()
부_큐  = deque()
모_큐  = deque()

이름_카운트 = {}      # 자녀 이름 사용 누적(성 제외)
부모_이름_카운트 = {}  # 부모 이름 사용 누적(성 제외, 부/모 통합 카운트)
사용_전체이름 = set()
사용_아이디 = set()

남아_확장_음절 = [
    "건", "겸", "결", "도", "민", "서", "선", "성", "시", "안", "우", "율",
    "윤", "이", "재", "준", "지", "찬", "태", "하", "현", "호", "후", "휘"
]
여아_확장_음절 = [
    "가", "나", "다", "라", "린", "민", "봄", "서", "설", "소", "수", "시",
    "아", "연", "예", "온", "유", "윤", "은", "이", "지", "채", "하", "현"
]
ID_EXTRA_WORDS = [
    "nova", "pixel", "orbit", "logic", "bright", "coding", "future", "vector",
    "spark", "matrix", "lambda", "studio", "canvas", "signal", "planet", "rocket",
    "bridge", "forest", "river", "silver", "crystal", "marble", "pepper", "mint",
    "cloud", "alpha", "delta", "omega", "prism", "focus", "dream", "value"
]

# --- 성씨 및 확률 ----------------------------------------------------------
성 = [
    ("김", 0.21782), ("이", 0.14882), ("박", 0.08536), ("최", 0.04754), ("정", 0.04402),
    ("강", 0.02355), ("조", 0.02161), ("윤", 0.02077), ("장", 0.02017), ("임", 0.01674),
    ("한", 0.01560), ("오", 0.01548), ("서", 0.01524), ("신", 0.01526), ("권", 0.01427),
    ("황", 0.01412), ("안", 0.01393), ("송", 0.01389), ("류", 0.01300), ("전", 0.01098),
    ("홍", 0.01137), ("고", 0.00957), ("문", 0.00938), ("양", 0.00884), ("손", 0.00919),
    ("배", 0.00812), ("백", 0.00773), ("허", 0.00660), ("유", 0.00566), ("남", 0.00559),
    ("심", 0.00551), ("노", 0.00498), ("정", 0.00442), ("하", 0.00463), ("곽", 0.00410),
    ("성", 0.00405), ("차", 0.00396), ("주", 0.00389), ("우", 0.00390), ("구", 0.00392),
    ("신", 0.00369), ("임", 0.00379), ("라", 0.00377), ("전", 0.00386), ("민", 0.00350),
    ("유", 0.00381), ("진", 0.00315), ("지", 0.00311), ("엄", 0.00291), ("채", 0.00254),
    ("원", 0.00262), ("천", 0.00239), ("방", 0.00191), ("공", 0.00183), ("강", 0.00201),
    ("현", 0.00180), ("함", 0.00164), ("변", 0.00164), ("염", 0.00139), ("양", 0.00181),
    ("변", 0.00121), ("여", 0.00123), ("추", 0.00121), ("노", 0.00134), ("도", 0.00115),
    ("소", 0.00098), ("신", 0.00103), ("석", 0.00100), ("선", 0.00085), ("설", 0.00086),
    ("마", 0.00078), ("길", 0.00075), ("주", 0.00081), ("연", 0.00068), ("방", 0.00068),
    ("위", 0.00062), ("표", 0.00062), ("명", 0.00059), ("기", 0.00056), ("반", 0.00054),
    ("왕", 0.00051), ("금", 0.00051), ("옥", 0.00051), ("육", 0.00047), ("인", 0.00045),
    ("맹", 0.00044), ("제", 0.00044), ("모", 0.00042), ("장", 0.00041), ("남", 0.00042),
    ("탁", 0.00043), ("국", 0.00039), ("여", 0.00039), ("진", 0.00042), ("어", 0.00038),
    ("은", 0.00034), ("편", 0.00033), ("구", 0.00029), ("용", 0.00016), ("남궁", 0.00015)
]
성씨, weights = zip(*성)

# --- 고정 비밀번호 ----------------------------------------------------------
비밀번호 = 'enoz7223!'

# --- 정보 시트에서 불러올 변수들 -------------------------------------------
반복횟수     = 0
랜덤최소     = 0
랜덤최대     = 0
gender_plan_queue = deque()
day_plan_queue = deque()

def build_weighted_gender_plan(total_count):
    total_count = max(0, int(total_count or 0))
    female_count = min(total_count, int(total_count * 0.55 + 0.5))
    male_count = total_count - female_count
    plan = (["여자"] * female_count) + (["남자"] * male_count)
    random.shuffle(plan)
    return deque(plan)

def build_even_day_plan(total_count):
    total_count = max(0, int(total_count or 0))
    month_wed_count = total_count // 2
    tue_thu_count = total_count - month_wed_count

    if total_count % 2 == 1 and random.choice([True, False]):
        month_wed_count, tue_thu_count = tue_thu_count, month_wed_count

    plan = (["월수"] * month_wed_count) + (["화목"] * tue_thu_count)
    random.shuffle(plan)
    return deque(plan)

def reset_signup_plans(total_count):
    global gender_plan_queue, day_plan_queue
    gender_plan_queue = build_weighted_gender_plan(total_count)
    day_plan_queue = build_even_day_plan(total_count)

def pop_signup_plan():
    if not gender_plan_queue or not day_plan_queue:
        raise RuntimeError("가입 계획 큐가 비어 있습니다. 초기화를 다시 실행하세요.")

    return {
        "sex": gender_plan_queue.popleft(),
        "day": URGENT_FORCE_DAY if RUN_WITHOUT_DELAY and URGENT_FORCE_DAY else day_plan_queue.popleft(),
    }

def load_run_state():
    if not os.path.exists(run_state_path):
        return {}

    try:
        with open(run_state_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        debug_log(f"상태 파일 로드 실패. 새 상태로 시작합니다. error={repr(e)}")
        return {}

def save_run_state(state):
    temp_path = run_state_path + ".tmp"
    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)
    os.replace(temp_path, run_state_path)

def get_status_counts(sheet_name=TARGET_SHEET_NAME):
    if not os.path.exists(status_file_path):
        return 0, 0

    status_sheet_name = resolve_status_sheet_name(sheet_name)
    today_str = datetime.now().strftime("%Y-%m-%d")
    total_count = 0
    today_count = 0

    workbook = openpyxl.load_workbook(status_file_path, data_only=True)
    try:
        if status_sheet_name not in workbook.sheetnames:
            return 0, 0

        worksheet = workbook[status_sheet_name]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(value is not None for value in row):
                continue
            total_count += 1
            if str(row[-1] or "").startswith(today_str):
                today_count += 1
    finally:
        workbook.close()

    return total_count, today_count

def parse_generation_number(value):
    text = text_or_empty(value)
    if not text:
        return ""
    match = re.search(r"\d+", text)
    return match.group(0) if match else ""

def load_generation_number(sheet_name):
    if not os.path.exists(excel_file_path):
        return ""

    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return ""

        worksheet = workbook[sheet_name]
        for row_index in range(1, worksheet.max_row + 1):
            for col_index in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row=row_index, column=col_index).value
                text = text_or_empty(value)
                if "기수" not in text:
                    continue

                own_generation = parse_generation_number(text)
                if own_generation:
                    return own_generation

                for dr, dc in ((0, 1), (1, 0), (0, 2), (2, 0)):
                    next_row = row_index + dr
                    next_col = col_index + dc
                    if next_row > worksheet.max_row or next_col > worksheet.max_column:
                        continue
                    generation = parse_generation_number(
                        worksheet.cell(row=next_row, column=next_col).value
                    )
                    if generation:
                        return generation

        # Convention: S3/T3 can be used as a compact settings pair.
        if text_or_empty(worksheet.cell(row=3, column=19).value) == "기수":
            return parse_generation_number(worksheet.cell(row=3, column=20).value)
        return parse_generation_number(worksheet.cell(row=3, column=20).value)
    finally:
        workbook.close()

def infer_generation_from_status_file(sheet_name):
    if not os.path.exists(status_file_path):
        return ""

    workbook = openpyxl.load_workbook(status_file_path, data_only=True)
    try:
        candidates = []
        for status_sheet in workbook.sheetnames:
            if not status_sheet.startswith(sheet_name):
                continue
            suffix = status_sheet[len(sheet_name):]
            if suffix.isdigit():
                candidates.append(suffix)
        if len(candidates) == 1:
            return candidates[0]
        return ""
    finally:
        workbook.close()

def resolve_status_sheet_name(sheet_name):
    if sheet_name in STATUS_SHEET_NAME_CACHE:
        return STATUS_SHEET_NAME_CACHE[sheet_name]

    generation = TARGET_GENERATION if sheet_name == TARGET_SHEET_NAME else ""
    if not generation:
        generation = load_generation_number(sheet_name)
    if not generation:
        generation = infer_generation_from_status_file(sheet_name)
    status_sheet_name = f"{sheet_name}{generation}" if generation else sheet_name
    STATUS_SHEET_NAME_CACHE[sheet_name] = status_sheet_name
    if generation:
        debug_log(f"현황 저장 시트 결정: {sheet_name} + {generation}기 -> {status_sheet_name}")
    else:
        debug_log(f"현황 저장 시트 결정: {sheet_name} 기수 미설정 -> {status_sheet_name}")
    return status_sheet_name

def active_window_bounds(now=None):
    now = now or datetime.now()
    start = now.replace(hour=ACTIVE_START_HOUR, minute=0, second=0, microsecond=0)
    end = now.replace(hour=ACTIVE_END_HOUR, minute=0, second=0, microsecond=0)
    return start, end

def schedule_window_end(now=None):
    now = now or datetime.now()
    _, active_end = active_window_bounds(now)
    if now.date() == DEADLINE_AT.date() and now < DEADLINE_AT:
        buffered_deadline = DEADLINE_AT - timedelta(minutes=SCHEDULE_DEADLINE_BUFFER_MINUTES)
        return min(active_end, buffered_deadline)
    return active_end

def is_active_time(now=None):
    now = now or datetime.now()
    start, end = active_window_bounds(now)
    return start <= now < end

def next_active_start(now=None):
    now = now or datetime.now()
    start, end = active_window_bounds(now)
    if now < start:
        return start
    return (now + timedelta(days=1)).replace(
        hour=ACTIVE_START_HOUR, minute=0, second=0, microsecond=0
    )

def sleep_until(target_time, reason):
    debug_log(f"{reason}: {target_time.strftime('%Y-%m-%d %H:%M:%S')}까지 대기")
    while True:
        remaining_seconds = (target_time - datetime.now()).total_seconds()
        if remaining_seconds <= 0:
            return
        time.sleep(min(remaining_seconds, 300))

def parse_schedule_time(value):
    try:
        return datetime.fromisoformat(value)
    except Exception:
        return None

def generate_random_schedule_times(count, start_time, end_time):
    if count <= 0:
        return []

    seconds = max(0, int((end_time - start_time).total_seconds()))
    if seconds <= 0:
        return [start_time.isoformat(timespec="seconds") for _ in range(count)]

    interval = seconds / count
    offsets = []
    for index in range(count):
        center = interval * (index + 0.5)
        jitter = random.uniform(-interval * 0.35, interval * 0.35)
        offsets.append(int(max(0, min(seconds, center + jitter))))
    offsets = sorted(offsets)
    return [
        (start_time + timedelta(seconds=offset)).isoformat(timespec="seconds")
        for offset in offsets
    ]

def remaining_work_days_until_deadline(now=None):
    now = now or datetime.now()
    if now >= DEADLINE_AT:
        return 1

    days = 0
    current_day = now.date()
    deadline_day = DEADLINE_AT.date()
    while current_day <= deadline_day:
        days += 1
        current_day += timedelta(days=1)
    return max(1, days)

def calculate_daily_target(remaining_total, today_count, now=None):
    now = now or datetime.now()
    max_possible_today = max(0, today_count + remaining_total)
    if max_possible_today <= 0:
        return today_count
    if now >= DEADLINE_AT or now.date() == DEADLINE_AT.date():
        return max_possible_today

    remaining_days = remaining_work_days_until_deadline(now)
    required_daily = (max_possible_today + remaining_days - 1) // remaining_days
    target = max(DAILY_TARGET_MIN, required_daily + DAILY_TARGET_BUFFER)
    return min(max_possible_today, min(target, DAILY_TARGET_MAX))

def apply_today_remaining_override(state, daily_target, today_count, max_possible_today, today):
    if today != TODAY_REMAINING_TARGET_OVERRIDE_DATE:
        return daily_target
    if state.get(TODAY_REMAINING_TARGET_OVERRIDE_KEY):
        return daily_target

    boosted_target = min(max_possible_today, today_count + TODAY_REMAINING_TARGET_OVERRIDE_COUNT)
    if boosted_target > daily_target:
        debug_log(
            f"오늘 남은 목표 강제 보정: {daily_target} -> {boosted_target} "
            f"(현재 {today_count}명 + 남은 {TODAY_REMAINING_TARGET_OVERRIDE_COUNT}명)"
        )
        daily_target = boosted_target

    state[TODAY_REMAINING_TARGET_OVERRIDE_KEY] = True
    return daily_target

def get_daily_plan(remaining_total, today_count):
    today = datetime.now().strftime("%Y-%m-%d")
    state = load_run_state()
    plan_key = {
        "target_sheet": TARGET_SHEET_NAME,
        "target_total": TOTAL_TARGET_COUNT,
        "deadline_at": DEADLINE_AT.isoformat(timespec="seconds"),
        "run_without_delay": RUN_WITHOUT_DELAY,
    }
    max_possible_today = max(0, today_count + remaining_total)
    required_daily_target = calculate_daily_target(remaining_total, today_count)

    state_matches_plan = all(state.get(key) == value for key, value in plan_key.items())

    if state.get("date") != today or not state_matches_plan:
        daily_target = required_daily_target
        state = {
            "date": today,
            "daily_target": daily_target,
            "schedule_times": [],
            **plan_key,
        }
        debug_log(f"오늘 목표 설정: {daily_target}명")
    else:
        daily_target = int(state.get("daily_target") or 0)
        if daily_target <= 0:
            daily_target = required_daily_target
        elif daily_target < required_daily_target:
            debug_log(f"오늘 목표 상향: {daily_target} -> {required_daily_target}")
            daily_target = required_daily_target

    daily_target = apply_today_remaining_override(
        state, daily_target, today_count, max_possible_today, today
    )

    if max_possible_today > 0 and daily_target > max_possible_today:
        daily_target = max_possible_today
        state["daily_target"] = daily_target

    remaining_today = max(0, daily_target - today_count)
    schedule_times = []
    now = datetime.now()
    for value in state.get("schedule_times", []):
        parsed = parse_schedule_time(value)
        if parsed and parsed.date() == now.date() and parsed >= now:
            schedule_times.append(parsed.isoformat(timespec="seconds"))

    schedule_times = sorted(schedule_times)
    if len(schedule_times) > remaining_today:
        schedule_times = schedule_times[:remaining_today]

    if RUN_WITHOUT_DELAY:
        schedule_times = [
            now.isoformat(timespec="seconds")
            for _ in range(remaining_today)
        ]
    elif len(schedule_times) < remaining_today:
        start, _ = active_window_bounds(now)
        end = schedule_window_end(now)
        if end <= now:
            end = active_window_bounds(now)[1]
        schedule_start = max(now, start)
        missing = remaining_today - len(schedule_times)
        schedule_times.extend(generate_random_schedule_times(missing, schedule_start, end))
        schedule_times = sorted(schedule_times)

    state["date"] = today
    state["daily_target"] = daily_target
    state["schedule_times"] = schedule_times
    state.update(plan_key)
    save_run_state(state)
    return daily_target, schedule_times

def consume_schedule_slot():
    today = datetime.now().strftime("%Y-%m-%d")
    state = load_run_state()
    if state.get("date") != today:
        return

    schedule_times = list(state.get("schedule_times", []))
    if schedule_times:
        schedule_times.pop(0)
        state["schedule_times"] = schedule_times
        save_run_state(state)

def update_dashboard_progress(run_status="", latest_log=""):
    if not os.path.exists(excel_file_path):
        return

    try:
        status_sheet_name = resolve_status_sheet_name(TARGET_SHEET_NAME)
        total_count, today_count = get_status_counts(TARGET_SHEET_NAME)
        remaining_total = max(0, TOTAL_TARGET_COUNT - total_count)
        state = load_run_state()
        daily_target = int(state.get("daily_target") or 0)
        if daily_target <= 0:
            daily_target = calculate_daily_target(remaining_total, today_count)

        now = datetime.now()
        future_schedules = []
        for value in state.get("schedule_times", []):
            parsed = parse_schedule_time(value)
            if parsed and parsed >= now:
                future_schedules.append(parsed)
        future_schedules.sort()

        workbook = openpyxl.load_workbook(excel_file_path)
        try:
            if DASHBOARD_SHEET_NAME not in workbook.sheetnames:
                return
            worksheet = workbook[DASHBOARD_SHEET_NAME]
            values = {
                "status_sheet": status_sheet_name,
                "current_total": total_count,
                "today_count": today_count,
                "remaining_total": remaining_total,
                "today_target": daily_target,
                "today_remaining": max(0, daily_target - today_count),
                "schedule_count": len(future_schedules),
                "next_schedule": future_schedules[0].strftime("%Y-%m-%d %H:%M:%S") if future_schedules else "",
                "last_schedule": future_schedules[-1].strftime("%Y-%m-%d %H:%M:%S") if future_schedules else "",
                "last_updated": now.strftime("%Y-%m-%d %H:%M:%S"),
                "run_status": run_status,
                "latest_log": latest_log,
            }
            for key, cell in DASHBOARD_PROGRESS_CELLS.items():
                worksheet[cell] = values.get(key, "")
            workbook.save(excel_file_path)
        finally:
            workbook.close()
    except PermissionError as exc:
        debug_log(f"대시보드 갱신 실패: 엑셀 파일이 열려 있거나 잠겨 있습니다. error={repr(exc)}")
    except Exception as exc:
        debug_log(f"대시보드 갱신 실패: error={repr(exc)}")

# --- 유틸: 복성 처리 --------------------------------------------------------
복성_목록 = {'남궁','황보','제갈','선우','서문','독고','동방','사공','사마','소봉','어금','장곡','탁안'}

def 이름부분만(전체이름: str) -> str:
    """성(1글자 또는 복성 2글자)을 제거한 '이름'만 반환."""
    if not 전체이름:
        return ""
    if len(전체이름) >= 3 and 전체이름[:2] in 복성_목록:
        return 전체이름[2:]
    return 전체이름[1:]

# --- 히스토리 로딩: 이름/부모이름 카운트 -----------------------------------
def 이름_카운트_로드():
    """가입 현황 파일에서 자녀/부모 이름(성 제외) 사용 누적 카운트를 읽어옴."""
    global 사용_전체이름, 사용_아이디
    child_counts = {}
    parent_counts = {}
    used_names = set()
    used_ids = set()

    현황_경로 = status_file_path
    if not os.path.exists(현황_경로):
        사용_전체이름 = used_names
        사용_아이디 = used_ids
        return child_counts, parent_counts

    wb = openpyxl.load_workbook(현황_경로, data_only=True)
    for 시트 in wb.sheetnames:
        ws = wb[시트]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            전체이름 = row[0]
            if 전체이름:
                used_names.add(str(전체이름).strip())
                nm = 이름부분만(전체이름)
                if nm:
                    child_counts[nm] = child_counts.get(nm, 0) + 1
            if len(row) >= 2 and row[1]:
                used_ids.add(str(row[1]).strip().lower())
            if len(row) >= 8:
                부모이름 = row[-2]
                if 부모이름:
                    pnm = 이름부분만(부모이름)
                    if pnm:
                        parent_counts[pnm] = parent_counts.get(pnm, 0) + 1
    wb.close()
    사용_전체이름 = used_names
    사용_아이디 = used_ids
    return child_counts, parent_counts

# --- 큐 초기화 --------------------------------------------------------------
def 큐_초기화():
    """남/여/부/모 이름 큐 섞어서 준비(1회전 중복 없음)."""
    global 남아_큐, 여아_큐, 부_큐, 모_큐
    남 = [x for x in 남아_리스트 if x]
    여 = [x for x in 여아_리스트 if x]
    부 = [x for x in 부_리스트 if x]
    모 = [x for x in 모_리스트 if x]
    random.shuffle(남)
    random.shuffle(여)
    random.shuffle(부)
    random.shuffle(모)
    남아_큐 = deque(남)
    여아_큐 = deque(여)
    부_큐  = deque(부)
    모_큐  = deque(모)

# --- 분산 추출 --------------------------------------------------------------
def 분산_선택(데크: deque, 카운트_딕트: dict) -> str:
    """카운트가 가장 적은 후보군에서 랜덤 1개 선택 후 카운트 +1."""
    if not 데크:
        return ""
    풀 = list(데크)
    최소 = min(카운트_딕트.get(n, 0) for n in 풀)
    후보 = [n for n in 풀 if 카운트_딕트.get(n, 0) == 최소]
    선택 = random.choice(후보)
    데크.remove(선택)
    카운트_딕트[선택] = 카운트_딕트.get(선택, 0) + 1
    return 선택

def 확장_이름후보(성별값):
    pool = 남아_확장_음절 if 성별값 == "남자" else 여아_확장_음절
    return random.choice(pool) + random.choice(pool)

def 사용가능_학생이름_생성(성별값):
    global 남아_큐, 여아_큐, 이름_카운트, 사용_전체이름
    source = 남아_리스트 if 성별값 == "남자" else 여아_리스트
    queue_name = "남자" if 성별값 == "남자" else "여자"

    for _ in range(300):
        성한자 = random.choices(성씨, weights=weights, k=1)[0]
        if 성별값 == "남자":
            if not 남아_큐:
                큐_초기화()
            이름후 = 분산_선택(남아_큐, 이름_카운트)
        else:
            if not 여아_큐:
                큐_초기화()
            이름후 = 분산_선택(여아_큐, 이름_카운트)
        if not 이름후 and source:
            이름후 = random.choice(source)
        if not 이름후:
            이름후 = 확장_이름후보(성별값)

        이름 = 성한자 + 이름후
        if 이름 not in 사용_전체이름:
            사용_전체이름.add(이름)
            return 성한자, 이름후, 이름

    for _ in range(1000):
        성한자 = random.choices(성씨, weights=weights, k=1)[0]
        이름후 = 확장_이름후보(성별값)
        이름 = 성한자 + 이름후
        if 이름 not in 사용_전체이름:
            이름_카운트[이름후] = 이름_카운트.get(이름후, 0) + 1
            사용_전체이름.add(이름)
            return 성한자, 이름후, 이름

    raise RuntimeError(f"{queue_name} 학생 이름 후보를 더 만들 수 없습니다.")

def 아이디_정리(value):
    return ''.join(
        ch.lower() for ch in str(value)
        if ('a' <= ch.lower() <= 'z') or ch.isdigit()
    )

def 사용가능_아이디_생성():
    global 사용_아이디
    words = []
    for pool in [아이디1_리스트, 아이디2_리스트, 아이디3_리스트, ID_EXTRA_WORDS]:
        for item in pool:
            cleaned = 아이디_정리(item)
            if cleaned:
                words.append(cleaned)
    words = sorted(set(words))

    for _ in range(1000):
        if words:
            part_count = min(len(words), random.choice([2, 2, 3]))
            base = ''.join(random.sample(words, part_count))
        else:
            base = "user"

        suffix_options = [
            f"{random.randint(10, 9999)}",
            f"{random.choice('abcdefghijklmnopqrstuvwxyz')}{random.randint(10, 999)}",
            f"{random.randint(10, 99)}{random.choice('abcdefghijklmnopqrstuvwxyz')}",
        ]
        suffix = random.choice(suffix_options)
        max_base_len = max(1, 15 - len(suffix))
        candidate = 아이디_정리(base[:max_base_len] + suffix)[:15]
        if len(candidate) < 4:
            continue
        if candidate not in 사용_아이디:
            사용_아이디.add(candidate)
            return candidate

    while True:
        candidate = f"u{int(time.time() * 1000) % 100000000}{random.randint(10, 99)}"[:15]
        if candidate not in 사용_아이디:
            사용_아이디.add(candidate)
            return candidate

def 약관_통과_후_가입폼_진입(page, timeout=15000, max_retry=3):
    """
    약관 페이지에서 모든 동의 체크 후 '다음'을 눌러 실제 가입 폼으로 진입합니다.
    - 체크박스(id^="cbAgree") 전부 강제 체크
    - '다음' 클릭 후 URL/필드 존재로 폼 진입 검증
    - 실패 시 최대 max_retry회 재시도
    사용 예:
        page.goto(회원가입사이트, wait_until="domcontentloaded")
        약관_통과_후_가입폼_진입(page)
    """
    import re, time

    # 초기 안정화
    try:
        page.wait_for_load_state('domcontentloaded', timeout=timeout)
        page.wait_for_load_state('networkidle', timeout=timeout)
    except Exception:
        pass

    # 공통 셀렉터들
    checks_selector = 'input[type="checkbox"][id^="cbAgree"]'
    next_selector = ".btn_type4.c1:has-text('다음'), a:has-text('다음'), button:has-text('다음')"

    for attempt in range(1, max_retry + 1):
        # 체크박스 기다렸다가 모두 체크
        page.wait_for_selector(checks_selector, timeout=timeout)
        checks = page.locator(checks_selector)
        cnt = checks.count()
        for i in range(cnt):
            try:
                checks.nth(i).set_checked(True, force=True)
            except Exception:
                try:
                    checks.nth(i).click(force=True)
                except Exception:
                    pass

        # 혹시 더 아래에도 체크박스가 있으면 대비해 스크롤
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        except Exception:
            pass
        time.sleep(0.3)

        # '다음' 클릭 (네비게이션 유무에 상관없이 후속 검증을 수행)
        next_btn = page.locator(next_selector).first
        try:
            with page.expect_navigation(wait_until="domcontentloaded", timeout=timeout):
                next_btn.click(force=True)
        except Exception:
            # 네비게이션 이벤트가 없을 수도 있으니 무시하고 상태 검증 진행
            try:
                next_btn.click(force=True)
            except Exception:
                pass

        # 실제 가입 폼 진입 확인: URL 또는 입력 필드 존재
        if page.locator('input[name="tbMemName"]').count() > 0:
            page.wait_for_selector('input[name="tbMemName"]', timeout=timeout)
            return
        if re.search(r"/Account/.*(MemJoin|Join)", page.url):
            page.wait_for_selector('input[name="tbMemName"]', timeout=timeout)
            return

        # 여전히 약관 페이지라면 다시 체크 후 재시도
        time.sleep(0.8)

    # 여기까지 오면 실패
    raise RuntimeError("약관 통과 실패: 가입 폼으로 이동하지 못했습니다.")


# --- 초기화: 정보 시트 읽기 -------------------------------------------------
def 초기화():
    """'정보' 시트에서 반복/랜덤값과 이름 풀을 읽고, 카운트/큐 준비."""
    global 반복횟수, 랜덤최소, 랜덤최대
    global 아이디1_리스트, 아이디2_리스트, 아이디3_리스트
    global 남아_리스트, 여아_리스트, 부_리스트, 모_리스트
    global 이름_카운트, 부모_이름_카운트

    아이디1_리스트.clear()
    아이디2_리스트.clear()
    아이디3_리스트.clear()
    남아_리스트.clear()
    여아_리스트.clear()
    부_리스트.clear()
    모_리스트.clear()

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb["정보"]

    반복횟수   = ws.cell(row=1, column=20).value
    랜덤최소   = ws.cell(row=2, column=20).value
    랜덤최대   = ws.cell(row=3, column=20).value

    아이디마지막행 = 92
    자녀마지막행   = 501
    부모마지막행   = 78

    for row in ws.iter_rows(min_row=2, max_row=아이디마지막행, values_only=True):
        아이디1_리스트.append(row[5])
        아이디2_리스트.append(row[6])
        아이디3_리스트.append(row[7])

    for row in ws.iter_rows(min_row=2, max_row=자녀마지막행, values_only=True):
        남아_리스트.append(row[10])
        여아_리스트.append(row[11])

    for row in ws.iter_rows(min_row=2, max_row=부모마지막행, values_only=True):
        부_리스트.append(row[14])
        모_리스트.append(row[15])
    wb.close()

    # 문자열만 유지하고 None/공백 제거
    def only_str(a):
        out = []
        for x in a:
            if x is None:
                continue
            s = str(x).strip()
            if s:
                out.append(s)
        return out

    아이디1_리스트[:] = only_str(아이디1_리스트)
    아이디2_리스트[:] = only_str(아이디2_리스트)
    아이디3_리스트[:] = only_str(아이디3_리스트)
    남아_리스트[:]    = only_str(남아_리스트)
    여아_리스트[:]    = only_str(여아_리스트)
    부_리스트[:]      = only_str(부_리스트)
    모_리스트[:]      = only_str(모_리스트)

    이름_카운트, 부모_이름_카운트 = 이름_카운트_로드()
    큐_초기화()
    reset_signup_plans(반복횟수)

# --- 시트별 주소 목록 로더 ---------------------------------------------------
def load_lists(sheet_name):
    """학교/아파트/동 리스트 로드 + URL 문자열화 + 최소요건 필터링."""
    global 학교_리스트, 아파트_리스트, 동_리스트
    global 회원가입사이트, 수강신청사이트

    학교_리스트.clear()
    아파트_리스트.clear()
    동_리스트.clear()

    def s(x):
        return "" if x is None else str(x).strip()

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        학교 = s(row[0]); 아파트 = s(row[1]); 동 = s(row[2])
        if 학교 and (아파트 or 동):
            rows.append((학교, 아파트, 동))

    학교_리스트[:]   = [r[0] for r in rows]
    아파트_리스트[:] = [r[1] for r in rows]
    동_리스트[:]     = [r[2] for r in rows]

    회원가입사이트 = s(ws.cell(row=1, column=20).value)
    수강신청사이트 = s(ws.cell(row=2, column=20).value)
    wb.close()

    if not 학교_리스트:
        raise RuntimeError(f"{sheet_name} 시트에 (학교 + 아파트/동) 데이터가 없습니다.")

# --- 현황 파일 저장 ---------------------------------------------------------
def 엑셀_초기화_및_데이터_저장(sheet_name, 이름, 아이디, 나이, 요일, 학교, 주소, 부모이름):
    """'추가인원 가입 현황.xlsx'에 가입 기록 저장(부모이름 포함)."""
    global 사용_전체이름, 사용_아이디
    status_sheet_name = resolve_status_sheet_name(sheet_name)
    현황_경로 = status_file_path
    if not os.path.exists(현황_경로):
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = status_sheet_name
        ws2.append(["이름","아이디","나이","요일","학교","주소","부모이름","가입일시"])
    else:
        wb2 = openpyxl.load_workbook(현황_경로)
        ws2 = (
            wb2[status_sheet_name]
            if status_sheet_name in wb2.sheetnames
            else wb2.create_sheet(title=status_sheet_name)
        )
        if ws2.max_row == 1 and ws2.max_column < 8:
            ws2.append(["이름","아이디","나이","요일","학교","주소","부모이름","가입일시"])

    ws2.append([
        이름, 아이디, 나이, 요일, 학교, 주소, 부모이름,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ])
    wb2.save(현황_경로)
    wb2.close()
    사용_전체이름.add(str(이름).strip())
    사용_아이디.add(str(아이디).strip().lower())

# --- 팝업 처리 --------------------------------------------------------------
def handle_alert(dialog):
    global 클릭
    time.sleep(0.5)
    클릭 = 1
    dialog.accept()

def 우편번호_주소_입력(page, queries, required_keyword=None):
    queries = [str(query).strip() for query in queries if query and str(query).strip()]
    if not queries:
        raise RuntimeError("우편번호 검색어가 없습니다.")
    search_queries = []
    for query in queries:
        search_queries.extend([query, query])

    def postcode_frame():
        for frame in page.frames:
            if "postcode.map.kakao.com" in frame.url:
                return frame
        return None

    page.click("span:text('우편번호 찾기')")
    time.sleep(1.2)

    last_error = None
    for idx, query in enumerate(search_queries):
        try:
            frame = postcode_frame()
            search_input = frame.locator('input[name="region_name"]') if frame else None
            if search_input and search_input.count() > 0:
                search_input.first.fill(query)
                frame.locator("button.btn_search").click()
            else:
                page.keyboard.type(query)
                page.keyboard.press("Enter")

            debug_log(f"우편번호 검색 query={query}")
            selected_frame = None
            for _ in range(24):
                frame = postcode_frame()
                if frame and frame.locator("button.link_post").count() > 0:
                    selected_frame = frame
                    break
                time.sleep(0.5)

            if not selected_frame:
                last_error = f"검색 결과 없음 query={query}"
                continue

            result_button = selected_frame.locator("button.link_post").first
            result_text = result_button.inner_text(timeout=5000)
            if required_keyword and required_keyword not in result_text:
                last_error = f"주소 지역 불일치 query={query} result={result_text}"
                debug_log(last_error)
                continue

            result_button.click()
            for _ in range(20):
                zip_code = page.locator('input[name="tbZipCode"]').input_value().strip()
                addr1 = page.locator('input[name="tbAddr1"]').input_value().strip()
                if zip_code and addr1 and (not required_keyword or required_keyword in addr1):
                    debug_log(f"우편번호 입력 완료 zip={zip_code} addr={addr1}")
                    return addr1
                time.sleep(0.3)

            last_error = f"주소 필드 미입력 query={query}"
        except Exception as e:
            last_error = repr(e)
            debug_log(f"우편번호 검색 실패 query={query} error={repr(e)}")

        if idx < len(search_queries) - 1:
            time.sleep(0.5)

    raise RuntimeError(f"우편번호/주소 입력 실패 queries={queries} required={required_keyword} last_error={last_error}")

def ensure_logged_in_for_course(page, user_id):
    dialog_messages = []

    def collect_dialog(dialog):
        message = dialog.message
        dialog_messages.append(message)
        debug_log(f"로그인 알림 id={user_id} message={message}")
        dialog.accept()

    page.on("dialog", collect_dialog)
    try:
        for nav_try_idx in range(2):
            try:
                page.goto(수강신청사이트, wait_until="domcontentloaded")
                break
            except Exception as e:
                if nav_try_idx == 0 and "NS_BINDING_ABORTED" in str(e):
                    debug_log(f"수강신청 페이지 이동 재시도 id={user_id} error={repr(e)}")
                    time.sleep(1.0)
                    continue
                raise

        try:
            page.wait_for_selector('input[name="tbID"]', timeout=5000)
        except Exception:
            pass

        login_id = page.locator('input[name="tbID"]')
        login_pass = page.locator('input[name="tbPass"]')
        if login_id.count() > 0 and login_pass.count() > 0:
            debug_log(f"로그인 페이지 감지 id={user_id}")
            login_id.fill(user_id)
            login_pass.fill(비밀번호)
            page.locator('a.btn_type2:has-text("LOGIN")').click()

            for _ in range(20):
                time.sleep(0.5)
                body_text = page.locator("body").inner_text(timeout=5000)
                if "LOGOUT" in body_text or "내 강의실" in body_text:
                    debug_log(f"로그인 완료 id={user_id} url={page.url}")
                    return
                if any("일치하는 정보" in message for message in dialog_messages):
                    break

            body_text = page.locator("body").inner_text(timeout=5000)
            raise RuntimeError(
                f"로그인 실패 id={user_id} url={page.url} dialogs={dialog_messages} "
                f"body={body_text[:120]!r}"
            )
    finally:
        try:
            page.remove_listener("dialog", collect_dialog)
        except Exception:
            pass

# --- 회원가입 ---------------------------------------------------------------
# --- 회원가입 ---------------------------------------------------------------
def 회원가입(page, sheet_name, signup_plan):
    """깨끗한(비로그인) 컨텍스트 기준 회원가입 1회 수행."""
    global 이름, 요일

    debug_log(f"회원가입 시작 sheet={sheet_name} plan={signup_plan}")
    load_lists(sheet_name)

    # 회원가입 첫 화면 진입 + 약관 통과(새 유틸 사용)
    page.goto(회원가입사이트, wait_until="domcontentloaded")
    약관_통과_후_가입폼_진입(page)  # ★ 이 한 줄이 핵심입니다.

    # 폼 필드 로드 대기 (클래스 의존 제거)
    page.wait_for_selector('input[name="tbMemName"]', timeout=15000)
    debug_log("회원가입 폼 진입 완료")

    # --- 이름 & 성별(분산 추출) ---
    성별값 = signup_plan["sex"]
    page.select_option('select[name="ddlSex"]', value=성별값)

    성한자, 이름후, 이름 = 사용가능_학생이름_생성(성별값)
    아이성 = 성한자
    page.fill('input[name="tbMemName"]', 이름)

    # --- 아이디 생성 ---
    아이디 = 사용가능_아이디_생성()
    page.fill('input[name="tbMemID"]', 아이디)
    page.locator("text=중복확인").click()
    page.fill('input[name="tbMemPass"]', 비밀번호)
    page.fill('input[name="tbMemPass2"]', 비밀번호)

    # --- 전화번호 ---
    page.fill('input[name="tbMobile1"]', '010')
    중간번호 = ''.join(str(random.randint(0,9)) for _ in range(4))
    끝번호   = ''.join(str(random.randint(0,9)) for _ in range(4))
    page.fill('input[name="tbMobile2"]', 중간번호)
    page.fill('input[name="tbMobile3"]', 끝번호)

    # --- 생년월일/나이 ---
    나이     = random.randint(11, 13)
    출생년도 = 2026 - 나이
    page.select_option('select[name="ddlBirthDay1"]', value=str(출생년도))
    page.select_option('select[name="ddlBirthDay2"]', value=f"{random.randint(1,12):02}")
    page.select_option('select[name="ddlBirthDay3"]', value=f"{random.randint(1,28):02}")

    # --- 학교 & 주소 ---
    학교선택 = random.randint(0, len(학교_리스트)-1)
    학교     = 학교_리스트[학교선택]

    # 옵션 value가 라벨과 다르면 실패할 수 있어 보완
    try:
        page.select_option('select[name="ddlAcaName1"]', value=학교)
    except Exception:
        page.locator('select[name="ddlAcaName1"]').select_option(label=학교)

    page.fill('input[name="tbGrade1"]', str(나이 - 7))
    page.fill('input[name="tbClass1"]', str(random.randint(1,5)))
    addr_queries = [
        아파트_리스트[학교선택],
        학교,
    ]
    주소1 = 우편번호_주소_입력(
        page,
        addr_queries,
        required_keyword=ADDRESS_REQUIRED_KEYWORDS.get(sheet_name),
    )
    호   = f"{random.randint(1,5)}0{random.randint(1,9)}"
    주소2 = f"{(동_리스트[학교선택] or 학교)} {호}호"
    page.fill('input[name="tbAddr2"]', 주소2)
    주소 = f"{주소1} {주소2}".strip()

    # --- 부모 정보(분산 추출) ---
    global 부_큐, 모_큐, 부모_이름_카운트
    부모성별 = random.choice(['부','모'])
    if 부모성별 == '부':
        if not 부_큐:
            큐_초기화()
        부이름후 = 분산_선택(부_큐, 부모_이름_카운트) or random.choice(부_리스트)
        부모이름 = f"{아이성}{부이름후}"
    else:
        if not 모_큐:
            큐_초기화()
        모이름후 = 분산_선택(모_큐, 부모_이름_카운트) or random.choice(모_리스트)
        부모이름 = f"{random.choices(성씨, weights=weights, k=1)[0]}{모이름후}"

    page.fill('input[name="tbPName"]', 부모이름)
    page.fill('input[name="tbPMobile1"]', '010')
    page.fill('input[name="tbPMobile2"]', 중간번호)
    page.fill('input[name="tbPMobile3"]', 끝번호)

    # --- 이메일 ---
    포털 = ["@gmail.com","@naver.com","@daum.net","@nate.com"]
    page.fill('input[name="tbMemEmail"]', f"{아이디}{random.choice(포털)}")

    page.click("#rdRoute1")
    time.sleep(0.4)
    page.wait_for_selector("span:text('회원가입')", timeout=10000)
    signup_dialog_messages = []

    def capture_signup_dialog(dialog):
        message = dialog.message
        signup_dialog_messages.append(message)
        debug_log(f"회원가입 알림 name={이름} id={아이디} message={message}")
        dialog.accept()

    page.on("dialog", capture_signup_dialog)
    try:
        page.click("span:text('회원가입')")
        debug_log(f"회원가입 제출 클릭 name={이름} id={아이디} school={학교}")
        for _ in range(10):
            time.sleep(0.5)
            page.keyboard.press('Enter')
    finally:
        try:
            page.remove_listener("dialog", capture_signup_dialog)
        except Exception:
            pass

    # 수강 요일
    요일 = signup_plan["day"]
    member_info = {
        "name": 이름,
        "user_id": 아이디,
        "age": 나이,
        "day": 요일,
        "school": 학교,
        "address": 주소,
        "parent_name": 부모이름,
        "sex": 성별값,
    }
    signup_confirmed = any("감사" in message or "완료" in message for message in signup_dialog_messages)
    if not signup_confirmed:
        raise RuntimeError(f"회원가입 성공 확인 실패 name={이름} id={아이디} dialogs={signup_dialog_messages}")

    try:
        ensure_logged_in_for_course(page, 아이디)
    except Exception as e:
        debug_log(
            f"회원가입은 확인됐지만 로그인 검증 실패. 수강신청 단계에서 같은 계정 재시도 "
            f"name={이름} id={아이디} error={repr(e)}"
        )

    debug_log(f"회원가입 검증 완료 name={이름} day={요일} dialogs={signup_dialog_messages}")
    return member_info

# --- 수강신청 ---------------------------------------------------------------
def course_label_matches_day(label_text, day):
    text = re.sub(r"\s+", "", label_text or "")
    if day == "월수":
        return ("월수" in text) or ("월" in text and "수" in text and "화" not in text and "목" not in text)
    if day == "화목":
        return ("화목" in text) or ("화" in text and "목" in text and "월" not in text and "수" not in text)
    return False

def select_course_option(page, sheet_name, day):
    clicked_week_filter = False
    week_value = {"월수": "2,4", "화목": "3,5"}.get(day)
    if week_value:
        week_button = page.locator(f'button.btn_week[data-week="{week_value}"]')
        if week_button.count() > 0:
            week_button.first.click()
            clicked_week_filter = True
            try:
                page.wait_for_load_state("domcontentloaded", timeout=10000)
            except Exception:
                pass
            time.sleep(1.0)

    page.wait_for_selector('input[name="rdIDX"]', state="attached", timeout=15000)
    radios = page.locator('input[name="rdIDX"]')
    radio_count = radios.count()
    options = []
    for idx in range(radio_count):
        radio = radios.nth(idx)
        course_id = radio.get_attribute("id") or ""
        label_text = ""
        if course_id:
            label = page.locator(f'label[for="{course_id}"]')
            if label.count() > 0:
                try:
                    label_text = label.first.inner_text(timeout=1500).strip()
                except Exception:
                    label_text = ""
        options.append((idx, course_id, label_text))

    candidate_indexes = [
        idx for idx, course_id, label_text in options
        if course_label_matches_day(label_text, day)
    ]

    if not candidate_indexes:
        legacy_ids = {
            ("BG", "월수"): ["3_25", "3_29"],
            ("BG", "화목"): ["4_26", "4_30"],
            ("PH", "월수"): ["1_55", "1_59"],
            ("PH", "화목"): ["2_56", "2_60"],
        }.get((sheet_name, day), [])
        for legacy_id in legacy_ids:
            for idx, course_id, _ in options:
                if course_id == legacy_id:
                    candidate_indexes.append(idx)
                    break
            if candidate_indexes:
                break

    if not candidate_indexes:
        legacy_prefixes = {
            ("BG", "월수"): ["3_"],
            ("BG", "화목"): ["4_"],
            ("PH", "월수"): ["1_"],
            ("PH", "화목"): ["2_"],
        }.get((sheet_name, day), [])
        for prefix in legacy_prefixes:
            for idx, course_id, _ in options:
                if course_id.startswith(prefix):
                    candidate_indexes.append(idx)
                    break
            if candidate_indexes:
                break

    if not candidate_indexes and (clicked_week_filter or radio_count == 1):
        candidate_indexes = [0]

    if not candidate_indexes:
        option_summary = [
            {"id": course_id, "label": label_text[:80]}
            for _, course_id, label_text in options
        ]
        raise RuntimeError(f"{sheet_name} {day} 과정 선택 후보를 찾지 못했습니다. options={option_summary}")

    selected_index = candidate_indexes[0]
    course_radio = radios.nth(selected_index)
    course_id = course_radio.get_attribute("id") or ""
    label_text = ""
    if course_id and page.locator(f'label[for="{course_id}"]').count() > 0:
        label = page.locator(f'label[for="{course_id}"]').first
        try:
            label_text = label.inner_text(timeout=1500).strip()
        except Exception:
            label_text = ""
        try:
            label.click(timeout=3000)
        except Exception:
            course_radio.check(force=True)
    else:
        course_radio.check(force=True)

    debug_log(
        f"{sheet_name} 과정 선택 완료 day={day} course_id={course_id} "
        f"label={label_text[:80]!r} week_filter={clicked_week_filter}"
    )

def 신청(page, sheet_name, member_info):
    요일 = member_info["day"]
    debug_log(f"수강신청 시작 name={member_info['name']} sheet={sheet_name} day={요일}")
    time.sleep(1.0)
    ensure_logged_in_for_course(page, member_info["user_id"])
    if page.locator('input#q1_2').count() > 0:
        page.locator('input#q1_2').click()
        time.sleep(0.4)

    select_course_option(page, sheet_name, 요일)
    debug_log(f"수강신청 요일 선택 완료 name={member_info['name']} day={요일}")

    page.once('dialog', handle_alert)
    page.locator('a.btn_type5.mb30:has-text("수강신청")').click()
    time.sleep(0.5)
    page.keyboard.press('Enter')
    time.sleep(0.5)
    debug_log(f"수강신청 1차 제출 완료 name={member_info['name']}")

    page.wait_for_selector('input#btn_poll', timeout=10000)
    poll_buttons = page.locator('input[id^="q"][id$="_1"]')
    total = poll_buttons.count()
    for idx in range(total):
        poll_buttons.nth(idx).click()

    page.click('input#btn_poll')
    time.sleep(0.3)
    page.click('input#cbAddress'); time.sleep(0.2)
    page.click('input#cbPc');      time.sleep(0.2)
    page.once('dialog', handle_alert)
    page.click('a:has(span:text("신청"))')
    time.sleep(0.5)
    page.keyboard.press('Enter')
    time.sleep(0.5)
    debug_log(f"수강신청 최종 제출 완료 name={member_info['name']}")

    # 온라인 신청이 완료된 뒤에만 현황 파일에 기록한다.
    엑셀_초기화_및_데이터_저장(
        sheet_name,
        member_info["name"],
        member_info["user_id"],
        member_info["age"],
        member_info["day"],
        member_info["school"],
        member_info["address"],
        member_info["parent_name"],
    )

    현황_경로 = status_file_path
    status_sheet_name = resolve_status_sheet_name(sheet_name)
    wb2 = openpyxl.load_workbook(현황_경로, data_only=True)
    ws2 = wb2[status_sheet_name]
    total = ws2.max_row - 1
    today_str = datetime.now().strftime('%Y-%m-%d')
    today_count = 0
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if str(row[-1]).startswith(today_str):
            today_count += 1
    wb2.close()
    print(f"{member_info['name']} 신청 완료 / 총 {total}명, 오늘 {today_count}명 추가되었습니다.")
    debug_log(f"기록 저장 완료 name={member_info['name']} total={total} today={today_count}")
    time.sleep(random.randint(POST_SUCCESS_SLEEP_MIN_SECONDS, POST_SUCCESS_SLEEP_MAX_SECONDS))

# --- 전체 동작 --------------------------------------------------------------
def 신규_가입_신청_1건(p, signup_plan):
    debug_log(f"신규 대상 시작 plan={signup_plan}")
    browser = p.firefox.launch(headless=True, args=['--disable-popup-blocking'])
    context = browser.new_context()
    context.set_default_timeout(20000)
    context.set_default_navigation_timeout(45000)
    page = context.new_page()
    member_info = None

    try:
        # 회원가입 재시도는 새 페이지로 다시 시도한다.
        for signup_try_idx in range(SIGNUP_MAX_ATTEMPTS):
            try:
                debug_log(f"회원가입 시도 {signup_try_idx + 1}/{SIGNUP_MAX_ATTEMPTS}")
                member_info = 회원가입(page, TARGET_SHEET_NAME, signup_plan)
                break
            except Exception as e:
                debug_log(f"회원가입 실패 try={signup_try_idx + 1} error={repr(e)}")
                if signup_try_idx < SIGNUP_MAX_ATTEMPTS - 1:
                    try:
                        page.close()
                    except Exception:
                        pass
                    page = context.new_page()
                    continue
                raise e

        # 회원가입이 끝난 계정은 신청 완료 전까지 다음 가입으로 넘어가지 않는다.
        apply_try_idx = 0
        while True:
            apply_try_idx += 1
            try:
                debug_log(f"수강신청 시도 {apply_try_idx} name={member_info['name']} id={member_info['user_id']}")
                신청(page, TARGET_SHEET_NAME, member_info)
                return member_info
            except Exception as e:
                debug_log(
                    f"수강신청 실패. 같은 계정으로 재시도 name={member_info['name']} "
                    f"id={member_info['user_id']} try={apply_try_idx} error={repr(e)}"
                )
                if apply_try_idx >= MAX_APPLY_ATTEMPTS_PER_MEMBER:
                    raise RuntimeError(
                        f"수강신청 최대 재시도 초과. 이 계정은 기록하지 않고 다음 대상으로 진행 "
                        f"name={member_info['name']} id={member_info['user_id']}"
                    ) from e
                try:
                    page.close()
                except Exception:
                    pass
                page = context.new_page()
                if APPLY_RETRY_DELAY_SECONDS > 0:
                    time.sleep(APPLY_RETRY_DELAY_SECONDS)
    finally:
        try:
            context.close()
        except Exception:
            pass
        try:
            browser.close()
        except Exception:
            pass

def 동작():
    apply_dashboard_settings()
    초기화()
    total_count, today_count = get_status_counts(TARGET_SHEET_NAME)
    remaining_total = max(0, TOTAL_TARGET_COUNT - total_count)
    reset_signup_plans(remaining_total)
    debug_log(
        f"동작 시작 sheet={TARGET_SHEET_NAME} total={total_count}/{TOTAL_TARGET_COUNT} "
        f"today={today_count} active={ACTIVE_START_HOUR}:00-{ACTIVE_END_HOUR}:00"
    )
    update_dashboard_progress("실행 시작", "매크로 시작")

    if remaining_total <= 0:
        debug_log(f"목표 달성 완료. 현재 {TARGET_SHEET_NAME} 누적 {total_count}명")
        update_dashboard_progress("목표 달성", "목표 달성 완료")
        return

    with sync_playwright() as p:
        while True:
            apply_dashboard_settings()
            total_count, today_count = get_status_counts(TARGET_SHEET_NAME)
            remaining_total = max(0, TOTAL_TARGET_COUNT - total_count)
            if remaining_total <= 0:
                debug_log(f"목표 달성 완료. 현재 {TARGET_SHEET_NAME} 누적 {total_count}명")
                update_dashboard_progress("목표 달성", "목표 달성 완료")
                return

            now = datetime.now()
            if not is_active_time(now):
                update_dashboard_progress("운영시간 외", "다음 운영시간까지 대기")
                sleep_until(next_active_start(now), "운영 시간 외")
                continue

            daily_target, schedule_times = get_daily_plan(remaining_total, today_count)
            update_dashboard_progress("실행 중", "오늘 목표/예약 갱신")
            if today_count >= daily_target:
                debug_log(f"오늘 목표 달성 today={today_count}/{daily_target}")
                update_dashboard_progress("오늘 목표 완료", "다음 운영일 대기")
                sleep_until(next_active_start(now), "오늘 목표 완료")
                continue

            if schedule_times:
                next_time = parse_schedule_time(schedule_times[0])
                _, active_end = active_window_bounds(now)
                if next_time and next_time >= active_end:
                    update_dashboard_progress("예약 대기", "다음 예약 시간이 운영시간을 벗어남")
                    sleep_until(next_active_start(now), "다음 예약 시간이 운영 시간을 벗어남")
                    continue
                if next_time and next_time > datetime.now():
                    update_dashboard_progress("예약 대기", f"다음 가입 예약 {next_time.strftime('%Y-%m-%d %H:%M:%S')}")
                    sleep_until(next_time, "다음 가입 예약")

            if not gender_plan_queue or not day_plan_queue:
                reset_signup_plans(remaining_total)

            signup_plan = pop_signup_plan()
            try:
                member_info = 신규_가입_신청_1건(p, signup_plan)
            except Exception as e:
                debug_log(f"회원가입 전 단계 실패. {OUTER_RETRY_DELAY_SECONDS}초 후 새 대상 재시도 plan={signup_plan} error={repr(e)}")
                reset_signup_plans(remaining_total)
                if OUTER_RETRY_DELAY_SECONDS > 0:
                    time.sleep(OUTER_RETRY_DELAY_SECONDS)
                continue

            consume_schedule_slot()

            total_count, today_count = get_status_counts(TARGET_SHEET_NAME)
            debug_log(
                f"대상 완료 name={member_info['name']} total={total_count}/{TOTAL_TARGET_COUNT} "
                f"today={today_count}/{daily_target}"
            )
            update_dashboard_progress("실행 중", f"대상 완료 total={total_count}/{TOTAL_TARGET_COUNT}")

if __name__ == "__main__":
    동작()
