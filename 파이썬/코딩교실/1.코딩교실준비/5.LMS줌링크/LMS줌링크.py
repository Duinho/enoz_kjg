import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.sync_api import Error as PlaywrightError
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE_PATH = BASE_DIR / "LMS줌링크.xlsx"
LOG_DIR = BASE_DIR / "logs"


@dataclass(frozen=True)
class Settings:
    login_url: str
    teacher_list_url: str
    admin_id: str
    admin_pw: str


@dataclass(frozen=True)
class ZoomTask:
    row: int
    teacher_id: str
    zoom_link: str


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def load_settings_and_tasks(excel_path=EXCEL_FILE_PATH):
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb.active

    settings = Settings(
        login_url=clean(ws.cell(row=1, column=17).value),
        teacher_list_url=clean(ws.cell(row=2, column=17).value),
        admin_id=clean(ws.cell(row=3, column=17).value),
        admin_pw=clean(ws.cell(row=4, column=17).value),
    )

    tasks = []
    trailing_data = []
    stopped_at_blank_row = None

    for row in range(2, ws.max_row + 1):
        teacher_id = clean(ws.cell(row=row, column=1).value)
        zoom_link = clean(ws.cell(row=row, column=2).value)

        if not teacher_id:
            if stopped_at_blank_row is None:
                stopped_at_blank_row = row
            if zoom_link:
                trailing_data.append((row, teacher_id, zoom_link))
            continue

        if stopped_at_blank_row is not None:
            trailing_data.append((row, teacher_id, zoom_link))
            continue

        tasks.append(ZoomTask(row=row, teacher_id=teacher_id, zoom_link=zoom_link))

    return settings, tasks, trailing_data


def validate_inputs(settings, tasks, trailing_data):
    errors = []
    warnings = []

    required_settings = {
        "Q1 로그인 사이트 링크": settings.login_url,
        "Q2 강사 관리 링크": settings.teacher_list_url,
        "Q3 admin ID": settings.admin_id,
        "Q4 admin PW": settings.admin_pw,
    }
    for label, value in required_settings.items():
        if not value:
            errors.append(f"{label}가 비어 있습니다.")

    for label, value in {
        "Q1 로그인 사이트 링크": settings.login_url,
        "Q2 강사 관리 링크": settings.teacher_list_url,
    }.items():
        if value and not value.startswith("https://"):
            errors.append(f"{label}가 https:// URL이 아닙니다: {value}")

    if not tasks:
        errors.append("A:B에 처리할 줌링크 목록이 없습니다.")

    seen_ids = {}
    seen_links = {}
    for task in tasks:
        if not task.zoom_link:
            errors.append(f"{task.row}행 {task.teacher_id}: 링크가 비어 있습니다.")
        elif not task.zoom_link.startswith("https://"):
            errors.append(f"{task.row}행 {task.teacher_id}: 링크가 https:// URL이 아닙니다.")
        elif "zoom.us/" not in task.zoom_link:
            warnings.append(f"{task.row}행 {task.teacher_id}: zoom.us 링크가 아닙니다.")

        seen_ids.setdefault(task.teacher_id, []).append(task.row)
        if task.zoom_link:
            seen_links.setdefault(task.zoom_link, []).append(task.row)

    duplicate_ids = {key: rows for key, rows in seen_ids.items() if len(rows) > 1}
    for teacher_id, rows in duplicate_ids.items():
        errors.append(f"강사/반 코드 중복: {teacher_id} rows={rows}")

    duplicate_links = {key: rows for key, rows in seen_links.items() if len(rows) > 1}
    for zoom_link, rows in duplicate_links.items():
        warnings.append(f"동일 줌링크 중복 사용 rows={rows}: {zoom_link}")

    if trailing_data:
        errors.append(f"빈 ID 행 이후 데이터가 있습니다: {trailing_data[:5]}")

    return errors, warnings


def write_log(log_path, rows):
    LOG_DIR.mkdir(exist_ok=True)
    fieldnames = ["phase", "row", "teacher_id", "status", "message"]
    with log_path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def accept_dialog(dialog):
    try:
        dialog.accept()
    except Exception:
        pass


def login(page, settings):
    page.goto(settings.login_url)
    page.fill('input[name="tbAdminId"]', settings.admin_id)
    page.fill('input[name="tbAdminPass"]', settings.admin_pw)
    page.press('input[name="tbAdminPass"]', "Enter")
    page.goto(settings.teacher_list_url)
    ensure_search_form(page)


def ensure_search_form(page):
    search_input = page.locator('input[name="teacher_name"]')
    try:
        search_input.wait_for(state="visible", timeout=1500)
        return
    except PlaywrightTimeoutError:
        pass

    page.locator('a.button:has-text("검색 보이기 / 감추기")').click(timeout=5000)
    search_input.wait_for(state="visible", timeout=10000)


def search_teacher(page, teacher_id):
    ensure_search_form(page)
    search_input = page.locator('input[name="teacher_name"]')
    search_input.fill(teacher_id)
    search_input.press("Enter")

    page.wait_for_function(
        """(teacherId) => {
            return [...document.querySelectorAll('span.font_b')]
                .some((element) => (element.textContent || '').trim() === teacherId);
        }""",
        arg=teacher_id,
        timeout=15000,
    )


def open_teacher_popup(page, teacher_id):
    search_teacher(page, teacher_id)
    with page.expect_popup(timeout=15000) as popup_info:
        page.evaluate(
            """(teacherId) => {
                const target = [...document.querySelectorAll('span.font_b')]
                    .find((element) => (element.textContent || '').trim() === teacherId);
                if (!target) throw new Error(`Teacher not found: ${teacherId}`);
                target.click();
            }""",
            teacher_id,
        )

    popup = popup_info.value
    popup.wait_for_load_state("domcontentloaded", timeout=15000)
    popup.wait_for_selector('input[name="video_link"]', timeout=15000)
    return popup


def set_zoom_link(page, task):
    popup = open_teacher_popup(page, task.teacher_id)
    try:
        video_input = popup.locator('input[name="video_link"]')
        video_input.fill(task.zoom_link)
        popup.locator('input[name="Submit"]').click(timeout=15000)
        try:
            popup.wait_for_event("close", timeout=5000)
        except PlaywrightTimeoutError:
            pass
    finally:
        try:
            if not popup.is_closed():
                popup.close()
        except PlaywrightError:
            pass


def read_zoom_link(page, task):
    popup = open_teacher_popup(page, task.teacher_id)
    try:
        current = popup.locator('input[name="video_link"]').input_value(timeout=15000)
        return clean(current)
    finally:
        try:
            if not popup.is_closed():
                popup.close()
        except PlaywrightError:
            pass


def run_apply(page, tasks, start_row, end_row):
    rows = []
    for task in tasks:
        if task.row < start_row or (end_row is not None and task.row > end_row):
            continue

        try:
            set_zoom_link(page, task)
            rows.append(
                {
                    "phase": "apply",
                    "row": task.row,
                    "teacher_id": task.teacher_id,
                    "status": "OK",
                    "message": "saved",
                }
            )
            print(f"{task.row}행 {task.teacher_id} 줌링크 저장 완료")
        except Exception as exc:
            rows.append(
                {
                    "phase": "apply",
                    "row": task.row,
                    "teacher_id": task.teacher_id,
                    "status": "FAIL",
                    "message": repr(exc),
                }
            )
            print(f"{task.row}행 {task.teacher_id} 줌링크 저장 실패: {exc}")
    return rows


def run_verify(page, tasks, start_row=2, end_row=None):
    rows = []
    for task in tasks:
        if task.row < start_row or (end_row is not None and task.row > end_row):
            continue

        try:
            current = read_zoom_link(page, task)
            if current == task.zoom_link:
                status = "OK"
                message = "matched"
            else:
                status = "FAIL"
                message = f"expected={task.zoom_link} current={current}"
            rows.append(
                {
                    "phase": "verify",
                    "row": task.row,
                    "teacher_id": task.teacher_id,
                    "status": status,
                    "message": message,
                }
            )
            print(f"{task.row}행 {task.teacher_id} 검증 {status}")
        except Exception as exc:
            rows.append(
                {
                    "phase": "verify",
                    "row": task.row,
                    "teacher_id": task.teacher_id,
                    "status": "FAIL",
                    "message": repr(exc),
                }
            )
            print(f"{task.row}행 {task.teacher_id} 검증 실패: {exc}")
    return rows


def parse_args():
    parser = argparse.ArgumentParser(description="LMS 강사/반 줌링크 저장 자동화")
    parser.add_argument("--dry-run", action="store_true", help="엑셀 검증만 수행")
    parser.add_argument("--verify-only", action="store_true", help="저장 없이 LMS 값만 검증")
    parser.add_argument("--headless", action="store_true", help="브라우저 창을 숨김")
    parser.add_argument("--start-row", type=int, default=2, help="엑셀 시작 행")
    parser.add_argument("--end-row", type=int, default=None, help="엑셀 종료 행")
    return parser.parse_args()


def main():
    args = parse_args()
    settings, tasks, trailing_data = load_settings_and_tasks()
    errors, warnings = validate_inputs(settings, tasks, trailing_data)

    print(f"엑셀 대상: {len(tasks)}건")
    for warning in warnings:
        print(f"경고: {warning}")

    if errors:
        for error in errors:
            print(f"오류: {error}")
        return 1

    if args.dry_run:
        print("dry-run 완료: LMS에는 접속하지 않았습니다.")
        return 0

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = LOG_DIR / f"zoomlink_{timestamp}.csv"
    log_rows = []

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=args.headless, args=["--disable-popup-blocking"]
        )
        context = browser.new_context()
        context.on("dialog", accept_dialog)
        page = context.new_page()
        page.set_default_timeout(15000)

        try:
            login(page, settings)
            if not args.verify_only:
                log_rows.extend(run_apply(page, tasks, args.start_row, args.end_row))
            log_rows.extend(run_verify(page, tasks, args.start_row, args.end_row))
        finally:
            browser.close()

    write_log(log_path, log_rows)
    failures = [row for row in log_rows if row["status"] != "OK"]
    print(f"로그 저장: {log_path}")

    if failures:
        print(f"실패 {len(failures)}건")
        return 1

    print("전체 작업 성공")
    return 0


if __name__ == "__main__":
    sys.exit(main())
