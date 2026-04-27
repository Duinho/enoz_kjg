from __future__ import annotations

import io
import os
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
INPUT_WORKBOOK_PATH = BASE_DIR / f"{Path(__file__).stem}.xlsx"
OPENAPI_URL = "https://www.schoolinfo.go.kr/openApi.do"
REGION_CODE_URL = "https://www.schoolinfo.go.kr/download/sido_sggCode.xlsx"

INFO_SHEET_NAME = "정보"
HEADER_TOP_ROW = 1
HEADER_BOTTOM_ROW = 2
DATA_START_ROW = 3
CONFIG_ROWS = range(1, 7)
DEFAULT_YEAR_CANDIDATES = 3
MAX_SHEET_NAME_LENGTH = 31
RESULT_FILE_TEMPLATE = "학급별 학생수 결과_{year}.xlsx"
RESULT_FALLBACK_TEMPLATE = "학급별 학생수 결과_{year}_{stamp}.xlsx"

HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(bold=True)
SIDO_SUFFIXES = ("특별자치도", "특별자치시", "특별시", "광역시", "자치도", "자치시", "도", "시")
SGG_SUFFIXES = ("특별자치시", "시", "군", "구")
SCHOOL_KIND_RULES = (
    ("초", "02", (1, 2, 3, 4, 5, 6)),
    ("중", "03", (1, 2, 3)),
    ("고", "04", (1, 2, 3)),
    ("특수", "05", (1, 2, 3)),
)
DEFAULT_SCHOOL_KIND = ("06", (1, 2, 3))


@dataclass(frozen=True)
class 조회설정:
    열번호: int
    api_key: str
    학교급명: str
    학교급코드: str
    시도명: str
    시군구명: str
    학년목록: tuple[int, ...]
    공시연도: int | None


@dataclass(frozen=True)
class 지역코드:
    시도명: str
    시도코드: str
    시군구명: str
    시군구코드: str


@dataclass(frozen=True)
class 작업결과:
    시트명원본: str
    학년목록: tuple[int, ...]
    학교데이터: list[dict]


def 문자열(value) -> str:
    return str(value or "").strip()


def 값있음(value) -> bool:
    return 문자열(value) not in {"", "0"}


def 공백제거(value) -> str:
    return 문자열(value).replace(" ", "")


def 접미사제거(value, suffixes) -> str:
    normalized = 공백제거(value)
    for suffix in suffixes:
        if normalized.endswith(suffix) and len(normalized) > len(suffix):
            return normalized[: -len(suffix)]
    return normalized


def 시도핵심(시도명) -> str:
    return 접미사제거(시도명, SIDO_SUFFIXES)


def 시군구핵심(시군구명) -> str:
    return 접미사제거(시군구명, SGG_SUFFIXES)


def 시도축약(시도명) -> str:
    core = 시도핵심(시도명)
    if len(core) >= 2 and core[-1] in {"남", "북"}:
        return f"{core[0]}{core[-1]}"
    return core


def 시군구축약(시군구명) -> str:
    return 공백제거(시군구명)


def 학교급정보(학교급명) -> tuple[str, tuple[int, ...], str]:
    text = 문자열(학교급명)
    for keyword, code, grades in SCHOOL_KIND_RULES:
        if keyword in text:
            return code, grades, keyword
    return DEFAULT_SCHOOL_KIND[0], DEFAULT_SCHOOL_KIND[1], text


def 숫자값(value):
    if value in (None, "", "-"):
        return 0
    text = str(value).replace(",", "").strip()
    try:
        return float(text) if "." in text else int(text)
    except ValueError:
        return 0


def 반평균(학생수, 학급수) -> float:
    if not 학급수:
        return 0
    return round(학생수 / 학급수, 1)


def 시트명정리(name) -> str:
    invalid_chars = set("[]:*?/\\")
    cleaned = "".join("_" if ch in invalid_chars else ch for ch in 문자열(name))
    cleaned = cleaned.strip().strip("'")
    return (cleaned or "결과")[:MAX_SHEET_NAME_LENGTH]


def 중복없는시트명목록(raw_names) -> list[str]:
    normalized = []
    used = set()
    for raw_name in raw_names:
        base_name = 시트명정리(raw_name)
        candidate = base_name
        suffix_number = 2
        while candidate in used:
            suffix = f"_{suffix_number}"
            candidate = f"{base_name[: MAX_SHEET_NAME_LENGTH - len(suffix)]}{suffix}"
            suffix_number += 1
        used.add(candidate)
        normalized.append(candidate)
    return normalized


def 시트명원본(시도명, 시군구명, 학교급명) -> str:
    _, _, 학교급축약 = 학교급정보(학교급명)
    return f"{시도축약(시도명)}_{시군구축약(시군구명)}_{학교급축약}"


def 정수파싱(value, label: str) -> int:
    if isinstance(value, bool):
        raise ValueError(f"{label} 값이 잘못되었습니다: {value}")
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        raise ValueError(f"{label} 값이 잘못되었습니다: {value}")

    text = 문자열(value).replace(",", "")
    try:
        return int(text)
    except ValueError as exc:
        raise ValueError(f"{label} 값이 잘못되었습니다: {value}") from exc


def 선택연도파싱(value, label: str) -> int | None:
    if not 값있음(value):
        return None
    return 정수파싱(value, label)


def 학년목록파싱(학년텍스트, 허용학년목록: tuple[int, ...]) -> tuple[int, ...]:
    text = 공백제거(학년텍스트)
    if not text:
        raise ValueError("타겟 학년이 비어 있습니다. 예: 1,2,3")

    for separator in ("/", "|", ";"):
        text = text.replace(separator, ",")

    result = []
    seen = set()
    허용집합 = set(허용학년목록)

    for token in text.split(","):
        token = token.strip()
        if not token:
            continue
        if not token.isdigit():
            raise ValueError(f"학년 형식이 잘못되었습니다: {token}")
        grade = int(token)
        if grade not in 허용집합:
            allow_text = ",".join(str(value) for value in 허용학년목록)
            raise ValueError(f"지원하지 않는 학년입니다: {grade} (허용: {allow_text})")
        if grade not in seen:
            seen.add(grade)
            result.append(grade)

    if not result:
        raise ValueError("타겟 학년이 비어 있습니다. 예: 1,2,3")

    return tuple(result)


def 마지막설정열찾기(ws) -> int:
    last_col = 1
    for col in range(2, ws.max_column + 1):
        if any(값있음(ws.cell(row=row, column=col).value) for row in CONFIG_ROWS):
            last_col = col
    return last_col


def 입력설정읽기() -> list[조회설정]:
    if not INPUT_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"입력 엑셀 파일이 없습니다: {INPUT_WORKBOOK_PATH}")

    workbook = openpyxl.load_workbook(INPUT_WORKBOOK_PATH, data_only=True)
    try:
        sheet = workbook[INFO_SHEET_NAME] if INFO_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        last_col = 마지막설정열찾기(sheet)

        env_api_key = 문자열(os.getenv("SCHOOLINFO_API_KEY"))
        sheet_api_key = next(
            (
                문자열(sheet.cell(row=1, column=col).value)
                for col in range(2, last_col + 1)
                if 값있음(sheet.cell(row=1, column=col).value)
            ),
            "",
        )
        default_api_key = env_api_key or sheet_api_key
        if not default_api_key:
            raise ValueError("API 인증키가 없습니다. 1행(B열 이후) 또는 환경변수 SCHOOLINFO_API_KEY를 설정하세요.")

        env_year = 선택연도파싱(os.getenv("SCHOOL_PBAN_YEAR"), "환경변수 SCHOOL_PBAN_YEAR")
        configs = []

        for col in range(2, last_col + 1):
            학교급명 = 문자열(sheet.cell(row=2, column=col).value)
            시도명 = 문자열(sheet.cell(row=3, column=col).value)
            시군구명 = 문자열(sheet.cell(row=4, column=col).value)

            if not any((값있음(학교급명), 값있음(시도명), 값있음(시군구명))):
                continue

            if not all((값있음(학교급명), 값있음(시도명), 값있음(시군구명))):
                raise ValueError(
                    f"{get_column_letter(col)}열 설정이 불완전합니다. 학교급/시도/시군구를 모두 입력하세요."
                )

            학교급코드, 허용학년목록, _ = 학교급정보(학교급명)
            try:
                학년목록 = 학년목록파싱(sheet.cell(row=5, column=col).value, 허용학년목록)
            except ValueError as exc:
                raise ValueError(f"{get_column_letter(col)}열 학년 설정 오류: {exc}") from exc

            공시연도 = env_year
            if 공시연도 is None:
                공시연도 = 선택연도파싱(sheet.cell(row=6, column=col).value, f"{get_column_letter(col)}열 공시연도")

            api_key = 문자열(sheet.cell(row=1, column=col).value) or default_api_key
            configs.append(
                조회설정(
                    열번호=col,
                    api_key=api_key,
                    학교급명=학교급명,
                    학교급코드=학교급코드,
                    시도명=시도명,
                    시군구명=시군구명,
                    학년목록=학년목록,
                    공시연도=공시연도,
                )
            )

        if not configs:
            raise ValueError("2행~6행(B열 이후)에 검색 조건이 없습니다.")
        return configs
    finally:
        workbook.close()


def 시도시군구코드가져오기(session: requests.Session) -> list[지역코드]:
    response = session.get(REGION_CODE_URL, timeout=30)
    response.raise_for_status()

    workbook = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        codes = []
        for row in range(4, sheet.max_row + 1):
            시도명 = 문자열(sheet.cell(row=row, column=1).value)
            시도코드 = 문자열(sheet.cell(row=row, column=2).value)
            시군구명 = 문자열(sheet.cell(row=row, column=3).value)
            시군구코드 = 문자열(sheet.cell(row=row, column=4).value)

            if not all((시도명, 시도코드, 시군구명, 시군구코드)):
                continue

            codes.append(
                지역코드(
                    시도명=시도명,
                    시도코드=시도코드.zfill(2),
                    시군구명=시군구명,
                    시군구코드=시군구코드.zfill(5),
                )
            )
        return codes
    finally:
        workbook.close()


def 지역코드찾기(코드목록: list[지역코드], 시도명: str, 시군구명: str) -> tuple[str, list[str]]:
    target_sido = 시도핵심(시도명)
    target_sgg = 시군구핵심(시군구명)

    candidates = [code for code in 코드목록 if 시도핵심(code.시도명) == target_sido]
    if not candidates:
        raise ValueError(f"시도 코드를 찾지 못했습니다: {시도명}")

    if target_sgg in {"", "전체"}:
        sgg_codes = sorted({code.시군구코드 for code in candidates if code.시군구코드 != "00000"})
        if not sgg_codes:
            raise ValueError(f"시군구 코드 목록을 찾지 못했습니다: {시도명}")
        return candidates[0].시도코드, sgg_codes

    exact = [code for code in candidates if 시군구핵심(code.시군구명) == target_sgg]
    if exact:
        related = [
            code
            for code in candidates
            if target_sgg in 시군구핵심(code.시군구명) or 시군구핵심(code.시군구명) in target_sgg
        ]
        selected = related if len(related) > len(exact) else exact
        return candidates[0].시도코드, sorted({code.시군구코드 for code in selected})

    partial = [
        code
        for code in candidates
        if target_sgg in 시군구핵심(code.시군구명) or 시군구핵심(code.시군구명) in target_sgg
    ]
    if partial:
        return candidates[0].시도코드, sorted({code.시군구코드 for code in partial})

    raise ValueError(f"시군구 코드를 찾지 못했습니다: {시도명} {시군구명}")


def 학교정보요청(
    session: requests.Session,
    api_key: str,
    api_type: str,
    공시연도: int,
    학교급코드: str,
    시도코드: str,
    시군구코드: str,
) -> list[dict]:
    response = session.get(
        OPENAPI_URL,
        params={
            "apiKey": api_key,
            "apiType": api_type,
            "pbanYr": str(공시연도),
            "schulKndCode": 학교급코드,
            "sidoCode": 시도코드,
            "sggCode": 시군구코드,
        },
        timeout=40,
    )
    response.raise_for_status()
    data = response.json()

    if data.get("resultCode") != "success":
        raise RuntimeError(data.get("resultMsg", "학교알리미 API 호출 실패"))

    return data.get("list", [])


def 공시연도후보(수동연도: int | None) -> list[int]:
    if 수동연도 is not None:
        return [수동연도]
    현재연도 = datetime.now().year
    return [현재연도 - offset for offset in range(DEFAULT_YEAR_CANDIDATES)]


def 중복제거학교데이터(rows: list[dict]) -> list[dict]:
    dedup = {}
    for row in rows:
        school_code = 문자열(row.get("SCHUL_CODE"))
        key = school_code or 문자열(row.get("SCHUL_NM"))
        dedup[key] = row
    return sorted(dedup.values(), key=lambda row: 문자열(row.get("SCHUL_NM")))


def 학교데이터가져오기(
    session: requests.Session,
    설정: 조회설정,
    시도코드: str,
    시군구코드목록: list[str],
    연도후보목록: list[int],
) -> tuple[int, list[dict]]:
    사용연도 = None
    전체행 = []
    마지막오류 = None

    for 시군구코드 in 시군구코드목록:
        rows = None
        if 사용연도 is None:
            for 공시연도 in 연도후보목록:
                try:
                    rows = 학교정보요청(
                        session=session,
                        api_key=설정.api_key,
                        api_type="09",
                        공시연도=공시연도,
                        학교급코드=설정.학교급코드,
                        시도코드=시도코드,
                        시군구코드=시군구코드,
                    )
                    사용연도 = 공시연도
                    break
                except RuntimeError as exc:
                    마지막오류 = exc
        else:
            rows = 학교정보요청(
                session=session,
                api_key=설정.api_key,
                api_type="09",
                공시연도=사용연도,
                학교급코드=설정.학교급코드,
                시도코드=시도코드,
                시군구코드=시군구코드,
            )

        if rows is not None:
            전체행.extend(rows)

    if 사용연도 is None:
        raise RuntimeError(
            "조회 가능한 공시연도를 찾지 못했습니다. "
            f"sggCodes={','.join(시군구코드목록)}, years={','.join(map(str, 연도후보목록))}, "
            f"마지막 오류: {마지막오류}"
        )

    return 사용연도, 중복제거학교데이터(전체행)


def 결과엑셀열기(출력경로: Path, 시트명목록: list[str]):
    normalized_names = 중복없는시트명목록(시트명목록)
    새파일생성 = not 출력경로.exists()
    workbook = openpyxl.Workbook() if 새파일생성 else openpyxl.load_workbook(출력경로)

    if 새파일생성 and "Sheet" in workbook.sheetnames and len(workbook.sheetnames) == 1:
        기본시트 = workbook["Sheet"]
    else:
        기본시트 = None

    for name in normalized_names:
        if name in workbook.sheetnames:
            sheet = workbook[name]
            for merged_range in list(sheet.merged_cells.ranges):
                sheet.unmerge_cells(str(merged_range))
            sheet.freeze_panes = None
            sheet.delete_rows(1, max(sheet.max_row, 1))
        else:
            workbook.create_sheet(name)

    if 기본시트 is not None and len(workbook.sheetnames) > len(normalized_names):
        del workbook["Sheet"]

    for index, name in enumerate(normalized_names):
        current_index = workbook.sheetnames.index(name)
        workbook.move_sheet(workbook[name], offset=index - current_index)

    return workbook, {name: workbook[name] for name in normalized_names}, normalized_names


def 헤더추가(ws, 학년목록: tuple[int, ...]) -> None:
    ws.freeze_panes = "B3"

    ws.cell(row=HEADER_TOP_ROW, column=1, value="학교명")
    ws.merge_cells(start_row=HEADER_TOP_ROW, start_column=1, end_row=HEADER_BOTTOM_ROW, end_column=1)
    ws.cell(row=HEADER_TOP_ROW, column=1).alignment = HEADER_ALIGNMENT
    ws.cell(row=HEADER_TOP_ROW, column=1).font = HEADER_FONT

    current_col = 2
    grade_headers = [f"{grade}학년" for grade in 학년목록]

    for title in ("학급수", "학생수", "반평균"):
        sub_headers = grade_headers + ["합계"]
        ws.cell(row=HEADER_TOP_ROW, column=current_col, value=title)
        ws.merge_cells(
            start_row=HEADER_TOP_ROW,
            start_column=current_col,
            end_row=HEADER_TOP_ROW,
            end_column=current_col + len(sub_headers) - 1,
        )
        ws.cell(row=HEADER_TOP_ROW, column=current_col).alignment = HEADER_ALIGNMENT
        ws.cell(row=HEADER_TOP_ROW, column=current_col).font = HEADER_FONT

        for sub_header in sub_headers:
            ws.cell(row=HEADER_BOTTOM_ROW, column=current_col, value=sub_header)
            ws.cell(row=HEADER_BOTTOM_ROW, column=current_col).alignment = HEADER_ALIGNMENT
            ws.cell(row=HEADER_BOTTOM_ROW, column=current_col).font = HEADER_FONT
            current_col += 1

    ws.column_dimensions["A"].width = 26
    for col in range(2, current_col):
        ws.column_dimensions[get_column_letter(col)].width = 10


def 결과시트작성(ws, 학교데이터: list[dict], 학년목록: tuple[int, ...]) -> None:
    학급수합계 = [0] * len(학년목록)
    학생수합계 = [0] * len(학년목록)
    current_row = DATA_START_ROW

    for 학교행 in 학교데이터:
        학년별학급수 = [숫자값(학교행.get(f"COL_C{grade}")) for grade in 학년목록]
        학년별학생수 = [숫자값(학교행.get(f"COL_S{grade}")) for grade in 학년목록]
        학년별반평균 = [반평균(학생수, 학급수) for 학생수, 학급수 in zip(학년별학생수, 학년별학급수)]
        학급수합 = sum(학년별학급수)
        학생수합 = sum(학년별학생수)
        row_values = [
            문자열(학교행.get("SCHUL_NM")),
            *학년별학급수,
            학급수합,
            *학년별학생수,
            학생수합,
            *학년별반평균,
            반평균(학생수합, 학급수합),
        ]
        for column, value in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=column, value=value)
            if column > 1 + (len(학년목록) + 1) * 2:
                cell.number_format = "0.0"

        for index, value in enumerate(학년별학급수):
            학급수합계[index] += value
        for index, value in enumerate(학년별학생수):
            학생수합계[index] += value
        current_row += 1

    summary_values = [
        "학년별 합계",
        *학급수합계,
        sum(학급수합계),
        *학생수합계,
        sum(학생수합계),
        *[반평균(학생수, 학급수) for 학생수, 학급수 in zip(학생수합계, 학급수합계)],
        반평균(sum(학생수합계), sum(학급수합계)),
    ]
    for column, value in enumerate(summary_values, start=1):
        cell = ws.cell(row=current_row, column=column, value=value)
        if column > 1 + (len(학년목록) + 1) * 2:
            cell.number_format = "0.0"

    summary_row = current_row
    for cell in ws[summary_row]:
        cell.font = HEADER_FONT


def 결과파일저장(workbook, 공시연도: int) -> Path:
    출력경로 = BASE_DIR / RESULT_FILE_TEMPLATE.format(year=공시연도)
    try:
        workbook.save(출력경로)
        return 출력경로
    except PermissionError:
        대체경로 = BASE_DIR / RESULT_FALLBACK_TEMPLATE.format(
            year=공시연도,
            stamp=datetime.now().strftime("%Y%m%d_%H%M%S"),
        )
        workbook.save(대체경로)
        print(f"결과 엑셀 파일이 열려 있어 새 파일로 저장했습니다: {대체경로}")
        return 대체경로


def 결과파일열기(파일경로: Path) -> None:
    try:
        if hasattr(os, "startfile"):
            os.startfile(str(파일경로))
            return

        if sys.platform == "darwin":
            subprocess.Popen(["open", str(파일경로)])
        else:
            subprocess.Popen(["xdg-open", str(파일경로)])
    except Exception as exc:
        print(f"결과 파일 자동 열기에 실패했습니다: {exc}")


def 실행() -> None:
    설정목록 = 입력설정읽기()

    with requests.Session() as session:
        코드목록 = 시도시군구코드가져오기(session)
        작업목록 = []
        공통공시연도 = None

        for 설정 in 설정목록:
            시도코드, 시군구코드목록 = 지역코드찾기(코드목록, 설정.시도명, 설정.시군구명)
            연도후보목록 = [공통공시연도] if 공통공시연도 is not None and 설정.공시연도 is None else 공시연도후보(설정.공시연도)
            사용연도, 학교데이터 = 학교데이터가져오기(session, 설정, 시도코드, 시군구코드목록, 연도후보목록)

            if 공통공시연도 is None:
                공통공시연도 = 사용연도
            elif 사용연도 != 공통공시연도:
                raise RuntimeError(
                    "검색별 공시연도가 다릅니다. "
                    f"({공통공시연도} vs {사용연도}) "
                    f"시트={설정.시도명}_{설정.시군구명}_{설정.학교급명}"
                )

            작업목록.append(
                작업결과(
                    시트명원본=시트명원본(설정.시도명, 설정.시군구명, 설정.학교급명),
                    학년목록=설정.학년목록,
                    학교데이터=학교데이터,
                )
            )

    if 공통공시연도 is None:
        raise RuntimeError("유효한 검색 결과가 없습니다.")

    workbook, 시트맵, 실제시트명목록 = 결과엑셀열기(
        BASE_DIR / RESULT_FILE_TEMPLATE.format(year=공통공시연도),
        [작업.시트명원본 for 작업 in 작업목록],
    )

    try:
        for 시트명, 작업 in zip(실제시트명목록, 작업목록):
            ws = 시트맵[시트명]
            헤더추가(ws, 작업.학년목록)
            결과시트작성(ws, 작업.학교데이터, 작업.학년목록)

        저장경로 = 결과파일저장(workbook, 공통공시연도)
    finally:
        workbook.close()

    요약 = ", ".join(f"{시트명}:{len(작업.학교데이터)}개" for 시트명, 작업 in zip(실제시트명목록, 작업목록))
    print(f"공시연도 {공통공시연도}, 시트 {len(작업목록)}개 저장 완료 ({저장경로})")
    print(f"시트별 건수: {요약}")
    결과파일열기(저장경로)


if __name__ == "__main__":
    실행()
