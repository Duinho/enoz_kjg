import importlib.util
import json
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


BASE_DIR = Path(__file__).resolve().parent
MAIN_SCRIPT = BASE_DIR / "LMS전산반배정.py"
WORKBOOK = BASE_DIR / "LMS전산반배정.xlsx"


def load_main_module():
    spec = importlib.util.spec_from_file_location("lms_class_assignment", MAIN_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def main():
    module = load_main_module()
    module.초기화()
    login_url = module.ws.cell(row=1, column=17).value
    course_url = module.ws.cell(row=2, column=17).value
    admin_id = module.ws.cell(row=3, column=17).value
    admin_pw = module.ws.cell(row=4, column=17).value
    target_date = module.ws.cell(row=5, column=17).value

    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["메인"]
    target_class = str(ws.cell(row=27, column=6).value).strip()
    target_user = str(ws.cell(row=27, column=8).value).strip()
    target_group_no = "202606_S2_2_10_1900_35_11"

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=["--disable-popup-blocking"])
        try:
            page = browser.new_page()
            page.goto(login_url)
            page.fill('input[name="tbAdminId"]', str(admin_id))
            page.fill('input[name="tbAdminPass"]', str(admin_pw))
            page.press('input[name="tbAdminPass"]', 'Enter')
            page.wait_for_load_state("load")
            page.wait_for_timeout(1500)
            page.goto(course_url)
            page.wait_for_load_state("load")
            page.select_option('select[name="ddlTargetDate"]', value=str(target_date))
            page.select_option('select[name="ddlKeyField"]', value='b.m_id')
            page.fill('input[name="tbKeyWord"].font_blue', target_user)
            page.press('input[name="tbKeyWord"].font_blue', 'Enter')
            with page.context.expect_page() as page_info:
                page.click(f'a[href*="{target_date}"].button_red_small')
            detail = page_info.value
            detail.wait_for_load_state("load")
            with detail.expect_navigation(wait_until="load"):
                detail.evaluate(
                    """(value) => {
                        const select = document.querySelector('select[name="ddlTargetGroupNo"]');
                        const opt = document.createElement('option');
                        opt.value = value;
                        opt.textContent = value + ' injected dry-run';
                        select.appendChild(opt);
                        select.value = value;
                        document.frm.action = 'popCourseEdit.asp';
                        document.frm.submit();
                    }""",
                    target_group_no,
                )
            detail.wait_for_load_state("load")
            data = detail.evaluate(
                """(expectedClass) => {
                    const select = document.querySelector('select[name="ddlTargetGroupNo"]');
                    const form = document.forms.frm;
                    const getValue = (name) => {
                        const el = form ? form.querySelector(`[name="${name}"]`) : null;
                        return el ? el.value : null;
                    };
                    return {
                        url: location.href,
                        expectedClass,
                        selectedValue: select ? select.value : null,
                        optionExists: select ? Array.from(select.options).some((opt) => opt.value === getValue('hhdTargetGroupNo')) : false,
                        hiddenTargetGroupNo: getValue('hhdTargetGroupNo'),
                        currentGroupNo: getValue('group_no'),
                        hasTargetOption: select ? Array.from(select.options).some((opt) => opt.value === '202606_S2_2_10_1900_35_11') : false,
                        optionTexts: select ? Array.from(select.options).filter((opt) => /35_1[1-6]|PH_TT_1[1-6]/.test(opt.textContent || opt.value)).map((opt) => ({value: opt.value, text: opt.textContent.trim()})) : [],
                        bodyHasSqlError: /오류|SQL|DBHelper|잘못/.test(document.body.innerText || ''),
                        bodySnippet: (document.body.innerText || '').slice(0, 600),
                    };
                }""",
                target_class,
            )
            print(json.dumps(data, ensure_ascii=False, indent=2))
            detail.close()
        finally:
            browser.close()


if __name__ == "__main__":
    main()
