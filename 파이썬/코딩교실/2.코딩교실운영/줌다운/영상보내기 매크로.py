import json
import os
import re
import time
from dataclasses import dataclass
from datetime import datetime, timedelta

import openpyxl
import pytz
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright


폴더경로 = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(폴더경로, "영상보내기.xlsx")
기본_서비스계정_json = os.path.join(폴더경로, "auto-send-link-74f62fdbda52.json")


@dataclass(frozen=True)
class 공통설정:
    알람받을번호: str
    api키: str
    멘트1: str
    멘트2: str
    문자박스링크: str
    문자박스아이디: str
    문자박스비번: str


@dataclass(frozen=True)
class 세트설정:
    이름: str
    시트링크_월수: str
    시트링크_화목: str
    드라이브링크: str
    회신번호: str


@dataclass(frozen=True)
class 처리문맥:
    날짜: str
    요일: str
    시트이름: str
    시트링크: str
    드라이브링크: str
    회신번호: str


@dataclass(frozen=True)
class 대상자:
    행번호: int
    반번호: str
    학생이름: str
    학생번호: str
    학부모번호: str


class 드라이브링크해결기:
    def __init__(self, drive_service):
        self.drive_service = drive_service
        self.folder_cache = {}
        self.subfolder_cache = {}
        self.video_link_cache = {}

    def find_folder_id(self, folder_name, parent_folder_id=None):
        cache_key = (folder_name, parent_folder_id)
        if cache_key in self.folder_cache:
            return self.folder_cache[cache_key]

        query = f"name = '{folder_name}' and mimeType = 'application/vnd.google-apps.folder'"
        if parent_folder_id:
            query += f" and '{parent_folder_id}' in parents"

        results = self.drive_service.files().list(
            q=query,
            fields="files(id, name, webViewLink)",
        ).execute()
        items = results.get("files", [])
        if not items:
            raise ValueError(f"No folder named '{folder_name}' found.")

        value = (items[0]["id"], items[0].get("webViewLink"))
        self.folder_cache[cache_key] = value
        return value

    def get_subfolders(self, parent_folder_id):
        if parent_folder_id in self.subfolder_cache:
            return self.subfolder_cache[parent_folder_id]

        query = f"'{parent_folder_id}' in parents and mimeType = 'application/vnd.google-apps.folder'"
        results = self.drive_service.files().list(
            q=query,
            fields="files(id, name)",
        ).execute()
        items = results.get("files", [])
        self.subfolder_cache[parent_folder_id] = items
        return items

    def get_video_link(self, 드라이브링크, 요일, 날짜, 반번호):
        cache_key = (드라이브링크, 요일, 날짜, 반번호)
        if cache_key in self.video_link_cache:
            return self.video_link_cache[cache_key]

        folder_match = re.search(r"/folders/([a-zA-Z0-9-_]+)", 드라이브링크 or "")
        if not folder_match:
            raise ValueError("Invalid Google Drive folder URL")

        root_folder_id = folder_match.group(1)
        날짜가공 = format_date(날짜)
        반이름 = format_class_name(반번호)

        요일폴더_id, _ = self.find_folder_id(요일, parent_folder_id=root_folder_id)
        날짜폴더_id = None
        for subfolder in self.get_subfolders(요일폴더_id):
            try:
                if format_date(subfolder["name"]) == 날짜가공:
                    날짜폴더_id = subfolder["id"]
                    break
            except ValueError:
                continue

        if not 날짜폴더_id:
            self.video_link_cache[cache_key] = None
            return None

        for subfolder in self.get_subfolders(날짜폴더_id):
            if 반이름 and 반이름 in subfolder["name"]:
                _, 영상링크 = self.find_folder_id(
                    subfolder["name"],
                    parent_folder_id=날짜폴더_id,
                )
                self.video_link_cache[cache_key] = 영상링크
                return 영상링크

        self.video_link_cache[cache_key] = None
        return None


def text_or_empty(value):
    if value is None:
        return ""
    return str(value).strip()


def format_date(date_str):
    match = re.findall(r"\d+", str(date_str))
    if len(match) < 2:
        raise ValueError("Invalid date format")
    return "".join(match[-2:])


def extract_spreadsheet_id(url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", url or "")
    if not match:
        raise ValueError("Invalid Google Sheets URL")
    return match.group(1)


def format_class_name(raw_value):
    raw_text = text_or_empty(raw_value)
    if not raw_text:
        return ""

    match = re.match(r"^(\d+)(반)?$", raw_text)
    if match:
        return f"{match.group(1).zfill(2)}반"
    return raw_text


def 어제날짜_문자열():
    kst = pytz.timezone("Asia/Seoul")
    yesterday = datetime.now(kst) - timedelta(days=1)
    weekday_dict = {
        "Monday": "월",
        "Tuesday": "화",
        "Wednesday": "수",
        "Thursday": "목",
        "Friday": "금",
        "Saturday": "토",
        "Sunday": "일",
    }
    weekday_kor = weekday_dict[yesterday.strftime("%A")]
    return f"{yesterday.month}/{yesterday.day}({weekday_kor})", weekday_kor


def resolve_context(raw_date, 세트):
    날짜 = text_or_empty(raw_date)
    if 날짜 == "자동":
        날짜, weekday_kor = 어제날짜_문자열()
    else:
        weekday_kor = None

    if "월" in 날짜 or "수" in 날짜:
        return 처리문맥(
            날짜=날짜,
            요일="월수",
            시트이름="출석부(월/수)",
            시트링크=세트.시트링크_월수,
            드라이브링크=세트.드라이브링크,
            회신번호=세트.회신번호,
        )
    if "화" in 날짜 or "목" in 날짜:
        return 처리문맥(
            날짜=날짜,
            요일="화목",
            시트이름="출석부(화/목)",
            시트링크=세트.시트링크_화목,
            드라이브링크=세트.드라이브링크,
            회신번호=세트.회신번호,
        )

    if weekday_kor:
        raise ValueError(f"자동 날짜가 발송 요일과 맞지 않습니다: {날짜}")
    raise ValueError(f"지원하지 않는 날짜 형식입니다: {날짜}")


def load_settings():
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active

    common = 공통설정(
        알람받을번호=text_or_empty(worksheet.cell(row=5, column=15).value),
        api키=text_or_empty(worksheet.cell(row=2, column=17).value),
        멘트1=text_or_empty(worksheet.cell(row=6, column=15).value),
        멘트2=text_or_empty(worksheet.cell(row=7, column=15).value),
        문자박스링크=text_or_empty(worksheet.cell(row=8, column=15).value),
        문자박스아이디=text_or_empty(worksheet.cell(row=9, column=15).value),
        문자박스비번=text_or_empty(worksheet.cell(row=10, column=15).value),
    )

    세트행정보 = [
        ("세트1", 3, 4, 5, 6),
        ("세트2", 7, 8, 9, 10),
        ("세트3", 11, 12, 13, 14),
    ]
    세트목록 = []
    for 이름, 월수행, 화목행, 드라이브행, 회신행 in 세트행정보:
        세트목록.append(
            세트설정(
                이름=이름,
                시트링크_월수=text_or_empty(worksheet.cell(row=월수행, column=17).value),
                시트링크_화목=text_or_empty(worksheet.cell(row=화목행, column=17).value),
                드라이브링크=text_or_empty(worksheet.cell(row=드라이브행, column=17).value),
                회신번호=f"010{text_or_empty(worksheet.cell(row=회신행, column=17).value)}",
            )
        )

    raw_date = text_or_empty(worksheet.cell(row=1, column=17).value)
    workbook.close()
    return common, raw_date, 세트목록


def find_date_column(시트데이터, 날짜):
    if len(시트데이터) < 2:
        return None
    for col_index, value in enumerate(시트데이터[1], start=1):
        if value == 날짜:
            return col_index
    return None


def find_target_rows(시트데이터, column_index, search_value):
    rows = []
    for row_index, row in enumerate(시트데이터[1:], start=2):
        if len(row) >= column_index and row[column_index - 1] == search_value:
            rows.append(row_index)
    return rows


def get_cell_value(row_values, column_index):
    if column_index is None or column_index <= 0 or len(row_values) < column_index:
        return ""
    return text_or_empty(row_values[column_index - 1])


def find_class_number(시트데이터, row_index):
    current_row = row_index
    while current_row > 0:
        row_values = 시트데이터[current_row - 1]
        if row_values:
            formatted = format_class_name(row_values[0])
            if formatted:
                return formatted
        current_row -= 1
    return ""


def parse_sheet_targets(sheets_service, 문맥):
    시트아이디 = extract_spreadsheet_id(문맥.시트링크)
    result = sheets_service.spreadsheets().values().get(
        spreadsheetId=시트아이디,
        range=문맥.시트이름,
    ).execute()
    시트데이터 = result.get("values", [])
    if not 시트데이터:
        raise ValueError("No data found.")

    header_row = 시트데이터[1] if len(시트데이터) > 1 else []
    전화번호열 = [
        col_index
        for col_index, value in enumerate(header_row, start=1)
        if value == "전화번호"
    ]
    이름열 = next(
        (col_index for col_index, value in enumerate(header_row, start=1) if value == "이름"),
        None,
    )

    if len(전화번호열) < 2:
        raise ValueError("전화번호 열이 두 개 발견되지 않았습니다.")
    if 이름열 is None:
        raise ValueError("이름 열이 발견되지 않았습니다.")

    날짜위치 = find_date_column(시트데이터, 문맥.날짜)
    if 날짜위치 is None:
        raise ValueError(f'날짜 "{문맥.날짜}" not found in the second row')

    영상행목록 = find_target_rows(시트데이터, 날짜위치, "영상")
    대상목록 = []
    for row_index in 영상행목록:
        row_values = 시트데이터[row_index - 1]
        대상목록.append(
            대상자(
                행번호=row_index,
                반번호=find_class_number(시트데이터, row_index),
                학생이름=get_cell_value(row_values, 이름열),
                학생번호=get_cell_value(row_values, 전화번호열[0]),
                학부모번호=get_cell_value(row_values, 전화번호열[1]),
            )
        )

    return 시트아이디, 날짜위치, 대상목록


def build_recipient_text(대상):
    recipients = []
    for number in [대상.학생번호, 대상.학부모번호]:
        if number and number not in recipients:
            recipients.append(number)
    return "\n".join(recipients)


def send_message(page, recipients, message, callback_number):
    if not recipients:
        raise ValueError("수신번호가 없습니다.")

    page.fill("textarea#recvList", recipients)
    page.fill("textarea#msg", message)
    page.locator("a.hand.openLayer", has_text="선택").click()

    frame = page.frame(name="callbackFrame")
    frame.wait_for_selector("span:has-text('전체 회신번호')", timeout=5000)
    frame.locator("span:has-text('전체 회신번호')").click()
    frame.wait_for_selector("#bulkCallbackNum", timeout=3000)
    frame.fill("#bulkCallbackNum", callback_number)
    frame.click("button[onclick*='callbackCheckForm']")

    time.sleep(0.3)
    page.locator("a.hand.openLayer", has_text="전송하기").click()
    time.sleep(0.3)
    page.keyboard.press("Enter")
    time.sleep(0.3)
    page.keyboard.press("Enter")


def load_service_account_credentials():
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    json_text = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if json_text:
        return Credentials.from_service_account_info(json.loads(json_text), scopes=scopes)

    json_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", 기본_서비스계정_json)
    if os.path.exists(json_path):
        return Credentials.from_service_account_file(json_path, scopes=scopes)
    return None


def update_sheet(sheets_write_service, 시트아이디, 시트이름, 날짜위치, 행번호목록):
    if not 행번호목록:
        return

    if sheets_write_service is None:
        print("서비스 계정 JSON이 없어 Google Sheets 수정은 건너뜁니다.")
        return

    spreadsheet = sheets_write_service.spreadsheets().get(spreadsheetId=시트아이디).execute()
    sheet_id = None
    for sheet in spreadsheet["sheets"]:
        if sheet["properties"]["title"] == 시트이름:
            sheet_id = sheet["properties"]["sheetId"]
            break

    if sheet_id is None:
        raise ValueError(f"Sheet name '{시트이름}' not found in the spreadsheet")

    value_updates = []
    note_updates = []
    column_letter = get_column_letter(날짜위치)
    for row_index in 행번호목록:
        value_updates.append(
            {
                "range": f"{시트이름}!{column_letter}{row_index}",
                "values": [["O"]],
            }
        )
        note_updates.append(
            {
                "updateCells": {
                    "range": {
                        "sheetId": sheet_id,
                        "startRowIndex": row_index - 1,
                        "endRowIndex": row_index,
                        "startColumnIndex": 날짜위치 - 1,
                        "endColumnIndex": 날짜위치,
                    },
                    "rows": [
                        {
                            "values": [
                                {
                                    "userEnteredValue": {"stringValue": "O"},
                                    "note": "영상 발송 완료",
                                }
                            ]
                        }
                    ],
                    "fields": "note,userEnteredValue",
                }
            }
        )

    sheets_write_service.spreadsheets().values().batchUpdate(
        spreadsheetId=시트아이디,
        body={"valueInputOption": "RAW", "data": value_updates},
    ).execute()
    sheets_write_service.spreadsheets().batchUpdate(
        spreadsheetId=시트아이디,
        body={"requests": note_updates},
    ).execute()


def 로그인(page, common):
    page.goto(common.문자박스링크)
    page.fill("input[name='id']", common.문자박스아이디)
    page.fill("input[name='pwd']", common.문자박스비번)
    page.press("input[name='pwd']", "Enter")
    time.sleep(1)
    try:
        page.click("button[onclick*='contentsLayerClose']")
    except Exception:
        pass


def process_set(page, common, set_config, raw_date, sheets_read_service, sheets_write_service, drive_resolver):
    문맥 = resolve_context(raw_date, set_config)
    시트아이디, 날짜위치, 대상목록 = parse_sheet_targets(sheets_read_service, 문맥)

    if not 대상목록:
        print(f"{set_config.이름}: 발송 대상이 없습니다.")
        return 문맥.회신번호

    성공행목록 = []
    for 대상 in 대상목록:
        recipients = build_recipient_text(대상)
        if not recipients:
            print(f"{대상.행번호}행: 수신번호가 없어 건너뜁니다.")
            continue

        영상링크 = drive_resolver.get_video_link(
            문맥.드라이브링크,
            문맥.요일,
            문맥.날짜,
            대상.반번호,
        )
        if not 영상링크:
            print(f"{대상.반번호} {대상.학생이름}: 영상 링크를 찾지 못해 건너뜁니다.")
            continue

        문자내용 = f"{common.멘트1}\n영상 링크 : {영상링크}\n{common.멘트2}"
        send_message(page, recipients, 문자내용, 문맥.회신번호)
        성공행목록.append(대상.행번호)
        print(f"{대상.반번호} {대상.학생이름} 학생 영상 발송 완료")

    update_sheet(
        sheets_write_service,
        시트아이디,
        문맥.시트이름,
        날짜위치,
        성공행목록,
    )
    return 문맥.회신번호


def 완료(page, 알람받을번호, 회신번호):
    send_message(page, 알람받을번호, "영상 발송 완료", 회신번호)


def 동작():
    common, raw_date, 세트목록 = load_settings()
    sheets_read_service = build("sheets", "v4", developerKey=common.api키)
    drive_service = build("drive", "v3", developerKey=common.api키)
    drive_resolver = 드라이브링크해결기(drive_service)

    credentials = load_service_account_credentials()
    sheets_write_service = None
    if credentials is not None:
        sheets_write_service = build("sheets", "v4", credentials=credentials)

    마지막_회신번호 = ""
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=["--disable-popup-blocking"])
        page = browser.new_page()

        로그인(page, common)
        for set_config in 세트목록:
            마지막_회신번호 = process_set(
                page,
                common,
                set_config,
                raw_date,
                sheets_read_service,
                sheets_write_service,
                drive_resolver,
            )

        if common.알람받을번호 and 마지막_회신번호:
            완료(page, common.알람받을번호, 마지막_회신번호)
        print("모든 영상 링크 발송 완료")
        browser.close()


동작()
