import json
import os
import re
import shutil
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from threading import Lock, local
from urllib.parse import quote
from zoneinfo import ZoneInfo

import google.auth
import openpyxl
import requests
from google.auth.exceptions import DefaultCredentialsError
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials as OAuthCredentials
from google.oauth2.service_account import Credentials as ServiceAccountCredentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from playwright.sync_api import sync_playwright
from requests.auth import HTTPBasicAuth


script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, "config.json")
excel_file_path = os.path.join(script_dir, "영상보내기.xlsx")
sender_script_path = os.path.join(script_dir, "영상보내기 매크로.py")
기본_서비스계정_json = os.path.join(script_dir, "auto-send-link-74f62fdbda52.json")
기본_oauth_client_json = os.path.join(script_dir, "google_oauth_client.json")
기본_oauth_token_json = os.path.join(script_dir, "google_token.json")
BASE_DIR = os.path.join(script_dir, "녹화영상")
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

priority_types = [
    "shared_screen_with_gallery_view",
    "shared_screen_with_speaker_view",
    "shared_screen",
    "speaker_view",
    "gallery_view",
]

topic_aliases = {
    "향균 강의 개인 회의실": "2026 포항 SW 코딩교육 1기 11반",
}

processed_meetings = set()
processed_meetings_lock = Lock()


@dataclass(frozen=True)
class Zoom설정:
    client_id: str
    client_secret: str
    account_id: str
    start_date: str
    end_date: str
    delete_after_download: bool
    max_workers: int
    upload_workers: int
    upload_to_drive: bool
    send_video_links_after_upload: bool
    send_completion_sms: bool


@dataclass(frozen=True)
class Drive세트설정:
    이름: str
    사업명: str
    드라이브링크: str


def text_or_empty(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_bool(value, default):
    if value is None:
        return default
    if isinstance(value, bool):
        return value

    value_text = str(value).strip().lower()
    if value_text in {"1", "true", "yes", "y", "on"}:
        return True
    if value_text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def resolve_recording_date(value):
    value_text = text_or_empty(value).lower()
    if value_text in {"", "auto", "자동", "yesterday", "어제"}:
        yesterday = datetime.now(ZoneInfo("Asia/Seoul")).date() - timedelta(days=1)
        return yesterday.isoformat()
    return text_or_empty(value)


def load_local_config():
    if not os.path.exists(config_path):
        return {}

    with open(config_path, "r", encoding="utf-8-sig") as file:
        return json.load(file)


def load_zoom_config():
    config = load_local_config()
    zoom_section = config.get("Zoom", {})
    recordings_section = config.get("Recordings", {})
    automation_section = config.get("Automation", {})

    settings = Zoom설정(
        client_id=os.getenv("ZOOM_CLIENT_ID", text_or_empty(zoom_section.get("client_id"))),
        client_secret=os.getenv("ZOOM_CLIENT_SECRET", text_or_empty(zoom_section.get("client_secret"))),
        account_id=os.getenv("ZOOM_ACCOUNT_ID", text_or_empty(zoom_section.get("account_id"))),
        start_date=resolve_recording_date(
            os.getenv("ZOOM_RECORDINGS_START_DATE", recordings_section.get("start_date"))
        ),
        end_date=resolve_recording_date(
            os.getenv("ZOOM_RECORDINGS_END_DATE", recordings_section.get("end_date"))
        ),
        delete_after_download=parse_bool(
            os.getenv("ZOOM_DELETE_AFTER_DOWNLOAD", recordings_section.get("delete_after_download")),
            False,
        ),
        max_workers=max(
            1,
            int(os.getenv("ZOOM_MAX_WORKERS", recordings_section.get("max_workers", 4))),
        ),
        upload_workers=max(
            1,
            int(os.getenv("ZOOM_DRIVE_UPLOAD_WORKERS", recordings_section.get("upload_workers", 4))),
        ),
        upload_to_drive=parse_bool(
            os.getenv("ZOOM_UPLOAD_TO_DRIVE", recordings_section.get("upload_to_drive")),
            True,
        ),
        send_video_links_after_upload=parse_bool(
            os.getenv(
                "ZOOM_SEND_VIDEO_LINKS_AFTER_UPLOAD",
                automation_section.get("send_video_links_after_upload"),
            ),
            True,
        ),
        send_completion_sms=parse_bool(
            os.getenv("ZOOM_SEND_COMPLETION_SMS", automation_section.get("send_completion_sms")),
            False,
        ),
    )

    missing = []
    if not settings.client_id:
        missing.append("ZOOM_CLIENT_ID")
    if not settings.client_secret:
        missing.append("ZOOM_CLIENT_SECRET")
    if not settings.account_id:
        missing.append("ZOOM_ACCOUNT_ID")
    if not settings.start_date:
        missing.append("ZOOM_RECORDINGS_START_DATE")
    if not settings.end_date:
        missing.append("ZOOM_RECORDINGS_END_DATE")

    if missing:
        raise ValueError(
            "Zoom 설정이 부족합니다. config.json 또는 환경변수를 확인하세요: "
            + ", ".join(missing)
            + "\nZoom 로그인 ID/PW가 아니라 Zoom Marketplace의 Server-to-Server OAuth "
            "Account ID, Client ID, Client Secret이 필요합니다."
        )

    return settings


def save_oauth_credentials(credentials, token_path):
    with open(token_path, "w", encoding="utf-8") as token_file:
        token_file.write(credentials.to_json())


def load_google_credentials():
    json_text = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if json_text:
        print("Google 인증: 환경변수 GOOGLE_SERVICE_ACCOUNT_JSON 사용")
        return ServiceAccountCredentials.from_service_account_info(json.loads(json_text), scopes=GOOGLE_SCOPES)

    json_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", 기본_서비스계정_json)
    if os.path.exists(json_path):
        print(f"Google 인증: 서비스 계정 JSON 사용 - {json_path}")
        return ServiceAccountCredentials.from_service_account_file(json_path, scopes=GOOGLE_SCOPES)

    token_path = os.getenv("GOOGLE_OAUTH_TOKEN_FILE", 기본_oauth_token_json)
    if os.path.exists(token_path):
        credentials = OAuthCredentials.from_authorized_user_file(token_path, GOOGLE_SCOPES)
        if not credentials.has_scopes(GOOGLE_SCOPES):
            print(f"Google 인증: OAuth 토큰 권한이 부족해 재로그인합니다 - {token_path}")
        else:
            if credentials.expired and credentials.refresh_token:
                credentials.refresh(Request())
                save_oauth_credentials(credentials, token_path)
            if credentials.valid:
                print(f"Google 인증: 저장된 OAuth 토큰 사용 - {token_path}")
                return credentials
            print(f"Google 인증: OAuth 토큰이 만료/무효입니다 - {token_path}")

    client_path = os.getenv("GOOGLE_OAUTH_CLIENT_FILE", 기본_oauth_client_json)
    if os.path.exists(client_path):
        print(f"Google 인증: OAuth 클라이언트로 로그인 진행 - {client_path}")
        flow = InstalledAppFlow.from_client_secrets_file(client_path, GOOGLE_SCOPES)
        credentials = flow.run_local_server(port=0, prompt="consent")
        save_oauth_credentials(credentials, token_path)
        return credentials

    try:
        credentials, _ = google.auth.default(scopes=GOOGLE_SCOPES)
        if credentials.expired and getattr(credentials, "refresh_token", None):
            credentials.refresh(Request())
        print("Google 인증: Application Default Credentials 사용")
        return credentials
    except DefaultCredentialsError:
        return None


def load_drive_sets():
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active
    rows = [
        ("세트1", 3, 5),
        ("세트2", 7, 9),
        ("세트3", 11, 13),
    ]
    drive_sets = []
    for name, label_row, drive_row in rows:
        business_name = text_or_empty(worksheet.cell(row=label_row, column=18).value) or name
        drive_link = text_or_empty(worksheet.cell(row=drive_row, column=17).value)
        if drive_link:
            drive_sets.append(Drive세트설정(name, business_name, drive_link))
    workbook.close()
    return drive_sets


def extract_drive_folder_id(drive_link):
    folder_match = re.search(r"/folders/([a-zA-Z0-9-_]+)", drive_link or "")
    if not folder_match:
        raise ValueError("Invalid Google Drive folder URL")
    return folder_match.group(1)


def escape_drive_query_text(value):
    return text_or_empty(value).replace("\\", "\\\\").replace("'", "\\'")


def format_drive_date(dt):
    return f"{dt.year}년 {dt.month}월 {dt.day}일"


def parse_drive_date(date_text):
    match = re.search(r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일", text_or_empty(date_text))
    if not match:
        raise ValueError(f"날짜 폴더명을 해석하지 못했습니다: {date_text}")
    year, month, day = map(int, match.groups())
    return datetime(year, month, day)


def get_day_group(dt):
    weekday = dt.weekday()
    if weekday in {0, 2}:
        return "월수"
    if weekday in {1, 3}:
        return "화목"
    return None


def extract_class_name(topic):
    match = re.search(r"(?<!\d)(\d{1,2})\s*반", topic)
    if match:
        return f"{int(match.group(1)):02d}반"
    return None


def extract_topic_from_local_file(file_path):
    file_name = os.path.splitext(os.path.basename(file_path))[0]
    match = re.match(r"(.+?) _\d{4}년 \d{1,2}월 \d{1,2}일(?:_\d+)?$", file_name)
    if match:
        return match.group(1)
    parent_name = os.path.basename(os.path.dirname(file_path))
    match = re.match(r"(.+?) _\d{4}년 \d{1,2}월 \d{1,2}일$", parent_name)
    if match:
        return match.group(1)
    return file_name


def extract_date_from_local_file(file_path):
    parts = os.path.normpath(file_path).split(os.sep)
    for part in reversed(parts):
        if re.search(r"\d{4}년\s*\d{1,2}월\s*\d{1,2}일", part):
            return parse_drive_date(part)
    raise ValueError(f"파일 경로에서 날짜를 찾지 못했습니다: {file_path}")


class Drive업로더:
    def __init__(self, credentials, drive_sets):
        self.credentials = credentials
        self.drive_sets = drive_sets
        self.folder_cache = {}
        self.folder_lock = Lock()
        self.thread_local = local()

    def get_service(self):
        if not hasattr(self.thread_local, "drive_service"):
            self.thread_local.drive_service = build(
                "drive",
                "v3",
                credentials=self.credentials,
                cache_discovery=False,
            )
        return self.thread_local.drive_service

    def find_drive_set(self, topic):
        topic_text = text_or_empty(topic)
        for drive_set in self.drive_sets:
            if drive_set.사업명 and drive_set.사업명 in topic_text:
                return drive_set
        if len(self.drive_sets) == 1:
            return self.drive_sets[0]
        return None

    def get_or_create_folder(self, service, folder_name, parent_folder_id):
        cache_key = (parent_folder_id, folder_name)
        with self.folder_lock:
            if cache_key in self.folder_cache:
                return self.folder_cache[cache_key]

            safe_name = escape_drive_query_text(folder_name)
            query = (
                f"name = '{safe_name}' and mimeType = 'application/vnd.google-apps.folder' "
                f"and '{parent_folder_id}' in parents and trashed = false"
            )
            results = service.files().list(
                q=query,
                fields="files(id, name, webViewLink)",
                pageSize=10,
                supportsAllDrives=True,
                includeItemsFromAllDrives=True,
            ).execute()
            items = results.get("files", [])
            if items:
                value = (items[0]["id"], items[0].get("webViewLink"))
                self.folder_cache[cache_key] = value
                return value

            created = service.files().create(
                body={
                    "name": folder_name,
                    "mimeType": "application/vnd.google-apps.folder",
                    "parents": [parent_folder_id],
                },
                fields="id, webViewLink",
                supportsAllDrives=True,
            ).execute()
            value = (created["id"], created.get("webViewLink"))
            self.folder_cache[cache_key] = value
            return value

    def file_exists(self, service, file_name, parent_folder_id):
        safe_name = escape_drive_query_text(file_name)
        query = f"name = '{safe_name}' and '{parent_folder_id}' in parents and trashed = false"
        results = service.files().list(
            q=query,
            fields="files(id, name)",
            pageSize=1,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        return bool(results.get("files", []))

    def upload_recording(self, file_path, topic, dt):
        topic = canonical_topic(topic)
        drive_set = self.find_drive_set(topic)
        if not drive_set:
            print(f"⚠️ Drive 업로드 스킵: 사업명을 찾지 못했습니다 - {topic}")
            return None

        day_group = get_day_group(dt)
        if not day_group:
            print(f"⚠️ Drive 업로드 스킵: 월수/화목 수업일이 아닙니다 - {topic}")
            return None

        class_name = extract_class_name(topic)
        if not class_name:
            print(f"⚠️ Drive 업로드 스킵: 반 정보를 찾지 못했습니다 - {topic}")
            return None

        service = self.get_service()
        root_folder_id = extract_drive_folder_id(drive_set.드라이브링크)
        day_folder_id, _ = self.get_or_create_folder(service, day_group, root_folder_id)
        date_folder_id, _ = self.get_or_create_folder(service, format_drive_date(dt), day_folder_id)
        class_folder_name = sanitize_filename(topic)
        class_folder_id, class_folder_link = self.get_or_create_folder(
            service,
            class_folder_name,
            date_folder_id,
        )

        file_name = os.path.basename(file_path)
        if self.file_exists(service, file_name, class_folder_id):
            print(f"☁️ 이미 업로드된 파일 스킵: {drive_set.사업명}/{day_group}/{class_folder_name}/{file_name}")
            return class_folder_link

        media = MediaFileUpload(
            file_path,
            mimetype="video/mp4",
            chunksize=128 * 1024 * 1024,
            resumable=True,
        )
        request = service.files().create(
            body={"name": file_name, "parents": [class_folder_id]},
            media_body=media,
            fields="id, webViewLink",
            supportsAllDrives=True,
        )

        response = None
        while response is None:
            status, response = request.next_chunk(num_retries=3)
            if status:
                progress = int(status.progress() * 100)
                print(f"☁️ 업로드 중 {progress}%: {file_name}")

        print(f"☁️ 업로드 완료: {drive_set.사업명}/{day_group}/{class_folder_name}/{file_name}")
        return class_folder_link


def upload_recording_with_retry(drive_uploader, file_path, topic, dt, attempts=3):
    for attempt in range(1, attempts + 1):
        try:
            return drive_uploader.upload_recording(file_path, topic, dt)
        except Exception as error:
            if attempt >= attempts:
                raise
            wait_seconds = min(10, attempt * 3)
            print(f"⚠️ Drive 업로드 재시도 {attempt}/{attempts - 1}: {topic} - {error}")
            time.sleep(wait_seconds)


def get_access_token(session, settings):
    response = session.post(
        "https://zoom.us/oauth/token",
        params={
            "grant_type": "account_credentials",
            "account_id": settings.account_id,
        },
        auth=HTTPBasicAuth(settings.client_id, settings.client_secret),
        timeout=30,
    )
    try:
        response.raise_for_status()
    except requests.HTTPError as error:
        detail = response.text[:500]
        raise RuntimeError(
            "Zoom Access Token 발급 실패. Account ID, Client ID, Client Secret, "
            f"앱 활성화 상태와 Recording/User scope를 확인하세요. 응답: {detail}"
        ) from error
    return response.json()["access_token"]


def zoom_get_with_retry(session, url, description, attempts=5, **kwargs):
    last_error = None
    for attempt in range(1, attempts + 1):
        try:
            response = session.get(url, **kwargs)
            if response.status_code not in {429} and response.status_code < 500:
                response.raise_for_status()
                return response
            last_error = requests.HTTPError(f"{response.status_code} {response.text[:300]}")
        except requests.RequestException as error:
            last_error = error

        if attempt >= attempts:
            raise last_error

        wait_seconds = min(30, attempt * 5)
        print(f"Zoom API 재시도 {attempt}/{attempts - 1}: {description} - {last_error}")
        time.sleep(wait_seconds)


def get_all_users(session, token):
    headers = {"Authorization": f"Bearer {token}"}
    params = {"page_size": 100}
    users = []

    while True:
        response = zoom_get_with_retry(
            session,
            "https://api.zoom.us/v2/users",
            "사용자 목록 조회",
            headers=headers,
            params=params,
            timeout=30,
        )
        data = response.json()
        users.extend(data.get("users", []))
        next_page_token = data.get("next_page_token")
        if not next_page_token:
            return users
        params["next_page_token"] = next_page_token


def list_recordings(session, token, user_id, settings):
    headers = {"Authorization": f"Bearer {token}"}
    params = {
        "from": settings.start_date,
        "to": settings.end_date,
        "page_size": 50,
    }
    meetings = []

    while True:
        response = zoom_get_with_retry(
            session,
            f"https://api.zoom.us/v2/users/{user_id}/recordings",
            f"녹화 목록 조회 {user_id}",
            headers=headers,
            params=params,
            timeout=30,
        )
        data = response.json()
        meetings.extend(data.get("meetings", []))
        next_page_token = data.get("next_page_token")
        if not next_page_token:
            return meetings
        params["next_page_token"] = next_page_token


def zoom_recording_lookup_id(meeting):
    value = text_or_empty(meeting.get("uuid")) or str(meeting["id"])
    encoded = quote(value, safe="")
    if value.startswith("/") or value.startswith("+"):
        return quote(encoded, safe="")
    return encoded


def mark_meeting_processed(meeting_key):
    with processed_meetings_lock:
        if meeting_key in processed_meetings:
            return False
        processed_meetings.add(meeting_key)
        return True


def sanitize_filename(name):
    sanitized = re.sub(r'[<>:"/\\|?*]+', "_", text_or_empty(name))
    sanitized = sanitized.rstrip(". ")
    return sanitized or "NoTopic"


def canonical_topic(topic):
    topic_text = text_or_empty(topic)
    return topic_aliases.get(topic_text, topic_text)


def get_region(topic):
    topic = canonical_topic(topic)
    if "경산" in topic:
        return "경산"
    if "포항" in topic:
        return "포항"
    if "구미" in topic:
        return "구미"
    if "온라인SW" in topic:
        return "대구"
    return "기타"


def select_recording_file(files):
    selected_files = select_recording_files(files)
    if selected_files:
        return selected_files[0]
    return None


def recording_duration_seconds(file_info):
    duration = file_info.get("duration")
    if duration not in {None, ""}:
        try:
            return int(float(duration))
        except (TypeError, ValueError):
            pass

    start_text = text_or_empty(file_info.get("recording_start"))
    end_text = text_or_empty(file_info.get("recording_end"))
    if start_text and end_text:
        try:
            start_dt = datetime.fromisoformat(start_text.rstrip("Z"))
            end_dt = datetime.fromisoformat(end_text.rstrip("Z"))
            return int((end_dt - start_dt).total_seconds())
        except ValueError:
            return None

    return None


def select_recording_files(files):
    mp4_files = []
    for file_info in files:
        if file_info.get("file_type") != "MP4" or not file_info.get("download_url"):
            continue
        duration_seconds = recording_duration_seconds(file_info)
        if duration_seconds is not None and duration_seconds < 60:
            print(
                f"⏭️ 1분 미만 MP4 스킵: {file_info.get('recording_type')} "
                f"({duration_seconds}초, {file_info.get('file_size')} bytes)"
            )
            continue
        mp4_files.append(file_info)

    for recording_type in priority_types:
        matched_files = [
            file_info
            for file_info in mp4_files
            if re.sub(r"\s*\([^)]*\)\s*$", "", text_or_empty(file_info.get("recording_type"))) == recording_type
        ]
        if matched_files:
            return sorted(matched_files, key=lambda file_info: int(file_info.get("file_size") or 0), reverse=True)

    return sorted(mp4_files, key=lambda file_info: int(file_info.get("file_size") or 0), reverse=True)


def build_recording_file_name(safe_topic, date_str, index, total_count):
    if total_count <= 1:
        return f"{safe_topic} _{date_str}.mp4"
    return f"{safe_topic} _{date_str}_{index}.mp4"


def download_one_recording_file(session, token, selected, file_path, topic, dt, recording_lookup_id, drive_uploader):
    download_url = selected["download_url"] + f"?access_token={token}"
    drive_ready = drive_uploader is None
    selected_size = int(selected.get("file_size") or 0)

    if os.path.exists(file_path):
        local_size = os.path.getsize(file_path)
        if selected_size and local_size < selected_size * 0.9:
            print(
                f"♻️ 기존 파일이 너무 작아 재다운로드: {os.path.basename(file_path)} "
                f"({local_size} / {selected_size} bytes)"
            )
        else:
            print(f"⏭️ 이미 다운로드된 파일 스킵: {os.path.basename(file_path)}")
            if drive_uploader:
                try:
                    drive_ready = bool(upload_recording_with_retry(drive_uploader, file_path, topic, dt))
                except Exception as error:
                    print(f"⚠️ Drive 업로드 실패: {topic} ({recording_lookup_id}) - {error}")
            return drive_ready

    print(f"🔽 다운로드: {os.path.basename(file_path)}")
    temp_path = f"{file_path}.download"
    with session.get(download_url, stream=True, timeout=120) as download_response:
        download_response.raise_for_status()
        with open(temp_path, "wb") as output_file:
            shutil.copyfileobj(download_response.raw, output_file)

    if selected_size and os.path.getsize(temp_path) < selected_size * 0.9:
        raise RuntimeError(
            f"다운로드 파일 크기가 너무 작습니다: {os.path.basename(file_path)} "
            f"({os.path.getsize(temp_path)} / {selected_size} bytes)"
        )

    if os.path.exists(file_path):
        os.remove(file_path)
    os.replace(temp_path, file_path)

    if drive_uploader:
        try:
            drive_ready = bool(upload_recording_with_retry(drive_uploader, file_path, topic, dt))
        except Exception as error:
            print(f"⚠️ Drive 업로드 실패: {topic} ({recording_lookup_id}) - {error}")

    return drive_ready


def trash_zoom_recording(session, token, recording_lookup_id, topic):
    delete_response = session.delete(
        f"https://api.zoom.us/v2/meetings/{recording_lookup_id}/recordings",
        headers={"Authorization": f"Bearer {token}"},
        params={"action": "trash"},
        timeout=30,
    )
    if delete_response.status_code == 204:
        print(f"🧹 Zoom 원본 휴지통 이동 완료: {topic}")
        return True
    print(f"❌ Zoom 원본 휴지통 이동 실패 ({delete_response.status_code}): {topic} - {delete_response.text}")
    return False


def download_and_delete(meeting, token, user_email, settings, drive_uploader=None):
    recording_lookup_id = zoom_recording_lookup_id(meeting)
    if not mark_meeting_processed(recording_lookup_id):
        print(f"⏭️ 이미 처리된 회의 스킵: {recording_lookup_id}")
        return

    topic = canonical_topic(meeting.get("topic")) or "NoTopic"
    safe_topic = sanitize_filename(topic)
    start_time = text_or_empty(meeting.get("start_time"))
    dt = datetime.fromisoformat(start_time.rstrip("Z"))
    date_str = format_drive_date(dt)
    region = get_region(topic)

    folder_path = os.path.join(BASE_DIR, region, date_str, f"{safe_topic} _{date_str}")

    with requests.Session() as session:
        headers = {"Authorization": f"Bearer {token}"}
        response = zoom_get_with_retry(
            session,
            f"https://api.zoom.us/v2/meetings/{recording_lookup_id}/recordings",
            f"녹화 상세 조회 {topic}",
            headers=headers,
            timeout=30,
        )
        files = response.json().get("recording_files", [])

        selected_files = select_recording_files(files)
        if not selected_files:
            print(f"⚠️ MP4 파일 없음: {topic} ({recording_lookup_id})")
            return

        os.makedirs(folder_path, exist_ok=True)
        if len(selected_files) > 1:
            print(f"🎞️ MP4 여러 개 감지: {topic} - {len(selected_files)}개 저장")

        drive_ready = True
        total_count = len(selected_files)
        for index, selected in enumerate(selected_files, start=1):
            file_name = build_recording_file_name(safe_topic, date_str, index, total_count)
            file_path = os.path.join(folder_path, file_name)
            file_ready = download_one_recording_file(
                session,
                token,
                selected,
                file_path,
                topic,
                dt,
                recording_lookup_id,
                drive_uploader,
            )
            drive_ready = drive_ready and file_ready

        if not settings.delete_after_download:
            print(f"🗂️ 다운로드 완료, 삭제는 건너뜀: {topic} ({recording_lookup_id})")
            return

        if not drive_ready:
            print(f"🛑 Zoom 원본 삭제 스킵: Drive 업로드 확인 실패 - {topic}")
            return

        trash_zoom_recording(session, token, recording_lookup_id, topic)


def get_sms_config_from_excel():
    workbook = openpyxl.load_workbook(excel_file_path, data_only=True)
    worksheet = workbook.active
    문자박스링크 = text_or_empty(worksheet.cell(row=8, column=15).value)
    문자박스아이디 = text_or_empty(worksheet.cell(row=9, column=15).value)
    문자박스비번 = text_or_empty(worksheet.cell(row=10, column=15).value)
    회신번호 = "010" + text_or_empty(worksheet.cell(row=6, column=17).value)
    알람받을번호 = text_or_empty(worksheet.cell(row=5, column=15).value)
    workbook.close()
    return 문자박스링크, 문자박스아이디, 문자박스비번, 회신번호, 알람받을번호


def send_sms_via_playwright(link, user_id, user_pwd, recv_number, callback_number):
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False, args=["--disable-popup-blocking"])
        page = browser.new_page()
        page.goto(link)
        page.fill("input[name='id']", user_id)
        page.fill("input[name='pwd']", user_pwd)
        page.press("input[name='pwd']", "Enter")
        time.sleep(1)
        try:
            page.click("button[onclick*='contentsLayerClose']")
        except Exception:
            pass

        page.fill("textarea#recvList", recv_number)
        page.fill("textarea#msg", "영상 발송 완료")
        page.locator("a.hand.openLayer", has_text="선택").click()
        frame = page.frame(name="callbackFrame")
        frame.wait_for_selector("span:has-text('전체 회신번호')", timeout=5000)
        frame.locator("span:has-text('전체 회신번호')").click()
        frame.wait_for_selector("#bulkCallbackNum", timeout=3000)
        frame.fill("#bulkCallbackNum", callback_number)
        frame.click("button[onclick*='callbackCheckForm']")
        time.sleep(0.3)
        page.locator("a.hand.openLayer", has_text="전송하기").click()
        time.sleep(0.3)
        page.keyboard.press("Enter")
        time.sleep(0.3)
        page.keyboard.press("Enter")
        print("📨 영상 발송 완료 메시지 전송 완료")
        browser.close()


def build_drive_uploader(settings):
    if not settings.upload_to_drive:
        print("Drive 업로드: 설정상 비활성화")
        return None

    credentials = load_google_credentials()
    if credentials is None:
        print("Drive 업로드: Google 인증이 없어 비활성화")
        return None

    drive_sets = load_drive_sets()
    if not drive_sets:
        print("Drive 업로드: 영상보내기.xlsx에서 Drive 세트 설정을 찾지 못했습니다.")
        return None

    print("Drive 업로드 세트: " + ", ".join(f"{item.이름}={item.사업명}" for item in drive_sets))
    return Drive업로더(credentials, drive_sets)


def upload_local_recordings(drive_uploader, upload_workers=4):
    if drive_uploader is None:
        print("로컬 영상 업로드 스킵: Drive 업로더가 준비되지 않았습니다.")
        return

    if not os.path.exists(BASE_DIR):
        print(f"로컬 영상 폴더가 없습니다: {BASE_DIR}")
        return

    upload_tasks = []
    for root, _, files in os.walk(BASE_DIR):
        for file_name in files:
            if not file_name.lower().endswith(".mp4"):
                continue
            file_path = os.path.join(root, file_name)
            try:
                topic = extract_topic_from_local_file(file_path)
                dt = extract_date_from_local_file(file_path)
            except Exception as error:
                print(f"⚠️ 로컬 영상 업로드 실패: {file_path} - {error}")
                continue
            upload_tasks.append((file_path, topic, dt))

    uploaded = 0
    skipped = 0
    failed = 0
    if not upload_tasks:
        print("로컬 영상 업로드 대상 MP4가 없습니다.")
        return

    print(f"로컬 영상 업로드 시작: 대상 {len(upload_tasks)}개, 동시 업로드 {upload_workers}개")
    with ThreadPoolExecutor(max_workers=upload_workers) as executor:
        futures = {
            executor.submit(upload_recording_with_retry, drive_uploader, file_path, topic, dt): file_path
            for file_path, topic, dt in upload_tasks
        }
        for future in as_completed(futures):
            file_path = futures[future]
            try:
                if future.result():
                    uploaded += 1
                else:
                    skipped += 1
            except Exception as error:
                failed += 1
                print(f"⚠️ 로컬 영상 업로드 실패: {file_path} - {error}")

    print(f"로컬 영상 업로드 완료: 성공 {uploaded}, 스킵 {skipped}, 실패 {failed}")


def run_video_sender():
    if not os.path.exists(sender_script_path):
        raise FileNotFoundError(f"영상 발송 스크립트를 찾지 못했습니다: {sender_script_path}")
    print("📨 업로드 완료 후 영상 링크 발송 스크립트 실행")
    subprocess.run([sys.executable, sender_script_path], cwd=script_dir, check=True)


def main():
    try:
        settings = load_zoom_config()
    except Exception as error:
        print(f"설정 로드 실패: {error}")
        return

    try:
        drive_uploader = build_drive_uploader(settings)
    except Exception as error:
        drive_uploader = None
        print(f"⚠️ Drive 업로드 준비 실패: {error}")

    if parse_bool(os.getenv("ZOOM_UPLOAD_LOCAL_ONLY"), False):
        upload_local_recordings(drive_uploader, settings.upload_workers)
        if settings.send_video_links_after_upload:
            if drive_uploader is None:
                print("영상 링크 발송 스킵: Drive 업로드가 준비되지 않았습니다.")
            else:
                run_video_sender()
        return

    with requests.Session() as session:
        print("🔐 Access Token 발급 중...")
        token = get_access_token(session, settings)

        print("👥 사용자 목록 조회 중...")
        users = get_all_users(session, token)
        print(f"총 사용자 수: {len(users)}")

        download_tasks = []
        with ThreadPoolExecutor(max_workers=settings.max_workers) as executor:
            for user in users:
                user_id = user["id"]
                user_email = user["email"]
                print(f"\n📁 [{user_email}] 회의 목록 조회 중...")

                try:
                    meetings = list_recordings(session, token, user_id, settings)
                    print(f"📋 회의 수: {len(meetings)}")
                    for meeting in meetings:
                        download_tasks.append(
                            executor.submit(download_and_delete, meeting, token, user_email, settings, drive_uploader)
                        )
                except Exception as error:
                    print(f"⚠️ 사용자 {user_email} 오류: {error}")

            for future in as_completed(download_tasks):
                try:
                    future.result()
                except Exception as error:
                    print(f"❌ 병렬 다운로드 오류: {error}")

    print("\n🎉 모든 병렬 다운로드 완료")
    if settings.send_video_links_after_upload:
        if drive_uploader is None:
            print("영상 링크 발송 스킵: Drive 업로드가 준비되지 않았습니다.")
        else:
            run_video_sender()

    if settings.send_completion_sms:
        문자박스링크, 문자박스아이디, 문자박스비번, 회신번호, 알람받을번호 = get_sms_config_from_excel()
        send_sms_via_playwright(
            문자박스링크,
            문자박스아이디,
            문자박스비번,
            알람받을번호,
            회신번호,
        )


if __name__ == "__main__":
    main()
