# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import os
import shutil
from tkinter import filedialog, messagebox


def process_names_in_excel(filepath, 저장_경로):
    """
    엑셀 파일의 B열 이름을 규칙에 따라 변경한 뒤 저장합니다.
    - A열: 순번(1,2,3,...) 자동 채우기 (비어 있을 때만)
    - C열: 원래 이름 백업 (비어 있을 때만)
    - B열: 마스킹
        * 이름이 2글자: 두 번째 글자를 'O'로 변경 (ex. 김수 -> 김O)
        * 이름이 3글자 이상: 마지막에서 두 번째 글자를 'O'로 변경 (ex. 제갈재균 -> 제갈O균)
    """
    wb = load_workbook(filepath)
    ws = wb.active

    # A열 자동 번호를 위해 시작 번호
    seq = 1

    # B열(B column)만 처리 (2열)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell, MergedCell):
            continue
        name = cell.value

        if name and isinstance(name, str) and len(name) > 1:
            # --- 1) A열: 순번 자동 채우기 (비어 있을 때만) ---
            a_cell = ws.cell(row=cell.row, column=1)  # A열
            if not isinstance(a_cell, MergedCell) and a_cell.value in (None, ""):
                a_cell.value = seq
                seq += 1

            # --- 2) C열: 원래 이름 백업 (비어 있을 때만) ---
            c_cell = ws.cell(row=cell.row, column=3)  # C열
            if not isinstance(c_cell, MergedCell) and c_cell.value in (None, ""):
                c_cell.value = name

            # --- 3) B열: 이름 마스킹 ---
            length = len(name)
            if length == 2:
                # 2글자 이름: 두 번째 글자를 'O'
                idx = 1
            else:
                # 3글자 이상: 마지막에서 두 번째 글자를 'O'
                idx = length - 2

            cell.value = name[:idx] + 'O' + name[idx+1:]

    wb.save(저장_경로)
    os.startfile(저장_경로)


# ===== 템플릿 복사용 설정 =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "엑셀양식")

# 실제 템플릿 파일명에 맞게 수정해서 사용하세요.
TEMPLATE_FILENAME_NOC = "이름 O 처리 양식.xlsx"  # 예: 엑셀양식\이름O처리_양식.xlsx


def download_template_noc():
    """
    엑셀양식 폴더에 있는 '이름O처리_양식.xlsx' 파일을
    사용자가 지정한 위치에 복사(저장)해 주는 함수.
    원본 템플릿은 항상 깨끗하게 유지됩니다.
    """
    src = os.path.join(TEMPLATE_DIR, TEMPLATE_FILENAME_NOC)

    if not os.path.exists(src):
        messagebox.showerror(
            "오류",
            f"이름 O 처리 템플릿 파일을 찾을 수 없습니다.\n\n"
            f"경로: {src}\n\n"
            f"※ '엑셀양식' 폴더 안에 '{TEMPLATE_FILENAME_NOC}' 파일이 있는지 확인해 주세요."
        )
        return

    dest = filedialog.asksaveasfilename(
        title="이름 O 처리 양식 저장 위치 선택",
        defaultextension=".xlsx",
        initialfile=TEMPLATE_FILENAME_NOC,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not dest:
        return

    try:
        shutil.copy2(src, dest)
        messagebox.showinfo("완료", f"양식 파일이 저장되었습니다:\n{dest}")
        try:
            os.startfile(dest)
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"양식 복사 중 오류가 발생했습니다:\n{e}")
