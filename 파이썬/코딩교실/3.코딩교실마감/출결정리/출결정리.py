# -*- coding: utf-8 -*-
# pip install openpyxl
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import os
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter

BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "\ucd9c\uacb0\uc815\ub9ac.xlsx"

SHEET_MAIN = "\uba54\uc778"
SHEET_DAYS: list[tuple[str, str]] = [("\uc6d4\uc218", "MW"), ("\ud654\ubaa9", "TT")]
OUTPUT_SHEET_NAME = ""  # Leave empty to use the active sheet in the output template.
OUTPUT_FILE_NAME = "포항시 코딩교육 출석률_{target_ym}.xlsx"
OUTPUT_FILE_SUFFIX = ""  # "" = use OUTPUT_FILE_NAME or overwrite template.
ALWAYS_COPY_TEMPLATE = True  # Copy template every run when OUTPUT_FILE_SUFFIX is set.
OPEN_RESULT_FILE = True
CLEAR_EXTRA_ROWS = True  # Remove rows beyond the written range.

TARGET_YEAR_FALLBACK = 0  # Set to 2025 if main sheet is missing target year.
REGION_CODE_FALLBACK = "PH"  # Set region code if main sheet does not include it.

MAIN_LABEL_TARGET_YM = "\ub300\uc0c1\uc5f0\uc6d4"
MAIN_LABEL_GROUP_BASE = "\uadf8\ub8f9\ub118\ubc84"
MAIN_LABEL_REGION = "\uc9c0\uc5ed"

# Group No day code mapping (customizable)
GROUP_NO_DAY_CODE = {"MW": "24", "TT": "34"}
COURSE_NAME_BY_DAY = {
    "MW": "1\uac1c\uc6d4/\uc8fc2\ud68c/\uc6d4\uc218/90\ubd84",
    "TT": "1\uac1c\uc6d4/\uc8fc2\ud68c/\ud654\ubaa9/90\ubd84",
}
CLASS_TIME_TEXT = "19:00 ~ 20:30"

REMOVE_HEADERS = ["\uac15\uc88c\uad6c\ubd84", "\uc7ac\uc218\uac15", "\uacb0\uc81c\ud560\uc778"]
REMOVE_COLUMNS = True  # True = delete columns (shift left), False = hide columns.
HEADER_ALIASES = {
    "\ucd9c\uc11d\uc728": "\ucd9c\uc11d\ub960",
    "\uac15\uc758\ub808\ubca8": "\uac15\uc758\uae30\uc218",
}
USE_FIXED_COLUMN_WIDTHS = True
COLUMN_WIDTHS = {
    "A": 5,
    "B": 8.5,
    "C": 8,
    "D": 18,
    "E": 26,
    "F": 10,
    "G": 21,
    "H": 12.5,
    "I": 13,
    "J": 6.75,
    "K": 6.75,
    "L": 6.75,
    "M": 6.75,
    "N": 6.75,
}
MAX_OUTPUT_COLUMN = "N"

HEADER_ROW = 2
DATA_START_ROW = 3
COL_CLASS = 1
COL_NAME = 5
COL_ID = 6
COL_DATE_START = 13  # M
COL_DATE_END = 20    # T

# Status mapping
PRESENT_MARKS = {"O"}  # O = 출석
ABSENT_MARKS = {"X"}   # X = 결석
COUNT_MISSING_AS_ABSENT = True

OUTPUT_HEADERS = [
    "No.",
    "\ub300\uc0c1\ub144\uc6d4",
    "\ud68c\uc6d0\uba85",
    "\ud68c\uc6d0 ID",
    "Group No",
    "\uac15\uc758\uae30\uc218",
    "\uac15\uc88c\uba85",
    "\uc218\uc5c5\uc2dc\uac04",
    "\ub2f4\ub2f9\uac15\uc0ac",
    "\ucd1d\uc218\uc5c5",
    "\ucd9c\uc11d",
    "\uacb0\uc11d",
    "\ucd9c\uc11d\ub960",
]

# Attendance rate color rules
COLOR_BLUE = "0000FF"
COLOR_RED = "FF0000"
COLOR_BLACK = "000000"
RATE_BLUE_THRESHOLD = 100.0
RATE_RED_THRESHOLD = 70.0


@dataclass
class DateInfo:
    display: str      # 2025.05.12
    yyyymmdd: str     # 20250512
    label: str        # 5/12(월)


@dataclass
class OutputRow:
    name: str
    student_id: str
    target_ym: str
    group_no: str
    lecture_term: str
    course_name: str
    class_time: str
    teacher_id: str
    total_classes: int
    present_count: int
    absent_count: int
    attendance_rate: int
    sequence: int
    order_index: int | None


def _to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, int):
        return str(value).strip()
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def _parse_year(value: Any) -> int:
    text = _to_str(value)
    if len(text) >= 4 and text[:4].isdigit():
        return int(text[:4])
    raise ValueError("Q3 (YYYYMM) is missing or invalid.")


def _parse_class_no(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    return int(digits) if digits else None


def _parse_date_label(text: str, year: int) -> DateInfo | None:
    if not text:
        return None
    parts = text.replace(" ", "").split("/")
    if len(parts) < 2:
        return None
    try:
        month = int(parts[0])
        day = int(parts[1].split("(")[0])
    except ValueError:
        return None
    yyyymmdd = f"{year:04d}{month:02d}{day:02d}"
    display = f"{year:04d}.{month:02d}.{day:02d}"
    return DateInfo(display=display, yyyymmdd=yyyymmdd, label=text)


def _first_four_digits(text: str) -> int | None:
    digits = ""
    for ch in text:
        if ch.isdigit():
            digits += ch
            if len(digits) == 4:
                return int(digits)
        else:
            digits = ""
    return None


def _parse_month_from_label(text: str) -> int | None:
    if not text:
        return None
    parts = text.replace(" ", "").split("/")
    if len(parts) < 2:
        return None
    try:
        return int(parts[0])
    except ValueError:
        return None


def _infer_month_from_headers(wb) -> int | None:
    for sheet_name, _ in SHEET_DAYS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for col in range(COL_DATE_START, COL_DATE_END + 1):
            header_text = _to_str(ws.cell(row=HEADER_ROW, column=col).value)
            month = _parse_month_from_label(header_text)
            if month:
                return month
    return None


def _derive_lecture_term(group_base: str) -> str:
    text = _to_str(group_base)
    if not text:
        return ""
    first = text.split("_")[0]
    digits = "".join(ch for ch in first if ch.isdigit())
    if digits:
        return f"{int(digits)}\uae30"
    return text


def _normalize_id(text: str) -> str:
    return text.strip().lower()


def _load_student_order(wb) -> tuple[dict[str, int], dict[str, int]]:
    order_map: dict[str, int] = {}
    name_map: dict[str, int] = {}
    sheet_name = "\ud559\uc0dd\uc21c\uc11c"
    if sheet_name not in wb.sheetnames:
        return order_map, name_map

    ws = wb[sheet_name]
    header_map = _build_header_map(ws, 1)
    order_col = header_map.get("\uad6c\ubd84") or header_map.get("\uc21c\uc11c") or header_map.get("No.")
    name_col = header_map.get("\uc774\ub984") or header_map.get("\ud68c\uc6d0\uba85")
    id_col = header_map.get("ID") or header_map.get("\uc544\uc774\ub514") or header_map.get("\ud68c\uc6d0 ID")

    seq = 0
    for row in range(2, ws.max_row + 1):
        raw_id = _to_str(ws.cell(row=row, column=id_col).value) if id_col else ""
        raw_name = _to_str(ws.cell(row=row, column=name_col).value) if name_col else ""
        if not raw_id and not raw_name:
            continue

        order_idx = None
        if order_col:
            val = ws.cell(row=row, column=order_col).value
            try:
                order_idx = int(val)
            except (TypeError, ValueError):
                order_idx = None
        if order_idx is None:
            seq += 1
            order_idx = seq

        if raw_id:
            order_map[_normalize_id(raw_id)] = order_idx
        if raw_name:
            name_map[raw_name] = order_idx

    return order_map, name_map


def _find_main_value(ws, label: str) -> str:
    for row in range(1, ws.max_row + 1):
        key = _to_str(ws.cell(row=row, column=1).value)
        if key == label:
            return _to_str(ws.cell(row=row, column=2).value)
    return ""


def _resolve_main_info(wb) -> tuple[str, str, int, str]:
    target_ym = ""
    region = ""
    group_base = ""

    if SHEET_MAIN in wb.sheetnames:
        ws_main = wb[SHEET_MAIN]
        target_ym = _find_main_value(ws_main, MAIN_LABEL_TARGET_YM)
        group_base = _find_main_value(ws_main, MAIN_LABEL_GROUP_BASE)
        region = _find_main_value(ws_main, MAIN_LABEL_REGION)

        if not target_ym:
            target_ym = _to_str(ws_main["Q3"].value)
        if not region:
            region = _to_str(ws_main["Q4"].value)
        if not group_base:
            group_base = _to_str(ws_main["Q2"].value)

    if not target_ym:
        year = TARGET_YEAR_FALLBACK
        if not year:
            raise ValueError("Target year is missing. Fill MAIN sheet or set TARGET_YEAR_FALLBACK.")
        month = _infer_month_from_headers(wb)
        if not month:
            raise ValueError("Unable to infer month from date headers (M2~T2).")
        target_ym = f"{year:04d}{month:02d}"

    if not region:
        region = REGION_CODE_FALLBACK.strip()
    if not region:
        raise ValueError("REGION_CODE_FALLBACK is empty.")

    if not group_base:
        raise ValueError("MAIN sheet is missing \uadf8\ub8f9\ub118\ubc84.")

    year = _parse_year(target_ym)
    return target_ym, region, year, group_base


def _build_output_rows(
    wb,
    target_ym: str,
    region: str,
    year: int,
    group_base: str,
    order_map: dict[str, int],
    name_map: dict[str, int],
) -> tuple[list[OutputRow], list[str]]:

    missing_cells: list[str] = []
    output_rows: list[OutputRow] = []
    sequence = 0

    for sheet_name, day_code in SHEET_DAYS:
        ws = wb[sheet_name]
        date_headers: dict[int, DateInfo] = {}
        for col in range(COL_DATE_START, COL_DATE_END + 1):
            header_text = _to_str(ws.cell(row=HEADER_ROW, column=col).value)
            parsed = _parse_date_label(header_text, year)
            if parsed:
                date_headers[col] = parsed

        total_classes = len(date_headers)
        if total_classes == 0:
            total_classes = COL_DATE_END - COL_DATE_START + 1
        current_class_no = None

        for row in range(DATA_START_ROW, ws.max_row + 1):
            class_val = ws.cell(row=row, column=COL_CLASS).value
            if class_val not in (None, ""):
                current_class_no = _parse_class_no(class_val)

            if current_class_no is None:
                continue

            name = _to_str(ws.cell(row=row, column=COL_NAME).value)
            student_id = _to_str(ws.cell(row=row, column=COL_ID).value)
            if not name and not student_id:
                continue

            present_count = 0
            absent_count = 0
            missing_count = 0
            for col in range(COL_DATE_START, COL_DATE_END + 1):
                status = _to_str(ws.cell(row=row, column=col).value)
                if not status:
                    col_letter = get_column_letter(col)
                    header = _to_str(ws.cell(row=HEADER_ROW, column=col).value)
                    if student_id:
                        missing_cells.append(
                            f"{sheet_name} {student_id} R{row} {col_letter}({header}) is empty"
                        )
                    missing_count += 1
                    continue
                if status in PRESENT_MARKS:
                    present_count += 1
                elif status in ABSENT_MARKS:
                    absent_count += 1

            if COUNT_MISSING_AS_ABSENT:
                absent_count += missing_count

            attendance_rate = round(present_count / total_classes * 100) if total_classes else 0

            group_day = GROUP_NO_DAY_CODE.get(day_code, day_code)
            group_no = f"{target_ym}_{group_base}_{group_day}_{current_class_no:02d}"
            teacher_id = f"{region}_{day_code}_{current_class_no:02d}"
            lecture_term = _derive_lecture_term(group_base)
            course_name = COURSE_NAME_BY_DAY.get(day_code, "")
            class_time = CLASS_TIME_TEXT
            sequence += 1
            order_index = None
            if student_id:
                order_index = order_map.get(_normalize_id(student_id))
            if order_index is None and name:
                order_index = name_map.get(name)

            output_rows.append(
                OutputRow(
                    name=name,
                    student_id=student_id,
                    target_ym=target_ym,
                    group_no=group_no,
                    lecture_term=lecture_term,
                    course_name=course_name,
                    class_time=class_time,
                    teacher_id=teacher_id,
                    total_classes=total_classes,
                    present_count=present_count,
                    absent_count=absent_count,
                    attendance_rate=int(attendance_rate),
                    sequence=sequence,
                    order_index=order_index,
                )
            )

    if order_map or name_map:
        output_rows.sort(
            key=lambda r: (r.order_index if r.order_index is not None else 10**9, r.sequence)
        )
    else:
        output_rows.sort(key=lambda r: r.sequence)
    return output_rows, missing_cells


def _find_header_row(ws) -> tuple[int, dict[str, int]] | None:
    for row in range(1, 8):
        header_map: dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            val = _to_str(ws.cell(row=row, column=col).value)
            if val:
                header_map[val] = col
        if "\ud68c\uc6d0 ID" in header_map and "\ud68c\uc6d0\uba85" in header_map:
            return row, header_map
    return None


def _build_header_map(ws, header_row: int) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = _to_str(ws.cell(row=header_row, column=col).value)
        if val:
            header_map[val] = col
    return header_map


def _capture_header_widths(ws, header_map: dict[str, int]) -> dict[str, float | None]:
    widths: dict[str, float | None] = {}
    for header, col in header_map.items():
        col_letter = get_column_letter(col)
        widths[header] = ws.column_dimensions[col_letter].width
    return widths


def _apply_header_widths(ws, header_map: dict[str, int], widths: dict[str, float | None]) -> None:
    for header, col in header_map.items():
        width = widths.get(header)
        if width is not None:
            ws.column_dimensions[get_column_letter(col)].width = width


def _apply_fixed_column_widths(ws) -> None:
    for col_letter, width in COLUMN_WIDTHS.items():
        dim = ws.column_dimensions[col_letter]
        dim.width = width
        dim.bestFit = False


def _normalize_template(ws) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == 1 or merged.max_row == 1:
            ws.unmerge_cells(str(merged))


def _merge_title_row(ws) -> None:
    max_col = column_index_from_string(MAX_OUTPUT_COLUMN)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)


def _trim_to_max_column(ws) -> None:
    max_col = column_index_from_string(MAX_OUTPUT_COLUMN)
    if ws.max_column > max_col:
        ws.delete_cols(max_col + 1, ws.max_column - max_col)


def _get_output_sheet(wb):
    if OUTPUT_SHEET_NAME and OUTPUT_SHEET_NAME in wb.sheetnames:
        ws = wb[OUTPUT_SHEET_NAME]
    else:
        ws = wb.active
    _normalize_template(ws)

    header_info = _find_header_row(ws)
    if not header_info:
        raise ValueError("Output template header row not found. Check header text.")

    header_row, header_map = header_info
    header_widths = _capture_header_widths(ws, header_map)
    remove_cols = [col for header, col in header_map.items() if header in REMOVE_HEADERS]
    if remove_cols:
        if REMOVE_COLUMNS:
            for col in sorted(remove_cols, reverse=True):
                ws.delete_cols(col)
        else:
            for col in remove_cols:
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].hidden = True
                ws.column_dimensions[col_letter].width = 0

        header_map = _build_header_map(ws, header_row)

    for alias, target in HEADER_ALIASES.items():
        if alias in header_map and target not in header_map:
            col = header_map.pop(alias)
            ws.cell(row=header_row, column=col).value = target
            header_map[target] = col
        if alias in header_widths and target not in header_widths:
            header_widths[target] = header_widths[alias]

    _trim_to_max_column(ws)
    _merge_title_row(ws)
    _apply_header_widths(ws, header_map, header_widths)
    if USE_FIXED_COLUMN_WIDTHS:
        _apply_fixed_column_widths(ws)
    missing = [h for h in OUTPUT_HEADERS if h not in header_map]
    if missing:
        raise ValueError(f"Output template is missing headers: {missing}")

    return ws, header_map, header_row


def _rate_font(rate: float, present_count: int, total_classes: int) -> Font:
    if total_classes > 0 and present_count == total_classes:
        return Font(color=COLOR_BLUE)
    if rate <= RATE_RED_THRESHOLD:
        return Font(color=COLOR_RED)
    return Font(color=COLOR_BLACK)


def _resolve_template_file(base_dir: Path, target_ym: str) -> Path:
    direct = base_dir / f"\ud3ec\ud56d\uc2dc_{target_ym}_\ucd9c\uc11d\ub960.xlsx"
    if direct.exists():
        return direct

    candidates = sorted(base_dir.glob(f"\ud3ec\ud56d\uc2dc_{target_ym}_\ucd9c\uc11d\ub960*.xlsx"))
    if candidates:
        return candidates[0]

    fallback = base_dir / "2025 \ud3ec\ud56d\uc2dc \ucf54\ub529\uad50\uc721_\ucd9c\uc11d\ub960.xlsx"
    if fallback.exists():
        return fallback

    return direct


def _resolve_output_file(template_file: Path, target_ym: str) -> Path:
    if OUTPUT_FILE_NAME:
        return template_file.with_name(OUTPUT_FILE_NAME.format(target_ym=target_ym))
    if OUTPUT_FILE_SUFFIX:
        return template_file.with_name(f"{template_file.stem}{OUTPUT_FILE_SUFFIX}{template_file.suffix}")
    return template_file


def _write_output(
    ws,
    header_map: dict[str, int],
    header_row: int,
    rows: list[OutputRow],
) -> None:
    start_row = header_row + 1
    base_height = ws.row_dimensions[start_row].height
    for idx, row in enumerate(rows, start=1):
        excel_row = start_row + idx - 1
        if base_height is not None:
            ws.row_dimensions[excel_row].height = base_height
        ws.cell(row=excel_row, column=header_map["No."]).value = idx
        ws.cell(row=excel_row, column=header_map["\ub300\uc0c1\ub144\uc6d4"]).value = row.target_ym
        ws.cell(row=excel_row, column=header_map["\ud68c\uc6d0\uba85"]).value = row.name
        ws.cell(row=excel_row, column=header_map["\ud68c\uc6d0 ID"]).value = row.student_id
        ws.cell(row=excel_row, column=header_map["Group No"]).value = row.group_no
        ws.cell(row=excel_row, column=header_map["\uac15\uc758\uae30\uc218"]).value = row.lecture_term
        if "\uac15\uc88c\uba85" in header_map:
            ws.cell(row=excel_row, column=header_map["\uac15\uc88c\uba85"]).value = row.course_name
        if "\uc218\uc5c5\uc2dc\uac04" in header_map:
            ws.cell(row=excel_row, column=header_map["\uc218\uc5c5\uc2dc\uac04"]).value = row.class_time
        ws.cell(row=excel_row, column=header_map["\ub2f4\ub2f9\uac15\uc0ac"]).value = row.teacher_id
        ws.cell(row=excel_row, column=header_map["\ucd1d\uc218\uc5c5"]).value = row.total_classes
        ws.cell(row=excel_row, column=header_map["\ucd9c\uc11d"]).value = row.present_count
        ws.cell(row=excel_row, column=header_map["\uacb0\uc11d"]).value = row.absent_count
        if "\ub300\uae30" in header_map:
            ws.cell(row=excel_row, column=header_map["\ub300\uae30"]).value = 0

        rate_cell = ws.cell(row=excel_row, column=header_map["\ucd9c\uc11d\ub960"])
        rate_cell.value = row.attendance_rate / 100
        rate_cell.number_format = "0%"
        rate_cell.font = _rate_font(row.attendance_rate, row.present_count, row.total_classes)

    if CLEAR_EXTRA_ROWS:
        last_row = start_row + len(rows) - 1
        if ws.max_row > last_row:
            ws.delete_rows(last_row + 1, ws.max_row - last_row)


def main() -> None:
    input_wb = load_workbook(INPUT_FILE)
    target_ym, region, year, group_base = _resolve_main_info(input_wb)
    order_map, name_map = _load_student_order(input_wb)
    rows, missing = _build_output_rows(
        input_wb, target_ym, region, year, group_base, order_map, name_map
    )

    template_file = _resolve_template_file(BASE_DIR, target_ym)
    output_file = _resolve_output_file(template_file, target_ym)
    if output_file != template_file:
        if ALWAYS_COPY_TEMPLATE or not output_file.exists():
            shutil.copy2(template_file, output_file)
    output_wb = load_workbook(output_file)
    ws, header_map, header_row = _get_output_sheet(output_wb)
    _write_output(ws, header_map, header_row, rows)
    if USE_FIXED_COLUMN_WIDTHS:
        _apply_fixed_column_widths(ws)
    output_wb.save(output_file)
    if OPEN_RESULT_FILE:
        os.startfile(output_file)

    if missing:
        print("[WARN] Empty attendance cells found:")
        for msg in missing:
            print(" -", msg)

    print(f"[OK] Wrote {len(rows)} rows to '{output_file.name}'.")


if __name__ == "__main__":
    main()
