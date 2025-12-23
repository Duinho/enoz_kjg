# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import pandas as pd
import time
import random
import os

# 점수 순서: 5점, 4점, 3점, 2점, 1점
SCORES = [5, 4, 3, 2, 1]
DELAY_BETWEEN_SUBMITS = 0.05  # 제출 간 최소 지연


# =========================
#  양식 생성
# =========================

def download_template_iloom():
    """
    이룸캠프 만족도 설문 자동화를 위한 양식을 저장합니다.
    A1: 구분, B1~F1: 1번~5번,
    A2: 유형1, A7: 유형2 (사용자가 나중에 '학생', '선생님' 등으로 직접 수정)
    K1: 링크, L1: (사용자가 링크 입력)
    """
    from tkinter import filedialog, messagebox

    filepath = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="이룸캠프_만족도_양식.xlsx"
    )

    if not filepath:
        return

    # 유형1(행 2~6), 유형2(행 7~11) 구조
    data = [
        ["유형1", 0, 0, 0, 0, 0],  # A2, B2~F2 : 5점 개수 (나중에 '학생' 등으로 수정)
        ["",      0, 0, 0, 0, 0],  # A3        : 4점 개수
        ["",      0, 0, 0, 0, 0],  # A4        : 3점 개수
        ["",      0, 0, 0, 0, 0],  # A5        : 2점 개수
        ["",      0, 0, 0, 0, 0],  # A6        : 1점 개수
        ["유형2", 0, 0, 0, 0, 0],  # A7, B7~F7 : 5점 개수 (나중에 '선생님' 등으로 수정)
        ["",      0, 0, 0, 0, 0],  # A8        : 4점 개수
        ["",      0, 0, 0, 0, 0],  # A9        : 3점 개수
        ["",      0, 0, 0, 0, 0],  # A10       : 2점 개수
        ["",      0, 0, 0, 0, 0],  # A11       : 1점 개수
    ]
    df = pd.DataFrame(data, columns=["구분", "1번", "2번", "3번", "4번", "5번"])

    with pd.ExcelWriter(filepath, engine="xlsxwriter") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1", startrow=0, startcol=0)
        workbook = writer.book
        ws = writer.sheets["Sheet1"]

        # K1에 '링크', L1에 링크 입력 칸
        ws.write("K1", "링크")
        ws.write("L1", "")

        # 안내 문구
        ws.write("K2", "A2/A7 셀에 구글폼 참여 유형 텍스트를 적으세요")
        ws.write("K3", "(예: '학생', '선생님' 그대로 입력)")
        ws.write("K4", "B2~F2 : 5점 응답 수")
        ws.write("K5", "B3~F3 : 4점 응답 수")
        ws.write("K6", "B4~F4 : 3점 응답 수")
        ws.write("K7", "B5~F5 : 2점 응답 수")
        ws.write("K8", "B6~F6 : 1점 응답 수 (유형2도 동일)")
        ws.write("K9", "L1 셀에 구글폼 링크 입력")

    messagebox.showinfo("저장 완료", f"양식 파일이 저장되었습니다:\n{filepath}")
    try:
        os.startfile(filepath)
    except Exception:
        pass


# =========================
#  엑셀 읽기 유틸
# =========================

def _read_link(ws):
    """
    L1에서 링크를 읽습니다.
    L1이 비어있으면 1행 전체에서 http로 시작하는 문자열을 한 번 더 찾아봅니다.
    """
    link = ws["L1"].value
    if isinstance(link, str) and link.startswith("http"):
        return link

    for cell in ws[1]:
        val = cell.value
        if isinstance(val, str) and val.startswith("http"):
            return val

    raise ValueError("L1 셀 또는 1행에서 설문 링크(URL)를 찾을 수 없습니다.")


def _read_group(ws, start_row: int):
    """
    한 그룹(유형1 또는 유형2)의 분포를 읽습니다.
    start_row 기준:
      start_row     : 5점 개수 행
      start_row + 1 : 4점 개수 행
      ...
      start_row + 4 : 1점 개수 행
    열은 B(1번) ~ F(5번).

    return:
      questions: list[ list[int] ]  # 질문별 [5점, 4점, 3점, 2점, 1점] 개수
      total_n : int                 # 응답자 수 (문항별 총합)
    """
    questions = []
    totals = []

    for col in range(2, 7):  # B=2 ~ F=6
        counts = []
        for offset in range(5):  # 5,4,3,2,1 점
            row = start_row + offset
            cell = ws.cell(row=row, column=col)
            val = cell.value

            if val is None or val == "":
                v = 0
            else:
                try:
                    v = int(val)
                except Exception:
                    raise ValueError(f"{cell.coordinate} 셀 값 '{val}'는(은) 정수가 아닙니다.")
            if v < 0:
                raise ValueError(f"{cell.coordinate} 셀 값은 음수일 수 없습니다.")
            counts.append(v)

        questions.append(counts)
        totals.append(sum(counts))

    if not totals:
        return questions, 0

    max_total = max(totals)
    if max_total == 0:
        # 이 그룹은 응답이 없는 것으로 간주
        return questions, 0

    if len(set(totals)) != 1:
        raise ValueError(
            "문항별 총 응답 수가 서로 다릅니다.\n"
            "각 열(B~F)의 (5점~1점) 합이 모두 같도록 수정해 주세요."
        )

    return questions, totals[0]


def _read_group_label(ws, row: int, fallback: str) -> str:
    """
    A열(row행)에서 그룹 라벨을 읽습니다.
    비어 있으면 fallback 사용.
    """
    val = ws.cell(row=row, column=1).value
    if isinstance(val, str) and val.strip():
        return val.strip()
    return fallback


def _build_schedules(questions, total_n: int):
    """
    questions: 질문별 [5점, 4점, 3점, 2점, 1점] 개수
    total_n : 응답자 수

    return:
      schedules: list[ list[int] ]
        schedules[q_idx][i] = i번째 응답자가 q_idx번째 문항에서 줄 점수
    """
    schedules = []

    for q_idx, counts in enumerate(questions):
        seq = []
        for score, cnt in zip(SCORES, counts):
            seq.extend([score] * cnt)

        if len(seq) != total_n:
            raise ValueError(
                f"{q_idx + 1}번 문항의 점수 개수 합({len(seq)})이 "
                f"응답자 수({total_n})와 일치하지 않습니다."
            )

        random.shuffle(seq)
        schedules.append(seq)

    return schedules


# =========================
#  Playwright 유틸
# =========================

def _score_to_index(score: int, n_options: int) -> int:
    """
    점수(1~5) → 보기 인덱스(0~n_options-1) 매핑.
    좌→우 = 1점→5점이라고 가정.
    """
    if n_options <= 1:
        return 0
    pos = round((score - 1) * (n_options - 1) / 4)
    pos = max(0, min(pos, n_options - 1))
    return int(pos)


def _query_radio_items(group):
    """
    라디오 버튼 요소 목록 조회.
    구글폼 테마에 따라 클래스가 달라질 수 있어 보조 셀렉터도 사용.
    """
    radios = group.query_selector_all(".AB7Lab.Id5V1")
    if not radios:
        radios = group.query_selector_all('[role="radio"]')
    return radios


def _click_by_text(page, text: str):
    """텍스트로 요소 찾아 클릭. 실패해도 예외를 밖으로 올리지 않음."""
    try:
        page.get_by_text(text, exact=True).click()
        page.wait_for_load_state("networkidle")
    except PlaywrightTimeoutError:
        pass
    except Exception:
        pass


def _fill_likert_from_scores(page, scores_per_question):
    """
    현재 페이지에서 라디오 문항들을 scores_per_question에 맞게 선택.
    scores_per_question: [문항1점수, 문항2점수, ...]
    """
    page.wait_for_selector('[role="radiogroup"]', timeout=10000)
    groups_all = page.query_selector_all('[role="radiogroup"]')

    # 라디오가 실제로 있는 그룹만 모음
    valid_groups = []
    for g in groups_all:
        radios = _query_radio_items(g)
        if radios:
            valid_groups.append(g)

    n = min(len(scores_per_question), len(valid_groups))

    for i in range(n):
        group = valid_groups[i]
        radios = _query_radio_items(group)
        if not radios:
            continue

        idx = _score_to_index(int(scores_per_question[i]), len(radios))
        if 0 <= idx < len(radios):
            radios[idx].click()


def _submit_once(page, link: str, group_label: str, scores_per_question):
    """
    한 응답(한 사람분)을 실제로 제출.
    - link: 구글폼 링크
    - group_label: 폼에 표시되는 참여 유형 텍스트 (예: '학생', '선생님')
    - scores_per_question: [문항1점수, 문항2점수, ...]
    """
    page.goto(link)
    page.wait_for_load_state("networkidle")

    # 1페이지: 설명만 있는 페이지 → '다음'
    _click_by_text(page, "다음")

    # 2페이지: 참여자 확인(유형 선택)
    _click_by_text(page, group_label)  # A2/A7에서 읽은 텍스트 그대로 사용
    _click_by_text(page, "다음")

    # 3페이지: 실제 1~5번 만족도 문항
    _fill_likert_from_scores(page, scores_per_question)

    # 제출
    _click_by_text(page, "제출")

    # 너무 과도한 연속 제출 방지용 약간의 지연
    time.sleep(DELAY_BETWEEN_SUBMITS)


def _run_group(page, link: str, group_label: str, questions, total_n: int):
    """
    한 그룹(유형1 또는 유형2)에 대해 total_n개의 응답을 제출.
    """
    if total_n <= 0:
        return

    schedules = _build_schedules(questions, total_n)
    num_questions = len(questions)

    for i in range(total_n):
        scores = [schedules[q][i] for q in range(num_questions)]
        _submit_once(page, link, group_label, scores)


# =========================
#  외부에서 호출할 메인 함수
# =========================

def run_from_excel(filepath: str):
    """
    엑셀 파일(이룸캠프 만족도 양식)에 맞춰
    유형1/유형2(예: 학생/선생님) 응답을 구글폼에 정확히 입력합니다.
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    link = _read_link(ws)

    # 그룹1: 행 2~6, 라벨은 A2
    student_label = _read_group_label(ws, row=2, fallback="학생")
    student_questions, student_total = _read_group(ws, start_row=2)

    # 그룹2: 행 7~11, 라벨은 A7
    teacher_label = _read_group_label(ws, row=7, fallback="선생님")
    teacher_questions, teacher_total = _read_group(ws, start_row=7)

    if student_total == 0 and teacher_total == 0:
        raise ValueError("두 그룹 모두 응답 개수가 0입니다. 최소 한 명 이상 응답 수를 입력해 주세요.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        if student_total > 0:
            _run_group(page, link, student_label, student_questions, student_total)

        if teacher_total > 0:
            _run_group(page, link, teacher_label, teacher_questions, teacher_total)

        browser.close()
