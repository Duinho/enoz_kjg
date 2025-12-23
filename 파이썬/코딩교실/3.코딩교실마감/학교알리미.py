# 라이브러리 설치를 위해 터미널에 아래의 pip 명령어 입력
# 먼저 pip install openpyxl playwright 입력
# 그 다음 playwright install 입력
# 위 터미널 창에 붙여넣기하여 라이브러리 설치
import os
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
import time
from tkinter import filedialog, messagebox

폴더경로 = os.path.dirname(os.path.abspath(__file__))  # 폴더경로를 코드가 있는 디렉토리로 저장
템플릿폴더 = os.path.join(폴더경로, "엑셀양식")
템플릿파일명 = "학교알리미 양식.xlsx"

def _default_excel_path():
    template = os.path.join(템플릿폴더, 템플릿파일명)
    if os.path.exists(template):
        return template
    fallback = os.path.join(폴더경로, 템플릿파일명)
    if os.path.exists(fallback):
        return fallback
    legacy = os.path.join(폴더경로, "학교알리미학생수.xlsx")
    return legacy

excel_file_path = _default_excel_path()  # 엑셀 파일 경로

def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()

def _collect_level_labels(page, level_name):
    labels = []
    radio_buttons = page.query_selector_all(f'input[name="{level_name}"]')
    for radio in radio_buttons:
        label_id = radio.get_attribute("id")
        if not label_id:
            continue
        label = page.query_selector(f'label[for="{label_id}"]')
        if label:
            labels.append({"id": label_id, "text": label.inner_text().strip()})
    return labels

def _filter_labels(labels, keyword):
    if keyword:
        return [label for label in labels if keyword in label["text"]]
    return labels

def _click_label_by_id(page, label_id):
    page.click(f'label[for="{label_id}"]')
    page.wait_for_load_state('load')

def _select_region(page, sido_text, sigungu_text):
    if sido_text:
        level2_labels = _collect_level_labels(page, "level2")
        for label in level2_labels:
            if sido_text in label["text"]:
                _click_label_by_id(page, label["id"])
                break
    if sigungu_text:
        level3_labels = _collect_level_labels(page, "level3")
        for label in level3_labels:
            if sigungu_text in label["text"]:
                _click_label_by_id(page, label["id"])
                break


def download_template_school():
    src = _default_excel_path()
    if not os.path.exists(src):
        messagebox.showerror("오류", f"학교알리미 양식 파일을 찾을 수 없습니다.\n경로: {src}")
        return

    dest = filedialog.asksaveasfilename(
        title="학교알리미 양식 저장 위치 선택",
        defaultextension=".xlsx",
        initialfile=템플릿파일명,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not dest:
        return

    try:
        shutil.copy2(src, dest)
        messagebox.showinfo("완료", f"학교알리미 양식 파일이 저장되었습니다:\n{dest}")
        try:
            os.startfile(dest)
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"학교알리미 양식 복사 중 오류가 발생했습니다:\n{e}")

def 엑셀(): # 엑셀을 열어 정보를 가져오는 함수
    global 학교알리미주소, 학교, 시도, 시군구, 학년 
    wb = openpyxl.load_workbook(excel_file_path)    # 엑셀파일 로드
    ws = wb['동작']                                 # 엑셀파일 동작을 ws에 저장
    학교알리미주소 = _clean_text(ws.cell(row=1, column=17).value) # 각 행과열에서 필요한 정보 변수에 저장
    학교 = _clean_text(ws.cell(row=2, column=17).value)
    시도 = _clean_text(ws.cell(row=3, column=17).value)
    시군구 = _clean_text(ws.cell(row=4, column=17).value)
    if "초등학교" in 학교:                           # 분류가 초등학교면 1~6학년 그게 아니면 (중,고) 1~3학년
        학년 = ['1', '2', '3', '4', '5', '6']
    else:
        학년 = ['1', '2', '3']
    wb.close() # 엑셀파일 닫기
    return ws  # ws값 리턴

def 헤더추가(ws): # 엑셀에 제일 위에 헤더를 미리 추가하는 함수
    headers = [
        '시군구', '학교명',
        '1학년', '2학년', '3학년', '4학년', '5학년', '6학년',
        '1학년', '2학년', '3학년', '4학년', '5학년', '6학년'
    ]
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_index, value=header)

def 학교받아오기(page):       
    # 엑셀에 저장한 학교알리미 주소로 접속
    page.goto(학교알리미주소)

    # level1: 학교급 (초/중/고)
    radio_buttons = page.query_selector_all('input[name="level1"]')
    for radio in radio_buttons:
        label = page.query_selector(f'label[for="{radio.get_attribute("id")}"]')
        if label and 학교 in label.inner_text():
            label.click()
            page.wait_for_load_state('load')
            break

    # level2: 시·도
    level2_labels = _filter_labels(_collect_level_labels(page, "level2"), 시도)
    if not level2_labels:
        print('시/도 조건에 맞는 항목을 찾지 못했습니다.')
        return []

    학교목록 = []
    for level2 in level2_labels:
        _click_label_by_id(page, level2["id"])

        # level3: 시·군·구
        level3_labels = _filter_labels(_collect_level_labels(page, "level3"), 시군구)
        if not level3_labels:
            print(f'시/군/구 조건에 맞는 항목 없음: {level2["text"]}')
            continue

        for level3 in level3_labels:
            _click_label_by_id(page, level3["id"])

            labels = page.query_selector_all('label[for^="shlCd"]')
            for label in labels:
                학교목록.append((label.inner_text(), level2["text"], level3["text"]))

    return 학교목록

def 학교정보검색(page, 학교목록, 학년, ws, start_row=2):  
    row_index = start_row  # 첫 번째 줄은 헤더이므로 두 번째 줄부터 시작
    current_region = None

    for 학교, 시도명, 시군구명 in 학교목록:  
        region_key = (시도명, 시군구명)
        if region_key != current_region:
            _select_region(page, 시도명, 시군구명)
            current_region = region_key

        label = page.query_selector(f'label:has-text("{학교}")')  
        if label:
            # 학교 클릭 및 이동
            label.click()
            page.wait_for_load_state('load') 
            page.click('a[data-tab-id="tabSel"]')
            page.wait_for_load_state('load')  
            page.click('a.accordian_title:has-text("학생현황")') 
            page.wait_for_load_state('load')                    
            page.click('label[for="hangmok01"]')   # 학생현황 선택
            page.wait_for_load_state('load')   

            # 검색 버튼 클릭
            page.click('#webSearchButton')                  
            page.wait_for_load_state('load')

            # "입력된 데이터가 없습니다." 체크
            no_data = page.query_selector('p:has-text("입력된 데이터가 없습니다.")')
            if no_data:
                print(f'{학교} → 데이터 없음, 건너뜀')
                # 다음 학교로 이동
                page.click('a.slidedown[href="javascript:research();"]') 
                page.wait_for_load_state('load')
                continue  # 이 학교는 패스

            # 데이터가 있는 경우에만 처리
            page.wait_for_selector('xpath=//tr[th[normalize-space(.)="학급수"]]', timeout=10000)
            page.wait_for_selector('xpath=//tr[th[normalize-space(.)="학생수"]]', timeout=10000)

            # 학급수 행
            class_counts = []
            class_row = page.query_selector('xpath=//tr[th[normalize-space(.)="학급수"]]')
            if class_row:
                tds = class_row.query_selector_all('td')
                for i, grade in enumerate(학년):
                    if i < len(tds):
                        raw_value = tds[i].inner_text().strip()
                        if raw_value in ('', '-'):
                            class_counts.append(0)
                        else:
                            class_counts.append(int(raw_value.split('(')[0].replace(',', '')))
                    else:
                        class_counts.append(0)

            # 학생수 행
            values = []
            student_row = page.query_selector('xpath=//tr[th[normalize-space(.)="학생수"]]')
            if student_row:
                tds = student_row.query_selector_all('td')
                for i, grade in enumerate(학년):
                    if i < len(tds):
                        raw_value = tds[i].inner_text().strip()
                        if raw_value in ('', '-'):
                            values.append(0)
                        else:
                            values.append(int(raw_value.replace(',', '').split('(')[0]))
                    else:
                        values.append(0)

            # ----------------------------
            # 엑셀 저장
            # ----------------------------
            region_label = f'{시도명} {시군구명}'.strip()
            ws.cell(row=row_index, column=1, value=region_label)
            ws.cell(row=row_index, column=2, value=학교)
            for col_index, value in enumerate(values, start=3):
                ws.cell(row=row_index, column=col_index, value=value)  # 학생 수
            for col_index, class_count in enumerate(class_counts, start=9):  
                ws.cell(row=row_index, column=col_index, value=class_count)  # 반 수

            row_index += 1  
            print(f'{학교} 학생수 및 반 수 저장 완료')

            # 다음 학교로 이동
            page.click('a.slidedown[href="javascript:research();"]') 
            page.wait_for_load_state('load')

    return row_index


def 동작():
    global browser, page  # 전역 변수로 사용
    with sync_playwright() as p:  # 크로미움 브라우저 열기
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()           # 새로운 창이 열리면 page에 저장
        ws = 엑셀()                         # 엑셀을 통해 ws에 엑셀값 저장
        헤더추가(ws)                        # ws에 헤더추가하여 저장
        학교목록 = 학교받아오기(page)        # 학교목록에 지역별 학교 정보를 저장
        학교정보검색(page, 학교목록, 학년, ws) # 각 정보를 통해 
        ws.parent.save(excel_file_path)    # 동작이 끝나면 엑셀 파일 저장
        try:
            os.startfile(excel_file_path)
        except Exception:
            pass
        print('모든 학교 저장 완료')
        browser.close()  # 브라우저 닫기

def run_from_excel(filepath: str):
    global excel_file_path
    if filepath:
        excel_file_path = filepath
    동작()


if __name__ == "__main__":
    동작()
