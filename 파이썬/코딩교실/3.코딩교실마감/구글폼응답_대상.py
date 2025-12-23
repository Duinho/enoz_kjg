# 구글폼응답_대상.py  (구글폼 응답 - 대상용)
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import time
import random
import os
import shutil
from tkinter import filedialog, messagebox

# 점수 순서: 5점, 4점, 3점, 2점, 1점
SCORES = [5, 4, 3, 2, 1]
DELAY_BETWEEN_SUBMITS = 0.0  # AB2 값을 사용하므로 기본값은 0
ALLOW_MISMATCHED_TOTALS = True

# =========================
#  템플릿(양식) 배포
# =========================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "엑셀양식")
TEMPLATE_FILENAME_TARGET = "구글폼응답(대상) 양식.xlsx"


def download_template_iloom():
    """
    엑셀양식 폴더에 있는 '구글폼응답(대상) 양식.xlsx' 파일을
    사용자가 선택한 위치에 복사(저장)하는 함수입니다.
    """
    src = os.path.join(TEMPLATE_DIR, TEMPLATE_FILENAME_TARGET)

    if not os.path.exists(src):
        messagebox.showerror(
            "오류",
            f"구글폼응답(대상) 템플릿 파일을 찾을 수 없습니다.\n\n"
            f"경로: {src}\n\n"
            f"'엑셀양식' 폴더 안에 '{TEMPLATE_FILENAME_TARGET}' 파일이 있는지 확인해 주세요."
        )
        return

    dest = filedialog.asksaveasfilename(
        title="구글폼응답(대상) 양식 저장 위치 선택",
        defaultextension=".xlsx",
        initialfile=TEMPLATE_FILENAME_TARGET,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not dest:
        return

    try:
        shutil.copy2(src, dest)
        messagebox.showinfo("완료", f"구글폼응답(대상) 양식 파일이 저장되었습니다:\n{dest}")
        try:
            os.startfile(dest)
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"구글폼응답(대상) 양식 복사 중 오류가 발생했습니다:\n{e}")


# =========================
#  엑셀 읽기 + 딜레이 유틸
# =========================

def _read_link(ws):
    """
    1행에서 http로 시작하는 문자열을 찾아 설문 링크로 사용합니다.
    (예: AA1, AB1 어디에 있어도 인식)
    """
    for cell in ws[1]:
        val = cell.value
        if isinstance(val, str) and val.strip().startswith("http"):
            return val.strip()
    raise ValueError("1행에서 설문 링크(URL)를 찾을 수 없습니다. (http로 시작하는 값 필요)")


def _read_delay_base(ws):
    """
    동작 시트의 AB2에서 기본 딜레이(초)를 읽습니다.
    - 비어 있거나 0 이하이거나 숫자가 아니면 0으로 간주 (딜레이 없음)
    """
    cell = ws["AB2"].value
    if cell is None or cell == "":
        return 0.0

    try:
        delay = float(cell)
    except Exception:
        return 0.0

    if delay <= 0:
        return 0.0
    return delay


def _rand_sleep(base_delay: float):
    """
    base_delay(초)를 기준으로 ±30% 범위에서 랜덤 딜레이를 줍니다.
    base_delay가 0 이하면 아무 것도 하지 않습니다.
    """
    if not base_delay or base_delay <= 0:
        return
    jitter = base_delay * 0.3
    low = max(0.0, base_delay - jitter)
    high = base_delay + jitter
    time.sleep(random.uniform(low, high))


def _read_group(ws, start_row: int, start_col: int = 2):
    """
    한 그룹(유형1 또는 유형2)의 분포를 읽습니다.

    start_row 기준:
      start_row     : 5점 개수 행
      start_row + 1 : 4점 개수 행
      start_row + 2 : 3점 개수 행
      start_row + 3 : 2점 개수 행
      start_row + 4 : 1점 개수 행

    열은 start_col(B=2) 이상, "값이 한 번도 안 쓰인 열"은 뒤에서 자동으로 무시합니다.

    return:
      questions: list[ list[int] ]  # 질문별 [5점, 4점, 3점, 2점, 1점] 개수
      total_n : int                 # 응답자 수 (문항별 총합)
    """
    max_col = ws.max_column
    questions = []
    totals = []

    for col in range(start_col, max_col + 1):
        # 이 열 전체가 비어 있다면 이 이후로는 문항이 없는 것으로 보고 중단
        all_empty = True
        for offset in range(5):
            row = start_row + offset
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                all_empty = False
                break
        if all_empty:
            break

        counts = [];
        for offset in range(5):  # 5,4,3,2,1 점
            row = start_row + offset
            cell = ws.cell(row=row, column=col)
            val = cell.value

            if val is None or val == "":
                v = 0
            else:
                try:
                    v = int(val)
                except Exception:
                    raise ValueError(f"{cell.coordinate} 셀 값 '{val}'는(은) 정수가 아닙니다.")
            if v < 0:
                raise ValueError(f"{cell.coordinate} 셀 값은 음수일 수 없습니다.")
            counts.append(v)

        questions.append(counts)
        totals.append(sum(counts))

    if not totals:
        # 이 그룹은 응답이 없는 것으로 간주
        return questions, 0

    max_total = max(totals)
    if max_total == 0:
        return questions, 0

    if len(set(totals)) != 1 and not ALLOW_MISMATCHED_TOTALS:
        raise ValueError(
            "문항별 총 응답 수가 서로 다릅니다.\n"
            "각 문항 열의 (5점~1점) 합이 모두 같도록 수정해 주세요."
        )

    return questions, max_total


def _read_group_label(ws, row: int):
    """
    A열(row행)에서 그룹 라벨을 읽습니다.
    - 공백/None이면 None을 반환 → '참여 유형 선택' 페이지 없이 진행.
    """
    val = ws.cell(row=row, column=1).value
    if isinstance(val, str) and val.strip():
        return val.strip()
    return None


def _build_schedules(questions, total_n: int):
    """
    questions: 질문별 [5점, 4점, 3점, 2점, 1점] 개수
    total_n : 응답자 수

    return:
      schedules: list[ list[int] ]
        schedules[q_idx][i] = i번째 응답자가 q_idx번째 문항에서 줄 점수
    """
    schedules = []

    for q_idx, counts in enumerate(questions):
        seq = []
        for score, cnt in zip(SCORES, counts):
            seq.extend([score] * cnt)

        if len(seq) != total_n:
            if not ALLOW_MISMATCHED_TOTALS:
                raise ValueError(
                    f"{q_idx + 1}번 문항의 점수 개수 합({len(seq)})이 "
                    f"응답자 수({total_n})와 일치하지 않습니다."
                )
            if len(seq) < total_n:
                weights = counts
                if sum(weights) <= 0:
                    weights = None
                seq.extend(random.choices(SCORES, weights=weights, k=total_n - len(seq)))
            else:
                random.shuffle(seq)
                seq = seq[:total_n]

        random.shuffle(seq)
        schedules.append(seq)

    return schedules


# =========================
#  Playwright 유틸
# =========================

def _score_to_index(score: int, n_options: int) -> int:
    """
    점수(1~5) → 보기 인덱스(0~n_options-1) 매핑.
    좌→우 = 1점→5점이라고 가정.
    """
    if n_options <= 1:
        return 0
    pos = round((score - 1) * (n_options - 1) / 4)
    pos = max(0, min(pos, n_options - 1))
    return int(pos)


def _query_radio_items(group):
    """
    라디오 버튼 요소 목록 조회.
    구글폼 테마에 따라 클래스가 달라질 수 있어 보조 셀렉터도 사용.
    """
    radios = group.query_selector_all(".AB7Lab.Id5V1")
    if not radios:
        radios = group.query_selector_all('[role="radio"]')
    return radios


def _click_by_text(page, text: str):
    """텍스트로 요소 찾아 클릭. 실패해도 예외를 밖으로 올리지 않음."""
    try:
        page.get_by_text(text, exact=True).click()
        page.wait_for_load_state("networkidle")
    except PlaywrightTimeoutError:
        pass
    except Exception:
        pass


def _fill_likert_from_scores(page, scores_per_question):
    """
    현재 페이지에서 라디오 문항들을 scores_per_question에 맞게 선택.
    대상용에서도, 폼용과 마찬가지로
    'radiogroup이 여러 개 있을 때 뒤에서부터 필요한 개수만 사용'합니다.
    """
    page.wait_for_selector('[role="radiogroup"]', timeout=10000)
    groups_all = page.query_selector_all('[role="radiogroup"]')

    valid_groups = []
    for g in groups_all:
        radios = _query_radio_items(g)
        if radios:
            valid_groups.append(g)

    if not valid_groups:
        return

    n_q = len(scores_per_question)
    n_g = len(valid_groups)

    # radiogroup이 더 많으면 뒤에서부터 n_q개 사용
    if n_g >= n_q:
        target_groups = valid_groups[-n_q:]
    else:
        target_groups = valid_groups

    for i, group in enumerate(target_groups):
        if i >= len(scores_per_question):
            break
        radios = _query_radio_items(group)
        if not radios:
            continue

        idx = _score_to_index(int(scores_per_question[i]), len(radios))
        if 0 <= idx < len(radios):
            radios[idx].click()


def _submit_once(page, link: str, group_label, scores_per_question, base_delay: float):
    """
    한 응답(한 사람분)을 실제로 제출.
    - link: 구글폼 링크
    - group_label:
        * 문자열(예: '학생', '선생님')이면 "참여 유형 선택" 페이지에서 해당 라벨 클릭
        * None이면 유형 선택 페이지를 스킵
    - scores_per_question: [문항1점수, 문항2점수, ...]
    - base_delay: AB2에서 읽은 기본 딜레이(초)
    """
    page.goto(link)
    page.wait_for_load_state("networkidle")

    # 첫 페이지: 안내만 있을 수 있으니 '다음' 한 번 시도
    _click_by_text(page, "다음")

    # 참여 유형 선택 페이지 (group_label 이 있을 때만)
    if group_label:
        _click_by_text(page, group_label)
        _click_by_text(page, "다음")

    # 만족도 문항 페이지
    _fill_likert_from_scores(page, scores_per_question)

    # 제출
    _click_by_text(page, "제출")

    # 응답 하나 끝난 뒤 랜덤 딜레이
    _rand_sleep(base_delay)


def _run_group(page, link: str, group_label, questions, total_n: int, base_delay: float):
    """
    한 그룹(유형1 또는 유형2)에 대해 total_n개의 응답을 제출.
    group_label 이 None이면 유형 선택 페이지 없이 진행.
    """
    if total_n <= 0:
        return

    schedules = _build_schedules(questions, total_n)
    num_questions = len(questions)

    for i in range(total_n):
        scores = [schedules[q][i] for q in range(num_questions)]
        _submit_once(page, link, group_label, scores, base_delay)


# =========================
#  외부에서 호출할 메인 함수
# =========================

def run_from_excel(filepath: str):
    """
    엑셀 파일(구글폼응답(대상) 양식)에 맞춰
    분포대로 만족도 응답을 구글폼에 입력합니다.
    AB2 셀에 기본 딜레이(초)를 넣으면, 각 응답 제출 후
    해당 값의 ±30% 범위에서 랜덤 대기합니다.
    """
    wb = load_workbook(filepath, data_only=True)

    # '동작' 시트를 우선 사용, 없으면 활성 시트 사용
    if "동작" in wb.sheetnames:
        ws = wb["동작"]
    else:
        ws = wb.active

    link = _read_link(ws)
    base_delay = _read_delay_base(ws)

    # 그룹1: 행 2~6, 라벨은 A2
    group1_label = _read_group_label(ws, row=2)      # None이면 유형 선택 페이지 스킵
    group1_questions, group1_total = _read_group(ws, start_row=2, start_col=2)

    # 그룹2: 행 8~12, 라벨은 A8
    group2_label = _read_group_label(ws, row=8)
    group2_questions, group2_total = _read_group(ws, start_row=8, start_col=2)

    if group1_total == 0 and group2_total == 0:
        raise ValueError("두 그룹 모두 응답 개수가 0입니다. 최소 한 명 이상 응답 수를 입력해 주세요.")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        if group1_total > 0:
            _run_group(page, link, group1_label, group1_questions, group1_total, base_delay)

        if group2_total > 0:
            _run_group(page, link, group2_label, group2_questions, group2_total, base_delay)

        browser.close()
