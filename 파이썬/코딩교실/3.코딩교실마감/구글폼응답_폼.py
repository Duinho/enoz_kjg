# 구글폼응답_폼.py  (구글폼 응답 - 폼 여러 개용)
# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import time
import random
import os
import shutil
from tkinter import filedialog, messagebox

# 점수 순서: 5점, 4점, 3점, 2점, 1점
SCORES = [5, 4, 3, 2, 1]
DELAY_BETWEEN_SUBMITS = 0.0  # AB2 값을 사용하므로 기본값은 0으로 둡니다.

# =========================
#  템플릿(양식) 배포
# =========================

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "엑셀양식")
TEMPLATE_FILENAME_FORM = "만족도조사(폼) 양식.xlsx"


def download_template_form():
    """
    엑셀양식 폴더에 있는 '만족도조사(폼) 양식.xlsx' 파일을
    사용자가 선택한 위치에 복사(저장)하는 함수입니다.
    """
    src = os.path.join(TEMPLATE_DIR, TEMPLATE_FILENAME_FORM)

    if not os.path.exists(src):
        messagebox.showerror(
            "오류",
            f"만족도조사(폼) 템플릿 파일을 찾을 수 없습니다.\n\n"
            f"경로: {src}\n\n"
            f"'엑셀양식' 폴더 안에 '{TEMPLATE_FILENAME_FORM}' 파일이 있는지 확인해 주세요."
        )
        return

    dest = filedialog.asksaveasfilename(
        title="만족도조사(폼) 양식 저장 위치 선택",
        defaultextension=".xlsx",
        initialfile=TEMPLATE_FILENAME_FORM,
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    if not dest:
        return

    try:
        shutil.copy2(src, dest)
        messagebox.showinfo("완료", f"만족도조사(폼) 양식 파일이 저장되었습니다:\n{dest}")
        try:
            os.startfile(dest)
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("오류", f"만족도조사(폼) 양식 복사 중 오류가 발생했습니다:\n{e}")


# 옛 이름과 호환용
download_template_iloom = download_template_form
download_template_combo = download_template_form  # 예전 combo 이름 사용 대비


# =========================
#  엑셀 읽기 + 딜레이 유틸
# =========================

def _read_link(ws):
    """
    1행에서 http로 시작하는 문자열을 찾아 설문 링크로 사용합니다.
    (예: AA1, AB1 어디에 있어도 인식)
    """
    for cell in ws[1]:
        val = cell.value
        if isinstance(val, str) and val.strip().startswith("http"):
            return val.strip()
    raise ValueError("1행에서 설문 링크(URL)를 찾을 수 없습니다. (http로 시작하는 값 필요)")


def _read_delay_base(ws):
    """
    동작 시트의 AB2에서 기본 딜레이(초)를 읽습니다.
    - 비어 있거나 0 이하이거나 숫자가 아니면 0으로 간주 (딜레이 없음)
    """
    cell = ws["AB2"].value
    if cell is None or cell == "":
        return 0.0

    try:
        delay = float(cell)
    except Exception:
        return 0.0

    if delay <= 0:
        return 0.0
    return delay


def _rand_sleep(base_delay: float):
    """
    base_delay(초)를 기준으로 ±30% 범위에서 랜덤 딜레이를 줍니다.
    base_delay가 0 이하면 아무 것도 하지 않습니다.
    """
    if not base_delay or base_delay <= 0:
        return
    jitter = base_delay * 0.3
    low = max(0.0, base_delay - jitter)
    high = base_delay + jitter
    time.sleep(random.uniform(low, high))


def _read_group(ws, start_row: int, start_col: int = 2):
    """
    한 폼 블록의 분포를 읽습니다.

    start_row 기준:
      start_row     : 5점 개수 행
      start_row + 1 : 4점 개수 행
      start_row + 2 : 3점 개수 행
      start_row + 3 : 2점 개수 행
      start_row + 4 : 1점 개수 행

    열은 start_col(B=2) 이상, "값이 한 번도 안 쓰인 열"은 뒤에서 자동으로 무시합니다.

    return:
      questions: list[ list[int] ]  # 질문별 [5점, 4점, 3점, 2점, 1점] 개수
      total_n : int                 # (분포에서 계산한) 응답자 수
    """
    max_col = ws.max_column
    questions = []
    totals = []

    for col in range(start_col, max_col + 1):
        # 이 열 전체가 비어 있다면 이 이후로는 문항이 없는 것으로 보고 중단
        all_empty = True
        for offset in range(5):  # 5,4,3,2,1 행
            row = start_row + offset
            cell = ws.cell(row=row, column=col)
            if cell.value not in (None, ""):
                all_empty = False
                break
        if all_empty:
            break

        counts = []
        for offset in range(5):  # 5,4,3,2,1 점
            row = start_row + offset
            cell = ws.cell(row=row, column=col)
            val = cell.value

            if val is None or val == "":
                v = 0
            else:
                try:
                    v = int(val)
                except Exception:
                    raise ValueError(f"{cell.coordinate} 셀 값 '{val}'는(은) 정수가 아닙니다.")
            if v < 0:
                raise ValueError(f"{cell.coordinate} 셀 값은 음수일 수 없습니다.")
            counts.append(v)

        questions.append(counts)
        totals.append(sum(counts))

    if not totals:
        # 이 폼 블록은 응답이 없는 것으로 간주
        return questions, 0

    max_total = max(totals)
    if max_total == 0:
        return questions, 0

    if len(set(totals)) != 1:
        raise ValueError(
            "문항별 총 응답 수가 서로 다릅니다.\n"
            "각 문항 열의 (5점~1점) 합이 모두 같도록 수정해 주세요."
        )

    return questions, totals[0]


def _build_schedules(questions, total_n: int):
    """
    questions: 질문별 [5점, 4점, 3점, 2점, 1점] 개수
    total_n : 응답자 수

    return:
      schedules: list[ list[int] ]
        schedules[q_idx][i] = i번째 응답자가 q_idx번째 문항에서 줄 점수
    """
    schedules = []

    for q_idx, counts in enumerate(questions):
        seq = []
        for score, cnt in zip(SCORES, counts):
            seq.extend([score] * cnt)

        if len(seq) != total_n:
            raise ValueError(
                f"{q_idx + 1}번 문항의 점수 개수 합({len(seq)})이 "
                f"응답자 수({total_n})와 일치하지 않습니다."
            )

        random.shuffle(seq)
        schedules.append(seq)

    return schedules


# =========================
#  Playwright 유틸
# =========================

def _score_to_index(score: int, n_options: int) -> int:
    """
    점수(1~5) → 보기 인덱스(0~n_options-1) 매핑.
    좌→우 = 1점→5점이라고 가정.
    """
    if n_options <= 1:
        return 0
    pos = round((score - 1) * (n_options - 1) / 4)
    pos = max(0, min(pos, n_options - 1))
    return int(pos)


def _query_radio_items(group):
    """
    라디오 버튼 요소 목록 조회.
    구글폼 테마에 따라 클래스가 달라질 수 있어 보조 셀렉터도 사용.
    """
    radios = group.query_selector_all(".AB7Lab.Id5V1")
    if not radios:
        radios = group.query_selector_all('[role="radio"]')
    return radios


def _click_by_text(page, text: str):
    """텍스트로 요소 찾아 클릭. 실패해도 예외를 밖으로 올리지 않음."""
    try:
        page.get_by_text(text, exact=True).click()
        page.wait_for_load_state("networkidle")
    except PlaywrightTimeoutError:
        pass
    except Exception:
        pass


def _fill_likert_from_scores(page, scores_per_question):
    """
    현재 페이지에서 라디오 문항들을 scores_per_question에 맞게 선택.
    scores_per_question: [문항1점수, 문항2점수, ...]
    """

    # 페이지 내 모든 'radiogroup' 수집
    page.wait_for_selector('[role="radiogroup"]', timeout=10000)
    groups_all = page.query_selector_all('[role="radiogroup"]')

    # 실제 라디오 버튼이 있는 그룹만 필터링
    valid_groups = []
    for g in groups_all:
        radios = _query_radio_items(g)
        if radios:
            valid_groups.append(g)

    if not valid_groups:
        return

    n_q = len(scores_per_question)
    n_g = len(valid_groups)

    # radiogroup이 더 많을 수 있으므로 "뒤에서부터" 필요한 개수(n_q)만 사용
    if n_g >= n_q:
        target_groups = valid_groups[-n_q:]
    else:
        # radiogroup이 질문 수보다 적으면 있는 것만 사용
        target_groups = valid_groups

    # target_groups[0] ↔ scores_per_question[0] 순서로 매핑
    for i, group in enumerate(target_groups):
        if i >= len(scores_per_question):
            break
        radios = _query_radio_items(group)
        if not radios:
            continue

        idx = _score_to_index(int(scores_per_question[i]), len(radios))
        if 0 <= idx < len(radios):
            radios[idx].click()


# =========================
#  메인 로직 (폼 여러 개, 한 설문 안에서 순서대로 진행)
# =========================

def run_from_excel_form(filepath: str):
    """
    엑셀 파일(만족도조사(폼) 양식)에 맞춰
    한 설문 안에서 여러 개의 '폼 페이지'를 순서대로 채웁니다.

    흐름 (한 번의 응답 기준):
      1페이지: 안내 → "다음"
      2페이지: A2 블록(첫 번째 폼) 문항들 → (폼이 2개 이상이면 "다음", 아니면 "제출")
      3페이지: A8 블록(두 번째 폼) 문항들 → (폼이 3개 이상이면 "다음", 아니면 "제출")
      4페이지: A14 블록(세 번째 폼) 문항들 → 항상 "제출"

    엑셀 '동작' 시트 구조:
      - 1행: 링크 (어느 셀에 있어도 http로 시작하면 인식)
      - 첫 번째 폼
        * A2 : 제목 (라벨용, 코드에서는 안 써도 됨)
        * B2~? : (5점 개수 또는 수식)
        * B3~B6 : 4/3/2/1점 개수
        * A7 : 첫 번째 폼의 전체 응답 수 (명수)
      - 두 번째 폼: A8, B8~?, B9~B12, A13
      - 세 번째 폼: A14, B14~?, B15~B18, A19

    A7/A13/A19 중 값이 있는 블록만 '폼 페이지'로 인식.
    여러 폼이 모두 존재하는 경우, A7/A13/A19의 값(응답 수)은 서로 같아야 함.

    AB2:
      - 기본 딜레이(초)
      - 값이 없거나 0 이하면 딜레이 없이 진행
      - 값이 양수면 각 응답(제출) 사이에 ±30% 랜덤 딜레이 적용
    """
    wb = load_workbook(filepath, data_only=True)

    # '동작' 시트를 우선 사용, 없으면 활성 시트 사용
    if "동작" in wb.sheetnames:
        ws = wb["동작"]
    else:
        ws = wb.active

    link = _read_link(ws)
    base_delay = _read_delay_base(ws)

    # 블록 시작 행 / 인원수(A열) 있는 행
    block_starts = [2, 8, 14]   # A2, A8, A14
    count_rows   = [7, 13, 19]  # A7, A13, A19

    blocks = []  # (schedules, total_n)
    totals_from_blocks = set()

    for start_row, cnt_row in zip(block_starts, count_rows):
        cnt_val = ws.cell(row=cnt_row, column=1).value  # A열: 해당 폼 전체 응답 수

        questions, total_from_dist = _read_group(ws, start_row=start_row, start_col=2)

        # 완전히 비어 있는 블록이면 건너뜀
        if cnt_val in (None, "", 0) and total_from_dist == 0:
            continue

        # cnt_val이 비었으면 분포 합을 사용
        if cnt_val in (None, ""):
            total_n = total_from_dist
        else:
            try:
                total_n = int(cnt_val)
            except Exception:
                raise ValueError(f"A{cnt_row} 셀 값 '{cnt_val}'는(은) 정수가 아닙니다.")

            # 분포 합이 0이 아니고, A7/A13/A19와 다르면 오류
            if total_from_dist not in (0, total_n):
                raise ValueError(
                    f"A{cnt_row}의 값({total_n})과 "
                    f"해당 블록 분포의 합({total_from_dist})이 일치하지 않습니다.\n"
                    f"분포(5점~1점 개수)와 A{cnt_row} 값을 다시 확인해 주세요."
                )

        if total_n <= 0:
            continue

        schedules = _build_schedules(questions, total_n)
        blocks.append((schedules, total_n))
        totals_from_blocks.add(total_n)

    if not blocks:
        raise ValueError("유효한 폼 블록을 찾지 못했습니다. A7/A13/A19에 응답 수를 입력해 주세요.")

    # 여러 폼이 모두 존재하는 경우, 응답 수(명수)는 같아야 함
    if len(totals_from_blocks) != 1:
        raise ValueError(
            "여러 폼(A7, A13, A19)의 응답 수가 서로 다릅니다.\n"
            "하나의 설문에서 1번폼→2번폼→3번폼 순서로 진행하려면\n"
            "A7, A13, A19 값(응답 수)을 모두 동일하게 맞춰 주세요."
        )

    total_runs = totals_from_blocks.pop()  # 전체 설문 제출 횟수 (예: 5번)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        # i번째 응답(1명 분) 처리
        for i in range(total_runs):
            # 설문 처음 페이지로 이동
            page.goto(link)
            page.wait_for_load_state("networkidle")

            # 1페이지: 안내 → '다음'
            _click_by_text(page, "다음")

            # 각 폼 페이지를 순서대로 처리
            for idx, (schedules, _) in enumerate(blocks):
                # 이 응답자(i)에 대해, 해당 폼 페이지의 각 문항 점수 구성
                scores_per_question = [
                    schedules[q_idx][i] for q_idx in range(len(schedules))
                ]

                # 현재 페이지(폼)의 문항 채우기
                _fill_likert_from_scores(page, scores_per_question)

                # 마지막 폼이면 '제출', 아니면 '다음'
                if idx == len(blocks) - 1:
                    _click_by_text(page, "제출")
                else:
                    _click_by_text(page, "다음")

            # 한 사람(한 번의 제출) 끝났을 때 AB2 기반 랜덤 딜레이
            _rand_sleep(base_delay)

        browser.close()


# 호환용 별칭들
def run_from_excel(filepath: str):
    return run_from_excel_form(filepath)


def run_from_excel_combo(filepath: str):
    return run_from_excel_form(filepath)
