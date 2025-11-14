# -*- coding: utf-8 -*-
import os, re, datetime, shutil
import openpyxl
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

# ──────────────────────────────────────────────────────────
# 경로
# ──────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, '보고서수정.xlsx')
REPORT_DIR = os.path.join(BASE_DIR, '보고서')
BACKUP_DIR = os.path.join(REPORT_DIR, '__BACKUP__')

# ──────────────────────────────────────────────────────────
# 전역 상태
# ──────────────────────────────────────────────────────────
cfg = {}
name_change_map = {}          # {orig -> mapped}
session_descriptions = []     # 1~8차 교육내용
roster = {}                   # { "MW_01": {"students": {"홍길동": {1:"출",..}}}, ... }
session_dates_map = {}        # { "MW_01": ["YYYY-MM-DD", ...], ... }

# ──────────────────────────────────────────────────────────
# 유틸
# ──────────────────────────────────────────────────────────
def _norm_name(s:str) -> str:
    return " ".join(str(s).strip().split())

def _ensure_dir(p:str):
    os.makedirs(p, exist_ok=True)

def _coerce_int(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, float):
        return int(val)
    s = str(val).strip()
    m = re.search(r'\d+', s)
    return int(m.group(0)) if m else None

def _backup_move(path:str):
    """
    path(결과 폴더 내 파일)를 __BACKUP__/ 동일한 하위 구조로 이동.
    """
    rel = os.path.relpath(path, REPORT_DIR)
    dst = os.path.join(BACKUP_DIR, rel)
    _ensure_dir(os.path.dirname(dst))
    if os.path.exists(dst):
        # 백업 대상도 충돌하면 타임스탬프 덧붙임
        base, ext = os.path.splitext(dst)
        n = 1
        while os.path.exists(f"{base}_{n}{ext}"):
            n += 1
        dst = f"{base}_{n}{ext}"
    shutil.move(path, dst)

def _suffix_change_in_html(path:str, old_num:str, new_num:str):
    """문서 내부 Group No/담당강사명의 접미사 _old → _new 교체"""
    with open(path, 'r', encoding='cp949', errors='ignore') as f:
        soup = BeautifulSoup(f, 'html.parser')
    for lbl in ('Group No', '담당강사명'):
        td0 = soup.find('td', string=lbl)
        if not td0: continue
        td1 = td0.find_next_sibling('td')
        if not td1: continue
        txt = td1.get_text()
        if '_' in txt:
            head = txt.rsplit('_', 1)[0]
            td1.string = f"{head}_{new_num}"
    with open(path, 'w', encoding='cp949', errors='ignore') as f:
        f.write(str(soup))

def _set_datetime_in_html(path:str, ymd:str):
    """수업일시의 날짜(YYYY.MM.DD(EEE))만 치환"""
    with open(path, 'r', encoding='cp949', errors='ignore') as f:
        soup = BeautifulSoup(f, 'html.parser')
    td0 = soup.find('td', string='수업일시')
    if td0:
        td1 = td0.find_next_sibling('td')
        if td1:
            raw = td1.get_text()
            dt = datetime.datetime.strptime(ymd, "%Y-%m-%d")
            date_str = dt.strftime("%Y.%m.%d")
            dow = dt.strftime("%a").upper()
            new_prefix = f"{date_str}({dow})"
            new_text = re.sub(r"\d{4}\.\d{2}\.\d{2}\([A-Z]{3}\)", new_prefix, raw)
            td1.clear()
            td1.append(new_text)
    with open(path, 'w', encoding='cp949', errors='ignore') as f:
        f.write(str(soup))

def _rewrite_attendance_html(path: str, attendance_list, education_idx=None, *, mode="attendance_only"):
    """
    mode:
      - "attendance_only": 출결만 덮어씀 + 특이사항 비움. 교육내용은 보존.
      - "stub_init": 스텁 생성용. 출결칸 비우고, 특이사항 비우며, 교육내용만 세팅.
    """
    with open(path, 'r', encoding='cp949', errors='ignore') as f:
        soup = BeautifulSoup(f, 'html.parser')

    # 1) 출결현황
    td0 = soup.find('td', string=lambda s: s and s.strip() == '출결현황')
    if td0:
        td1 = td0.find_next_sibling('td')
        if td1:
            td1.clear()
            if mode == "attendance_only":
                p = soup.new_tag('p', **{'class': 'MsoNormal'})
                for i, (nm, status) in enumerate(sorted(attendance_list, key=lambda x: x[0])):
                    if i:
                        p.append(', ')
                    s_name = soup.new_tag('span', **{'class': 'SpellE'}); s_name.string = nm
                    color  = 'blue' if status == '출' else '#EE0000'
                    s_stat = soup.new_tag('span', lang='EN-US', style=f'color:{color}')
                    s_stat.string = f'[{status}]'
                    p.append(s_name); p.append(s_stat)
                td1.append(p)
            # stub_init: 출결은 비워둠

    # 2) 특이사항
    nk = soup.find('td', string=lambda s: s and s.strip() == '특이사항')
    if nk:
        nv = nk.find_next_sibling('td')
        if nv:
            nv.clear()  # 두 모드 공통으로 '특이사항' 비움

    # 3) 교육내용
    if mode == "stub_init" and education_idx:
        ek = soup.find('td', string='교육내용')
        if ek:
            ev = ek.find_next_sibling('td')
            if ev:
                ev.clear()
                p = soup.new_tag('p', **{'class':'MsoNormal'})
                text = session_descriptions[education_idx-1] if 1 <= education_idx <= len(session_descriptions) else ''
                p.string = text or ''
                ev.append(p)

    with open(path, 'w', encoding='cp949', errors='ignore') as f:
        f.write(str(soup))

def _attendance_list_for(group:str, session_idx:int):
    """roster에서 해당 반·회차 출결 리스트 생성"""
    items = []
    g = roster.get(group, {})
    for nm, sess in g.get("students", {}).items():
        status = sess.get(session_idx, "출")
        items.append((nm, status))
    return items

# ──────────────────────────────────────────────────────────
# 설정·시트 로드
# ──────────────────────────────────────────────────────────
def 초기화():
    global cfg, name_change_map, session_descriptions
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['메인']

    # Q7,Q8 → Q5,Q6 반영(표시 겸 사용값)
    ws.cell(5,17, ws.cell(7,17).value)  # Q5 ← Q7
    ws.cell(6,17, ws.cell(8,17).value)  # Q6 ← Q8
    wb.save(EXCEL_PATH)

    region_code = (ws.cell(1,17).value or '').strip()    # Q1
    target_ym   = str(ws.cell(2,17).value or '').strip() # Q2

    link_cells = {'PH':((17,17),(18,17)), 'BG':((19,17),(20,17)), 'GM':((21,17),(22,17))}
    if region_code not in link_cells:
        raise RuntimeError(f"지원되지 않는 지역코드: {region_code} (Q1)")
    (lr,lc),(sr,sc) = link_cells[region_code]

    cfg.clear()
    cfg['region_code'] = region_code
    cfg['region']      = f"{region_code}_"
    cfg['target_ym']   = target_ym
    cfg['login_url']   = ws.cell(lr,lc).value
    cfg['sched_url']   = ws.cell(sr,sc).value
    cfg['admin_id']    = ws.cell(3,17).value
    cfg['admin_pw']    = ws.cell(4,17).value

    # 다운로드 반수: Q5/Q6
    cfg['mw_download_count'] = int(ws.cell(5,17).value or 0)
    cfg['tt_download_count'] = int(ws.cell(6,17).value or 0)

    # 교육내용
    session_descriptions[:] = [(ws.cell(i,19).value or '').strip() for i in range(1,9)]

    # 반이름변경 맵
    name_change_map.clear()
    if '반이름변경' in wb.sheetnames:
        for o,n in wb['반이름변경'].iter_rows(min_row=1, max_col=2, values_only=True):
            if o and n:
                name_change_map[str(o).strip()] = str(n).strip()

    # 로스터 구축(출결 덮어쓰기에 사용)
    build_full_roster(wb)

def build_full_roster(wb):
    """시트 값이 단일 진실. MW/TT 각각 A,E,M~T 읽어 roster 채움."""
    global roster
    roster = {}
    if '월수반' in wb.sheetnames:
        r = _build_roster_from_sheet(wb['월수반'], 'MW')
        _merge_roster(r)
    if '화목반' in wb.sheetnames:
        r = _build_roster_from_sheet(wb['화목반'], 'TT')
        _merge_roster(r)

def _build_roster_from_sheet(ws, prefix:str, start_row:int=2):
    """
    A열 병합 대응:
      - A가 비어 있으면 직전 행의 분반번호(last_grp_no)를 사용
      - E열 이름이 비어 있으면 스킵
      - M..T(13..20열)에서 'x'만 결, 그 외는 출
    """
    r = {}
    last_grp_no = None
    for rr in range(start_row, ws.max_row + 1):
        grp_no_candidate = _coerce_int(ws.cell(rr, 1).value)  # A
        if grp_no_candidate is not None:
            last_grp_no = grp_no_candidate
        grp_no = last_grp_no

        name_raw = ws.cell(rr, 5).value  # E
        if not grp_no or not name_raw or not str(name_raw).strip():
            continue

        group = f"{prefix}_{int(grp_no):02d}"
        name  = _norm_name(name_raw)

        sess = {}
        for i, cc in enumerate(range(13, 21), start=1):  # M..T = 1..8회차
            v = ws.cell(rr, cc).value
            absent = (isinstance(v, str) and v.strip().lower() == 'x')
            sess[i] = '결' if absent else '출'

        r.setdefault(group, {"students": {}})
        r[group]["students"].setdefault(name, {}).update(sess)
    return r

def _merge_roster(part):
    for g, payload in part.items():
        roster.setdefault(g, {"students": {}})
        for nm, sess in payload["students"].items():
            roster[g]["students"].setdefault(nm, {}).update(sess)

# ──────────────────────────────────────────────────────────
# 로그인·수집·다운로드(Q5/Q6 기반)
# ──────────────────────────────────────────────────────────
def 로그인(page):
    page.goto(cfg['login_url'])
    page.fill('input[name="tbAdminId"]', cfg['admin_id'])
    page.fill('input[name="tbAdminPass"]', cfg['admin_pw'])
    page.press('input[name="tbAdminPass"]', 'Enter')
    page.goto(cfg['sched_url'])
    page.wait_for_selector('select[name="ddlTargetDate"]')
    page.select_option('select[name="ddlTargetDate"]', cfg['target_ym'])
    page.wait_for_selector('select[name="ddlKeyField"]')
    page.select_option('select[name="ddlKeyField"]','a.tutor_id')

def _collect_dates_for_group(page, code:str):
    inp = 'input[name="tbKeyWord"][size="25"]'
    btn = 'a.button_gray_small:has-text("보고서"),a.button_red_small:has-text("보고서")'
    page.fill(inp, code); page.press(inp, 'Enter')
    loc = page.locator(btn)
    try:
        loc.nth(7).wait_for(timeout=30000)
    except PlaywrightTimeoutError:
        return [], []
    bts = loc.all()
    dates = []
    for b in bts[::-1]:
        href = b.get_attribute('href') or ''
        m = re.findall(r"'(.*?)'", href)
        if len(m) > 1:
            dates.append(m[1])
    return dates, bts

def download_all_originals():
    """
    사이트 스케쥴 관리에서 Q5/Q6에 지정된 반수만큼 전부 시도해 받는다.
    - MW_01..MW_{Q5}, TT_01..TT_{Q6}
    - 빨간 버튼/결과 없음은 건너뛰고 스텁으로 보충
    """
    _ensure_dir(REPORT_DIR)

    mw_n = int(cfg.get('mw_download_count', 0))
    tt_n = int(cfg.get('tt_download_count', 0))
    groups = [f"MW_{i:02d}" for i in range(1, mw_n+1)] + [f"TT_{i:02d}" for i in range(1, tt_n+1)]
    if not groups:
        print("[WARN] Q5/Q6 다운로드 반수가 0입니다.")
        return

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=False, args=['--disable-popup-blocking'])
        ctx = browser.new_context(accept_downloads=True)
        page = ctx.new_page()
        로그인(page)

        for orig in groups:
            _download_group_to_folder(page, orig)

        browser.close()

def _download_group_to_folder(page, orig:str):
    """원본명 폴더(지역+orig)에 다운로드 또는 그대로 두기(이미 있으면 스킵)"""
    grp_dir = os.path.join(REPORT_DIR, f"{cfg['region']}{orig}")
    _ensure_dir(grp_dir)

    code = f"{cfg['region']}{orig}"
    dates, buttons = _collect_dates_for_group(page, code)
    session_dates_map[orig] = dates[:]

    for idx, b in enumerate(buttons[::-1], start=1):
        date = dates[idx-1] if idx-1 < len(dates) else 'Unknown'
        dst  = os.path.join(grp_dir, f"{cfg['region']}{orig}_{date}.doc")
        cls = b.get_attribute('class') or ''
        if 'button_red_small' in cls:
            continue
        if not os.path.exists(dst):
            with page.expect_popup() as pp: b.click()
            pop = pp.value; pop.wait_for_selector('a.button_yellow')
            with pop.expect_download() as dl:
                pop.click('a.button_yellow')
            dl.value.save_as(dst); pop.close()

# ──────────────────────────────────────────────────────────
# 일괄 이관(체인 해소, 충돌 → 기존본은 BACKUP으로 이동)
# ──────────────────────────────────────────────────────────
def _resolve_final_mapping(mapping:dict):
    keys = set(mapping.keys())
    finals = {}
    for k in list(keys):
        seen = set()
        cur = k
        while cur in mapping and cur not in seen:
            seen.add(cur)
            cur = mapping[cur]
        finals[k] = cur
    return finals

def migrate_all_by_mapping():
    """
    temp 폴더로 먼저 이동 후 최종 폴더로 병합.
    동일 파일명이 있으면 기존 파일을 BACKUP으로 이동하고 새 파일을 정식 파일로 둠.
    """
    if not name_change_map:
        return

    _ensure_dir(BACKUP_DIR)
    finals = _resolve_final_mapping(name_change_map)

    # 1) 임시로 rename
    temp_map = {}  # orig -> tmp_dir
    for orig in set(list(finals.keys())):
        src_dir = os.path.join(REPORT_DIR, f"{cfg['region']}{orig}")
        if not os.path.isdir(src_dir):
            continue
        tmp_dir = os.path.join(REPORT_DIR, f"__TMP__{cfg['region']}{orig}")
        if os.path.isdir(tmp_dir):
            shutil.rmtree(tmp_dir, ignore_errors=True)
        shutil.move(src_dir, tmp_dir)
        temp_map[orig] = tmp_dir

    # 2) temp → final 병합
    for orig, tmp_dir in temp_map.items():
        final = finals.get(orig, orig)
        dst_dir = os.path.join(REPORT_DIR, f"{cfg['region']}{final}")
        _ensure_dir(dst_dir)

        for fn in os.listdir(tmp_dir):
            if not fn.endswith('.doc'):
                continue
            new_fn = re.sub(rf'^{re.escape(cfg["region"])}{re.escape(orig)}_', f'{cfg["region"]}{final}_', fn)
            src = os.path.join(tmp_dir, fn)
            dst = os.path.join(dst_dir, new_fn)

            if os.path.exists(dst):
                _backup_move(dst)  # 기존본 백업으로 이동
            shutil.move(src, dst)

            # 내부 접미사 old→new
            _suffix_change_in_html(dst, orig.split('_')[1], final.split('_')[1])

        shutil.rmtree(tmp_dir, ignore_errors=True)

    # 남은 빈 원본 폴더 정리
    for orig in list(finals.keys()):
        d = os.path.join(REPORT_DIR, f"{cfg['region']}{orig}")
        if os.path.isdir(d):
            shutil.rmtree(d, ignore_errors=True)

# ──────────────────────────────────────────────────────────
# 템플릿 복사 방식 스텁 생성(로스터 기반, 8회차 보장)
# ──────────────────────────────────────────────────────────
def _find_template_group(prefix: str):
    root = REPORT_DIR
    cand = []
    for fn in os.listdir(root):
        d = os.path.join(root, fn)
        if not os.path.isdir(d):
            continue
        if not fn.startswith(f"{cfg['region']}{prefix}_"):
            continue
        if any(f.endswith(".doc") for f in os.listdir(d)):
            group = fn.replace(cfg['region'], '')
            cand.append(group)
    cand.sort(key=lambda g: int(g.split("_")[1]))
    return cand[0] if cand else None

def _first_doc_of(group: str):
    grp_dir = os.path.join(REPORT_DIR, f"{cfg['region']}{group}")
    docs = sorted([f for f in os.listdir(grp_dir) if f.endswith(".doc")])
    return os.path.join(grp_dir, docs[0]) if docs else None

def _target_dates_for_group(group: str, template_dates: list, total_sessions: int = 8):
    seen, result = set(), []
    for d in session_dates_map.get(group, []):
        if d and d not in seen:
            seen.add(d); result.append(d)
    for d in template_dates:
        if len(result) >= total_sessions:
            break
        if d and d not in seen:
            seen.add(d); result.append(d)
    return result

def create_stubs_from_template():
    """
    시트(roster)에 존재하는 반만 스텁 생성.
    날짜 반영 + 교육내용만 세팅 + 특이사항 비움 + 출결칸 비워둠.
    """
    for prefix in ('MW', 'TT'):
        template_group = _find_template_group(prefix)
        if not template_group:
            print(f"[WARN] {prefix} 템플릿 반을 찾지 못했습니다. 스텁 생성 생략.")
            continue
        base_doc = _first_doc_of(template_group)
        if not base_doc or not os.path.exists(base_doc):
            print(f"[ERROR] 템플릿 파일 없음: {base_doc}")
            continue

        template_dates = session_dates_map.get(template_group, [])
        if not template_dates:
            tdir = os.path.join(REPORT_DIR, f"{cfg['region']}{template_group}")
            template_dates = []
            for fn in sorted(os.listdir(tdir)):
                m = re.search(r'_(\d{4}-\d{2}-\d{2})\.doc$', fn)
                if m: template_dates.append(m.group(1))

        groups = sorted([g for g in roster.keys() if g.startswith(prefix + '_')])
        for group in groups:
            grp_dir = os.path.join(REPORT_DIR, f"{cfg['region']}{group}")
            _ensure_dir(grp_dir)

            dates = _target_dates_for_group(group, template_dates, total_sessions=8)
            for idx, date in enumerate(dates, start=1):
                dst = os.path.join(grp_dir, f"{cfg['region']}{group}_{date}.doc")
                if os.path.exists(dst):
                    continue
                shutil.copy(base_doc, dst)
                try:
                    t_num = template_group.split('_')[1]
                    g_num = group.split('_')[1]
                    _suffix_change_in_html(dst, t_num, g_num)
                except Exception:
                    pass
                _set_datetime_in_html(dst, date)
                _rewrite_attendance_html(dst, [], education_idx=idx, mode="stub_init")
    print("▶ 템플릿 기반 스텁 생성 완료")

# ──────────────────────────────────────────────────────────
# 출결 전면 재작성(모든 파일에서 특이사항도 비움)
# ──────────────────────────────────────────────────────────
def rewrite_all_attendance():
    for group in sorted(roster.keys()):
        grp_dir = os.path.join(REPORT_DIR, f"{cfg['region']}{group}")
        if not os.path.isdir(grp_dir):
            continue
        files = sorted([fn for fn in os.listdir(grp_dir) if fn.endswith('.doc')])
        dates = []
        for fn in files:
            m = re.search(r'_(\d{4}-\d{2}-\d{2})\.doc$', fn)
            if m:
                dates.append(m.group(1))
        if not dates:
            dates = session_dates_map.get(group, [])
        for idx, date in enumerate(dates, start=1):
            path = os.path.join(grp_dir, f"{cfg['region']}{group}_{date}.doc")
            if not os.path.exists(path):
                continue
            _set_datetime_in_html(path, date)
            att = _attendance_list_for(group, idx)
            _rewrite_attendance_html(path, att, mode="attendance_only")  # 특이사항 자동 비움

# ──────────────────────────────────────────────────────────
# 중복(과거 *_dupN) 정리 → BACKUP 이관
# ──────────────────────────────────────────────────────────
def cleanup_old_dups():
    """
    과거 실행에서 생성된 *_dupN.doc 파일들을 모두 BACKUP으로 이동.
    """
    _ensure_dir(BACKUP_DIR)
    for root, _, files in os.walk(REPORT_DIR):
        # BACKUP 폴더는 스킵
        if BACKUP_DIR in os.path.abspath(root):
            continue
        for fn in files:
            if fn.endswith('.doc') and re.search(r'_dup\d+\.doc$', fn):
                src = os.path.join(root, fn)
                _backup_move(src)

# ──────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────
def main():
    _ensure_dir(REPORT_DIR)
    _ensure_dir(BACKUP_DIR)

    초기화()
    download_all_originals()      # 1) Q5/Q6 범위 다운로드
    migrate_all_by_mapping()      # 2) 반이름변경 체인 이관(충돌은 BACKUP으로 이관)
    create_stubs_from_template()  # 3) 누락 회차 템플릿 스텁 보충(최대 8회차)
    rewrite_all_attendance()      # 4) 모든 파일의 출결 갱신 + 특이사항 비우기
    cleanup_old_dups()            # 5) 과거 *_dupN 정리 → BACKUP 폴더로 이동

if __name__ == '__main__':
    main()
